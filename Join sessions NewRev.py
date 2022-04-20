#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Aug  3 12:33:45 2021

@author: lauramilton
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jul 28 15:27:07 2021

@author: lauramilton
"""

#date_column = 'Date'
time_column1 = 'Time bin (1800s each)'
time_column2 = 'Time bin (600s each)'
active_poke_column = 'Active Poke'
inactive_poke_column = 'Inactive Poke'
total_poke_column = 'Total Poke'
pellet_column = 'Pellet Count'
percent_active_column = '% Active Pokes'

session_length = '180' # session length in minutes
time_bin_length1 = '30'
time_bin_length2 = ''

import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/NewRev/'
export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/NewRev/'
# group_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/FED Data Labelled files/Sham/Output/Joined/'
cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/NewRev/Cohort/'

import pandas as pd
import numpy as np
import os
import openpyxl

num_of_bins1 = int(session_length) / int(time_bin_length1)
if time_bin_length2 != '':
    num_of_bins2 = int(session_length) / int(time_bin_length2)

for folder in sorted(os.listdir(import_location)):
    if folder.startswith('FED'):
        

        # New Rev time bin 1 within values
        #date_1w = []
        time_1w = []
        active_poke_1w = []
        inactive_poke_1w = []
        total_poke_1w = []
        pellet_1w = []
        percent_active_1w = []
        
        # New Rev time bin 1 cumulative values
        #date_1c = []
        time_1c = []
        active_poke_1c = []
        inactive_poke_1c = []
        total_poke_1c = []
        pellet_1c = []
        percent_active_1c = []

        # New Rev time bin 2 within values
        #date_2w = []
        time_2w = []
        active_poke_2w = []
        inactive_poke_2w = []
        total_poke_2w = []
        pellet_2w = []
        percent_active_2w = []
        
        # FR time bin 2 cumulative values
        #date_2c = []
        time_2c = []
        active_poke_2c = []
        inactive_poke_2c = []
        total_poke_2c = []
        pellet_2c = []
        percent_active_2c = []
        

        session_num = 1
        
        for filename in sorted(os.listdir(os.path.join(import_location, folder))):
            
            if filename.endswith(".xlsx"):
                if filename.startswith('FED'): # can probs turn it around to be if filename.startswith() and use variable/name from the code that generates those files
                    if not filename.endswith('Overview.xlsx'): # will be able to remove later once those files aren't in the folder anymore (i.e. when file naming procedure is finalized)
                        if not filename.endswith('Joined.xlsx'):    
                            print(filename)
                            import_name = filename
                            import_destination = import_location + folder + '/' + import_name
                            
                            # import_sheet1 = str(time_bin_length1) + 'min Binned Within'
                            import_sheet1 = 'Binned Within'
                            df1w = pd.read_excel(import_destination, sheet_name = import_sheet1)
                            
                            #import_sheet2 = str(time_bin_length1) + 'min Binned Cumulative'
                            import_sheet2 = 'Binned Cumulative'
                            df1c = pd.read_excel(import_destination, sheet_name = import_sheet2)
    
                            #date1w = df1w[date_column].tolist()
                            time1w = df1w[time_column1].tolist()
                            active_poke1w = df1w[active_poke_column].tolist()
                            inactive_poke1w = df1w[inactive_poke_column].tolist()
                            total_poke1w = df1w[total_poke_column].tolist()
                            pellet1w = df1w[pellet_column].tolist()
                            percent_active1w = df1w[percent_active_column].tolist()
                            
                            #date1c = df1c[date_column].tolist()
                            time1c = df1c[time_column1].tolist()
                            active_poke1c = df1c[active_poke_column].tolist()
                            inactive_poke1c = df1c[inactive_poke_column].tolist()
                            total_poke1c = df1c[total_poke_column].tolist()
                            pellet1c = df1c[pellet_column].tolist()
                            percent_active1c = df1c[percent_active_column].tolist()
                            
                            if time_bin_length2 != '':
                                import_sheet3 = str(time_bin_length2) + 'min Binned Within'
                                df2w = pd.read_excel(import_destination, sheet_name = import_sheet3)
                                import_sheet4 = str(time_bin_length2) + 'min Binned Cumulative'
                                df2c = pd.read_excel(import_destination, sheet_name = import_sheet4)

                                # date2w = df2w[date_column].tolist()
                                time2w = df2w[time_column2].tolist()
                                active_poke2w = df2w[active_poke_column].tolist()
                                inactive_poke2w = df2w[inactive_poke_column].tolist()
                                total_poke2w = df2w[total_poke_column].tolist()
                                pellet2w = df2w[pellet_column].tolist()
                                percent_active2w = df2w[percent_active_column].tolist()
                                
                                # date2c = df2c[date_column].tolist()
                                time2c = df2c[time_column2].tolist()
                                active_poke2c = df2c[active_poke_column].tolist()
                                inactive_poke2c = df2c[inactive_poke_column].tolist()
                                total_poke2c = df2c[total_poke_column].tolist()
                                pellet2c = df2c[pellet_column].tolist()
                                percent_active2c = df2c[percent_active_column].tolist()
                            
                            if len(time1w) > num_of_bins1:
                                
                                #del date1w[int(num_of_bins1):]
                                del time1w[int(num_of_bins1):]
                                del active_poke1w[int(num_of_bins1):]
                                del inactive_poke1w[int(num_of_bins1):]
                                del total_poke1w[int(num_of_bins1):]
                                del pellet1w[int(num_of_bins1):]
                                del percent_active1w[int(num_of_bins1):]
                                
                                #del date1c[int(num_of_bins1):]
                                del time1c[int(num_of_bins1):]
                                del active_poke1c[int(num_of_bins1):]
                                del inactive_poke1c[int(num_of_bins1):]
                                del total_poke1c[int(num_of_bins1):]
                                del pellet1c[int(num_of_bins1):]
                                del percent_active1c[int(num_of_bins1):]

                            if time_bin_length2 != '':
                                if len(time2w) > num_of_bins2:
                                    
                                    #del date2w[int(num_of_bins2):]
                                    del time2w[int(num_of_bins2):]
                                    del active_poke2w[int(num_of_bins2):]
                                    del inactive_poke2w[int(num_of_bins2):]
                                    del total_poke2w[int(num_of_bins2):]
                                    del pellet2w[int(num_of_bins2):]
                                    del percent_active2w[int(num_of_bins2):]
                                    
                                    #del date2c[int(num_of_bins2):]
                                    del time2c[int(num_of_bins2):]
                                    del active_poke2c[int(num_of_bins2):]
                                    del inactive_poke2c[int(num_of_bins2):]
                                    del total_poke2c[int(num_of_bins2):]
                                    del pellet2c[int(num_of_bins2):]
                                    del percent_active2c[int(num_of_bins2):]

                                    
                            if len(time1w) < num_of_bins1:
                                
                                bin_num1 = len(time1w) + 1
                                
                                while len(time1w) < num_of_bins1:
                                    #date1w.append(date1w[0])
                                    time1w.append(bin_num1)
                                    active_poke1w.append(np.nan)
                                    inactive_poke1w.append(np.nan)
                                    total_poke1w.append(np.nan)
                                    pellet1w.append(np.nan)
                                    percent_active1w.append(np.nan)
                                    
                                    #date1c.append(date1c[0])
                                    time1c.append(bin_num1)
                                    active_poke1c.append(active_poke1c[-1])
                                    inactive_poke1c.append(inactive_poke1c[-1])
                                    total_poke1c.append(total_poke1c[-1])
                                    pellet1c.append(pellet1c[-1])
                                    percent_active1c.append(percent_active1c[-1])

                                    bin_num1 +=1
                            
                            if time_bin_length2 != '':
                                if len(time2w) < num_of_bins2:
                                
                                    bin_num2 = len(time2w) + 1
                                    
                                    while len(time2w) < num_of_bins2:
                                        #date2w.append(date2w[0])
                                        time2w.append(bin_num2)
                                        active_poke2w.append(np.nan)
                                        inactive_poke2w.append(np.nan)
                                        total_poke2w.append(np.nan)
                                        pellet2w.append(np.nan)
                                        percent_active2w.append(np.nan)
                                        
                                        #date2c.append(date2c[0])
                                        time2c.append(bin_num2)
                                        active_poke2c.append(np.nan)
                                        inactive_poke2c.append(np.nan)
                                        total_poke2c.append(np.nan)
                                        pellet2c.append(np.nan)
                                        percent_active2c.append(np.nan)

                                        bin_num2 +=1
                                        
                            

                            
                            #for a, b, c, d, e, f, g in zip(time1w, active_poke1w, inactive_poke1w, total_poke1w, pellet1w, percent_active1w, date1w):
                            for a, b, c, d, e, f in zip(time1w, active_poke1w, inactive_poke1w, total_poke1w, pellet1w, percent_active1w):
                                    
                                    time_1w.append(str(session_num) + '.' + str(a))
                                    active_poke_1w.append(b)
                                    inactive_poke_1w.append(c)
                                    total_poke_1w.append(d)
                                    pellet_1w.append(e)
                                    percent_active_1w.append(f)
                                    #date_1w.append(g)
                                    
                            #for a, b, c, d, e, f, g in zip(time1c, active_poke1c, inactive_poke1c, total_poke1c, pellet1c, percent_active1c, date1c):
                            for a, b, c, d, e, f in zip(time1c, active_poke1c, inactive_poke1c, total_poke1c, pellet1c, percent_active1c):
                                    
                                    time_1c.append(str(session_num) + '.' + str(a))
                                    active_poke_1c.append(b)
                                    inactive_poke_1c.append(c)
                                    total_poke_1c.append(d)
                                    pellet_1c.append(e)
                                    percent_active_1c.append(f)
                                    #date_1c.append(g)

                            
                            if time_bin_length2 != '':
	                                #for a, b, c, d, e, f, g in zip(time2w, active_poke2w, inactive_poke2w, total_poke2w, pellet2w, percent_active2w, date2w):
                                    for a, b, c, d, e, f in zip(time2w, active_poke2w, inactive_poke2w, total_poke2w, pellet2w, percent_active2w):
                                    
                                        time_2w.append(str(session_num) + '.' + str(a))
                                        active_poke_2w.append(b)
                                        inactive_poke_2w.append(c)
                                        total_poke_2w.append(d)
                                        pellet_2w.append(e)
                                        percent_active_2w.append(f)
                                        #date_2w.append(g)
                                    
                                    #for a, b, c, d, e, f, g in zip(time2c, active_poke2c, inactive_poke2c, total_poke2c, pellet2c, percent_active2c, date2c):
                                    for a, b, c, d, e, f in zip(time2c, active_poke2c, inactive_poke2c, total_poke2c, pellet2c, percent_active2c):
                                        time_2c.append(str(session_num) + '.' + str(a))
                                        active_poke_2c.append(b)
                                        inactive_poke_2c.append(c)
                                        total_poke_2c.append(d)
                                        pellet_2c.append(e)
                                        percent_active_2c.append(f)
                                        #date_2c.append(g)

                            
                            
                            
                            active_poke_1w = ['NA' if x != x else x for x in active_poke_1w]
                            inactive_poke_1w = ['NA' if x != x else x for x in inactive_poke_1w]
                            total_poke_1w = ['NA' if x != x else x for x in total_poke_1w]
                            pellet_1w = ['NA' if x != x else x for x in pellet_1w]
                            percent_active_1w = ['NA' if x != x else x for x in percent_active_1w]
                            
                            active_poke_1c = ['NA' if x != x else x for x in active_poke_1c]
                            inactive_poke_1c = ['NA' if x != x else x for x in inactive_poke_1c]
                            total_poke_1c = ['NA' if x != x else x for x in total_poke_1c]
                            pellet_1c = ['NA' if x != x else x for x in pellet_1c]
                            percent_active_1c = ['NA' if x != x else x for x in percent_active_1c]

                            
                            if time_bin_length2 != '':
                                active_poke_2w = ['NA' if x != x else x for x in active_poke_2w]
                                inactive_poke_2w = ['NA' if x != x else x for x in inactive_poke_2w]
                                total_poke_2w = ['NA' if x != x else x for x in total_poke_2w]
                                pellet_2w = ['NA' if x != x else x for x in pellet_2w]
                                percent_active_2w = ['NA' if x != x else x for x in percent_active_2w]
                                
                                active_poke_2c = ['NA' if x != x else x for x in active_poke_2c]
                                inactive_poke_2c = ['NA' if x != x else x for x in inactive_poke_2c]
                                total_poke_2c = ['NA' if x != x else x for x in total_poke_2c]
                                pellet_2c = ['NA' if x != x else x for x in pellet_2c]
                                percent_active_2c = ['NA' if x != x else x for x in percent_active_2c]
                        
                            session_num += 1
                            
                        ##### Export results
                           
                        if 'New Reversal' in filename:
                            results1w = {'Time Bin': time_1w, 'Active Poke': active_poke_1w, 'Inactive Poke': inactive_poke_1w, 'Total Poke': total_poke_1w, 'Pellet Count': pellet_1w, 'Percent Active': percent_active_1w}
                            export_file1w = pd.DataFrame(results1w, columns = ['Time Bin', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])

                            results1c = {'Time Bin': time_1c, 'Active Poke': active_poke_1c, 'Inactive Poke': inactive_poke_1c, 'Total Poke': total_poke_1c, 'Pellet Count': pellet_1c, 'Percent Active': percent_active_1c}
                            export_file1c = pd.DataFrame(results1c, columns = ['Time Bin', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])

                            if time_bin_length2 != '':
                                results2w = {'Time Bin': time_2w, 'Active Poke': active_poke_2w, 'Inactive Poke': inactive_poke_2w, 'Total Poke': total_poke_2w, 'Pellet Count': pellet_2w, 'Percent Active': percent_active_2w}
                                export_file2w = pd.DataFrame(results2w, columns = ['Time Bin', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
                                
                                results2c = {'Time Bin': time_2c, 'Active Poke': active_poke_2c, 'Inactive Poke': inactive_poke_2c, 'Total Poke': total_poke_2c, 'Pellet Count': pellet_2c, 'Percent Active': percent_active_2c}
                                export_file2c = pd.DataFrame(results2c, columns = ['Time Bin', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
                            
                            from openpyxl import Workbook
                        
                            wb = Workbook()
                                                        
                            ws1 = wb.active
                            ws1.title = str(time_bin_length1) + 'min within joined'
                                                            
                            ws2 = wb.create_sheet()
                            ws2.title = str(time_bin_length1) + 'min cumulative joined'

                            if time_bin_length2 != '':
                                ws3 = wb.create_sheet()
                                ws3.title = str(time_bin_length2) + 'min within joined'
                                
                                ws4 = wb.create_sheet()
                                ws4.title = str(time_bin_length2) + 'min cumulative joined'
                            
                            
                            
                            results_to_export = []
                            
                            results_to_export.append(export_file1w)
                            results_to_export.append(export_file1c)
                        
                            if time_bin_length2 != '':
                                results_to_export.append(export_file2w)
                                results_to_export.append(export_file2c)
                        
                            
                            sheets_to_export = wb.sheetnames
                            
                            NewRev_name = folder + ' NewRev Joined.xlsx'
                            joined_destination = export_location + folder + '/' + NewRev_name
                                
        
                            with pd.ExcelWriter(joined_destination) as writer:
                                
                                for i in range(len(sheets_to_export)):
                                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
        
                            joined_name = NewRev_name
                                
        
                            # group_joined_destination = group_export_location + joined_name
                            
                            # with pd.ExcelWriter(group_joined_destination) as writer:
                                
                            #     for i in range(len(sheets_to_export)):
                            #         results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)    
                                    
                                    
                            cohort_joined_destination = cohort_export_location + joined_name
                            
                            with pd.ExcelWriter(cohort_joined_destination) as writer:
                                
                                for i in range(len(sheets_to_export)):
                                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)    

