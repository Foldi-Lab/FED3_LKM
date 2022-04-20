#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Aug  3 12:51:22 2021

@author: lauramilton
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon May 17 13:27:16 2021

@author: lauramilton
"""

time_column = 'MM:DD:YYYY hh:mm:ss'
event_column = 'Event'
active_poke_column = 'Active_Poke'
session_type_column = 'Session_type'
left_poke_column = 'Left_Poke_Count'
right_poke_column = 'Right_Poke_Count'
pellet_count_column = 'Pellet_Count'


retrieval_time_column = 'Retrieval_Time'
ipi_column = 'InterPelletInterval'
poke_time_column = 'Poke_Time'


seconds_in_bins1 = '600' # Enter how long you want the time bins to be in seconds; if not using bins enter '' which will give duration in seconds
seconds_in_bins2 = '1800'
initiation_poke = True

# Enter the FED import folder, FED export folder, cohort/experiment export folder and the FED number

import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/FED Data Labelled files/'
export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/Output/'
cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/Output/Cohort/'
# FED_num = 'Subset'

dates = ['FR5 D1', 'FR5 D2', 'FR5 D3', 'FR5 D4', 'FR5 D5']


#-----------------------------------------------------------------------------

# Import the revelant data: time, FR ratio, event, active port, left poke, right poke, and pellet count.

import pandas as pd
import numpy as np
import os
import openpyxl

#-----

for folder in sorted(os.listdir(import_location)):
    if folder.startswith('Mouse ID#21'):
        print(folder)
        for filename in sorted(os.listdir(os.path.join(import_location, folder))):
            
            if filename.endswith(".CSV"):
                
                for i in range(0, len(dates)):
                    if dates[i] in filename:
                        print(filename)
                        # Import the csv data

                        import_name = filename
                        import_destination = import_location + folder + '/' + import_name
                        
                
                #-----
                
                        df = pd.read_csv(import_destination)
                        
                        time = df[time_column].tolist()
                        session_type = df[session_type_column].tolist()
                        event = df[event_column].tolist()
                        active = df[active_poke_column].tolist()
                        left_poke = df[left_poke_column].tolist()
                        right_poke = df[right_poke_column].tolist()
                        pellet_count = df[pellet_count_column].tolist()
                        
                        retrieval_time = df[retrieval_time_column].tolist()
                        ipi = df[ipi_column].tolist()
                        poke_time = df[poke_time_column].tolist()

                        # if session_type[0] == 'Ext (rev)' or session_type[0] == 'Extinct':
                        #     export_name = folder + ' ' +  import_name[7:13] + ' Extinction' + '.xlsx'
                        # else:
                        #     export_name = folder + ' ' + import_name[7:13] + ' FR Training' + '.xlsx'
                        # export_destination = export_location + folder + '/' + export_name
                        
                        if session_type[0] == 'Ext (rev)' or session_type[0] == 'Extinct':
                            export_name = folder + ' ' +  import_name[7:13] + ' Extinction' + '.xlsx'
                        else:
                            export_name = folder + ' ' + import_name[15:21] + ' FR Training' + '.xlsx'
                        export_destination = export_location + folder + '/' + export_name


                        # Change the time column from strings to datetime format for calculating durations between timestamps and creating time bins
                        
                        import datetime as dt
                        
                        time_list = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time]
                        
                        # Start time of the session is the first timestamp
                        
                        start_time = time_list[0]
                        
                        # Remove the initiation poke/s and pellet data (if required)
                        
                        if initiation_poke == True:
                        
                            if session_type[0] == 'FR1' or session_type[0] == 'FR1 (reversed)':
                            
                                left_count = left_poke[0]
                                right_count = right_poke[0]
                                
                                del time[:2]
                                del session_type[:2]
                                del event[:2]
                                del left_poke[:2]
                                del right_poke[:2]
                                del pellet_count[:2]
                                del retrieval_time[:2]
                                del poke_time[:2]
                                del ipi[:2]
                                
                            elif session_type[0] == 'FR3' or session_type[0] == 'FR3 (reversed)':
                                
                                left_count = left_poke[3]
                                right_count = right_poke[3]
                                
                                del time[:4]
                                del session_type[:4]
                                del event[:4]
                                del left_poke[:4]
                                del right_poke[:4]
                                del pellet_count[:4]
                                del retrieval_time[:4]
                                del poke_time[:4]
                                del ipi[:4]

                                
                            elif session_type[0] == 'FR5' or session_type[0] == 'FR5 (reversed)' or session_type[0] == 'FR5 reversed':
                                
                                left_count = left_poke[5]
                                right_count = right_poke[5]
                                
                                del time[:6]
                                del session_type[:6]
                                del event[:6]
                                del left_poke[:6]
                                del right_poke[:6]
                                del pellet_count[:6]
                                del retrieval_time[:6]
                                del poke_time[:6]
                                del ipi[:6]
                                
                            elif session_type[0] == 'Ext (rev)' or session_type[0] == 'Extinct':
                                
                                left_count = left_poke[0]
                                right_count = right_poke[0]
                                
                                del time[:1]
                                del session_type[:1]
                                del event[:1]
                                del left_poke[:1]
                                del right_poke[:1]
                                del pellet_count[:1]
                                del retrieval_time[:1]
                                del poke_time[:1]
                                del ipi[:1]
                                        
                            # Subtract the poke and pellet from the subsequent cumulative data
                            
                            left_poke_shifted = []
                            right_poke_shifted = []
                            pellet_count_shifted = []
                            
                            for i in range(0, len(left_poke)):
                                left_poke_shifted.append(left_poke[i] - left_count)
                                right_poke_shifted.append(right_poke[i] - right_count)
                                if session_type[0] == 'Ext (rev)' or session_type[0] == 'Extinct':
                                    pellet_count_shifted.append(pellet_count[i])
                                else:   
                                    pellet_count_shifted.append(pellet_count[i] - 1)
                                        
                            # Shift the pellet_count lists one step backwards so that the count is in the same index as the corresponding poke

                            pellet_count_shifted.pop(0)
                            
                            retrieval_time.pop(0)
                            ipi.pop(0)
                            
                            # Add in a value to the end of each list to maintain list length so as not to lose the last row of other data
                            # for pellet_count this is a duplicate of preceding row, for retrieval_time this is np.nan
    
                            pellet_count_shifted.append(pellet_count_shifted[-1])
                            
                            retrieval_time.append(np.nan)
                            ipi.append(np.nan)

                            
                            # Create new lists that only include data from the lines where event is NOT Pellet
                            
                            time2 = []
                            session_type2 = []
                            event2 = []
                            active2 = []
                            left_poke2 = []
                            right_poke2 = []
                            pellet_count2 = []
                            retrieval_time2 = []
                            poke_time2 = []
                            ipi_2 = []

                            for a, b, c, d, e, f, g, h, i, j in zip(event, left_poke_shifted, right_poke_shifted, pellet_count_shifted, time, session_type, active, retrieval_time, poke_time, ipi):
                                if a != 'Pellet':
                                    event2.append(a)
                                    left_poke2.append(b)
                                    right_poke2.append(c) 
                                    pellet_count2.append(d)
                                    time2.append(e)
                                    session_type2.append(f)
                                    active2.append(g)
                                    retrieval_time2.append(h)
                                    poke_time2.append(i)
                                    ipi_2.append(j)
                        
                        else:
                            # Shift the pellet_count lists one step backwards so that the count is in the same index as the corresponding poke
                    
                            pellet_count.pop(0)
                            
                            retrieval_time.pop(0)
                            ipi.pop(0)

                            
                            # Add in a value to the end of each list to maintain list length so as not to lose the last row of other data
                            # for pellet_count this is a duplicate of preceding row, for retrieval_time this is np.nan
    
                            pellet_count.append(pellet_count[-1])

                            retrieval_time.append(np.nan)
                            ipi.append(np.nan)
                            
                            # Create new lists that only include data from the lines where event is NOT Pellet
                            
                            time2 = []
                            session_type2 = []
                            event2 = []
                            active2 = []
                            left_poke2 = []
                            right_poke2 = []
                            pellet_count2 = []
                            retrieval_time2 = []
                            poke_time2 = []
                            ipi_2 = []
                            
                            for a, b, c, d, e, f, g, h, i, j in zip(event, left_poke, right_poke, pellet_count, time, session_type, active, retrieval_time, poke_time, ipi):
                                if a != 'Pellet':
                                    event2.append(a)
                                    left_poke2.append(b)
                                    right_poke2.append(c) 
                                    pellet_count2.append(d)
                                    time2.append(e)
                                    session_type2.append(f)
                                    active2.append(g)
                                    retrieval_time2.append(h)
                                    poke_time2.append(i)
                                    ipi_2.append(j)
                        
                        

                        
                        # Change the time2 column from strings to datetime format for calculating durations between timestamps and creating time bins
                        
                        time2 = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time2]
                        
                        # Create a new column that is the total number of pokes
                        
                        total_pokes = []
                        
                        for i in range(0, len(active2)):
                            total_pokes.append(left_poke2[i] + right_poke2[i])
                        
                        # Create new column that is % of pokes that are into the active port. Needs to be based on active_poke value (i.e. Left or Right)
                        
                        percent_active = []
                        
                        for i in range(0, len(active2)):
                            if active2[i] == 'Left':
                                if left_poke2[i] != 0:
                                    percent_active.append(left_poke2[i] / total_pokes[i] * 100)
                                else:
                                    percent_active.append(0)
                            else:
                                if right_poke2[i] != 0:
                                    percent_active.append(right_poke2[i] / total_pokes[i] * 100)
                                else:
                                    percent_active.append(0)
                                
                        # Calculate how much time (in seconds) has elapsed from start time for each event (nose poke)
                        
                        seconds_elapsed = []
                        
                        for i in range(0, len(time2)):
                            time_from_start = time2[i] - start_time
                            result = int(time_from_start.total_seconds())
                            seconds_elapsed.append(result)
                        
                        # Chronological Poke Time Data split by side
                    
                        l_poke_time = [] # poke time for all left pokes
                        r_poke_time = [] # poke time for all right pokes
                        
                        l_poke_time_chron = [] # poke time for all left pokes with nan if it was a right poke to maintain array length the same as all pokes
                        r_poke_time_chron = [] # poke time for all right pokes with nan if it was a left poke to maintain array length the same as all pokes
                        
                        for i in range(0, len(poke_time2)): # Collates the poke_time by side
                           if event2[i] == 'Left':
                               l_poke_time.append(poke_time2[i])
                               l_poke_time_chron.append(poke_time2[i])
                               r_poke_time_chron.append(np.nan)
                           else:
                               r_poke_time.append(poke_time2[i])
                               r_poke_time_chron.append(poke_time2[i])
                               l_poke_time_chron.append(np.nan)

                        
                        # Create time bins of desired length (input for desired length is at the top of the code)
                        
                        import math
                        
                        if seconds_in_bins1 != '':
                        
                            end_time = seconds_elapsed[-1]
                            
                            # Create end value that is the next greatest multiple of the required time bin duration greater than the end_time so that those data points are retained in the output
                            
                            excess = math.ceil(end_time / int(seconds_in_bins1)) # Rounds up so that the number of bins is the multiple of bin length greater than the time point
                            
                            end_bin = excess * int(seconds_in_bins1) # Creates the end value for the last time bin
                            
                            interval_range = pd.interval_range(start=0, freq=int(seconds_in_bins1), end=end_bin, closed="left") #'left' means that it includes the first data point which is 0 because it's start time
                            
                            time_bins = df[seconds_elapsed] = pd.cut(seconds_elapsed, interval_range, include_lowest=True, ordered=True)
                            
                            # Create dictionary of time bins as keys with bin number as matching entry to label the bins
                            
                            # Create a list of integers of the same length as there are time bins
                            
                            bin_number = []
                            a = 1
                            for i in range(0, len(interval_range)):
                                bin_number.append(a)
                                a += 1
                            
                            make_dictionary = zip(interval_range, bin_number)
                            
                            bin_dictionary = dict(make_dictionary)
                            
                            bin_num = []
                            for i in range(0, len(time_bins)):
                                bin_num.append(bin_dictionary.get((time_bins[i])))

                            # Get the last entry into each bin for the rest of the columns required
                            
                            bin_num_binned1 = []
                            left_binned1 = []
                            right_binned1 = []
                            pellet_binned1 = []
                            active_binned1 = []
                            session_type_binned1 = []
                            percent_active_binned1 = []
                            total_pokes_binned1 = []
                            
                            # time binned
                            for i in range(1, len(time_bins)):
                                    if bin_num[i - 1] != bin_num[i]:
                                        bin_num_binned1.append(bin_num[i - 1])
                            bin_num_binned1.append(bin_num[i - 0])
                            
                            # left binned
                            for i in range(1, len(left_poke2)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        left_binned1.append(left_poke2[i - 1])
                            left_binned1.append(left_poke2[i - 0])
                                    
                            # right binned
                            for i in range(1, len(right_poke2)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        right_binned1.append(right_poke2[i - 1])
                            right_binned1.append(right_poke2[i - 0])
                                 
                            # pellet binned
                            for i in range(1, len(pellet_count2)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        pellet_binned1.append(pellet_count2[i - 1])
                            pellet_binned1.append(pellet_count2[i - 0])
                            
                            # active binned
                            for i in range(1, len(active2)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        active_binned1.append(active2[i - 1])
                            active_binned1.append(active2[i - 0])
                            
                            # session type binned
                            for i in range(1, len(session_type2)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        session_type_binned1.append(session_type2[i - 1])
                            session_type_binned1.append(session_type2[i - 0])
                            
                            # total pokes binned
                            for i in range(1, len(total_pokes)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        total_pokes_binned1.append(total_pokes[i - 1])
                            total_pokes_binned1.append(total_pokes[i - 0])
                            
                            # percent active binned
                            for i in range(0, len(active_binned1)):
                                if active_binned1[i] == 'Left':
                                    percent_active_binned1.append(left_binned1[i] / total_pokes_binned1[i] * 100)
                                else:
                                    percent_active_binned1.append(right_binned1[i] / total_pokes_binned1[i] * 100)
                            
                            # Account for potential empty first bin
                            
                            bin_counter = 1
                            bin_index = 0
                            
                            for i in range(0, len(bin_num)):
                                if bin_num_binned1[bin_index] != bin_counter:
                                    bin_num_binned1.insert(bin_counter - 1, bin_counter)
                                    left_binned1.insert(bin_counter - 1, 0)
                                    right_binned1.insert(bin_counter - 1, 0)
                                    pellet_binned1.insert(bin_counter - 1, 0)
                                    active_binned1.insert(bin_counter - 1, active2[i])
                                    session_type_binned1.insert(bin_counter - 1, session_type2[0])
                                    total_pokes_binned1.insert(bin_counter - 1, 0)
                                    percent_active_binned1.insert(bin_counter - 1, np.nan)
                            
                            # Enter new row for any bins that have no data in them (for cumulative bins this is a duplication of the preceding row)
                        
                            active_port = active_binned1[0]
                            session_type3 = session_type2[0]
                                  
                            for i in bin_dictionary:
                                if bin_dictionary[i] in bin_num_binned1:
                                    continue
                                else:
                                    bin_num_binned1.insert((bin_dictionary[i] - 1), bin_dictionary[i])
                                    left_binned1.insert((bin_dictionary[i] - 1), left_binned1[(bin_dictionary[i] - 2)])
                                    right_binned1.insert((bin_dictionary[i] - 1), right_binned1[(bin_dictionary[i] - 2)])
                                    pellet_binned1.insert((bin_dictionary[i] - 1), pellet_binned1[(bin_dictionary[i] - 2)])
                                    active_binned1.insert((bin_dictionary[i] - 1), active_port)
                                    percent_active_binned1.insert((bin_dictionary[i] - 1), percent_active_binned1[(bin_dictionary[i] - 2)])
                                    session_type_binned1.insert((bin_dictionary[i] - 1), session_type3)
                                    total_pokes_binned1.insert((bin_dictionary[i] - 1), total_pokes_binned1[(bin_dictionary[i] - 2)])
                                    
                            # Make new counts that are within the bins rather than cumulative
                            
                            left_binned_within1 = []
                            right_binned_within1 = []
                            pellet_binned_within1 = []
                            percent_active_binned_within1 = []
                            total_pokes_binned_within1 = []
                            
                            left_binned_within1.append(left_binned1[0])
                            for i in range(1, len(left_binned1)):
                                left_binned_within1.append(left_binned1[i] - left_binned1[i - 1])
                            
                            right_binned_within1.append(right_binned1[0])
                            for i in range(1, len(right_binned1)):
                                right_binned_within1.append(right_binned1[i] - right_binned1[i - 1])
                            
                            pellet_binned_within1.append(pellet_binned1[0])
                            for i in range(1, len(pellet_binned1)):
                                pellet_binned_within1.append(pellet_binned1[i] - pellet_binned1[i - 1])
                                
                            total_pokes_binned_within1.append(total_pokes_binned1[0])
                            for i in range (1, len(total_pokes_binned1)):
                                total_pokes_binned_within1.append(total_pokes_binned1[i] - total_pokes_binned1[i - 1])
                                
                            for i in range(0, len(active_binned1)):
                                if active_binned1[i] == 'Left':
                                    if left_binned_within1[i] + right_binned_within1[i] != 0:
                                        percent_active_binned_within1.append(left_binned_within1[i] / total_pokes_binned_within1[i] * 100)
                                    else:
                                        percent_active_binned_within1.append('NA')
                                else:
                                    if left_binned_within1[i] + right_binned_within1[i] != 0:
                                        percent_active_binned_within1.append(right_binned_within1[i] / total_pokes_binned_within1[i] * 100)
                                    else:
                                        percent_active_binned_within1.append('NA')
                                        
                                        
                            # Binned poke time, retrieval time and ipi data
                    
                            # Create mean poke and retrieval time columns for bins
                                        
                            binned1_sum_all_poke_time = [] # cumulative sum of all poke times
                            binned1_sum_l_poke_time = [] # cumulative sum of active poke times
                            binned1_sum_r_poke_time = [] # cumulative sum of inactive poke times
                            binned1_mean_all_poke_time = [] # cumulative mean of all poke times
                            binned1_mean_l_poke_time = [] # cumulative mean of active poke times
                            binned1_mean_r_poke_time = [] # cumulative mean of inactive poke times
                            
                            
                            sum_all_poke_time = 0
                            sum_l_poke_time = 0
                            sum_r_poke_time = 0
                            index = 0
                            poke_index = 0

                            for i in range(0, len(poke_time2)):
                                
                                while int(bin_num[i]) >= int(bin_num_binned1[index]) + 1:
                                    binned1_sum_all_poke_time.append(sum_all_poke_time)
                                    binned1_sum_l_poke_time.append(sum_l_poke_time)
                                    binned1_sum_r_poke_time.append(sum_r_poke_time)
                                    if int(total_pokes_binned1[index]) != 0:
                                        binned1_mean_all_poke_time.append((sum_all_poke_time / int(total_pokes_binned1[index])))
                                    else:
                                        binned1_mean_all_poke_time.append('N/A')
                                    if int(left_binned1[index]) != 0:
                                        binned1_mean_l_poke_time.append((sum_l_poke_time / int(left_binned1[index])))
                                    else:
                                        binned1_mean_l_poke_time.append('N/A')
                                    if int(right_binned1[index]) != 0:
                                        binned1_mean_r_poke_time.append((sum_r_poke_time / int(right_binned1[index])))
                                    else:
                                        binned1_mean_r_poke_time.append('N/A')

                                    index += 1
                                    
                                    if int(bin_num[i]) <= int(bin_num_binned1[index]):
                                        
                                        break
                                
                                sum_all_poke_time += poke_time2[poke_index]
                                if event2[i] == 'Left':
                                    sum_l_poke_time += poke_time2[poke_index]
                                elif event2[i] == 'Right':
                                    sum_r_poke_time += poke_time2[poke_index]
                                poke_index += 1
                                    
                            binned1_sum_all_poke_time.append(sum_all_poke_time)
                            binned1_sum_l_poke_time.append(sum_l_poke_time)
                            binned1_sum_r_poke_time.append(sum_r_poke_time)
                            binned1_mean_all_poke_time.append((sum_all_poke_time / int(total_pokes_binned1[index])))
                            if int(left_binned1[index]) != 0:
                                binned1_mean_l_poke_time.append((sum_l_poke_time / int(left_binned1[index])))
                            else:
                                binned1_mean_l_poke_time.append('N/A')
                            if int(right_binned1[index]) != 0:
                                binned1_mean_r_poke_time.append((sum_r_poke_time / int(right_binned1[index])))
                            else:
                                binned1_mean_r_poke_time.append('N/A')
                                
                                
                                
                            binned1_within_sum_all_poke_time = [] # within binned sum of all poke times
                            binned1_within_sum_l_poke_time = [] # within binned sum of active poke times
                            binned1_within_sum_r_poke_time = [] # within binned sum of inactive poke times
                            
                            binned1_within_sum_all_poke_time.append(binned1_sum_all_poke_time[0])
                            binned1_within_sum_l_poke_time.append(binned1_sum_l_poke_time[0])
                            binned1_within_sum_r_poke_time.append(binned1_sum_r_poke_time[0])  
                            
                            for i in range(1, len(binned1_sum_all_poke_time)):
                                binned1_within_sum_all_poke_time.append((binned1_sum_all_poke_time[i] - binned1_sum_all_poke_time[i - 1]))
                                binned1_within_sum_l_poke_time.append((binned1_sum_l_poke_time[i] - binned1_sum_l_poke_time[i - 1]))
                                binned1_within_sum_r_poke_time.append((binned1_sum_r_poke_time[i] - binned1_sum_r_poke_time[i - 1]))  
                            
                            binned1_within_mean_all_poke_time = [] # within binned mean of all poke times
                            binned1_within_mean_l_poke_time = [] # within binned mean of active poke times
                            binned1_within_mean_r_poke_time = [] # within binned mean of inactive poke times
                            
                            for i in range(0, len(binned1_within_sum_all_poke_time)):
                                if total_pokes_binned_within1[i] !=0:
                                    binned1_within_mean_all_poke_time.append((binned1_within_sum_all_poke_time[i] / total_pokes_binned_within1[i]))
                                else: 
                                    binned1_within_mean_all_poke_time.append('N/A')
                                if left_binned_within1[i] != 0:
                                    binned1_within_mean_l_poke_time.append((binned1_within_sum_l_poke_time[i] / left_binned_within1[i]))
                                else:
                                    binned1_within_mean_l_poke_time.append('N/A')
                                if right_binned_within1[i] != 0:
                                    binned1_within_mean_r_poke_time.append((binned1_within_sum_r_poke_time[i] / right_binned_within1[i]))
                                else:
                                    binned1_within_mean_r_poke_time.append('N/A')
                            
                            # # retrieval time and ipi bins will just be the lists with the nans removed as there is one value for each bin (PR step)
                            
                            # retrieval_time_binned = [x for x in retrieval_time2 if not math.isnan(x)] # gets rid of all the nan values
                            # ipi_binned = [x for x in ipi_2 if not math.isnan(x)] # gets rid of all the nan values
                            
                            # # if final step is incomplete need to add N/A to end or retrieval_time_binned and ipi_binned
                            
                            # if event[-1] != 'Pellet':
                            #     retrieval_time_binned.append('N/A')
                            #     ipi_binned.append('N/A')            
                            
                            
                            
                            
                            
                            
                        
                        if seconds_in_bins2 != '':
                        
                            end_time = seconds_elapsed[-1]
                            
                            # Create end value that is the next greatest multiple of the required time bin duration greater than the end_time so that those data points are retained in the output
                            
                            excess = math.ceil(end_time / int(seconds_in_bins2)) # Rounds up so that the number of bins is the multiple of bin length greater than the time point
                            
                            end_bin = excess * int(seconds_in_bins2) # Creates the end value for the last time bin
                            
                            interval_range = pd.interval_range(start=0, freq=int(seconds_in_bins2), end=end_bin, closed="left") #'left' means that it includes the first data point which is 0 because it's start time
                            
                            time_bins = df[seconds_elapsed] = pd.cut(seconds_elapsed, interval_range, include_lowest=True, ordered=True)
                            
                            # Create dictionary of time bins as keys with bin number as matching entry to label the bins
                            
                            # Create a list of integers of the same length as there are time bins
                            
                            bin_number = []
                            a = 1
                            for i in range(0, len(interval_range)):
                                bin_number.append(a)
                                a += 1
                            
                            make_dictionary = zip(interval_range, bin_number)
                            
                            bin_dictionary = dict(make_dictionary)
                            
                            bin_num = []
                            for i in range(0, len(time_bins)):
                                bin_num.append(bin_dictionary.get((time_bins[i])))
                            
                            # Get the last entry into each bin for the rest of the columns required
                            
                            bin_num_binned2 = []
                            left_binned2 = []
                            right_binned2 = []
                            pellet_binned2 = []
                            active_binned2 = []
                            session_type_binned2 = []
                            percent_active_binned2 = []
                            total_pokes_binned2 = []
                            
                            # time binned
                            for i in range(1, len(time_bins)):
                                    if bin_num[i - 1] != bin_num[i]:
                                        bin_num_binned2.append(bin_num[i - 1])
                            bin_num_binned2.append(bin_num[i - 0])
                            
                            # left binned
                            for i in range(1, len(left_poke2)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        left_binned2.append(left_poke2[i - 1])
                            left_binned2.append(left_poke2[i - 0])
                                    
                            # right binned
                            for i in range(1, len(right_poke2)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        right_binned2.append(right_poke2[i - 1])
                            right_binned2.append(right_poke2[i - 0])
                                 
                            # pellet binned
                            for i in range(1, len(pellet_count2)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        pellet_binned2.append(pellet_count2[i - 1])
                            pellet_binned2.append(pellet_count2[i - 0])
                            
                            # active binned
                            for i in range(1, len(active2)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        active_binned2.append(active2[i - 1])
                            active_binned2.append(active2[i - 0])
                            
                            # session type binned
                            for i in range(1, len(session_type2)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        session_type_binned2.append(session_type2[i - 1])
                            session_type_binned2.append(session_type2[i - 0])
                            
                            # total pokes binned
                            for i in range(1, len(total_pokes)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        total_pokes_binned2.append(total_pokes[i - 1])
                            total_pokes_binned2.append(total_pokes[i - 0])
                            
                            # percent active binned
                            for i in range(0, len(active_binned2)):
                                if active_binned2[i] == 'Left':
                                    percent_active_binned2.append(left_binned2[i] / total_pokes_binned2[i] * 100)
                                else:
                                    percent_active_binned2.append(right_binned2[i] / total_pokes_binned2[i] * 100)
                            
                            # Account for potential empty first bin
                            
                            bin_counter = 1
                            bin_index = 0
                            
                            for i in range(0, len(bin_num)):
                                if bin_num_binned2[bin_index] != bin_counter:
                                    bin_num_binned2.insert(bin_counter - 1, bin_counter)
                                    left_binned2.insert(bin_counter - 1, 0)
                                    right_binned2.insert(bin_counter - 1, 0)
                                    pellet_binned2.insert(bin_counter - 1, 0)
                                    active_binned2.insert(bin_counter - 1, active2[i])
                                    session_type_binned2.insert(bin_counter - 1, session_type2[0])
                                    total_pokes_binned2.insert(bin_counter - 1, 0)
                                    percent_active_binned2.insert(bin_counter - 1, np.nan)
                            
                            # Enter new row for any bins that have no data in them (for cumulative bins this is a duplication of the preceding row)
                        
                            active_port = active_binned2[0]
                            session_type3 = session_type2[0]
                                  
                            for i in bin_dictionary:
                                if bin_dictionary[i] in bin_num_binned2:
                                    continue
                                else:
                                    bin_num_binned2.insert((bin_dictionary[i] - 1), bin_dictionary[i])
                                    left_binned2.insert((bin_dictionary[i] - 1), left_binned2[(bin_dictionary[i] - 2)])
                                    right_binned2.insert((bin_dictionary[i] - 1), right_binned2[(bin_dictionary[i] - 2)])
                                    pellet_binned2.insert((bin_dictionary[i] - 1), pellet_binned2[(bin_dictionary[i] - 2)])
                                    active_binned2.insert((bin_dictionary[i] - 1), active_port)
                                    percent_active_binned2.insert((bin_dictionary[i] - 1), percent_active_binned2[(bin_dictionary[i] - 2)])
                                    session_type_binned2.insert((bin_dictionary[i] - 1), session_type3)
                                    total_pokes_binned2.insert((bin_dictionary[i] - 1), total_pokes_binned2[(bin_dictionary[i] - 2)])
                                    
                            # Make new counts that are within the bins rather than cumulative
                            
                            left_binned_within2 = []
                            right_binned_within2 = []
                            pellet_binned_within2 = []
                            percent_active_binned_within2 = []
                            total_pokes_binned_within2 = []
                            
                            left_binned_within2.append(left_binned2[0])
                            for i in range(1, len(left_binned2)):
                                left_binned_within2.append(left_binned2[i] - left_binned2[i - 1])
                            
                            right_binned_within2.append(right_binned2[0])
                            for i in range(1, len(right_binned2)):
                                right_binned_within2.append(right_binned2[i] - right_binned2[i - 1])
                            
                            pellet_binned_within2.append(pellet_binned2[0])
                            for i in range(1, len(pellet_binned2)):
                                pellet_binned_within2.append(pellet_binned2[i] - pellet_binned2[i - 1])
                                
                            total_pokes_binned_within2.append(total_pokes_binned2[0])
                            for i in range (1, len(total_pokes_binned2)):
                                total_pokes_binned_within2.append(total_pokes_binned2[i] - total_pokes_binned2[i - 1])
                                
                            for i in range(0, len(active_binned2)):
                                if active_binned2[i] == 'Left':
                                    if left_binned_within2[i] + right_binned_within2[i] != 0:
                                        percent_active_binned_within2.append(left_binned_within2[i] / total_pokes_binned_within2[i] * 100)
                                    else:
                                        percent_active_binned_within2.append('NA')
                                else:
                                    if left_binned_within2[i] + right_binned_within2[i] != 0:
                                        percent_active_binned_within2.append(right_binned_within2[i] / total_pokes_binned_within2[i] * 100)
                                    else:
                                        percent_active_binned_within2.append('NA')
                        
                            # Binned poke time, retrieval time and ipi data
                    
                            # Create mean poke and retrieval time columns for bins
                                        
                            binned2_sum_all_poke_time = [] # cumulative sum of all poke times
                            binned2_sum_l_poke_time = [] # cumulative sum of active poke times
                            binned2_sum_r_poke_time = [] # cumulative sum of inactive poke times
                            binned2_mean_all_poke_time = [] # cumulative mean of all poke times
                            binned2_mean_l_poke_time = [] # cumulative mean of active poke times
                            binned2_mean_r_poke_time = [] # cumulative mean of inactive poke times
                            
                            
                            sum_all_poke_time = 0
                            sum_l_poke_time = 0
                            sum_r_poke_time = 0
                            index = 0
                            poke_index = 0
                            
                            for i in range(0, len(poke_time2)):
                                
                                while int(bin_num[i]) >= int(bin_num_binned2[index]) + 1:
                                    binned2_sum_all_poke_time.append(sum_all_poke_time)
                                    binned2_sum_l_poke_time.append(sum_l_poke_time)
                                    binned2_sum_r_poke_time.append(sum_r_poke_time)
                                    if int(total_pokes_binned2[index]) != 0:
                                        binned2_mean_all_poke_time.append((sum_all_poke_time / int(total_pokes_binned2[index])))
                                    else:
                                        binned2_mean_all_poke_time.append('N/A')
                                    if int(left_binned2[index]) != 0:
                                        binned2_mean_l_poke_time.append((sum_l_poke_time / int(left_binned2[index])))
                                    else:
                                        binned2_mean_l_poke_time.append('N/A')
                                    if int(right_binned2[index]) != 0:
                                        binned2_mean_r_poke_time.append((sum_r_poke_time / int(right_binned2[index])))
                                    else:
                                        binned2_mean_r_poke_time.append('N/A')

                                    index += 1
                                    
                                    if int(bin_num[i]) <= int(bin_num_binned2[index]):
                                        
                                        break
                                
                                sum_all_poke_time += poke_time2[poke_index]
                                if event2[i] == 'Left':
                                    sum_l_poke_time += poke_time2[poke_index]
                                elif event2[i] == 'Right':
                                    sum_r_poke_time += poke_time2[poke_index]
                                poke_index += 1
                                    
                            binned2_sum_all_poke_time.append(sum_all_poke_time)
                            binned2_sum_l_poke_time.append(sum_l_poke_time)
                            binned2_sum_r_poke_time.append(sum_r_poke_time)
                            binned2_mean_all_poke_time.append((sum_all_poke_time / int(total_pokes_binned2[index])))
                            if int(left_binned2[index]) != 0:
                                binned2_mean_l_poke_time.append((sum_l_poke_time / int(left_binned2[index])))
                            else:
                                binned2_mean_l_poke_time.append('N/A')
                            if int(right_binned2[index]) != 0:
                                binned2_mean_r_poke_time.append((sum_r_poke_time / int(right_binned2[index])))
                            else:
                                binned2_mean_r_poke_time.append('N/A')
                                
                                
                                
                                
                                
                            binned2_within_sum_all_poke_time = [] # within binned sum of all poke times
                            binned2_within_sum_l_poke_time = [] # within binned sum of active poke times
                            binned2_within_sum_r_poke_time = [] # within binned sum of inactive poke times
                            
                            binned2_within_sum_all_poke_time.append(binned2_sum_all_poke_time[0])
                            binned2_within_sum_l_poke_time.append(binned2_sum_l_poke_time[0])
                            binned2_within_sum_r_poke_time.append(binned2_sum_r_poke_time[0])  
                            
                            for i in range(1, len(binned2_sum_all_poke_time)):
                                binned2_within_sum_all_poke_time.append((binned2_sum_all_poke_time[i] - binned2_sum_all_poke_time[i - 1]))
                                binned2_within_sum_l_poke_time.append((binned2_sum_l_poke_time[i] - binned2_sum_l_poke_time[i - 1]))
                                binned2_within_sum_r_poke_time.append((binned2_sum_r_poke_time[i] - binned2_sum_r_poke_time[i - 1]))  
                            
                            binned2_within_mean_all_poke_time = [] # within binned mean of all poke times
                            binned2_within_mean_l_poke_time = [] # within binned mean of active poke times
                            binned2_within_mean_r_poke_time = [] # within binned mean of inactive poke times
                            
                            for i in range(0, len(binned2_within_sum_all_poke_time)):
                                if total_pokes_binned_within2[i] !=0:
                                    binned2_within_mean_all_poke_time.append((binned2_within_sum_all_poke_time[i] / total_pokes_binned_within2[i]))
                                else: 
                                    binned2_within_mean_all_poke_time.append('N/A')
                                if left_binned_within2[i] != 0:
                                    binned2_within_mean_l_poke_time.append((binned2_within_sum_l_poke_time[i] / left_binned_within2[i]))
                                else:
                                    binned2_within_mean_l_poke_time.append('N/A')
                                if right_binned_within2[i] != 0:
                                    binned2_within_mean_r_poke_time.append((binned2_within_sum_r_poke_time[i] / right_binned_within2[i]))
                                else:
                                    binned2_within_mean_r_poke_time.append('N/A')
                            
                            # # retrieval time and ipi bins will just be the lists with the nans removed as there is one value for each bin (PR step)
                            
                            # retrieval_time_binned = [x for x in retrieval_time2 if not math.isnan(x)] # gets rid of all the nan values
                            # ipi_binned = [x for x in ipi_2 if not math.isnan(x)] # gets rid of all the nan values
                            
                            # # if final step is incomplete need to add N/A to end or retrieval_time_binned and ipi_binned
                            
                            # if event[-1] != 'Pellet':
                            #     retrieval_time_binned.append('N/A')
                            #     ipi_binned.append('N/A')                             
                            
                        #####----- Create session summary data-----#####
                        
                        task = session_type2[0]
                        
                        session_hours = str(math.floor(seconds_elapsed[-1] / 3600))
                        
                        session_mins = math.floor((seconds_elapsed[-1] % 3600) / 60)
                        
                        session_secs = seconds_elapsed[-1] % 60
                        
                        if session_mins < 10:
                            mins = '0' + str(session_mins)
                        else:
                            mins = str(session_mins)
                            
                        if session_secs < 10:
                            secs = '0' + str(session_secs)
                        else:
                            secs = str(session_secs)
                            
                        session_duration = str(session_hours) + ':' + mins + ':' + secs
                        
                        # Assign left and right pokes as active or inactive based on the active port
                        
                        if active2[0] == 'Left':
                            active_pokes = left_poke2[-1]
                            inactive_pokes = right_poke2[-1]
                        elif active2[0] == 'Right':
                            active_pokes = right_poke2[-1]
                            inactive_pokes = left_poke2[-1]
                        
                        active_port = active2[0]
                        
                        # date = import_name[7:13]
                        date = import_name[15:21]

                        aus_date = date[2:4] + '/' + date[0:2] + '/20' + date[4:]
        
                        variable = ['Filename', 'Date', 'Task', 'Duration', 'Active port', 'Total Pokes', 'Active Pokes', 'Inactive Pokes', '% Active Pokes', 'Pellets']
                        
                        value = [import_name.strip('.CSV'), aus_date, task, session_duration, active2[0], total_pokes[-1], active_pokes, inactive_pokes, percent_active[-1], pellet_count2[-1]]
                        
                        #####-----
                        
                        # Create aus_date column for bins
                        
                        if seconds_in_bins1 != '':
                            
                            aus_date_binned1 = []
                            
                            for i in range(0, len(bin_num_binned1)):
                                
                                aus_date_binned1.append(aus_date)
                                
                        if seconds_in_bins2 != '':
                        
                            aus_date_binned2 = []
                            
                            for i in range(0, len(bin_num_binned2)):
                                
                                aus_date_binned2.append(aus_date)

                        #####----- Export the data -----#####
                        
                        export_results = 'Y'
                        
                        if export_results == 'Y':
                            
                            # Always export Summary and Chronological Data
                            
                            results_summary = {'Variable': variable, 'Value': value}
                            export_file_summary = pd.DataFrame(results_summary, columns = ['Variable', 'Value'])
                            
                            # Assign left and right pokes as active or inactive based on the active port
                            
                            if active2[0] == 'Left':
                                active_poke = left_poke2
                                inactive_poke = right_poke2
                                active_poke_time_chron = l_poke_time_chron
                                inactive_poke_time_chron = r_poke_time_chron
                                
                                if seconds_in_bins1 != '':
                                    active_poke_binned1 = left_binned1
                                    active_poke_binned_within1 = left_binned_within1
                                    inactive_poke_binned1 = right_binned1
                                    inactive_poke_binned_within1 = right_binned_within1
                                    binned1_active_poke_time = binned1_mean_l_poke_time
                                    binned1_inactive_poke_time = binned1_mean_r_poke_time
                                    binned1_within_active_poke_time = binned1_within_mean_l_poke_time
                                    binned1_within_inactive_poke_time = binned1_within_mean_r_poke_time
                                    
                                if seconds_in_bins2 != '':
                                    active_poke_binned2 = left_binned2
                                    active_poke_binned_within2 = left_binned_within2
                                    inactive_poke_binned2 = right_binned2
                                    inactive_poke_binned_within2 = right_binned_within2
                                    binned2_active_poke_time = binned2_mean_l_poke_time
                                    binned2_inactive_poke_time = binned2_mean_r_poke_time
                                    binned2_within_active_poke_time = binned2_within_mean_l_poke_time
                                    binned2_within_inactive_poke_time = binned2_within_mean_r_poke_time
                                    
                            elif active2[0] == 'Right':
                                active_poke = right_poke2
                                inactive_poke = left_poke2
                                active_poke_time_chron = r_poke_time_chron
                                inactive_poke_time_chron = l_poke_time_chron
                                
                                if seconds_in_bins1 != '':
                                    active_poke_binned1 = right_binned1
                                    active_poke_binned_within1 = right_binned_within1
                                    inactive_poke_binned1 = left_binned1
                                    inactive_poke_binned_within1 = left_binned_within1
                                    binned1_active_poke_time = binned1_mean_r_poke_time
                                    binned1_inactive_poke_time = binned1_mean_l_poke_time
                                    binned1_within_active_poke_time = binned1_within_mean_r_poke_time
                                    binned1_within_inactive_poke_time = binned1_within_mean_l_poke_time
                                    
                                if seconds_in_bins2 != '':
                                    active_poke_binned2 = right_binned2
                                    active_poke_binned_within2 = right_binned_within2
                                    inactive_poke_binned2 = left_binned2
                                    inactive_poke_binned_within2 = left_binned_within2
                                    binned2_active_poke_time = binned2_mean_r_poke_time
                                    binned2_inactive_poke_time = binned2_mean_l_poke_time
                                    binned2_within_active_poke_time = binned2_within_mean_r_poke_time
                                    binned2_within_inactive_poke_time = binned2_within_mean_l_poke_time
                            
                            
                            results_chronological = {'Time (seconds)': seconds_elapsed, 'Session Type': session_type2, 'Active port': active2, 'Active Poke': active_poke, 'Inactive Poke': inactive_poke, 
                                                      'Total Poke': total_pokes, 'Pellet Count': pellet_count2, '% Active Pokes': percent_active,
                                                      'Poke Time': poke_time2, 'Active Poke Time': active_poke_time_chron, 'Inactive Poke Time': inactive_poke_time_chron, 'Pellet Retrieval Time': retrieval_time2, 'Inter-Pellet Interval': ipi_2}
                            export_file_chronological = pd.DataFrame(results_chronological, columns = ['Time (seconds)', 'Session Type', 'Active port', 'Active Poke', 'Inactive Poke',
                                                                                                        'Total Poke', 'Pellet Count', '% Active Pokes',
                                                                                                        'Poke Time', 'Active Poke Time', 'Inactive Poke Time', 'Pellet Retrieval Time', 'Inter-Pellet Interval'])
                            # print(len(aus_date_binned1), len(bin_num_binned1), len(session_type_binned1), len(active_binned1), len(active_poke_binned1), len(inactive_poke_binned1), len( total_pokes_binned1), len(pellet_binned1), len(percent_active_binned1), len(binned1_mean_all_poke_time), len(binned1_active_poke_time), len(binned1_inactive_poke_time))
                                
                            # If using time bins also export binned data (cumulative and within)
                            
                            if seconds_in_bins1 != '':
                            
                                results_binned_c1 = {'Date': aus_date_binned1, 'Time bin (' + seconds_in_bins1 + 's each)': bin_num_binned1, 'Session Type': session_type_binned1, 'Active Port': active_binned1, 
                                                    'Active Poke': active_poke_binned1, 'Inactive Poke': inactive_poke_binned1, 'Total Poke': total_pokes_binned1, 'Pellet Count': pellet_binned1, '% Active Pokes': percent_active_binned1,
                                                    'Mean Poke Time': binned1_mean_all_poke_time, 'Mean Active Poke Time': binned1_active_poke_time, 'Mean Inactive Poke Time': binned1_inactive_poke_time}
                                export_file_binned_c1 = pd.DataFrame(results_binned_c1, columns = ['Date', 'Time bin (' + seconds_in_bins1 + 's each)', 'Session Type', 'Active Port', 
                                                                                                  'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', '% Active Pokes',
                                                                                                  'Mean Poke Time', 'Mean Active Poke Time', 'Mean Inactive Poke Time'])
                                
                                results_binned_w1 = {'Date': aus_date_binned1, 'Time bin (' + seconds_in_bins1 + 's each)': bin_num_binned1, 'Session Type': session_type_binned1, 'Active Port': active_binned1, 
                                                    'Active Poke': active_poke_binned_within1, 'Inactive Poke': inactive_poke_binned_within1, 'Total Poke': total_pokes_binned_within1, 'Pellet Count': pellet_binned_within1, '% Active Pokes': percent_active_binned_within1,
                                                    'Mean Poke Time': binned1_within_mean_all_poke_time, 'Mean Active Poke Time': binned1_within_active_poke_time, 'Mean Inactive Poke Time': binned1_within_inactive_poke_time}
                                export_file_binned_w1 = pd.DataFrame(results_binned_w1, columns = ['Date', 'Time bin (' + seconds_in_bins1 + 's each)', 'Session Type', 'Active Port', 
                                                                                                  'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', '% Active Pokes',
                                                                                                  'Mean Poke Time', 'Mean Active Poke Time', 'Mean Inactive Poke Time'])
                            
                            if seconds_in_bins2 != '':
                            
                                results_binned_c2 = {'Date': aus_date_binned2, 'Time bin (' + seconds_in_bins2 + 's each)': bin_num_binned2, 'Session Type': session_type_binned2, 'Active Port': active_binned2, 
                                                    'Active Poke': active_poke_binned2, 'Inactive Poke': inactive_poke_binned2, 'Total Poke': total_pokes_binned2, 'Pellet Count': pellet_binned2, '% Active Pokes': percent_active_binned2,
                                                    'Mean Poke Time': binned2_mean_all_poke_time, 'Mean Active Poke Time': binned2_active_poke_time, 'Mean Inactive Poke Time': binned2_inactive_poke_time}
                                export_file_binned_c2 = pd.DataFrame(results_binned_c2, columns = ['Date', 'Time bin (' + seconds_in_bins2 + 's each)', 'Session Type', 'Active Port', 
                                                                                                  'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', '% Active Pokes',
                                                                                                  'Mean Poke Time', 'Mean Active Poke Time', 'Mean Inactive Poke Time'])
                                
                                results_binned_w2 = {'Date': aus_date_binned2, 'Time bin (' + seconds_in_bins2 + 's each)': bin_num_binned2, 'Session Type': session_type_binned2, 'Active Port': active_binned2, 
                                                    'Active Poke': active_poke_binned_within2, 'Inactive Poke': inactive_poke_binned_within2, 'Total Poke': total_pokes_binned_within2, 'Pellet Count': pellet_binned_within2, '% Active Pokes': percent_active_binned_within2,
                                                    'Mean Poke Time': binned2_within_mean_all_poke_time, 'Mean Active Poke Time': binned2_within_active_poke_time, 'Mean Inactive Poke Time': binned2_within_inactive_poke_time}
                                export_file_binned_w2 = pd.DataFrame(results_binned_w2, columns = ['Date', 'Time bin (' + seconds_in_bins2 + 's each)', 'Session Type', 'Active Port', 
                                                                                                  'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', '% Active Pokes',
                                                                                                  'Mean Poke Time', 'Mean Active Poke Time', 'Mean Inactive Poke Time'])
        
                            
                            # Export to excel
                            
                            from openpyxl import Workbook
                        
                            wb = Workbook()
                            
                            ws1 = wb.active
                            ws1.title = 'Summary'
                            
                            ws2 = wb.create_sheet()
                            ws2.title = 'Chronological'
                            
                            results_to_export = [export_file_summary, export_file_chronological]
                            
                            if seconds_in_bins1 != '':
                                
                                bin_min_length1 = int(seconds_in_bins1) / 60
                                
                                ws3 = wb.create_sheet()
                                ws3.title = str(int(bin_min_length1)) + 'min Binned Cumulative'
                                
                                ws4 = wb.create_sheet()
                                ws4.title = str(int(bin_min_length1)) + 'min Binned Within'
                                
                                results_to_export.append(export_file_binned_c1)
                                results_to_export.append(export_file_binned_w1)
                                
                            if seconds_in_bins2 != '':
                                
                                bin_min_length2 = int(seconds_in_bins2) / 60
                                
                                ws5 = wb.create_sheet()
                                ws5.title = str(int(bin_min_length2)) + 'min Binned Cumulative'
                                
                                ws6 = wb.create_sheet()
                                ws6.title = str(int(bin_min_length2)) + 'min Binned Within'
                                
                                results_to_export.append(export_file_binned_c2)
                                results_to_export.append(export_file_binned_w2)
                            
                            sheets_to_export = wb.sheetnames
                            
                            with pd.ExcelWriter(export_destination) as writer:
                                
                                for i in range(len(sheets_to_export)):
                                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
                            
                            print(filename + ' exported')
                            
        print(folder + ' COMPLETE')
                    
        ##########---------- Overview Sheets----------##########
        
        from openpyxl import Workbook
                
        wb = Workbook()
        
        ws = wb.active
        ws.title = 'Summary Overview'
        
        value_column = 'Value'
        variable_column = 'Variable'
        
        value_summary = []
        session_num = []
        counter = 1
        
        # export_location_folder = export_location + folder + '/'
        
        for filename in sorted(os.listdir(os.path.join(export_location, folder))):
            
            if filename.endswith("Training.xlsx"):
                
                if counter == 1:
                    
                    export_name = filename
                    export_destination = export_location + folder + '/' + export_name
                    
                    df_overview = pd.read_excel(export_destination, sheet_name = 'Summary')
                    
                    df_overview.drop(columns='Value', inplace=True)
                    
                export_name = filename
                export_destination = export_location + folder + '/' + export_name
                
                name = 'Session ' + str(counter)
                session_num.append(name)
                
                df = pd.read_excel(export_destination, sheet_name = 'Summary')
                
                values = df[value_column].tolist()
                
                df_overview.insert(counter, name, values)
                
                counter += 1
                    
            overview_name = folder + ' FR Training Overview.xlsx'
            overview_destination = export_location + folder + '/' + overview_name
            
            with pd.ExcelWriter(overview_destination) as writer:
                            
                df_overview.to_excel(writer, sheet_name='Overview', engine='openpyxl', index=False, header=True)
            
            cohort_overview_destination = cohort_export_location + overview_name
            
            with pd.ExcelWriter(cohort_overview_destination) as writer:
                            
                df_overview.to_excel(writer, sheet_name='Overview', engine='openpyxl', index=False, header=True)
            
        print(filename, 'Overview complete')

        