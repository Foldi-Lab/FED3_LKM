#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Feb 23 17:51:03 2022

@author: lauramilton
"""



import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/NewRev/Cohort/PSI/'
export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/NewRev/Cohort/PSI/'
cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/NewRev/Cohort/'
# FED_num = 'FED3'

sheet_to_join1 = 'Summary Completed Joined'
sheet_to_join2 = 'Summary ALL Joined'

sheet_to_join = sheet_to_join2

group_name = 'PSI Summary ALL Joined.xlsx'

import pandas as pd
import numpy as np
import os
import openpyxl

from openpyxl import Workbook
        
wb = Workbook()

ws = wb.active
ws.title = 'Reversals Attempted'

ws2 = wb.create_sheet()
ws2.title = 'Reversals Achieved'

ws3 = wb.create_sheet()
ws3.title = 'Reversals Completed'

ws4 = wb.create_sheet()
ws4.title = 'Win-Stay'

ws5 = wb.create_sheet()
ws5.title = 'Lose-Shift'

ws6 = wb.create_sheet()
ws6.title = 'Total Pokes'

ws7 = wb.create_sheet()
ws7.title = 'Active Pokes'

ws8 = wb.create_sheet()
ws8.title = 'Inactive Pokes'

ws9 = wb.create_sheet()
ws9.title = 'Percent Active'

ws10 = wb.create_sheet()
ws10.title = 'Pellets'

ws11 = wb.create_sheet()
ws11.title = 'Left Pokes'

ws12 = wb.create_sheet()
ws12.title = 'Left Active Pokes'

ws13 = wb.create_sheet()
ws13.title = 'Left Inactive Pokes'

ws14 = wb.create_sheet()
ws14.title = 'Right Pokes'

ws15 = wb.create_sheet()
ws15.title = 'Right Active Pokes'

ws16 = wb.create_sheet()
ws16.title = 'Right Inactive Pokes'

ws17 = wb.create_sheet()
ws17.title = 'Inactive Pokes Per Rev'

ws18 = wb.create_sheet()
ws18.title = 'Left Inactive Pokes Per Rev'

ws19 = wb.create_sheet()
ws19.title = 'Right Inactive Pokes Per Rev'



date_column = 'Date'
attempted_column = 'Reversals Attempted'
achieved_column = 'Reversals Achieved'
completed_column = 'Reversals Completed'
winstay_column = 'Win-Stay'
loseshift_column = 'Lose-Shift'
total_column = 'Total Pokes'
active_column = 'Active Pokes'
inactive_column = 'Inactive Pokes'
percent_column = 'Percent Active'
pellet_column = 'Pellets'
left_column = 'Left Pokes'
left_active_column = 'Left Active Pokes'
left_inactive_column = 'Left Inactive Pokes'
right_column = 'Right Pokes'
right_active_column = 'Right Active Pokes'
right_inactive_column = 'Right Inactive Pokes'
inactive_per_rev_column = 'Inactive Pokes Per Reversal'
left_inactive_per_rev_column = 'Left Inactive Pokes Per Reversal'
right_inactive_per_rev_column = 'Right Inactive Pokes Per Reversal'


counter = 1

for filename in sorted(list(os.listdir(import_location))):
    
    if filename.endswith("Summary Joined.xlsx"):
        print(filename)
        if counter == 1:
            
            import_name = filename
            import_destination = import_location + import_name
            
            df_attempted = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_achieved = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_completed = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_winstay = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_loseshift = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_total = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_active = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_inactive = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_percent = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_pellet = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_left = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_left_active = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_left_inactive = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_right = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_right_active = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_right_inactive = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_inactive_per = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_left_inactive_per = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_right_inactive_per = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            
            cols_to_drop = ['Reversals Attempted','Reversals Achieved', 'Reversals Completed', 'Win-Stay', 'Lose-Shift',	'Total Pokes',	'Active Pokes',	'Inactive Pokes', 'Percent Active',	'Pellets', 'Left Pokes', 'Left Active Pokes', 'Left Inactive Pokes', 'Right Pokes', 'Right Active Pokes', 'Right Inactive Pokes', 'Inactive Pokes Per Reversal', 'Left Inactive Pokes Per Reversal', 'Right Inactive Pokes Per Reversal']
            df_attempted.drop(columns= cols_to_drop, inplace=True)
            df_achieved.drop(columns=cols_to_drop, inplace=True)
            df_completed.drop(columns=cols_to_drop, inplace=True)
            df_winstay.drop(columns=cols_to_drop, inplace=True)
            df_loseshift.drop(columns=cols_to_drop, inplace=True)
            df_total.drop(columns=cols_to_drop, inplace=True)
            df_active.drop(columns=cols_to_drop, inplace=True)
            df_inactive.drop(columns=cols_to_drop, inplace=True)
            df_percent.drop(columns=cols_to_drop, inplace=True)
            df_pellet.drop(columns=cols_to_drop, inplace=True)
            df_left.drop(columns=cols_to_drop, inplace=True)
            df_left_active.drop(columns=cols_to_drop, inplace=True)
            df_left_inactive.drop(columns=cols_to_drop, inplace=True)
            df_right.drop(columns=cols_to_drop, inplace=True)
            df_right_active.drop(columns=cols_to_drop, inplace=True)
            df_right_inactive.drop(columns=cols_to_drop, inplace=True)
            df_inactive_per.drop(columns=cols_to_drop, inplace=True)
            df_left_inactive_per.drop(columns=cols_to_drop, inplace=True)
            df_right_inactive_per.drop(columns=cols_to_drop, inplace=True)
            
        import_name = filename
        import_destination = import_location + import_name
        
        name = filename.strip(' Summary Joined.xlsx')
        
        df = pd.read_excel(import_destination, sheet_name = sheet_to_join)
        
        date = df[date_column].tolist()
        attempted = df[attempted_column].tolist()
        achieved = df[achieved_column].tolist()
        completed = df[completed_column].tolist()
        winstay = df[winstay_column].tolist()
        loseshift = df[loseshift_column].tolist()
        total = df[total_column].tolist()
        active = df[active_column].tolist()
        inactive = df[inactive_column].tolist()
        percent = df[percent_column].tolist()
        pellet = df[pellet_column].tolist()
        left = df[left_column].tolist()
        left_active = df[left_active_column].tolist()
        left_inactive = df[left_inactive_column].tolist()
        right = df[right_column].tolist()
        right_active = df[right_active_column].tolist()
        right_inactive = df[right_inactive_column].tolist()
        inactive_per_rev = df[inactive_per_rev_column].tolist()
        left_inactive_per_rev = df[left_inactive_per_rev_column].tolist()
        right_inactive_per_rev = df[right_inactive_per_rev_column].tolist()

        df_attempted.insert((counter), name, attempted)
        df_achieved.insert((counter), name, achieved)
        df_completed.insert((counter), name, completed)
        df_winstay.insert((counter), name, winstay)
        df_loseshift.insert((counter), name, loseshift)
        df_total.insert((counter), name, total)
        df_active.insert((counter), name, active)
        df_inactive.insert((counter), name, inactive)
        df_percent.insert((counter), name, percent)
        df_pellet.insert((counter), name, pellet)
        df_left.insert((counter), name, left)
        df_left_active.insert((counter), name, left_active)
        df_left_inactive.insert((counter), name, left_inactive)
        df_right.insert((counter), name, right)
        df_right_active.insert((counter), name, right_active)
        df_right_inactive.insert((counter), name, right_inactive)
        df_inactive_per.insert((counter), name, inactive_per_rev)
        df_left_inactive_per.insert((counter), name, left_inactive_per_rev)
        df_right_inactive_per.insert((counter), name, right_inactive_per_rev)
            
        counter += 1

# df_active = df_active.replace(np.nan, 'NA')
# df_inactive = df_inactive.replace(np.nan, 'NA')
# df_total = df_total.replace(np.nan, 'NA')
# df_pellet = df_pellet.replace(np.nan, 'NA')
# df_percent = df_percent.replace(np.nan, 'NA')
        
group_destination = export_location + group_name

results_to_export = [df_attempted, df_achieved, df_completed, df_winstay, df_loseshift, df_total, df_active, df_inactive, df_percent, df_pellet, df_left, df_left_active, df_left_inactive, df_right, df_right_active, df_right_inactive, df_inactive_per, df_left_inactive_per, df_right_inactive_per]

sheets_to_export = wb.sheetnames

with pd.ExcelWriter(group_destination) as writer:
    
    for i in range(len(sheets_to_export)):
                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
                    
cohort_destination = cohort_export_location + group_name

with pd.ExcelWriter(cohort_destination) as writer:
    
    for i in range(len(sheets_to_export)):
                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)

