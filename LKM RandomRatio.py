#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Sep 20 11:14:54 2021

@author: lauramilton
"""

# PROGRESSIVE RATIO TASK

time_column = 'MM:DD:YYYY hh:mm:ss'
event_column = 'Event'
active_poke_column = 'Active_Poke'
session_type_column = 'FR'
schedule_column = 'Session_type'
left_poke_column = 'Left_Poke_Count'
right_poke_column = 'Right_Poke_Count'
pellet_count_column = 'Pellet_Count'

retrieval_time_column = 'Retrieval_Time'
ipi_column = 'InterPelletInterval'
poke_time_column = 'Poke_Time'

import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/RRatio Pilot/'
export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/RRatio Pilot/'
cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/RRatio Pilot/'
ratio_list_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/RandRatio test/Fixed Ratios.xlsx'

initiation_poke_active = True
session_duration_mins = '180'
seconds_in_bins = [60, 600, 1800, 3600]

export_results = 'N'

#-----------------------------------------------------------------------------

# Import the revelant data: time, FR ratio, event, active port, left poke, right poke, and pellet count.

import pandas as pd
import numpy as np
import os
import openpyxl
import math


#-----

for folder in sorted(os.listdir(import_location)):
    if folder.startswith('RR'):
        print(folder)
        
        Date_summary = []
        Task_summary = []
        Duration_summary = []
        Active_port_summary = []
        ALL_Total_Pokes_summary = []
        ALL_Active_Pokes_summary = []
        ALL_Inactive_Pokes_summary = []
        ALL_Pokes_Percent_Active_summary = []
        Empty_Total_Pokes_summary = []
        Empty_Active_Pokes_summary = []
        Empty_Inactive_Pokes_summary = []
        Empty_Pokes_Percent_Active_summary = []
        W_Pellet_Total_Pokes_summary = []
        W_Pellet_Active_Pokes_summary = []
        W_Pellet_Inactive_Pokes_summary = []
        W_Pellet_Pokes_Percent_Active_summary = []
        Active_Pokes_Percent_Empty_summary = []
        Active_Pokes_Percent_W_Pellet_summary = []
        Pellets_summary = []
        Magazine_Checks_summary = []
        
        
        
        
        for filename in sorted(os.listdir(os.path.join(import_location, folder))):
            
            if filename.endswith(".CSV"):
                
                if '_00' in filename:
                        
            # Import the csv data
            
                    import_name = filename
                    import_destination = import_location + folder + '/' + import_name
                    export_name = folder + ' ' + import_name[7:13] + ' Random Ratio.xlsx'
                    export_destination = export_location + folder + '/' + export_name
                    print(filename)
            #-----
            
                    df = pd.read_csv(import_destination)
                    
                    time = df[time_column].tolist()
                    session_type = df[session_type_column].tolist()
                    schedule = df[schedule_column].tolist()
                    event = df[event_column].tolist()
                    active = df[active_poke_column].tolist()
                    left_poke = df[left_poke_column].tolist()
                    right_poke = df[right_poke_column].tolist()
                    pellet_count = df[pellet_count_column].tolist()
                    retrieval_time = df[retrieval_time_column].tolist()
                    ipi = df[ipi_column].tolist()
                    poke_time = df[poke_time_column].tolist()

                    # Change the time column from strings to datetime format for calculating durations between timestamps and creating time bins
                    
                    import datetime as dt
                    
                    time_list = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time]
                    
                    # Start time of the session is the first timestamp
                    
                    start_time = time_list[0]
                    
                    # Remove the initiation poke/s and pellet data (if required)
                                
                    if schedule[0] == 'RRatio5':
                        
                        del time[:4]
                        del session_type[:4]
                        del schedule[:4]
                        del event[:4]
                        del active[:4]
                        del left_poke[:4]
                        del right_poke[:4]
                        del pellet_count[:4]
                        del retrieval_time[:4]
                        del poke_time[:4]
                        del ipi[:4]

                    elif schedule[0] == 'RRatio10':

                        del time[:10]
                        del session_type[:10]
                        del schedule[:10]
                        del event[:10]
                        del active[:10]
                        del left_poke[:10]
                        del right_poke[:10]
                        del pellet_count[:10]
                        del retrieval_time[:10]
                        del poke_time[:10]
                        del ipi[:10]

                        
                    elif schedule[0] == 'RRatio20':
                        
                        del time[:15]
                        del session_type[:15]
                        del schedule[:15]
                        del event[:15]
                        del active[:15]
                        del left_poke[:15]
                        del right_poke[:15]
                        del pellet_count[:15]
                        del retrieval_time[:15]
                        del poke_time[:15]
                        del ipi[:15]
                    
                    # Subtract the pellet from the subsequent cumulative data
                    
                    pellet_count_shifted = []
                    
                    for i in range(0, len(left_poke)):
                        pellet_count_shifted.append(pellet_count[i] - 1)
                    
                    # Shift the retrieval_time and pellet_count lists one step backwards so that the count/time is in the same index as the corresponding poke
                    
                    retrieval_time.pop(0)
                    pellet_count_shifted.pop(0)
                    ipi.pop(0)
                    
                    # Add in a value to the end of each list to maintain list length so as not to lose the last row of other data
                    # for pellet_count this is a duplicate of preceding row, for retrieval_time and ipi this is np.nan
                    
                    retrieval_time.append(np.nan)
                    pellet_count_shifted.append(pellet_count[-1])
                    ipi.append(np.nan)
                    
                    # Create a MagCheck count column and columns that split left and right pokes into those with and without a pellet in the well
                    
                    mag_check = []
                    left_all = []
                    left_empty = []
                    left_w_pellet = []
                    right_all = []
                    right_empty = []
                    right_w_pellet = []

                    mag_check_counter = 0
                    left_empty_counter = 0
                    left_w_pellet_counter = 0
                    right_empty_counter = 0
                    right_w_pellet_counter = 0
                    
                    for i in range(0, len(event)):
                        if event[i] == 'MagCheck':
                            mag_check_counter += 1
                        elif event[i] == 'Left':
                            left_empty_counter += 1
                        elif event[i] == "LeftWithPellet":
                            left_w_pellet_counter += 1
                        elif event[i] == "Right":
                            right_empty_counter += 1
                        elif event[i] == "RightWithPellet":
                            right_w_pellet_counter += 1
                        mag_check.append(mag_check_counter)
                        left_empty.append(left_empty_counter)
                        left_w_pellet.append(left_w_pellet_counter)
                        right_empty.append(right_empty_counter)
                        right_w_pellet.append(right_w_pellet_counter)
                        left_all.append(left_empty_counter + left_w_pellet_counter)
                        right_all.append(right_empty_counter + right_w_pellet_counter)
                    
                    #####-----##### CHRONOLOGICAL DATA #####-----#####
                    
                    # Create new lists that only include data from the lines where the event is a Poke, in ReversalTask this means != 'Pellet' (i.e. Poke, Left, or Right)
                    
                    time2 = []
                    session_type2 = []
                    schedule2 = []
                    event2 = []
                    active2 = []
                    left_empty2 = []
                    left_w_pellet2 = []
                    right_empty2 = []
                    right_w_pellet2 = []
                    pellet_count2 = []
                    retrieval_time2 = []
                    poke_time2 = []
                    ipi_2 = []
                    mag_check2 = []
                    left_all2 = []
                    right_all2 = []
                    
                    for a, b, c, d, e, f, g, h, i, j, k, l, m, n, o, p in zip(event, left_empty, right_empty, pellet_count_shifted, time, session_type, active, retrieval_time, poke_time, ipi, left_w_pellet, right_w_pellet, schedule, mag_check, left_all, right_all):
                        if a != 'Pellet':
                            event2.append(a)
                            left_empty2.append(b)
                            right_empty2.append(c) 
                            pellet_count2.append(d)
                            time2.append(e)
                            session_type2.append(f)
                            active2.append(g)
                            retrieval_time2.append(h)
                            poke_time2.append(i)
                            ipi_2.append(j)
                            left_w_pellet2.append(k)
                            right_w_pellet2.append(l)
                            schedule2.append(m)
                            mag_check2.append(n)
                            left_all2.append(o)
                            right_all2.append(p)
                    
                    # print(len(time), len(session_type), len(event), len(active), len(left_poke), len(right_poke), len(pellet_count), len(retrieval_time),len(poke_time), len(ipi))                
                    # print(len(time2), len(session_type2), len(event2), len(active2), len(left_poke2), len(right_poke2), len(pellet_count2), len(retrieval_time2),len(poke_time2), len(ipi_2))                
                    ####################################################################################################
                    ####################################################################################################
                    
                    # Change the time2 column from strings to datetime format for calculating durations between timestamps
                    
                    time2 = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time2]
                    
                    # Calculate how much time (in seconds) has elapsed from start time for each event (nose poke)
                    
                    seconds_elapsed = []
                    
                    for i in range(0, len(time2)):
                        time_from_start = time2[i] - start_time
                        result = int(time_from_start.total_seconds())
                        seconds_elapsed.append(result)
                        
                    # Transform seconds_elapsed into duration in h:mm:ss
                                
                    duration_from_start = []
                    
                    for i in range(0, len(seconds_elapsed)):
                        time_hours = str(math.floor(seconds_elapsed[i] / 3600))
                        
                        time_mins = math.floor((seconds_elapsed[i] % 3600) / 60)
                        
                        time_secs = seconds_elapsed[i] % 60
                            
                        if time_mins < 10:
                            t_mins = '0' + str(time_mins)
                        else:
                            t_mins = str(time_mins)
                            
                        if time_secs < 10:
                            t_secs = '0' + str(time_secs)
                        else:
                            t_secs = str(time_secs)
                            
                        duration_from_start.append(str(time_hours) + ':' + t_mins + ':' + t_secs)
                    
                    ##### Remove any data that is from after max session duration
                    
                    end_seconds = int(session_duration_mins) * 60
                    
                    rows_to_keep = []                        
                    for i in range(0, len(seconds_elapsed)):
                        if seconds_elapsed[i] < end_seconds:
                            rows_to_keep.append(seconds_elapsed[i])
                    
                    rows_to_remove = len(seconds_elapsed) - len(rows_to_keep)

                    seconds_elapsed = seconds_elapsed[: (len(seconds_elapsed) - rows_to_remove)]
                    duration_from_start = duration_from_start[: (len(duration_from_start) - rows_to_remove)]
                    event2 = event2[: (len(event2) - rows_to_remove)]
                    left_empty2 = left_empty2[: (len(left_empty2) - rows_to_remove)]
                    right_empty2 = right_empty2[: (len(right_empty2) - rows_to_remove)]
                    pellet_count2 = pellet_count2[: (len(pellet_count2) - rows_to_remove)]
                    time2 = time2[: (len(time2) - rows_to_remove)]
                    session_type2 = session_type2[: (len(session_type2) - rows_to_remove)]
                    active2 = active2[: (len(active2) - rows_to_remove)]
                    retrieval_time2 = retrieval_time2[: (len(retrieval_time2) - rows_to_remove)]
                    poke_time2 = poke_time2[: (len(poke_time2) - rows_to_remove)]
                    ipi_2 = ipi_2[: (len(ipi_2) - rows_to_remove)]
                    left_w_pellet2 = left_w_pellet2[: (len(left_w_pellet2) - rows_to_remove)]
                    right_w_pellet2 = right_w_pellet2[: (len(right_w_pellet2) - rows_to_remove)]
                    schedule2 = schedule2[: (len(schedule2) - rows_to_remove)]
                    mag_check2 = mag_check2[: (len(mag_check2) - rows_to_remove)]
                    left_all2 = left_all2[: (len(left_all2) - rows_to_remove)]
                    right_all2 = right_all2[: (len(right_all2) - rows_to_remove)]
                    
                    # Create a new column that is the total number of pokes
                    
                    total_pokes_all = []
                    total_pokes_empty = []
                    total_pokes_w_pellet = []
                    
                    for i in range(0, len(active2)):
                        total_pokes_all.append(left_empty2[i] + right_empty2[i] + left_w_pellet2[i] + right_w_pellet2[i])
                        total_pokes_empty.append(left_empty2[i] + right_empty2[i])
                        total_pokes_w_pellet.append(left_w_pellet2[i] + right_w_pellet2[i])
                    
                    # Create new column that is % of pokes that are into the active port. Needs to be based on active_poke value (i.e. Left or Right)
                    
                    all_percent_active = []
                    empty_percent_active = []
                    w_pellet_percent_active = []
                    
                    active_percent_empty = []
                    active_percent_w_pellet = []
                    
                    for i in range(0, len(active2)):
                        if active2[i] == 'Left':
                            if total_pokes_all[i] != 0:
                                all_percent_active.append((left_empty2[i] + left_w_pellet2[i]) / total_pokes_all[i] * 100)
                            else:
                                all_percent_active.append('N/A')
                            if total_pokes_empty[i] != 0:
                                empty_percent_active.append(left_empty2[i] / total_pokes_empty[i] * 100)
                            else:
                                empty_percent_active.append('N/A')
                            if total_pokes_w_pellet[i] != 0:
                                w_pellet_percent_active.append(left_w_pellet2[i] / total_pokes_w_pellet[i] * 100)
                            else:
                                w_pellet_percent_active.append('N/A')
                            if left_all2[i] != 0:
                                active_percent_empty.append(left_empty2[i] / left_all2[i] * 100)
                                active_percent_w_pellet.append(left_w_pellet2[i] / left_all2[i] * 100)
                            else:
                                active_percent_empty.append('N/A')
                                active_percent_w_pellet.append('N/A')
                        elif active2[i] == 'Right':
                            if total_pokes_all[i] != 0:
                                all_percent_active.append((right_empty2[i] + right_w_pellet2[i]) / total_pokes_all[i] * 100)
                            else:
                                all_percent_active.append('N/A')
                            if total_pokes_empty[i] != 0:
                                empty_percent_active.append(right_empty2[i] / total_pokes_empty[i] * 100)
                            else:
                                empty_percent_active.append('N/A')
                            if total_pokes_w_pellet[i] != 0:
                                w_pellet_percent_active.append(right_w_pellet2[i] / total_pokes_w_pellet[i] * 100)
                            else:
                                w_pellet_percent_active.append('N/A')
                            if right_all2[i] != 0:
                                active_percent_empty.append(right_empty2[i] / right_all2[i] * 100)
                                active_percent_w_pellet.append(right_w_pellet2[i] / right_all2[i] * 100)
                            else:
                                active_percent_empty.append('N/A')
                                active_percent_w_pellet.append('N/A')


                    # Chronological Poke Time Data split by side
                    
                    l_poke_time_all = [] # poke time for all left pokes
                    l_poke_time_empty = []
                    l_poke_time_w_pellet = []
                    r_poke_time_all = [] # poke time for all right pokes
                    r_poke_time_empty = []
                    r_poke_time_w_pellet = []
                    
                    l_poke_time_all_chron = [] # poke time for all left pokes with nan if it was a right poke to maintain array length the same as all pokes
                    l_poke_time_empty_chron = []
                    l_poke_time_w_pellet_chron = []
                    
                    r_poke_time_all_chron = [] # poke time for all right pokes with nan if it was a left poke to maintain array length the same as all pokes
                    r_poke_time_empty_chron = []
                    r_poke_time_w_pellet_chron = []
                    
                    for i in range(0, len(poke_time2)): # Collates the poke_time by side
                       if event2[i] == 'Left':
                           l_poke_time_all.append(poke_time2[i])
                           l_poke_time_empty.append(poke_time2[i])
                           l_poke_time_w_pellet.append(np.nan)
                           l_poke_time_all_chron.append(poke_time2[i])
                           l_poke_time_empty_chron.append(poke_time2[i])
                           l_poke_time_w_pellet_chron.append(np.nan)
                           r_poke_time_all_chron.append(np.nan)
                           r_poke_time_empty_chron.append(np.nan)
                           r_poke_time_w_pellet_chron.append(np.nan)
                       elif event2[i] == 'LeftWithPellet':
                           l_poke_time_all.append(poke_time2[i])
                           l_poke_time_empty.append(np.nan)
                           l_poke_time_w_pellet.append(poke_time2[i])
                           l_poke_time_all_chron.append(poke_time2[i])
                           l_poke_time_empty_chron.append(np.nan)
                           l_poke_time_w_pellet_chron.append(poke_time2[i])
                           r_poke_time_all_chron.append(np.nan)
                           r_poke_time_empty_chron.append(np.nan)
                           r_poke_time_w_pellet_chron.append(np.nan)
                       elif event2[i] == 'Right':
                           r_poke_time_all.append(poke_time2[i])
                           r_poke_time_empty.append(poke_time2[i])
                           r_poke_time_w_pellet.append(np.nan)
                           l_poke_time_all_chron.append(np.nan)
                           l_poke_time_empty_chron.append(np.nan)
                           l_poke_time_w_pellet_chron.append(np.nan)
                           r_poke_time_all_chron.append(poke_time2[i])
                           r_poke_time_empty_chron.append(poke_time2[i])
                           r_poke_time_w_pellet_chron.append(np.nan)
                       elif event2[i] == 'RightWithPellet':
                           r_poke_time_all.append(poke_time2[i])
                           r_poke_time_empty.append(np.nan)
                           r_poke_time_w_pellet.append(poke_time2[i])
                           l_poke_time_all_chron.append(np.nan)
                           l_poke_time_empty_chron.append(np.nan)
                           l_poke_time_w_pellet_chron.append(np.nan)
                           r_poke_time_all_chron.append(poke_time2[i])
                           r_poke_time_empty_chron.append(np.nan)
                           r_poke_time_w_pellet_chron.append(poke_time2[i])
                       elif event2[i] == 'MagCheck':
                           l_poke_time_all_chron.append(np.nan)
                           l_poke_time_empty_chron.append(np.nan)
                           l_poke_time_w_pellet_chron.append(np.nan)
                           r_poke_time_all_chron.append(np.nan)
                           r_poke_time_empty_chron.append(np.nan)
                           r_poke_time_w_pellet_chron.append(np.nan)

                   
                    ####################################################################################################
                    ####################################################################################################

                    # Create blocks based on FR steps - These are cumulative
                    
                    step_block = []
                    seconds_elapsed_blocked = []
                    duration_blocked = []
                    left_all_blocked = []
                    left_empty_blocked = []
                    left_w_pellet_blocked = []
                    right_all_blocked = []
                    right_empty_blocked = []
                    right_w_pellet_blocked = []
                    pellet_blocked = []
                    active_blocked = []
                    all_percent_active_blocked = []
                    empty_percent_active_blocked = []
                    w_pellet_percent_active_blocked = []
                    total_pokes_all_blocked = []
                    total_pokes_empty_blocked = []
                    total_pokes_w_pellet_blocked = []
                    mag_check_blocked = []
                    active_percent_empty_blocked = []
                    active_percent_w_pellet_blocked = []
                    
                    for i in range(0, len(session_type2)):
                        if math.isnan(retrieval_time2[i]) == False:
                            step_block.append(session_type2[i])
                            seconds_elapsed_blocked.append(seconds_elapsed[i])
                            duration_blocked.append(duration_from_start[i])
                            left_all_blocked.append(left_empty2[i] + left_w_pellet2[i])
                            left_empty_blocked.append(left_empty2[i])
                            left_w_pellet_blocked.append(left_w_pellet2[i])
                            right_all_blocked.append(right_empty2[i] + right_w_pellet2[i])
                            right_empty_blocked.append(right_empty2[i])
                            right_w_pellet_blocked.append(right_w_pellet2[i])
                            pellet_blocked.append(pellet_count2[i])
                            active_blocked.append(active2[i])
                            total_pokes_all_blocked.append(total_pokes_all[i])
                            total_pokes_empty_blocked.append(total_pokes_empty[i])
                            total_pokes_w_pellet_blocked.append(total_pokes_w_pellet[i])
                            mag_check_blocked.append(mag_check2[i])
                    
                    if math.isnan(retrieval_time2[-1]) == True: # if the final block is incomplete need to add it
                        step_block.append(session_type2[-1])
                        seconds_elapsed_blocked.append(seconds_elapsed[-1])
                        duration_blocked.append(duration_from_start[-1])
                        left_all_blocked.append(left_empty2[-1] + left_w_pellet2[-1])
                        left_empty_blocked.append(left_empty2[-1])
                        left_w_pellet_blocked.append(left_w_pellet2[-1])
                        right_all_blocked.append(right_empty2[-1] + right_w_pellet2[-1])
                        right_empty_blocked.append(right_empty2[-1])
                        right_w_pellet_blocked.append(right_w_pellet2[-1])
                        pellet_blocked.append(pellet_count2[-1])
                        active_blocked.append(active2[-1])
                        total_pokes_all_blocked.append(total_pokes_all[-1])
                        total_pokes_empty_blocked.append(total_pokes_empty[-1])
                        total_pokes_w_pellet_blocked.append(total_pokes_w_pellet[-1])
                        mag_check_blocked.append(mag_check2[-1])
                                       
                    # percent active blocked
                    for i in range(0, len(active_blocked)):
                        if active_blocked[i] == 'Left':
                            all_percent_active_blocked.append(left_all_blocked[i] / total_pokes_all_blocked[i] * 100)
                            empty_percent_active_blocked.append(left_empty_blocked[i] / total_pokes_empty_blocked[i] * 100)
                            if total_pokes_w_pellet_blocked[i] != 0:
                                w_pellet_percent_active_blocked.append(left_w_pellet_blocked[i] / total_pokes_w_pellet_blocked[i] * 100)
                            else:
                                w_pellet_percent_active_blocked.append('N/A')
                            active_percent_empty_blocked.append(left_empty_blocked[i] /left_all_blocked[i] * 100)
                            active_percent_w_pellet_blocked.append(left_w_pellet_blocked[i] / left_all_blocked[i] * 100)

                        else:
                            all_percent_active_blocked.append(right_all_blocked[i] / total_pokes_all_blocked[i] * 100)
                            empty_percent_active_blocked.append(right_empty_blocked[i] / total_pokes_empty_blocked[i] * 100)
                            if total_pokes_w_pellet_blocked[i] != 0:
                                w_pellet_percent_active_blocked.append(right_w_pellet_blocked[i] / total_pokes_w_pellet_blocked[i] * 100)
                            else:
                                w_pellet_percent_active.blocked.append('N/A')
                            active_percent_empty_blocked.append(right_empty_blocked[i] /right_all_blocked[i] * 100)
                            active_percent_w_pellet_blocked.append(right_w_pellet_blocked[i] / right_all_blocked[i] * 100)

                    
                    # Create new bins where counts are within blocks rather than cumulative
                    
                    seconds_elapsed_blocked_within = []
                    left_all_blocked_within = []
                    left_empty_blocked_within = []
                    left_w_pellet_blocked_within = []
                    right_all_blocked_within = []
                    right_empty_blocked_within = []
                    right_w_pellet_blocked_within = []
                    pellet_blocked_within = []
                    all_percent_active_blocked_within = []
                    empty_percent_active_blocked_within = []
                    w_pellet_percent_active_blocked_within = []
                    total_pokes_all_blocked_within = []
                    total_pokes_empty_blocked_within = []
                    total_pokes_w_pellet_blocked_within = []
                    mag_check_blocked_within = []
                    active_percent_empty_blocked_within = []
                    active_percent_w_pellet_blocked_within = []                   

                    seconds_elapsed_blocked_within.append(seconds_elapsed_blocked[0])
                    for i in range(1, len(seconds_elapsed_blocked)):
                        seconds_elapsed_blocked_within.append(seconds_elapsed_blocked[i] - seconds_elapsed_blocked[i - 1])
                    
                    # Transform seconds_elapsed into duration in h:mm:ss
                                
                    duration_blocked_within = []
                    
                    for i in range(0, len(seconds_elapsed_blocked_within)):
                        time_hours = str(math.floor(seconds_elapsed_blocked_within[i] / 3600))
                        
                        time_mins = math.floor((seconds_elapsed_blocked_within[i] % 3600) / 60)
                        
                        time_secs = seconds_elapsed_blocked_within[i] % 60
                            
                        if time_mins < 10:
                            t_mins = '0' + str(time_mins)
                        else:
                            t_mins = str(time_mins)
                            
                        if time_secs < 10:
                            t_secs = '0' + str(time_secs)
                        else:
                            t_secs = str(time_secs)
                            
                        duration_blocked_within.append(str(time_hours) + ':' + t_mins + ':' + t_secs)
                    
                    left_all_blocked_within.append(left_all_blocked[0])
                    for i in range(1, len(left_all_blocked)):
                        left_all_blocked_within.append(left_all_blocked[i] - left_all_blocked[i - 1])

                    left_empty_blocked_within.append(left_empty_blocked[0])
                    for i in range(1, len(left_empty_blocked)):
                        left_empty_blocked_within.append(left_empty_blocked[i] - left_empty_blocked[i - 1])

                    left_w_pellet_blocked_within.append(left_w_pellet_blocked[0])
                    for i in range(1, len(left_w_pellet_blocked)):
                        left_w_pellet_blocked_within.append(left_w_pellet_blocked[i] - left_w_pellet_blocked[i - 1])

                    right_all_blocked_within.append(right_all_blocked[0])
                    for i in range(1, len(right_all_blocked)):
                        right_all_blocked_within.append(right_all_blocked[i] - right_all_blocked[i - 1])

                    right_empty_blocked_within.append(right_empty_blocked[0])
                    for i in range(1, len(right_empty_blocked)):
                        right_empty_blocked_within.append(right_empty_blocked[i] - right_empty_blocked[i - 1])

                    right_w_pellet_blocked_within.append(right_w_pellet_blocked[0])
                    for i in range(1, len(right_w_pellet_blocked)):
                        right_w_pellet_blocked_within.append(right_w_pellet_blocked[i] - right_w_pellet_blocked[i - 1])
                    
                    pellet_blocked_within.append(pellet_blocked[0])
                    for i in range(1, len(pellet_blocked)):
                        pellet_blocked_within.append(pellet_blocked[i] - pellet_blocked[i - 1])
                        
                    total_pokes_all_blocked_within.append(total_pokes_all_blocked[0])
                    for i in range (1, len(total_pokes_all_blocked)):
                        total_pokes_all_blocked_within.append(total_pokes_all_blocked[i] - total_pokes_all_blocked[i - 1])
                    
                    total_pokes_empty_blocked_within.append(total_pokes_empty_blocked[0])
                    for i in range (1, len(total_pokes_empty_blocked)):
                        total_pokes_empty_blocked_within.append(total_pokes_empty_blocked[i] - total_pokes_empty_blocked[i - 1])

                    total_pokes_w_pellet_blocked_within.append(total_pokes_w_pellet_blocked[0])
                    for i in range (1, len(total_pokes_w_pellet_blocked)):
                        total_pokes_w_pellet_blocked_within.append(total_pokes_w_pellet_blocked[i] - total_pokes_w_pellet_blocked[i - 1])
                        
                    mag_check_blocked_within.append(mag_check_blocked[0])
                    for i in range(1, len(mag_check_blocked)):
                        mag_check_blocked_within.append(mag_check_blocked[i] - mag_check_blocked[i - 1])

                    for i in range(0, len(active_blocked)):
                        if active_blocked[i] == 'Left':
                            if total_pokes_all_blocked_within[i] != 0:
                                all_percent_active_blocked_within.append(left_all_blocked_within[i] / total_pokes_all_blocked_within[i] * 100)
                            else:
                                all_percent_active_blocked_within.append('NA')
                            if total_pokes_empty_blocked_within[i] != 0:
                                empty_percent_active_blocked_within.append(left_empty_blocked_within[i] / total_pokes_empty_blocked_within[i] * 100)
                            else:
                                empty_percent_active_blocked_within.append('NA')
                            if total_pokes_w_pellet_blocked_within[i] != 0:
                                w_pellet_percent_active_blocked_within.append(left_w_pellet_blocked_within[i] / total_pokes_w_pellet_blocked_within[i] * 100)
                            else:
                                w_pellet_percent_active_blocked_within.append('NA')
                            if left_all_blocked_within[i] != 0:
                                active_percent_empty_blocked_within.append(left_empty_blocked_within[i] / left_all_blocked_within[i] * 100)
                                active_percent_w_pellet_blocked_within.append(left_w_pellet_blocked_within[i] / left_all_blocked_within[i] * 100)
                            else:
                                active_percent_empty_blocked_within.append('N/A')
                                active_percent_w_pellet_blocked_within.append('N/A')

                        else:
                            if total_pokes_all_blocked_within[i] != 0:
                                all_percent_active_blocked_within.append(right_all_blocked_within[i] / total_pokes_all_blocked_within[i] * 100)
                            else:
                                all_percent_active_blocked_within.append('NA')
                            if total_pokes_empty_blocked_within[i] != 0:
                                empty_percent_active_blocked_within.append(right_empty_blocked_within[i] / total_pokes_empty_blocked_within[i] * 100)
                            else:
                                empty_percent_active_blocked_within.append('NA')
                            if total_pokes_w_pellet_blocked_within[i] != 0:
                                w_pellet_percent_active_blocked_within.append(right_w_pellet_blocked_within[i] / total_pokes_w_pellet_blocked_within[i] * 100)
                            else:
                                w_pellet_percent_active_blocked_within.append('NA')    
                            if right_all_blocked_within[i] != 0:
                                active_percent_empty_blocked_within.append(right_empty_blocked_within[i] / right_all_blocked_within[i] * 100)
                                active_percent_w_pellet_blocked_within.append(right_w_pellet_blocked_within[i] / right_all_blocked_within[i] * 100)
                            else:
                                active_percent_empty_blocked_within.append('N/A')
                                active_percent_w_pellet_blocked_within.append('N/A')


                    
                    # Blocked poke time, retrieval time and ipi data
                    
                    # Create mean poke and retrieval time columns for bins
                                
                    blocked_sum_both_all_poke_time = []
                    blocked_sum_both_empty_poke_time = []
                    blocked_sum_both_w_pellet_poke_time = []
                    blocked_sum_l_all_poke_time = [] 
                    blocked_sum_l_empty_poke_time = []
                    blocked_sum_l_w_pellet_poke_time = []
                    blocked_sum_r_all_poke_time = [] 
                    blocked_sum_r_empty_poke_time = []
                    blocked_sum_r_w_pellet_poke_time = []
                    
                    blocked_mean_both_all_poke_time = [] 
                    blocked_mean_both_empty_poke_time = []
                    blocked_mean_both_w_pellet_poke_time = []
                    blocked_mean_l_all_poke_time = [] 
                    blocked_mean_l_empty_poke_time = []
                    blocked_mean_l_w_pellet_poke_time = []
                    blocked_mean_r_all_poke_time = [] 
                    blocked_mean_r_empty_poke_time = []
                    blocked_mean_r_w_pellet_poke_time = []
                    
                    sum_both_all_poke_time = 0
                    sum_both_empty_poke_time = 0
                    sum_both_w_pellet_poke_time = 0
                    sum_l_all_poke_time = 0
                    sum_l_empty_poke_time = 0
                    sum_l_w_pellet_poke_time = 0
                    sum_r_all_poke_time = 0
                    sum_r_empty_poke_time = 0
                    sum_r_w_pellet_poke_time = 0
                    
                    block_index = 0
                    row_index = 0
                                        
                    # print(seconds_elapsed, seconds_elapsed_binned, left_poke2, poke_time2)
                    # print(len(seconds_elapsed), len(poke_time2), len(seconds_elapsed_binned))
                    for i in range(0, len(poke_time2)):
                        if int(seconds_elapsed[i]) < int(seconds_elapsed_blocked[block_index]):
                            
                            if event2[i] != 'MagCheck':
                                sum_both_all_poke_time += poke_time2[row_index]
                            if event2[i] == 'Left':
                                sum_l_all_poke_time += poke_time2[row_index]
                                sum_l_empty_poke_time += poke_time2[row_index]
                                sum_both_empty_poke_time += poke_time2[row_index]
                            elif event2[i] == 'Right':
                                sum_r_all_poke_time += poke_time2[row_index]
                                sum_r_empty_poke_time += poke_time2[row_index]
                                sum_both_empty_poke_time += poke_time2[row_index]
                            elif event2[i] == 'LeftWithPellet':
                                sum_l_all_poke_time += poke_time2[row_index]
                                sum_l_w_pellet_poke_time += poke_time2[row_index]
                                sum_both_w_pellet_poke_time += poke_time2[row_index]
                            elif event2[i] == 'RightWithPellet':
                                sum_r_all_poke_time += poke_time2[row_index]
                                sum_r_w_pellet_poke_time += poke_time2[row_index]
                                sum_both_w_pellet_poke_time += poke_time2[row_index]
                            row_index += 1
                            
                        elif int(seconds_elapsed[i]) == int(seconds_elapsed_blocked[block_index]):
                            if event2[i] != 'MagCheck':
                                sum_both_all_poke_time += poke_time2[row_index]
                            if event2[i] == 'Left':
                                sum_l_all_poke_time += poke_time2[row_index]
                                sum_l_empty_poke_time += poke_time2[row_index]
                                sum_both_empty_poke_time += poke_time2[row_index]
                            elif event2[i] == 'Right':
                                sum_r_all_poke_time += poke_time2[row_index]
                                sum_r_empty_poke_time += poke_time2[row_index]
                                sum_both_empty_poke_time += poke_time2[row_index]
                            elif event2[i] == 'LeftWithPellet':
                                sum_l_all_poke_time += poke_time2[row_index]
                                sum_l_w_pellet_poke_time += poke_time2[row_index]
                                sum_both_w_pellet_poke_time += poke_time2[row_index]
                            elif event2[i] == 'RightWithPellet':
                                sum_r_all_poke_time += poke_time2[row_index]
                                sum_r_w_pellet_poke_time += poke_time2[row_index]
                                sum_both_w_pellet_poke_time += poke_time2[row_index]
                            row_index += 1
                            
                            blocked_sum_both_all_poke_time.append(sum_both_all_poke_time) 
                            blocked_sum_both_empty_poke_time.append(sum_both_empty_poke_time)
                            blocked_sum_both_w_pellet_poke_time.append(sum_both_w_pellet_poke_time)
                            blocked_sum_l_all_poke_time.append(sum_l_all_poke_time) 
                            blocked_sum_l_empty_poke_time.append(sum_l_empty_poke_time)
                            blocked_sum_l_w_pellet_poke_time.append(sum_l_w_pellet_poke_time)
                            blocked_sum_r_all_poke_time.append(sum_r_all_poke_time) 
                            blocked_sum_r_empty_poke_time.append(sum_r_empty_poke_time)
                            blocked_sum_r_w_pellet_poke_time.append(sum_r_w_pellet_poke_time)
                            
                            blocked_mean_both_all_poke_time.append(sum_both_all_poke_time / total_pokes_all_blocked[block_index]) 
                            blocked_mean_both_empty_poke_time.append(sum_both_empty_poke_time / total_pokes_empty_blocked[block_index])
                            if total_pokes_w_pellet_blocked[block_index] != 0:
                                blocked_mean_both_w_pellet_poke_time.append(sum_both_w_pellet_poke_time / total_pokes_w_pellet_blocked[block_index])
                            else:
                                blocked_mean_both_w_pellet_poke_time.append('N/A')
                            
                            if left_all_blocked[block_index] != 0:
                                blocked_mean_l_all_poke_time.append(sum_l_all_poke_time / left_all_blocked[block_index]) 
                            else:
                                blocked_mean_l_all_poke_time.append('N/A')
                            if left_empty_blocked[block_index] != 0:
                                blocked_mean_l_empty_poke_time.append(sum_l_empty_poke_time / left_empty_blocked[block_index])
                            else:
                                blocked_mean_l_empty_poke_time.append('N/A')
                            if left_w_pellet_blocked[block_index] != 0:
                                blocked_mean_l_w_pellet_poke_time.append(sum_l_w_pellet_poke_time / left_w_pellet_blocked[block_index])
                            else:
                                blocked_mean_l_w_pellet_poke_time.append('N/A')
                            
                            if right_all_blocked[block_index] != 0:
                                blocked_mean_r_all_poke_time.append(sum_r_all_poke_time / right_all_blocked[block_index]) 
                            else:
                                blocked_mean_r_all_poke_time.append('N/A')
                            if right_empty_blocked[block_index] != 0:
                                blocked_mean_r_empty_poke_time.append(sum_r_empty_poke_time / right_empty_blocked[block_index])
                            else:
                                blocked_mean_r_empty_poke_time.append('N/A')
                            if right_w_pellet_blocked[block_index] != 0:
                                blocked_mean_r_w_pellet_poke_time.append(sum_r_w_pellet_poke_time / right_w_pellet_blocked[block_index])
                            else:
                                blocked_mean_r_w_pellet_poke_time.append('N/A')
                            
                            block_index += 1
                        
                        elif int(seconds_elapsed[i]) > int(seconds_elapsed_blocked[block_index]):
                            
                            blocked_sum_both_all_poke_time.append(sum_both_all_poke_time) 
                            blocked_sum_both_empty_poke_time.append(sum_both_empty_poke_time)
                            blocked_sum_both_w_pellet_poke_time.append(sum_both_w_pellet_poke_time)
                            blocked_sum_l_all_poke_time.append(sum_l_all_poke_time) 
                            blocked_sum_l_empty_poke_time.append(sum_l_empty_poke_time)
                            blocked_sum_l_w_pellet_poke_time.append(sum_l_w_pellet_poke_time)
                            blocked_sum_r_all_poke_time.append(sum_r_all_poke_time) 
                            blocked_sum_r_empty_poke_time.append(sum_r_empty_poke_time)
                            blocked_sum_r_w_pellet_poke_time.append(sum_r_w_pellet_poke_time)
                            
                            blocked_mean_both_all_poke_time.append(sum_both_all_poke_time / total_pokes_all_blocked[block_index]) 
                            blocked_mean_both_empty_poke_time.append(sum_both_empty_poke_time / total_pokes_empty_blocked[block_index])
                            blocked_mean_both_w_pellet_poke_time.append(sum_both_w_pellet_poke_time / total_pokes_w_pellet_blocked[block_index])
                            
                            if left_all_blocked[block_index] != 0:
                                blocked_mean_l_all_poke_time.append(sum_l_all_poke_time / left_all_blocked[block_index]) 
                            else:
                                blocked_mean_l_all_poke_time.append('N/A')
                            if left_empty_blocked[block_index] != 0:
                                blocked_mean_l_empty_poke_time.append(sum_l_empty_poke_time / left_empty_blocked[block_index])
                            else:
                                blocked_mean_l_empty_poke_time.append('N/A')
                            if left_w_pellet_blocked[block_index] != 0:
                                blocked_mean_l_w_pellet_poke_time.append(sum_l_w_pellet_poke_time / left_w_pellet_blocked[block_index])
                            else:
                                blocked_mean_l_w_pellet_poke_time.append('N/A')
                            
                            if right_all_blocked[block_index] != 0:
                                blocked_mean_r_all_poke_time.append(sum_r_all_poke_time / right_all_blocked[block_index]) 
                            else:
                                blocked_mean_r_all_poke_time.append('N/A')
                            if right_empty_blocked[block_index] != 0:
                                blocked_mean_r_empty_poke_time.append(sum_r_empty_poke_time / right_empty_blocked[block_index])
                            else:
                                blocked_mean_r_empty_poke_time.append('N/A')
                            if right_w_pellet_blocked[block_index] != 0:
                                blocked_mean_r_w_pellet_poke_time.append(sum_r_w_pellet_poke_time / right_w_pellet_blocked[block_index])
                            else:
                                blocked_mean_r_w_pellet_poke_time.append('N/A')
                            
                            block_index += 1
                            
                            if event2[i] != 'MagCheck':
                                sum_both_all_poke_time += poke_time2[row_index]
                            if event2[i] == 'Left':
                                sum_l_all_poke_time += poke_time2[row_index]
                                sum_l_empty_poke_time += poke_time2[row_index]
                                sum_both_empty_poke_time += poke_time2[row_index]
                            elif event2[i] == 'Right':
                                sum_r_all_poke_time += poke_time2[row_index]
                                sum_r_empty_poke_time += poke_time2[row_index]
                                sum_both_empty_poke_time += poke_time2[row_index]
                            elif event2[i] == 'LeftWithPellet':
                                sum_l_all_poke_time += poke_time2[row_index]
                                sum_l_w_pellet_poke_time += poke_time2[row_index]
                                sum_both_w_pellet_poke_time += poke_time2[row_index]
                            elif event2[i] == 'RightWithPellet':
                                sum_r_all_poke_time += poke_time2[row_index]
                                sum_r_w_pellet_poke_time += poke_time2[row_index]
                                sum_both_w_pellet_poke_time += poke_time2[row_index]
                            row_index += 1

                    #####-----       
                    
                    blocked_within_sum_both_all_poke_time = [] 
                    blocked_within_sum_both_empty_poke_time = []
                    blocked_within_sum_both_w_pellet_poke_time = []
                    blocked_within_sum_l_all_poke_time = [] 
                    blocked_within_sum_l_empty_poke_time = []
                    blocked_within_sum_l_w_pellet_poke_time = []
                    blocked_within_sum_r_all_poke_time = [] 
                    blocked_within_sum_r_empty_poke_time = []
                    blocked_within_sum_r_w_pellet_poke_time = []
                    
                    blocked_within_sum_both_all_poke_time.append(blocked_sum_both_all_poke_time[0]) 
                    blocked_within_sum_both_empty_poke_time.append(blocked_sum_both_empty_poke_time[0])
                    blocked_within_sum_both_w_pellet_poke_time.append(blocked_sum_both_w_pellet_poke_time[0])
                    blocked_within_sum_l_all_poke_time.append(blocked_sum_l_all_poke_time[0]) 
                    blocked_within_sum_l_empty_poke_time.append(blocked_sum_l_empty_poke_time[0])
                    blocked_within_sum_l_w_pellet_poke_time.append(blocked_sum_l_w_pellet_poke_time[0])
                    blocked_within_sum_r_all_poke_time.append(blocked_sum_r_all_poke_time[0])
                    blocked_within_sum_r_empty_poke_time.append(blocked_sum_r_empty_poke_time[0])
                    blocked_within_sum_r_w_pellet_poke_time.append(blocked_sum_r_w_pellet_poke_time[0])
                                       
                    for i in range(1, len(blocked_sum_both_all_poke_time)):
                        blocked_within_sum_both_all_poke_time.append(blocked_sum_both_all_poke_time[i] - blocked_sum_both_all_poke_time[i - 1]) 
                        blocked_within_sum_both_empty_poke_time.append(blocked_sum_both_empty_poke_time[i] - blocked_sum_both_empty_poke_time[i - 1])
                        blocked_within_sum_both_w_pellet_poke_time.append(blocked_sum_both_w_pellet_poke_time[i] - blocked_sum_both_w_pellet_poke_time[i - 1])
                        blocked_within_sum_l_all_poke_time.append(blocked_sum_l_all_poke_time[i] - blocked_sum_l_all_poke_time[i - 1])
                        blocked_within_sum_l_empty_poke_time.append(blocked_sum_l_empty_poke_time[i] - blocked_sum_l_empty_poke_time[i - 1])
                        blocked_within_sum_l_w_pellet_poke_time.append(blocked_sum_l_w_pellet_poke_time[i] - blocked_sum_l_w_pellet_poke_time[i - 1])
                        blocked_within_sum_r_all_poke_time.append(blocked_sum_r_all_poke_time[i] - blocked_sum_r_all_poke_time[i - 1])
                        blocked_within_sum_r_empty_poke_time.append(blocked_sum_r_empty_poke_time[i] - blocked_sum_r_empty_poke_time[i] - 1)
                        blocked_within_sum_r_w_pellet_poke_time.append(blocked_sum_r_w_pellet_poke_time[i] - blocked_sum_r_w_pellet_poke_time[i] - 1)
                    
                    blocked_within_mean_both_all_poke_time = [] 
                    blocked_within_mean_both_empty_poke_time = []
                    blocked_within_mean_both_w_pellet_poke_time = []
                    blocked_within_mean_l_all_poke_time = [] 
                    blocked_within_mean_l_empty_poke_time = []
                    blocked_within_mean_l_w_pellet_poke_time = []
                    blocked_within_mean_r_all_poke_time = [] 
                    blocked_within_mean_r_empty_poke_time = []
                    blocked_within_mean_r_w_pellet_poke_time = []

                    
                    for i in range(0, len(blocked_within_sum_both_all_poke_time)):
                        blocked_within_mean_both_all_poke_time.append(blocked_within_sum_both_all_poke_time[i] / total_pokes_all_blocked_within[i])
                        blocked_within_mean_both_empty_poke_time.append(blocked_within_sum_both_empty_poke_time[i] / total_pokes_empty_blocked_within[i])
                        if total_pokes_w_pellet_blocked_within[i] != 0:
                            blocked_within_mean_both_w_pellet_poke_time.append(blocked_within_sum_both_w_pellet_poke_time[i] / total_pokes_w_pellet_blocked_within[i])
                        else:
                            blocked_within_mean_both_w_pellet_poke_time.append('N/A')
                        
                        if left_all_blocked_within[i] != 0:
                            blocked_within_mean_l_all_poke_time.append(blocked_within_sum_l_all_poke_time[i] / left_all_blocked_within[i])
                        else:
                            blocked_within_mean_l_all_poke_time.append('N/A')
                        
                        if left_empty_blocked_within[i] != 0:
                            blocked_within_mean_l_empty_poke_time.append(blocked_within_sum_l_empty_poke_time[i] / left_empty_blocked_within[i])
                        else:
                            blocked_within_mean_l_empty_poke_time.append('N/A')
                        
                        if left_w_pellet_blocked_within[i] != 0:
                            blocked_within_mean_l_w_pellet_poke_time.append(blocked_within_sum_l_w_pellet_poke_time[i] / left_w_pellet_blocked_within[i])
                        else:
                            blocked_within_mean_l_w_pellet_poke_time.append('N/A')
                        
                        if right_all_blocked_within[i] != 0:
                            blocked_within_mean_r_all_poke_time.append(blocked_within_sum_r_all_poke_time[i] / right_all_blocked_within[i])
                        else:
                            blocked_within_mean_r_all_poke_time.append('N/A')
                        
                        if right_empty_blocked_within[i] != 0:
                            blocked_within_mean_r_empty_poke_time.append(blocked_within_sum_r_empty_poke_time[i] / right_empty_blocked_within[i])
                        else:
                            blocked_within_mean_r_empty_poke_time.append('N/A')
                        
                        if right_w_pellet_blocked_within[i] != 0:
                            blocked_within_mean_r_w_pellet_poke_time.append(blocked_within_sum_r_w_pellet_poke_time[i] / right_w_pellet_blocked_within[i])
                        else:
                            blocked_within_mean_r_w_pellet_poke_time.append('N/A')

                    #####-----                   
                    
                    # retrieval time and ipi blocks will just be the lists with the nans removed as there is one value for each block (FR step)

                    retrieval_time2 = [i if i!='Timed_out' else 0 for i in retrieval_time2]

                    retrieval_time2 = [float(i) for i in retrieval_time2]

                    retrieval_time_blocked = [x for x in retrieval_time2 if not math.isnan(x)] # gets rid of all the nan values
                    ipi_blocked = [x for x in ipi_2 if not math.isnan(x)] # gets rid of all the nan values
                    
                    # # First ipi value is nan need to add in NA to compensate
                    # ipi_blocked.insert(0, 'N/A')
                                       
                    # if final step is incomplete need to add N/A to end or retrieval_time_blocked and ipi_blocked
                    
                    if event[-1] != 'Pellet':
                        retrieval_time_blocked.append('N/A')
                        if pellet_count2[-1] != 0:
                            ipi_blocked.append('N/A')
                    
                    ####################################################################################################
                    ####################################################################################################

                    # Create chronological data that is within each PR step
                    
                    # create binary poke columns
                                
                    left_poke_all_binary = []
                    left_poke_empty_binary = []
                    left_poke_w_pellet_binary = []
                    
                    if left_all2[0] == 1:
                        left_poke_all_binary.append(1)
                    else:
                        left_poke_all_binary.append(0)
                    
                    for i in range(1, len(left_all2)):
                        if left_all2[i - 1] != left_all2[i]:
                            left_poke_all_binary.append(1)
                        else:
                            left_poke_all_binary.append(0)
                    
                    if left_empty2[0] == 1:
                        left_poke_empty_binary.append(1)
                    else:
                        left_poke_empty_binary.append(0)
                    
                    for i in range(1, len(left_empty2)):
                        if left_empty2[i - 1] != left_empty2[i]:
                            left_poke_empty_binary.append(1)
                        else:
                            left_poke_empty_binary.append(0)

                    if left_w_pellet2[0] == 1:
                        left_poke_w_pellet_binary.append(1)
                    else:
                        left_poke_w_pellet_binary.append(0)
                    
                    for i in range(1, len(left_w_pellet2)):
                        if left_w_pellet2[i - 1] != left_w_pellet2[i]:
                            left_poke_w_pellet_binary.append(1)
                        else:
                            left_poke_w_pellet_binary.append(0)

                    left_poke_all_chron_within = []
                    left_poke_all_chron_within_binary = []
                    left_poke_all_counter = 0
                    
                    for i in range (0, (len(left_poke_all_binary) - 1)):
                        
                        if session_type2[i + 1] == session_type2[i]:
                            if left_poke_all_binary[i] == 0:
                                left_poke_all_chron_within.append(left_poke_all_counter)
                                left_poke_all_chron_within_binary.append(np.nan)
                            elif left_poke_all_binary[i] == 1:
                                left_poke_all_counter += 1
                                left_poke_all_chron_within.append(left_poke_all_counter)
                                left_poke_all_chron_within_binary.append(left_poke_all_counter)
                            
                        elif session_type2[i + 1] != session_type2[i]:
                            if left_poke_all_binary[i] == 0:
                                left_poke_all_chron_within.append(left_poke_all_counter)
                                left_poke_all_chron_within_binary.append(np.nan)
                            elif left_poke_all_binary[i] == 1:
                                left_poke_all_counter += 1
                                left_poke_all_chron_within.append(left_poke_all_counter)
                                left_poke_all_chron_within_binary.append(left_poke_all_counter)
                            left_poke_all_counter = 0
                    
                    left_poke_all_chron_within.append(left_all_blocked_within[-1])
                    
                    if left_poke_all_chron_within[-1] != left_poke_all_chron_within[-2]:
                        left_poke_all_chron_within_binary.append(left_all_blocked_within[-1])
                    else:
                        left_poke_all_chron_within_binary.append(np.nan)
                    
                    left_poke_empty_chron_within = []
                    left_poke_empty_chron_within_binary = []
                    left_poke_empty_counter = 0
                    
                    for i in range (0, (len(left_poke_empty_binary) - 1)):
                        
                        if session_type2[i + 1] == session_type2[i]:
                            if left_poke_empty_binary[i] == 0:
                                left_poke_empty_chron_within.append(left_poke_empty_counter)
                                left_poke_empty_chron_within_binary.append(np.nan)
                            elif left_poke_empty_binary[i] == 1:
                                left_poke_empty_counter += 1
                                left_poke_empty_chron_within.append(left_poke_empty_counter)
                                left_poke_empty_chron_within_binary.append(left_poke_empty_counter)
                            
                        elif session_type2[i + 1] != session_type2[i]:
                            if left_poke_empty_binary[i] == 0:
                                left_poke_empty_chron_within.append(left_poke_empty_counter)
                                left_poke_empty_chron_within_binary.append(np.nan)
                            elif left_poke_empty_binary[i] == 1:
                                left_poke_empty_counter += 1
                                left_poke_empty_chron_within.append(left_poke_empty_counter)
                                left_poke_empty_chron_within_binary.append(left_poke_empty_counter)
                            left_poke_empty_counter = 0
                    
                    left_poke_empty_chron_within.append(left_empty_blocked_within[-1])
                    
                    if left_poke_empty_chron_within[-1] != left_poke_empty_chron_within[-2]:
                        left_poke_empty_chron_within_binary.append(left_empty_blocked_within[-1])
                    else:
                        left_poke_empty_chron_within_binary.append(np.nan)
                    
                    left_poke_w_pellet_chron_within = []
                    left_poke_w_pellet_chron_within_binary = []
                    left_poke_w_pellet_counter = 0
                    
                    for i in range (0, (len(left_poke_w_pellet_binary) - 1)):
                        
                        if session_type2[i + 1] == session_type2[i]:
                            if left_poke_w_pellet_binary[i] == 0:
                                left_poke_w_pellet_chron_within.append(left_poke_w_pellet_counter)
                                left_poke_w_pellet_chron_within_binary.append(np.nan)
                            elif left_poke_w_pellet_binary[i] == 1:
                                left_poke_w_pellet_counter += 1
                                left_poke_w_pellet_chron_within.append(left_poke_w_pellet_counter)
                                left_poke_w_pellet_chron_within_binary.append(left_poke_w_pellet_counter)
                            
                        elif session_type2[i + 1] != session_type2[i]:
                            if left_poke_w_pellet_binary[i] == 0:
                                left_poke_w_pellet_chron_within.append(left_poke_w_pellet_counter)
                                left_poke_w_pellet_chron_within_binary.append(np.nan)
                            elif left_poke_w_pellet_binary[i] == 1:
                                left_poke_w_pellet_counter += 1
                                left_poke_w_pellet_chron_within.append(left_poke_w_pellet_counter)
                                left_poke_w_pellet_chron_within_binary.append(left_poke_w_pellet_counter)
                            left_poke_w_pellet_counter = 0
                    
                    left_poke_w_pellet_chron_within.append(left_w_pellet_blocked_within[-1])
                    
                    if left_poke_w_pellet_chron_within[-1] != left_poke_w_pellet_chron_within[-2]:
                        left_poke_w_pellet_chron_within_binary.append(left_w_pellet_blocked_within[-1])
                    else:
                        left_poke_w_pellet_chron_within_binary.append(np.nan)

                    right_poke_all_binary = []
                    right_poke_empty_binary = []
                    right_poke_w_pellet_binary = []
                    
                    if right_all2[0] == 1:
                        right_poke_all_binary.append(1)
                    else:
                        right_poke_all_binary.append(0)
                    
                    for i in range(1, len(right_all2)):
                        if right_all2[i - 1] != right_all2[i]:
                            right_poke_all_binary.append(1)
                        else:
                            right_poke_all_binary.append(0)
                    
                    if right_empty2[0] == 1:
                        right_poke_empty_binary.append(1)
                    else:
                        right_poke_empty_binary.append(0)
                    
                    for i in range(1, len(right_empty2)):
                        if right_empty2[i - 1] != right_empty2[i]:
                            right_poke_empty_binary.append(1)
                        else:
                            right_poke_empty_binary.append(0)

                    if right_w_pellet2[0] == 1:
                        right_poke_w_pellet_binary.append(1)
                    else:
                        right_poke_w_pellet_binary.append(0)
                    
                    for i in range(1, len(right_w_pellet2)):
                        if right_w_pellet2[i - 1] != right_w_pellet2[i]:
                            right_poke_w_pellet_binary.append(1)
                        else:
                            right_poke_w_pellet_binary.append(0)

                    right_poke_all_chron_within = []
                    right_poke_all_chron_within_binary = []
                    right_poke_all_counter = 0
                    
                    for i in range (0, (len(right_poke_all_binary) - 1)):
                        
                        if session_type2[i + 1] == session_type2[i]:
                            if right_poke_all_binary[i] == 0:
                                right_poke_all_chron_within.append(right_poke_all_counter)
                                right_poke_all_chron_within_binary.append(np.nan)
                            elif right_poke_all_binary[i] == 1:
                                right_poke_all_counter += 1
                                right_poke_all_chron_within.append(right_poke_all_counter)
                                right_poke_all_chron_within_binary.append(right_poke_all_counter)
                            
                        elif session_type2[i + 1] != session_type2[i]:
                            if right_poke_all_binary[i] == 0:
                                right_poke_all_chron_within.append(right_poke_all_counter)
                                right_poke_all_chron_within_binary.append(np.nan)
                            elif right_poke_all_binary[i] == 1:
                                right_poke_all_counter += 1
                                right_poke_all_chron_within.append(right_poke_all_counter)
                                right_poke_all_chron_within_binary.append(right_poke_all_counter)
                            right_poke_all_counter = 0
                    
                    right_poke_all_chron_within.append(right_all_blocked_within[-1])
                    
                    if right_poke_all_chron_within[-1] != right_poke_all_chron_within[-2]:
                        right_poke_all_chron_within_binary.append(right_all_blocked_within[-1])
                    else:
                        right_poke_all_chron_within_binary.append(np.nan)
                    
                    right_poke_empty_chron_within = []
                    right_poke_empty_chron_within_binary = []
                    right_poke_empty_counter = 0
                    
                    for i in range (0, (len(right_poke_empty_binary) - 1)):
                        
                        if session_type2[i + 1] == session_type2[i]:
                            if right_poke_empty_binary[i] == 0:
                                right_poke_empty_chron_within.append(right_poke_empty_counter)
                                right_poke_empty_chron_within_binary.append(np.nan)
                            elif right_poke_empty_binary[i] == 1:
                                right_poke_empty_counter += 1
                                right_poke_empty_chron_within.append(right_poke_empty_counter)
                                right_poke_empty_chron_within_binary.append(right_poke_empty_counter)
                            
                        elif session_type2[i + 1] != session_type2[i]:
                            if right_poke_empty_binary[i] == 0:
                                right_poke_empty_chron_within.append(right_poke_empty_counter)
                                right_poke_empty_chron_within_binary.append(np.nan)
                            elif right_poke_empty_binary[i] == 1:
                                right_poke_empty_counter += 1
                                right_poke_empty_chron_within.append(right_poke_empty_counter)
                                right_poke_empty_chron_within_binary.append(right_poke_empty_counter)
                            right_poke_empty_counter = 0
                    
                    right_poke_empty_chron_within.append(right_empty_blocked_within[-1])
                    
                    if right_poke_empty_chron_within[-1] != right_poke_empty_chron_within[-2]:
                        right_poke_empty_chron_within_binary.append(right_empty_blocked_within[-1])
                    else:
                        right_poke_empty_chron_within_binary.append(np.nan)
                    
                    right_poke_w_pellet_chron_within = []
                    right_poke_w_pellet_chron_within_binary = []
                    right_poke_w_pellet_counter = 0
                    
                    for i in range (0, (len(right_poke_w_pellet_binary) - 1)):
                        
                        if session_type2[i + 1] == session_type2[i]:
                            if right_poke_w_pellet_binary[i] == 0:
                                right_poke_w_pellet_chron_within.append(right_poke_w_pellet_counter)
                                right_poke_w_pellet_chron_within_binary.append(np.nan)
                            elif right_poke_w_pellet_binary[i] == 1:
                                right_poke_w_pellet_counter += 1
                                right_poke_w_pellet_chron_within.append(right_poke_w_pellet_counter)
                                right_poke_w_pellet_chron_within_binary.append(right_poke_w_pellet_counter)
                            
                        elif session_type2[i + 1] != session_type2[i]:
                            if right_poke_w_pellet_binary[i] == 0:
                                right_poke_w_pellet_chron_within.append(right_poke_w_pellet_counter)
                                right_poke_w_pellet_chron_within_binary.append(np.nan)
                            elif right_poke_w_pellet_binary[i] == 1:
                                right_poke_w_pellet_counter += 1
                                right_poke_w_pellet_chron_within.append(right_poke_w_pellet_counter)
                                right_poke_w_pellet_chron_within_binary.append(right_poke_w_pellet_counter)
                            right_poke_w_pellet_counter = 0
                    
                    right_poke_w_pellet_chron_within.append(right_w_pellet_blocked_within[-1])
                    
                    if right_poke_w_pellet_chron_within[-1] != right_poke_w_pellet_chron_within[-2]:
                        right_poke_w_pellet_chron_within_binary.append(right_w_pellet_blocked_within[-1])
                    else:
                        right_poke_w_pellet_chron_within_binary.append(np.nan)
    
                    # create chronological within data for total pokes and percent active
                    
                    total_pokes_all_chron_within = []
                    all_percent_active_chron_within = []
                    
                    for i in range(0, len(left_poke_all_chron_within)):
                        total_pokes_all_chron_within.append(left_poke_all_chron_within[i] + right_poke_all_chron_within[i])
                        if active2[0] == 'Left':
                            if left_poke_all_chron_within[i] + right_poke_all_chron_within[i] != 0:
                                all_percent_active_chron_within.append((left_poke_all_chron_within[i] / (left_poke_all_chron_within[i] + right_poke_all_chron_within[i]) * 100))
                            else:
                                all_percent_active_chron_within.append('N/A')
                        elif active2[0] == 'Right':
                            if left_poke_all_chron_within[i] + right_poke_all_chron_within[i] != 0:
                                all_percent_active_chron_within.append((right_poke_all_chron_within[i] / (left_poke_all_chron_within[i] + right_poke_all_chron_within[i]) * 100))
                            else:
                                all_percent_active_chron_within.append('N/A')
    
                    total_pokes_empty_chron_within = []
                    empty_percent_active_chron_within = []
                    active_percent_empty_chron_within = []
                    
                    for i in range(0, len(left_poke_empty_chron_within)):
                        total_pokes_empty_chron_within.append(left_poke_empty_chron_within[i] + right_poke_empty_chron_within[i])
                        if active2[0] == 'Left':
                            if left_poke_empty_chron_within[i] + right_poke_empty_chron_within[i] != 0:
                                empty_percent_active_chron_within.append((left_poke_empty_chron_within[i] / (left_poke_empty_chron_within[i] + right_poke_empty_chron_within[i]) * 100))
                            else:
                                empty_percent_active_chron_within.append('N/A')        
                            if left_poke_all_chron_within[i] != 0:
                                active_percent_empty_chron_within.append(left_poke_empty_chron_within[i] / left_poke_all_chron_within[i] * 100)
                            else:
                                active_percent_empty_chron_within.append('N/A')
                        elif active2[0] == 'Right':
                            if left_poke_empty_chron_within[i] + right_poke_empty_chron_within[i] != 0:
                                empty_percent_active_chron_within.append((right_poke_empty_chron_within[i] / (left_poke_empty_chron_within[i] + right_poke_empty_chron_within[i]) * 100))
                            else:
                                empty_percent_active_chron_within.append('N/A')      
                            if right_poke_all_chron_within[i] != 0:
                                active_percent_empty_chron_within.append(right_poke_empty_chron_within[i] / right_poke_all_chron_within[i] * 100)
                            else:
                                active_percent_empty_chron_within.append('N/A')


                    total_pokes_w_pellet_chron_within = []
                    w_pellet_percent_active_chron_within = []
                    active_percent_w_pellet_chron_within = []
                    
                    for i in range(0, len(left_poke_w_pellet_chron_within)):
                        total_pokes_w_pellet_chron_within.append(left_poke_w_pellet_chron_within[i] + right_poke_w_pellet_chron_within[i])
                        if active2[0] == 'Left':
                            if left_poke_w_pellet_chron_within[i] + right_poke_w_pellet_chron_within[i] != 0:
                                w_pellet_percent_active_chron_within.append((left_poke_w_pellet_chron_within[i] / (left_poke_w_pellet_chron_within[i] + right_poke_w_pellet_chron_within[i]) * 100))
                            else:
                                w_pellet_percent_active_chron_within.append('N/A')
                            if left_poke_all_chron_within[i] != 0:
                                active_percent_w_pellet_chron_within.append(left_poke_w_pellet_chron_within[i] / left_poke_all_chron_within[i] * 100)
                            else:
                                active_percent_w_pellet_chron_within.append('N/A')
                        elif active2[0] == 'Right':
                            if left_poke_w_pellet_chron_within[i] + right_poke_w_pellet_chron_within[i] != 0:
                                w_pellet_percent_active_chron_within.append((right_poke_w_pellet_chron_within[i] / (left_poke_w_pellet_chron_within[i] + right_poke_w_pellet_chron_within[i]) * 100))
                            else:
                                w_pellet_percent_active_chron_within.append('N/A')
                            if right_poke_all_chron_within[i] != 0:
                                active_percent_w_pellet_chron_within.append(right_poke_w_pellet_chron_within[i] / right_poke_all_chron_within[i] * 100)
                            else:
                                active_percent_w_pellet_chron_within.append('N/A')

############################################
# TIME BINS
############################################

                    if len(seconds_in_bins) != 0:
                        for i in range(0, len(seconds_in_bins)):
                            end_time = seconds_elapsed[-1]
                                
                            # Create end value that is the next greatest multiple of the required time bin duration greater than the end_time so that those data points are retained in the output
                            
                            excess = math.ceil(end_time / int(seconds_in_bins[i])) # Rounds up so that the number of bins is the multiple of bin length greater than the time point
                            
                            end_bin = excess * int(seconds_in_bins[i]) # Creates the end value for the last time bin
                            
                            interval_range = pd.interval_range(start=0, freq=int(seconds_in_bins[i]), end=end_bin, closed="left") #'left' means that it includes the first data point which is 0 because it's start time
                            
                            time_bins = df[seconds_elapsed] = pd.cut(seconds_elapsed, interval_range, include_lowest=True, ordered=True)
                        
                            # Create dictionary of time bins as keys with bin number as matching entry to label the bins
                            
                            # Create a list of integers of the same length as there are time bins
                            
                            bin_number = []
                            a = 1
                            for i in range(0, len(interval_range)):
                                bin_number.append(a)
                                a += 1
                            
                            make_dictionary = zip(interval_range, bin_number)
                            
                            bin_dictionary = dict(make_dictionary)
                            
                            bin_num = []
                            for i in range(0, len(time_bins)):
                                bin_num.append(bin_dictionary.get((time_bins[i])))
                                                        
                            bin_num_binned = []
                            step_bin = []
                            seconds_elapsed_binned = []
                            duration_binned = []
                            left_all_binned = []
                            left_empty_binned = []
                            left_w_pellet_binned = []
                            right_all_binned = []
                            right_empty_binned = []
                            right_w_pellet_binned = []
                            pellet_binned = []
                            active_binned = []
                            all_percent_active_binned = []
                            empty_percent_active_binned = []
                            w_pellet_percent_active_binned = []
                            total_pokes_all_binned = []
                            total_pokes_empty_binned = []
                            total_pokes_w_pellet_binned = []
                            mag_check_binned = []
                            active_percent_empty_binned = []
                            active_percent_w_pellet_binned = []

                            for i in range(1, len(bin_num)):
                                if bin_num[i - 1] != bin_num[i]:
                                    bin_num_binned.append(bin_num[i - 1])
                                    step_bin.append(session_type2[i - 1])
                                    seconds_elapsed_binned.append(seconds_elapsed[i - 1])
                                    duration_binned.append(duration_from_start[i - 1])
                                    left_all_binned.append(left_empty2[i - 1] + left_w_pellet2[i - 1])
                                    left_empty_binned.append(left_empty2[i - 1])
                                    left_w_pellet_binned.append(left_w_pellet2[i - 1])
                                    right_all_binned.append(right_empty2[i - 1] + right_w_pellet2[i - 1])
                                    right_empty_binned.append(right_empty2[i - 1])
                                    right_w_pellet_binned.append(right_w_pellet2[i - 1])
                                    pellet_binned.append(pellet_count2[i - 1])
                                    active_binned.append(active2[i - 1])
                                    total_pokes_all_binned.append(total_pokes_all[i - 1])
                                    total_pokes_empty_binned.append(total_pokes_empty[i - 1])
                                    total_pokes_w_pellet_binned.append(total_pokes_w_pellet[i - 1])
                                    mag_check_binned.append(mag_check2[i - 1])
                            bin_num_binned.append(bin_num[i - 0])
                            step_bin.append(session_type2[i - 0])
                            seconds_elapsed_binned.append(seconds_elapsed[i - 0])
                            duration_binned.append(duration_from_start[i - 0])
                            left_all_binned.append(left_empty2[i - 0] + left_w_pellet2[i - 0])
                            left_empty_binned.append(left_empty2[i - 0])
                            left_w_pellet_binned.append(left_w_pellet2[i - 0])
                            right_all_binned.append(right_empty2[i - 0] + right_w_pellet2[i - 0])
                            right_empty_binned.append(right_empty2[i - 0])
                            right_w_pellet_binned.append(right_w_pellet2[i - 0])
                            pellet_binned.append(pellet_count2[i - 0])
                            active_binned.append(active2[i - 0])
                            total_pokes_all_binned.append(total_pokes_all[i - 0])
                            total_pokes_empty_binned.append(total_pokes_empty[i - 0])
                            total_pokes_w_pellet_binned.append(total_pokes_w_pellet[i - 0])
                            mag_check_binned.append(mag_check2[i - 0])

                            # percent active binned
                            for i in range(0, len(active_binned)):
                                if active_binned[i] == 'Left':
                                    all_percent_active_binned.append(left_all_binned[i] / total_pokes_all_binned[i] * 100)
                                    empty_percent_active_binned.append(left_empty_binned[i] / total_pokes_empty_binned[i] * 100)
                                    if total_pokes_w_pellet_binned[i] != 0:
                                        w_pellet_percent_active_binned.append(left_w_pellet_binned[i] / total_pokes_w_pellet_binned[i] * 100)
                                    else:
                                        w_pellet_percent_active_binned.append('N/A')
                                    active_percent_empty_binned.append(left_empty_binned[i] /left_all_binned[i] * 100)
                                    active_percent_w_pellet_binned.append(left_w_pellet_binned[i] / left_all_binned[i] * 100)
        
                                else:
                                    all_percent_active_binned.append(right_all_binned[i] / total_pokes_all_binned[i] * 100)
                                    empty_percent_active_binned.append(right_empty_binned[i] / total_pokes_empty_binned[i] * 100)
                                    if total_pokes_w_pellet_binned[i] != 0:
                                        w_pellet_percent_active_binned.append(right_w_pellet_binned[i] / total_pokes_w_pellet_binned[i] * 100)
                                    else:
                                        w_pellet_percent_active.binned.append('N/A')
                                    active_percent_empty_binned.append(right_empty_binned[i] /right_all_binned[i] * 100)
                                    active_percent_w_pellet_binned.append(right_w_pellet_binned[i] / right_all_binned[i] * 100)

                            
                            ##### LEADING EMPTY TIME BINS #####
                            
                            bin_counter = 1

                            while bin_num[0] != bin_counter: # fills in any empty leading bins until it gets to the bin with the first data point
                                bin_num_binned.insert(bin_counter - 1, bin_counter)
                                step_bin.insert(bin_counter - 1, 0)
                                seconds_elapsed_binned.insert(bin_counter - 1, 0)
                                duration_binned.insert(bin_counter - 1, 0)
                                left_all_binned.insert(bin_counter - 1, 0)
                                left_empty_binned.insert(bin_counter - 1, 0)
                                left_w_pellet_binned.insert(bin_counter - 1, 0)
                                right_all_binned.insert(bin_counter - 1, 0)
                                right_empty_binned.insert(bin_counter - 1, 0)
                                right_w_pellet_binned.insert(bin_counter - 1, 0)
                                pellet_binned.insert(bin_counter - 1, 0)
                                active_binned.insert(bin_counter - 1, active_binned[0])
                                all_percent_active_binned.insert(bin_counter - 1, np.nan)
                                empty_percent_active_binned.insert(bin_counter - 1, np.nan)
                                w_pellet_percent_active_binned.insert(bin_counter - 1, np.nan)
                                total_pokes_all_binned.insert(bin_counter - 1, 0)
                                total_pokes_empty_binned.insert(bin_counter - 1, 0)
                                total_pokes_w_pellet_binned.insert(bin_counter - 1, 0)
                                mag_check_binned.insert(bin_counter - 1, 0)
                                active_percent_empty_binned.insert(bin_counter - 1, np.nan)
                                active_percent_w_pellet_binned.insert(bin_counter - 1, np.nan)
                                bin_counter += 1
                            
                            # binned poke time, retrieval time and ipi data
                            
                            # Create mean poke and retrieval time columns for bins
                                        
                            binned_sum_both_all_poke_time = []
                            binned_sum_both_empty_poke_time = []
                            binned_sum_both_w_pellet_poke_time = []
                            binned_sum_l_all_poke_time = [] 
                            binned_sum_l_empty_poke_time = []
                            binned_sum_l_w_pellet_poke_time = []
                            binned_sum_r_all_poke_time = [] 
                            binned_sum_r_empty_poke_time = []
                            binned_sum_r_w_pellet_poke_time = []
                            
                            binned_mean_both_all_poke_time = [] 
                            binned_mean_both_empty_poke_time = []
                            binned_mean_both_w_pellet_poke_time = []
                            binned_mean_l_all_poke_time = [] 
                            binned_mean_l_empty_poke_time = []
                            binned_mean_l_w_pellet_poke_time = []
                            binned_mean_r_all_poke_time = [] 
                            binned_mean_r_empty_poke_time = []
                            binned_mean_r_w_pellet_poke_time = []
                            
                            sum_both_all_poke_time = 0
                            sum_both_empty_poke_time = 0
                            sum_both_w_pellet_poke_time = 0
                            sum_l_all_poke_time = 0
                            sum_l_empty_poke_time = 0
                            sum_l_w_pellet_poke_time = 0
                            sum_r_all_poke_time = 0
                            sum_r_empty_poke_time = 0
                            sum_r_w_pellet_poke_time = 0
                            
                            bin_index = 0
                            row_index = 0
                            
                            index = 1
                            
                            while bin_num[0] > index: # if there are leading empty time bins need to fill them in
                                binned_sum_both_all_poke_time.append(0)
                                binned_sum_both_empty_poke_time.append(0)
                                binned_sum_both_w_pellet_poke_time.append(0)
                                binned_sum_l_all_poke_time.append(0) 
                                binned_sum_l_empty_poke_time.append(0)
                                binned_sum_l_w_pellet_poke_time.append(0)
                                binned_sum_r_all_poke_time.append(0) 
                                binned_sum_r_empty_poke_time.append(0)
                                binned_sum_r_w_pellet_poke_time.append(0)
                                
                                index += 1

                            # print(seconds_elapsed, seconds_elapsed_binned, left_poke2, poke_time2)
                            # print(len(seconds_elapsed), len(poke_time2), len(seconds_elapsed_binned))
                            for i in range(0, len(poke_time2)):
                                if int(seconds_elapsed[i]) < int(seconds_elapsed_binned[bin_index]):
                                    
                                    if event2[i] != 'MagCheck':
                                        sum_both_all_poke_time += poke_time2[row_index]
                                    if event2[i] == 'Left':
                                        sum_l_all_poke_time += poke_time2[row_index]
                                        sum_l_empty_poke_time += poke_time2[row_index]
                                        sum_both_empty_poke_time += poke_time2[row_index]
                                    elif event2[i] == 'Right':
                                        sum_r_all_poke_time += poke_time2[row_index]
                                        sum_r_empty_poke_time += poke_time2[row_index]
                                        sum_both_empty_poke_time += poke_time2[row_index]
                                    elif event2[i] == 'LeftWithPellet':
                                        sum_l_all_poke_time += poke_time2[row_index]
                                        sum_l_w_pellet_poke_time += poke_time2[row_index]
                                        sum_both_w_pellet_poke_time += poke_time2[row_index]
                                    elif event2[i] == 'RightWithPellet':
                                        sum_r_all_poke_time += poke_time2[row_index]
                                        sum_r_w_pellet_poke_time += poke_time2[row_index]
                                        sum_both_w_pellet_poke_time += poke_time2[row_index]
                                    row_index += 1
                                    
                                elif int(seconds_elapsed[i]) == int(seconds_elapsed_binned[bin_index]):
                                    if event2[i] != 'MagCheck':
                                        sum_both_all_poke_time += poke_time2[row_index]
                                    if event2[i] == 'Left':
                                        sum_l_all_poke_time += poke_time2[row_index]
                                        sum_l_empty_poke_time += poke_time2[row_index]
                                        sum_both_empty_poke_time += poke_time2[row_index]
                                    elif event2[i] == 'Right':
                                        sum_r_all_poke_time += poke_time2[row_index]
                                        sum_r_empty_poke_time += poke_time2[row_index]
                                        sum_both_empty_poke_time += poke_time2[row_index]
                                    elif event2[i] == 'LeftWithPellet':
                                        sum_l_all_poke_time += poke_time2[row_index]
                                        sum_l_w_pellet_poke_time += poke_time2[row_index]
                                        sum_both_w_pellet_poke_time += poke_time2[row_index]
                                    elif event2[i] == 'RightWithPellet':
                                        sum_r_all_poke_time += poke_time2[row_index]
                                        sum_r_w_pellet_poke_time += poke_time2[row_index]
                                        sum_both_w_pellet_poke_time += poke_time2[row_index]
                                    row_index += 1
                                    
                                    binned_sum_both_all_poke_time.append(sum_both_all_poke_time) 
                                    binned_sum_both_empty_poke_time.append(sum_both_empty_poke_time)
                                    binned_sum_both_w_pellet_poke_time.append(sum_both_w_pellet_poke_time)
                                    binned_sum_l_all_poke_time.append(sum_l_all_poke_time) 
                                    binned_sum_l_empty_poke_time.append(sum_l_empty_poke_time)
                                    binned_sum_l_w_pellet_poke_time.append(sum_l_w_pellet_poke_time)
                                    binned_sum_r_all_poke_time.append(sum_r_all_poke_time) 
                                    binned_sum_r_empty_poke_time.append(sum_r_empty_poke_time)
                                    binned_sum_r_w_pellet_poke_time.append(sum_r_w_pellet_poke_time)
                                    
                                    binned_mean_both_all_poke_time.append(sum_both_all_poke_time / total_pokes_all_binned[bin_index]) 
                                    binned_mean_both_empty_poke_time.append(sum_both_empty_poke_time / total_pokes_empty_binned[bin_index])
                                    if total_pokes_w_pellet_binned[bin_index] != 0:
                                        binned_mean_both_w_pellet_poke_time.append(sum_both_w_pellet_poke_time / total_pokes_w_pellet_binned[bin_index])
                                    else:
                                        binned_mean_both_w_pellet_poke_time.append('N/A')
                                    
                                    if left_all_binned[bin_index] != 0:
                                        binned_mean_l_all_poke_time.append(sum_l_all_poke_time / left_all_binned[bin_index]) 
                                    else:
                                        binned_mean_l_all_poke_time.append('N/A')
                                    if left_empty_binned[bin_index] != 0:
                                        binned_mean_l_empty_poke_time.append(sum_l_empty_poke_time / left_empty_binned[bin_index])
                                    else:
                                        binned_mean_l_empty_poke_time.append('N/A')
                                    if left_w_pellet_binned[bin_index] != 0:
                                        binned_mean_l_w_pellet_poke_time.append(sum_l_w_pellet_poke_time / left_w_pellet_binned[bin_index])
                                    else:
                                        binned_mean_l_w_pellet_poke_time.append('N/A')
                                    
                                    if right_all_binned[bin_index] != 0:
                                        binned_mean_r_all_poke_time.append(sum_r_all_poke_time / right_all_binned[bin_index]) 
                                    else:
                                        binned_mean_r_all_poke_time.append('N/A')
                                    if right_empty_binned[bin_index] != 0:
                                        binned_mean_r_empty_poke_time.append(sum_r_empty_poke_time / right_empty_binned[bin_index])
                                    else:
                                        binned_mean_r_empty_poke_time.append('N/A')
                                    if right_w_pellet_binned[bin_index] != 0:
                                        binned_mean_r_w_pellet_poke_time.append(sum_r_w_pellet_poke_time / right_w_pellet_binned[bin_index])
                                    else:
                                        binned_mean_r_w_pellet_poke_time.append('N/A')
                                    
                                    bin_index += 1
                                
                                elif int(seconds_elapsed[i]) > int(seconds_elapsed_binned[bin_index]):
                                    
                                    binned_sum_both_all_poke_time.append(sum_both_all_poke_time) 
                                    binned_sum_both_empty_poke_time.append(sum_both_empty_poke_time)
                                    binned_sum_both_w_pellet_poke_time.append(sum_both_w_pellet_poke_time)
                                    binned_sum_l_all_poke_time.append(sum_l_all_poke_time) 
                                    binned_sum_l_empty_poke_time.append(sum_l_empty_poke_time)
                                    binned_sum_l_w_pellet_poke_time.append(sum_l_w_pellet_poke_time)
                                    binned_sum_r_all_poke_time.append(sum_r_all_poke_time) 
                                    binned_sum_r_empty_poke_time.append(sum_r_empty_poke_time)
                                    binned_sum_r_w_pellet_poke_time.append(sum_r_w_pellet_poke_time)
                                    
                                    binned_mean_both_all_poke_time.append(sum_both_all_poke_time / total_pokes_all_binned[bin_index]) 
                                    binned_mean_both_empty_poke_time.append(sum_both_empty_poke_time / total_pokes_empty_binned[bin_index])
                                    binned_mean_both_w_pellet_poke_time.append(sum_both_w_pellet_poke_time / total_pokes_w_pellet_binned[bin_index])
                                    
                                    if left_all_binned[bin_index] != 0:
                                        binned_mean_l_all_poke_time.append(sum_l_all_poke_time / left_all_binned[bin_index]) 
                                    else:
                                        binned_mean_l_all_poke_time.append('N/A')
                                    if left_empty_binned[bin_index] != 0:
                                        binned_mean_l_empty_poke_time.append(sum_l_empty_poke_time / left_empty_binned[bin_index])
                                    else:
                                        binned_mean_l_empty_poke_time.append('N/A')
                                    if left_w_pellet_binned[bin_index] != 0:
                                        binned_mean_l_w_pellet_poke_time.append(sum_l_w_pellet_poke_time / left_w_pellet_binned[bin_index])
                                    else:
                                        binned_mean_l_w_pellet_poke_time.append('N/A')
                                    
                                    if right_all_binned[bin_index] != 0:
                                        binned_mean_r_all_poke_time.append(sum_r_all_poke_time / right_all_binned[bin_index]) 
                                    else:
                                        binned_mean_r_all_poke_time.append('N/A')
                                    if right_empty_binned[bin_index] != 0:
                                        binned_mean_r_empty_poke_time.append(sum_r_empty_poke_time / right_empty_binned[bin_index])
                                    else:
                                        binned_mean_r_empty_poke_time.append('N/A')
                                    if right_w_pellet_binned[bin_index] != 0:
                                        binned_mean_r_w_pellet_poke_time.append(sum_r_w_pellet_poke_time / right_w_pellet_binned[bin_index])
                                    else:
                                        binned_mean_r_w_pellet_poke_time.append('N/A')
                                    
                                    bin_index += 1
                                    
                                    if event2[i] != 'MagCheck':
                                        sum_both_all_poke_time += poke_time2[row_index]
                                    if event2[i] == 'Left':
                                        sum_l_all_poke_time += poke_time2[row_index]
                                        sum_l_empty_poke_time += poke_time2[row_index]
                                        sum_both_empty_poke_time += poke_time2[row_index]
                                    elif event2[i] == 'Right':
                                        sum_r_all_poke_time += poke_time2[row_index]
                                        sum_r_empty_poke_time += poke_time2[row_index]
                                        sum_both_empty_poke_time += poke_time2[row_index]
                                    elif event2[i] == 'LeftWithPellet':
                                        sum_l_all_poke_time += poke_time2[row_index]
                                        sum_l_w_pellet_poke_time += poke_time2[row_index]
                                        sum_both_w_pellet_poke_time += poke_time2[row_index]
                                    elif event2[i] == 'RightWithPellet':
                                        sum_r_all_poke_time += poke_time2[row_index]
                                        sum_r_w_pellet_poke_time += poke_time2[row_index]
                                        sum_both_w_pellet_poke_time += poke_time2[row_index]
                                    row_index += 1
        
                            
                            #####-----  
                            
                            # Enter new row for any bins that have no data in them (for cumulative bins this is a duplication of the preceding row)

                            for i in bin_dictionary:
                                if bin_dictionary[i] in bin_num_binned:
                                    continue
                                else:
                                    bin_num_binned.insert((bin_dictionary[i] - 1), bin_dictionary[i])
                                    step_bin.insert((bin_dictionary[i] - 1), step_bin[(bin_dictionary[i] - 2)])
                                    seconds_elapsed_binned.insert((bin_dictionary[i] - 1), seconds_elapsed_binned[(bin_dictionary[i] - 2)])
                                    duration_binned.insert((bin_dictionary[i] - 1), duration_binned[(bin_dictionary[i] - 2)])
                                    left_all_binned.insert((bin_dictionary[i] - 1), left_all_binned[(bin_dictionary[i] - 2)])
                                    left_empty_binned.insert((bin_dictionary[i] - 1), left_empty_binned[(bin_dictionary[i] - 2)])
                                    left_w_pellet_binned.insert((bin_dictionary[i] - 1), left_w_pellet_binned[(bin_dictionary[i] - 2)])
                                    right_all_binned.insert((bin_dictionary[i] - 1), right_all_binned[(bin_dictionary[i] - 2)])
                                    right_empty_binned.insert((bin_dictionary[i] - 1), right_empty_binned[(bin_dictionary[i] - 2)])
                                    right_w_pellet_binned.insert((bin_dictionary[i] - 1), right_w_pellet_binned[(bin_dictionary[i] - 2)])
                                    pellet_binned.insert((bin_dictionary[i] - 1), pellet_binned[(bin_dictionary[i] - 2)])
                                    active_binned.insert((bin_dictionary[i] - 1), active_binned[(bin_dictionary[i] - 2)])
                                    all_percent_active_binned.insert((bin_dictionary[i] - 1), all_percent_active_binned[(bin_dictionary[i] - 2)])
                                    empty_percent_active_binned.insert((bin_dictionary[i] - 1), empty_percent_active_binned[(bin_dictionary[i] - 2)])
                                    w_pellet_percent_active_binned.insert((bin_dictionary[i] - 1), w_pellet_percent_active_binned[(bin_dictionary[i] - 2)])
                                    total_pokes_all_binned.insert((bin_dictionary[i] - 1), total_pokes_all_binned[(bin_dictionary[i] - 2)])
                                    total_pokes_empty_binned.insert((bin_dictionary[i] - 1), total_pokes_empty_binned[(bin_dictionary[i] - 2)])
                                    total_pokes_w_pellet_binned.insert((bin_dictionary[i] - 1), total_pokes_w_pellet_binned[(bin_dictionary[i] - 2)])
                                    mag_check_binned.insert((bin_dictionary[i] - 1), mag_check_binned[(bin_dictionary[i] - 2)])
                                    active_percent_empty_binned.insert((bin_dictionary[i] - 1), active_percent_empty_binned[(bin_dictionary[i] - 2)])
                                    active_percent_w_pellet_binned.insert((bin_dictionary[i] - 1), active_percent_w_pellet_binned[(bin_dictionary[i] - 2)])
                                    
                                    binned_sum_both_all_poke_time.insert((bin_dictionary[i] - 1), binned_sum_both_all_poke_time[(bin_dictionary[i] - 2)])
                                    binned_sum_both_empty_poke_time.insert((bin_dictionary[i] - 1), binned_sum_both_empty_poke_time[(bin_dictionary[i] - 2)])
                                    binned_sum_both_w_pellet_poke_time.insert((bin_dictionary[i] - 1), binned_sum_both_w_pellet_poke_time[(bin_dictionary[i] - 2)])
                                    binned_sum_l_all_poke_time.insert((bin_dictionary[i] - 1), binned_sum_l_all_poke_time[(bin_dictionary[i] - 2)]) 
                                    binned_sum_l_empty_poke_time.insert((bin_dictionary[i] - 1), binned_sum_l_empty_poke_time[(bin_dictionary[i] - 2)])
                                    binned_sum_l_w_pellet_poke_time.insert((bin_dictionary[i] - 1), binned_sum_l_w_pellet_poke_time[(bin_dictionary[i] - 2)])
                                    binned_sum_r_all_poke_time.insert((bin_dictionary[i] - 1), binned_sum_r_all_poke_time[(bin_dictionary[i] - 2)]) 
                                    binned_sum_r_empty_poke_time.insert((bin_dictionary[i] - 1), binned_sum_r_empty_poke_time[(bin_dictionary[i] - 2)])
                                    binned_sum_r_w_pellet_poke_time.insert((bin_dictionary[i] - 1), binned_sum_r_w_pellet_poke_time[(bin_dictionary[i] - 2)])
                                    
                                    binned_mean_both_all_poke_time.insert((bin_dictionary[i] - 1), binned_mean_both_all_poke_time[(bin_dictionary[i] - 2)]) 
                                    binned_mean_both_empty_poke_time.insert((bin_dictionary[i] - 1), binned_mean_both_empty_poke_time[(bin_dictionary[i] - 2)])
                                    binned_mean_both_w_pellet_poke_time.insert((bin_dictionary[i] - 1), binned_mean_both_w_pellet_poke_time[(bin_dictionary[i] - 2)])
                                    binned_mean_l_all_poke_time.insert((bin_dictionary[i] - 1), binned_mean_l_all_poke_time[(bin_dictionary[i] - 2)]) 
                                    binned_mean_l_empty_poke_time.insert((bin_dictionary[i] - 1), binned_mean_l_empty_poke_time[(bin_dictionary[i] - 2)])
                                    binned_mean_l_w_pellet_poke_time.insert((bin_dictionary[i] - 1), binned_mean_l_w_pellet_poke_time[(bin_dictionary[i] - 2)])
                                    binned_mean_r_all_poke_time.insert((bin_dictionary[i] - 1), binned_mean_r_all_poke_time[(bin_dictionary[i] - 2)]) 
                                    binned_mean_r_empty_poke_time.insert((bin_dictionary[i] - 1), binned_mean_r_empty_poke_time[(bin_dictionary[i] - 2)])
                                    binned_mean_r_w_pellet_poke_time.insert((bin_dictionary[i] - 1), binned_mean_r_w_pellet_poke_time[(bin_dictionary[i] - 2)])

                            # Create new bins where counts are within bins rather than cumulative
                    
                            seconds_elapsed_binned_within = []
                            left_all_binned_within = []
                            left_empty_binned_within = []
                            left_w_pellet_binned_within = []
                            right_all_binned_within = []
                            right_empty_binned_within = []
                            right_w_pellet_binned_within = []
                            pellet_binned_within = []
                            all_percent_active_binned_within = []
                            empty_percent_active_binned_within = []
                            w_pellet_percent_active_binned_within = []
                            total_pokes_all_binned_within = []
                            total_pokes_empty_binned_within = []
                            total_pokes_w_pellet_binned_within = []
                            mag_check_binned_within = []
                            active_percent_empty_binned_within = []
                            active_percent_w_pellet_binned_within = []                   
        
                            seconds_elapsed_binned_within.append(seconds_elapsed_binned[0])
                            for i in range(1, len(seconds_elapsed_binned)):
                                seconds_elapsed_binned_within.append(seconds_elapsed_binned[i] - seconds_elapsed_binned[i - 1])
                            
                            # Transform seconds_elapsed into duration in h:mm:ss
                                        
                            duration_binned_within = []
                            
                            for i in range(0, len(seconds_elapsed_binned_within)):
                                time_hours = str(math.floor(seconds_elapsed_binned_within[i] / 3600))
                                
                                time_mins = math.floor((seconds_elapsed_binned_within[i] % 3600) / 60)
                                
                                time_secs = seconds_elapsed_binned_within[i] % 60
                                    
                                if time_mins < 10:
                                    t_mins = '0' + str(time_mins)
                                else:
                                    t_mins = str(time_mins)
                                    
                                if time_secs < 10:
                                    t_secs = '0' + str(time_secs)
                                else:
                                    t_secs = str(time_secs)
                                    
                                duration_binned_within.append(str(time_hours) + ':' + t_mins + ':' + t_secs)
                            
                            left_all_binned_within.append(left_all_binned[0])
                            left_empty_binned_within.append(left_empty_binned[0])
                            left_w_pellet_binned_within.append(left_w_pellet_binned[0])
                            right_all_binned_within.append(right_all_binned[0])
                            right_empty_binned_within.append(right_empty_binned[0])
                            right_w_pellet_binned_within.append(right_w_pellet_binned[0])
                            pellet_binned_within.append(pellet_binned[0])
                            total_pokes_all_binned_within.append(total_pokes_all_binned[0])
                            total_pokes_empty_binned_within.append(total_pokes_empty_binned[0])
                            total_pokes_w_pellet_binned_within.append(total_pokes_w_pellet_binned[0])
                            mag_check_binned_within.append(mag_check_binned[0])

                            for i in range(1, len(active_binned)):
                                left_all_binned_within.append(left_all_binned[i] - left_all_binned[i - 1])
                                left_empty_binned_within.append(left_empty_binned[i] - left_empty_binned[i - 1])
                                left_w_pellet_binned_within.append(left_w_pellet_binned[i] - left_w_pellet_binned[i - 1])
                                right_all_binned_within.append(right_all_binned[i] - right_all_binned[i - 1])
                                right_empty_binned_within.append(right_empty_binned[i] - right_empty_binned[i - 1])
                                right_w_pellet_binned_within.append(right_w_pellet_binned[i] - right_w_pellet_binned[i - 1])
                                pellet_binned_within.append(pellet_binned[i] - pellet_binned[i - 1])
                                total_pokes_all_binned_within.append(total_pokes_all_binned[i] - total_pokes_all_binned[i - 1])
                                total_pokes_empty_binned_within.append(total_pokes_empty_binned[i] - total_pokes_empty_binned[i - 1])
                                total_pokes_w_pellet_binned_within.append(total_pokes_w_pellet_binned[i] - total_pokes_w_pellet_binned[i - 1])
                                mag_check_binned_within.append(mag_check_binned[i] - mag_check_binned[i - 1])
                            
                            for i in range(0, len(active_binned)):
                                if active_binned[i] == 'Left':
                                    if total_pokes_all_binned_within[i] != 0:
                                        all_percent_active_binned_within.append(left_all_binned_within[i] / total_pokes_all_binned_within[i] * 100)
                                    else:
                                        all_percent_active_binned_within.append('NA')
                                    if total_pokes_empty_binned_within[i] != 0:
                                        empty_percent_active_binned_within.append(left_empty_binned_within[i] / total_pokes_empty_binned_within[i] * 100)
                                    else:
                                        empty_percent_active_binned_within.append('NA')
                                    if total_pokes_w_pellet_binned_within[i] != 0:
                                        w_pellet_percent_active_binned_within.append(left_w_pellet_binned_within[i] / total_pokes_w_pellet_binned_within[i] * 100)
                                    else:
                                        w_pellet_percent_active_binned_within.append('NA')
                                    if left_all_binned_within[i] != 0:
                                        active_percent_empty_binned_within.append(left_empty_binned_within[i] / left_all_binned_within[i] * 100)
                                        active_percent_w_pellet_binned_within.append(left_w_pellet_binned_within[i] / left_all_binned_within[i] * 100)
                                    else:
                                        active_percent_empty_binned_within.append('N/A')
                                        active_percent_w_pellet_binned_within.append('N/A')
        
                                else:
                                    if total_pokes_all_binned_within[i] != 0:
                                        all_percent_active_binned_within.append(right_all_binned_within[i] / total_pokes_all_binned_within[i] * 100)
                                    else:
                                        all_percent_active_binned_within.append('NA')
                                    if total_pokes_empty_binned_within[i] != 0:
                                        empty_percent_active_binned_within.append(right_empty_binned_within[i] / total_pokes_empty_binned_within[i] * 100)
                                    else:
                                        empty_percent_active_binned_within.append('NA')
                                    if total_pokes_w_pellet_binned_within[i] != 0:
                                        w_pellet_percent_active_binned_within.append(right_w_pellet_binned_within[i] / total_pokes_w_pellet_binned_within[i] * 100)
                                    else:
                                        w_pellet_percent_active_binned_within.append('NA')    
                                    if right_all_binned_within[i] != 0:
                                        active_percent_empty_binned_within.append(right_empty_binned_within[i] / right_all_binned_within[i] * 100)
                                        active_percent_w_pellet_binned_within.append(right_w_pellet_binned_within[i] / right_all_binned_within[i] * 100)
                                    else:
                                        active_percent_empty_binned_within.append('N/A')
                                        active_percent_w_pellet_binned_within.append('N/A')

                            binned_within_sum_both_all_poke_time = [] 
                            binned_within_sum_both_empty_poke_time = []
                            binned_within_sum_both_w_pellet_poke_time = []
                            binned_within_sum_l_all_poke_time = [] 
                            binned_within_sum_l_empty_poke_time = []
                            binned_within_sum_l_w_pellet_poke_time = []
                            binned_within_sum_r_all_poke_time = [] 
                            binned_within_sum_r_empty_poke_time = []
                            binned_within_sum_r_w_pellet_poke_time = []
                            
                            binned_within_sum_both_all_poke_time.append(binned_sum_both_all_poke_time[0]) 
                            binned_within_sum_both_empty_poke_time.append(binned_sum_both_empty_poke_time[0])
                            binned_within_sum_both_w_pellet_poke_time.append(binned_sum_both_w_pellet_poke_time[0])
                            binned_within_sum_l_all_poke_time.append(binned_sum_l_all_poke_time[0]) 
                            binned_within_sum_l_empty_poke_time.append(binned_sum_l_empty_poke_time[0])
                            binned_within_sum_l_w_pellet_poke_time.append(binned_sum_l_w_pellet_poke_time[0])
                            binned_within_sum_r_all_poke_time.append(binned_sum_r_all_poke_time[0])
                            binned_within_sum_r_empty_poke_time.append(binned_sum_r_empty_poke_time[0])
                            binned_within_sum_r_w_pellet_poke_time.append(binned_sum_r_w_pellet_poke_time[0])
                                               
                            for i in range(1, len(binned_sum_both_all_poke_time)):
                                binned_within_sum_both_all_poke_time.append(binned_sum_both_all_poke_time[i] - binned_sum_both_all_poke_time[i - 1]) 
                                binned_within_sum_both_empty_poke_time.append(binned_sum_both_empty_poke_time[i] - binned_sum_both_empty_poke_time[i - 1])
                                binned_within_sum_both_w_pellet_poke_time.append(binned_sum_both_w_pellet_poke_time[i] - binned_sum_both_w_pellet_poke_time[i - 1])
                                binned_within_sum_l_all_poke_time.append(binned_sum_l_all_poke_time[i] - binned_sum_l_all_poke_time[i - 1])
                                binned_within_sum_l_empty_poke_time.append(binned_sum_l_empty_poke_time[i] - binned_sum_l_empty_poke_time[i - 1])
                                binned_within_sum_l_w_pellet_poke_time.append(binned_sum_l_w_pellet_poke_time[i] - binned_sum_l_w_pellet_poke_time[i - 1])
                                binned_within_sum_r_all_poke_time.append(binned_sum_r_all_poke_time[i] - binned_sum_r_all_poke_time[i - 1])
                                binned_within_sum_r_empty_poke_time.append(binned_sum_r_empty_poke_time[i] - binned_sum_r_empty_poke_time[i] - 1)
                                binned_within_sum_r_w_pellet_poke_time.append(binned_sum_r_w_pellet_poke_time[i] - binned_sum_r_w_pellet_poke_time[i] - 1)
                            
                            binned_within_mean_both_all_poke_time = [] 
                            binned_within_mean_both_empty_poke_time = []
                            binned_within_mean_both_w_pellet_poke_time = []
                            binned_within_mean_l_all_poke_time = [] 
                            binned_within_mean_l_empty_poke_time = []
                            binned_within_mean_l_w_pellet_poke_time = []
                            binned_within_mean_r_all_poke_time = [] 
                            binned_within_mean_r_empty_poke_time = []
                            binned_within_mean_r_w_pellet_poke_time = []
        
                            for i in range(0, len(binned_within_sum_both_all_poke_time)):
                                if total_pokes_all_binned_within[i] != 0:
                                    binned_within_mean_both_all_poke_time.append(binned_within_sum_both_all_poke_time[i] / total_pokes_all_binned_within[i])
                                else:
                                    binned_within_mean_both_all_poke_time.append('N/A')
                                if total_pokes_empty_binned_within[i] != 0:
                                    binned_within_mean_both_empty_poke_time.append(binned_within_sum_both_empty_poke_time[i] / total_pokes_empty_binned_within[i])
                                else:
                                    binned_within_mean_both_empty_poke_time.append('N/A')
                                if total_pokes_w_pellet_binned_within[i] != 0:
                                    binned_within_mean_both_w_pellet_poke_time.append(binned_within_sum_both_w_pellet_poke_time[i] / total_pokes_w_pellet_binned_within[i])
                                else:
                                    binned_within_mean_both_w_pellet_poke_time.append('N/A')
                                
                                if left_all_binned_within[i] != 0:
                                    binned_within_mean_l_all_poke_time.append(binned_within_sum_l_all_poke_time[i] / left_all_binned_within[i])
                                else:
                                    binned_within_mean_l_all_poke_time.append('N/A')
                                
                                if left_empty_binned_within[i] != 0:
                                    binned_within_mean_l_empty_poke_time.append(binned_within_sum_l_empty_poke_time[i] / left_empty_binned_within[i])
                                else:
                                    binned_within_mean_l_empty_poke_time.append('N/A')
                                
                                if left_w_pellet_binned_within[i] != 0:
                                    binned_within_mean_l_w_pellet_poke_time.append(binned_within_sum_l_w_pellet_poke_time[i] / left_w_pellet_binned_within[i])
                                else:
                                    binned_within_mean_l_w_pellet_poke_time.append('N/A')
                                
                                if right_all_binned_within[i] != 0:
                                    binned_within_mean_r_all_poke_time.append(binned_within_sum_r_all_poke_time[i] / right_all_binned_within[i])
                                else:
                                    binned_within_mean_r_all_poke_time.append('N/A')
                                
                                if right_empty_binned_within[i] != 0:
                                    binned_within_mean_r_empty_poke_time.append(binned_within_sum_r_empty_poke_time[i] / right_empty_binned_within[i])
                                else:
                                    binned_within_mean_r_empty_poke_time.append('N/A')
                                
                                if right_w_pellet_binned_within[i] != 0:
                                    binned_within_mean_r_w_pellet_poke_time.append(binned_within_sum_r_w_pellet_poke_time[i] / right_w_pellet_binned_within[i])
                                else:
                                    binned_within_mean_r_w_pellet_poke_time.append('N/A')
                    

############################################
############################################

#####----- Create session summary data-----#####
                    
                    task = schedule[0]
                    
                    session_duration = duration_from_start[-1]
                        
                    # Assign left and right pokes as active or inactive based on the active port
                    
                    if active2[0] == 'Left':
                        active_pokes_all = left_all2[-1]
                        active_pokes_empty = left_empty2[-1]
                        active_pokes_w_pellet = left_w_pellet2[-1]
                        inactive_pokes_all = right_all2[-1]
                        inactive_pokes_empty = right_empty2[-1]
                        inactive_pokes_w_pellet = right_w_pellet2[-1]
                    elif active2[0] == 'Right':
                        active_pokes_all = right_all2[-1]
                        active_pokes_empty = right_empty2[-1]
                        active_pokes_w_pellet = right_w_pellet2[-1]
                        inactive_pokes_all = left_all2[-1]
                        inactive_pokes_empty = left_empty2[-1]
                        inactive_pokes_w_pellet = left_w_pellet2[-1]
                    
                    active_port = active2[0]
                    
                    date = import_name[7:13]
                    aus_date = date[2:4] + '/' + date[0:2] + '/20' + date[4:]
                    
                    mag_checks = mag_check2[-1]
                    
                    variable = ['Filename', 'Date', 'Task', 'Duration', 'Active port', 
                                'ALL Total Pokes', 'ALL Active Pokes', 'ALL Inactive Pokes', 'ALL Pokes % Active',
                                'Empty Total Pokes', 'Empty Active Pokes', 'Empty Inactive Pokes', 'Empty Pokes % Active',
                                'W Pellet Total Pokes', 'W Pellet Active Pokes', 'W Pellet Inactive Pokes', 'W Pellet Pokes % Active', 
                                'Active Pokes % Empty', 'Active Pokes % W Pellet', 'Pellets', 'Magazine Checks']
                    
                    value = [import_name.strip('.CSV'), aus_date, task, session_duration, active_port, 
                             total_pokes_all[-1], active_pokes_all, inactive_pokes_all, all_percent_active[-1],
                             total_pokes_empty[-1], active_pokes_empty, inactive_pokes_empty, empty_percent_active[-1],
                             total_pokes_w_pellet[-1], active_pokes_w_pellet, inactive_pokes_w_pellet, w_pellet_percent_active[-1], 
                             active_percent_empty[-1], active_percent_w_pellet[-1], pellet_count2[-1], mag_checks]

                    # for i in range(0, len(variable)):
                    #     print(variable[i], value[i])
                    
                    Date_summary.append(aus_date)
                    Task_summary.append(task)
                    Duration_summary.append(session_duration)
                    Active_port_summary.append(active_port)
                    ALL_Total_Pokes_summary.append(total_pokes_all[-1])
                    ALL_Active_Pokes_summary.append(active_pokes_all)
                    ALL_Inactive_Pokes_summary.append(inactive_pokes_all)
                    ALL_Pokes_Percent_Active_summary.append(all_percent_active[-1])
                    Empty_Total_Pokes_summary.append(total_pokes_empty[-1])
                    Empty_Active_Pokes_summary.append(active_pokes_empty)
                    Empty_Inactive_Pokes_summary.append(inactive_pokes_empty)
                    Empty_Pokes_Percent_Active_summary.append(empty_percent_active[-1])
                    W_Pellet_Total_Pokes_summary.append(total_pokes_w_pellet[-1])
                    W_Pellet_Active_Pokes_summary.append(active_pokes_w_pellet)
                    W_Pellet_Inactive_Pokes_summary.append(inactive_pokes_w_pellet)
                    W_Pellet_Pokes_Percent_Active_summary.append(w_pellet_percent_active[-1])
                    Active_Pokes_Percent_Empty_summary.append(active_percent_empty[-1])
                    Active_Pokes_Percent_W_Pellet_summary.append(active_percent_w_pellet[-1])
                    Pellets_summary.append(pellet_count2[-1])
                    Magazine_Checks_summary.append(mag_checks)

                    #####-----##### Create FR Step summary data #####-----#####
                    
                    if schedule[0] == 'RRatio5':
                        FR_steps = [1, 2, 3, 4, 5]
                    elif schedule[0] == 'RRatio10':
                        FR_steps = [6, 7, 8, 9, 10]
                    elif schedule[0] == 'RRatio20':
                        FR_steps = [11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
                                            
                    FR_freq = []
                    
                    from collections import Counter
                    freq = Counter(step_block)
                    for i in range(0, len(FR_steps)):
                        frequency = freq.get(FR_steps[i], 0)
                        FR_freq.append(frequency)
                    
                    # for i in range(0, len(FR_steps)):
                    #     print(FR_steps[i], FR_freq[i])
                        
                    FR_mean_inactive = []
                    FR_mean_mag_checks = []
                    
                    for i in range(0, len(FR_steps)):
                        step_count = 0
                        incorrect = 0
                        mag_check_count = 0
                        step = FR_steps[i]
                        for i in range(0, len(step_block)):
                            if step_block[i] == step:
                                step_count += 1
                                incorrect += int(right_all_blocked_within[i])
                                mag_check_count += int(mag_check_blocked_within[i])
                                # print(step, incorrect, mag_check_count)
                        if step_count != 0:
                            FR_mean_inactive.append(incorrect / step_count)
                            FR_mean_mag_checks.append(mag_check_count / step_count)
                        else:
                            FR_mean_inactive.append('N/A')
                            FR_mean_mag_checks.append('N/A')
                    
                    # print(FR_mean_inactive, FR_mean_mag_checks)
                                
                    results_steps = {'FR': FR_steps, 'Frequency': FR_freq, 'Mean Inactive Pokes': FR_mean_inactive, 'Mean Magazine Checks': FR_mean_mag_checks}
                    export_file_steps = pd.DataFrame(results_steps, columns = ['FR', 'Frequency', 'Mean Inactive Pokes', 'Mean Magazine Checks'])

                    #####-----
                    
                    # Create aus_date column for PR steps
                    
                    aus_date_blocked = []
                        
                    for i in range(0, len(step_block)):
                        aus_date_blocked.append(aus_date)
                   
                    # Always export Summary, Chronological and step/binned data
                    
                    # Summary data
                    
                    results_summary = {'Variable': variable, 'Value': value}
                    export_file_summary = pd.DataFrame(results_summary, columns = ['Variable', 'Value'])
                    
                    # Chronological data
                    
                    # Assign left and right pokes as active or inactive based on the active port
                        
                    if active2[0] == 'Left':
                        active_poke_all = left_all2
                        active_poke_empty = left_empty2
                        active_poke_w_pellet = left_w_pellet2
                        
                        active_poke_all_chron_within = left_poke_all_chron_within
                        active_poke_empty_chron_within = left_poke_empty_chron_within
                        active_poke_w_pellet_chron_within = left_poke_w_pellet_chron_within
                        
                        active_poke_all_chron_within_binary = left_poke_all_chron_within_binary
                        active_poke_empty_chron_within_binary = left_poke_empty_chron_within_binary
                        active_poke_w_pellet_chron_within_binary = left_poke_w_pellet_chron_within_binary
                        
                        active_poke_time_all_chron = l_poke_time_all_chron
                        active_poke_time_empty_chron = l_poke_time_empty_chron
                        active_poke_time_w_pellet_chron = l_poke_time_w_pellet_chron
                        
                        inactive_poke_all = right_all2
                        inactive_poke_empty = right_empty2
                        inactive_poke_w_pellet = right_w_pellet2
                        
                        inactive_poke_all_chron_within = right_poke_all_chron_within
                        inactive_poke_empty_chron_within = right_poke_empty_chron_within
                        inactive_poke_w_pellet_chron_within = right_poke_w_pellet_chron_within
                        
                        inactive_poke_all_chron_within_binary = right_poke_all_chron_within_binary
                        inactive_poke_empty_chron_within_binary = right_poke_empty_chron_within_binary
                        inactive_poke_w_pellet_chron_within_binary = right_poke_w_pellet_chron_within_binary
                        
                        inactive_poke_time_all_chron = r_poke_time_all_chron
                        inactive_poke_time_empty_chron = r_poke_time_empty_chron
                        inactive_poke_time_w_pellet_chron = r_poke_time_w_pellet_chron
                        
                        active_poke_all_blocked = left_all_blocked
                        active_poke_empty_blocked = left_empty_blocked
                        active_poke_w_pellet_blocked = left_w_pellet_blocked
                        
                        active_poke_all_blocked_within = left_all_blocked_within
                        active_poke_empty_blocked_within = left_empty_blocked_within
                        active_poke_w_pellet_blocked_within = left_w_pellet_blocked_within
                        
                        inactive_poke_all_blocked = right_all_blocked
                        inactive_poke_empty_blocked = right_empty_blocked
                        inactive_poke_w_pellet_blocked = right_w_pellet_blocked
                        
                        inactive_poke_all_blocked_within = right_all_blocked_within
                        inactive_poke_empty_blocked_within = right_empty_blocked_within
                        inactive_poke_w_pellet_blocked_within = right_w_pellet_blocked_within
                       
                        blocked_active_all_poke_time = blocked_mean_l_all_poke_time
                        blocked_active_empty_poke_time = blocked_mean_l_empty_poke_time
                        blocked_active_w_pellet_poke_time = blocked_mean_l_w_pellet_poke_time
                        
                        blocked_inactive_all_poke_time = blocked_mean_r_all_poke_time
                        blocked_inactive_empty_poke_time = blocked_mean_r_empty_poke_time
                        blocked_inactive_w_pellet_poke_time = blocked_mean_r_w_pellet_poke_time

                        blocked_within_active_all_poke_time = blocked_within_mean_l_all_poke_time
                        blocked_within_active_empty_poke_time = blocked_within_mean_l_empty_poke_time
                        blocked_within_active_w_pellet_poke_time = blocked_within_mean_l_w_pellet_poke_time
                        
                        blocked_within_inactive_all_poke_time = blocked_within_mean_r_all_poke_time
                        blocked_within_inactive_empty_poke_time = blocked_within_mean_r_empty_poke_time
                        blocked_within_inactive_w_pellet_poke_time = blocked_within_mean_r_w_pellet_poke_time
                   
                            
                    elif active2[0] == 'Right':
                        active_poke_all = right_all2
                        active_poke_empty = right_empty2
                        active_poke_w_pellet = right_w_pellet2
                        
                        active_poke_all_chron_within = right_poke_all_chron_within
                        active_poke_empty_chron_within = right_poke_empty_chron_within
                        active_poke_w_pellet_chron_within = right_poke_w_pellet_chron_within
                        
                        active_poke_all_chron_within_binary = right_poke_all_chron_within_binary
                        active_poke_empty_chron_within_binary = right_poke_empty_chron_within_binary
                        active_poke_w_pellet_chron_within_binary = right_poke_w_pellet_chron_within_binary
                        
                        active_poke_time_all_chron = r_poke_time_all_chron
                        active_poke_time_empty_chron = r_poke_time_empty_chron
                        active_poke_time_w_pellet_chron = r_poke_time_w_pellet_chron
                        
                        inactive_poke_all = left_all2
                        inactive_poke_empty = left_empty2
                        inactive_poke_w_pellet = left_w_pellet2
                        
                        inactive_poke_all_chron_within = left_poke_all_chron_within
                        inactive_poke_empty_chron_within = left_poke_empty_chron_within
                        inactive_poke_w_pellet_chron_within = left_poke_w_pellet_chron_within
                        
                        inactive_poke_all_chron_within_binary = left_poke_all_chron_within_binary
                        inactive_poke_empty_chron_within_binary = left_poke_empty_chron_within_binary
                        inactive_poke_w_pellet_chron_within_binary = left_poke_w_pellet_chron_within_binary
                        
                        inactive_poke_time_all_chron = l_poke_time_all_chron
                        inactive_poke_time_empty_chron = l_poke_time_empty_chron
                        inactive_poke_time_w_pellet_chron = l_poke_time_w_pellet_chron
                        
                        active_poke_all_blocked = right_all_blocked
                        active_poke_empty_blocked = right_empty_blocked
                        active_poke_w_pellet_blocked = right_w_pellet_blocked
                        
                        active_poke_all_blocked_within = right_all_blocked_within
                        active_poke_empty_blocked_within = right_empty_blocked_within
                        active_poke_w_pellet_blocked_within = right_w_pellet_blocked_within
                        
                        inactive_poke_all_blocked = left_all_blocked
                        inactive_poke_empty_blocked = left_empty_blocked
                        inactive_poke_w_pellet_blocked = left_w_pellet_blocked
                        
                        inactive_poke_all_blocked_within = left_all_blocked_within
                        inactive_poke_empty_blocked_within = left_empty_blocked_within
                        inactive_poke_w_pellet_blocked_within = left_w_pellet_blocked_within
                       
                        blocked_active_all_poke_time = blocked_mean_r_all_poke_time
                        blocked_active_empty_poke_time = blocked_mean_r_empty_poke_time
                        blocked_active_w_pellet_poke_time = blocked_mean_r_w_pellet_poke_time
                        
                        blocked_inactive_all_poke_time = blocked_mean_l_all_poke_time
                        blocked_inactive_empty_poke_time = blocked_mean_l_empty_poke_time
                        blocked_inactive_w_pellet_poke_time = blocked_mean_l_w_pellet_poke_time

                        blocked_within_active_all_poke_time = blocked_within_mean_r_all_poke_time
                        blocked_within_active_empty_poke_time = blocked_within_mean_r_empty_poke_time
                        blocked_within_active_w_pellet_poke_time = blocked_within_mean_r_w_pellet_poke_time
                        
                        blocked_within_inactive_all_poke_time = blocked_within_mean_l_all_poke_time
                        blocked_within_inactive_empty_poke_time = blocked_within_mean_l_empty_poke_time
                        blocked_within_inactive_w_pellet_poke_time = blocked_within_mean_l_w_pellet_poke_time
                   
                    
                    results_chronological_c = {'Time (seconds)': seconds_elapsed, 'FR Step': session_type2, 'Active port': active2, 
                                               'ALL Active Pokes': active_poke_all, 'Empty Active Pokes': active_poke_empty, 'W Pellet Active Pokes': active_poke_w_pellet,
                                               'ALL Inactive Pokes': inactive_poke_all, 'Empty Inactive Pokes': inactive_poke_empty, 'W Pellet Inactive Pokes': inactive_poke_w_pellet,
                                               'ALL Total Pokes': total_pokes_all, 'Empty Total Pokes': total_pokes_empty, 'W Pellet Total Pokes': total_pokes_w_pellet, 
                                               'Pellet Count': pellet_count2, 
                                               'ALL Pokes % Active': all_percent_active, 'Empty Pokes % Active': empty_percent_active, 'W Pellet Pokes % Active': w_pellet_percent_active,
                                               'Active Pokes % Empty': active_percent_empty, 'Active Pokes % W Pellet': active_percent_w_pellet,
                                               'Poke Time': poke_time2, 
                                               'ALL Active Poke Time': active_poke_time_all_chron, 'Empty Active Poke Time': active_poke_time_empty_chron, 'W Pellet Active Poke Time': active_poke_time_w_pellet_chron,
                                               'ALL Inactive Poke Time': inactive_poke_time_all_chron, 'Empty Inactive Poke Time': inactive_poke_time_empty_chron, 'W Pellet Inactive Poke Time': inactive_poke_time_w_pellet_chron, 
                                               'Pellet Retrieval Time': retrieval_time2, 'Inter-Pellet Interval': ipi_2}
                    export_file_chronological_c = pd.DataFrame(results_chronological_c, columns = ['Time (seconds)', 'FR Step', 'Active port', 
                                                                                                   'ALL Active Pokes', 'Empty Active Pokes', 'W Pellet Active Pokes',
                                                                                                   'ALL Inactive Pokes', 'Empty Inactive Pokes', 'W Pellet Inactive Pokes',
                                                                                                   'ALL Total Pokes', 'Empty Total Pokes', 'W PEllet Total Pokes', 
                                                                                                   '', 'Pellet Count', 
                                                                                                   'ALL Pokes % Active', 'Empty Pokes % Active', 'W Pellet Pokes % Active', '',
                                                                                                   'Active Pokes % Empty', 'Active Pokes % W Pellet',
                                                                                                   '', 'Poke Time', 
                                                                                                   'ALL Active Poke Time', 'Empty Active Poke Time', 'W Pellet Active Poke Time', 
                                                                                                   'ALL Inactive Poke Time', 'Empty Inactive Poke Time', 'W Pellet Inactive Poke Time', 
                                                                                                   'Pellet Retrieval Time', 'Inter-Pellet Interval'])
                    
                    
                    results_chronological_w = {'Time (seconds)': seconds_elapsed, 'FR Step': session_type2, 'Active port': active2, 
                                               'ALL Active Pokes': active_poke_all_chron_within, 'Empty Active Pokes': active_poke_empty_chron_within, 'W Pellet Active Pokes': active_poke_w_pellet_chron_within,
                                               'ALL Inactive Pokes': inactive_poke_all_chron_within, 'Empty Inactive Pokes': inactive_poke_empty_chron_within, 'W Pellet Inactive Pokes': inactive_poke_w_pellet_chron_within,
                                               'ALL Total Pokes': total_pokes_all_chron_within, 'Empty Total Pokes': total_pokes_empty_chron_within, 'W Pellet Total Pokes': total_pokes_w_pellet_chron_within, 
                                               'Pellet Count': pellet_count2, 
                                               'ALL % Active Pokes': all_percent_active_chron_within, 'Empty % Active Pokes': empty_percent_active_chron_within, 'W Pellet % Active Pokes': w_pellet_percent_active_chron_within,
                                               'Active Pokes % Empty': active_percent_empty_chron_within, 'Active Pokes % W Pellet': active_percent_w_pellet_chron_within,
                                               'Poke Time': poke_time2, 
                                               'ALL Active Poke Time': active_poke_time_all_chron, 'Empty Active Poke Time': active_poke_time_empty_chron, 'W Pellet Active Poke Time': active_poke_time_w_pellet_chron,
                                               'ALL Inactive Poke Time': inactive_poke_time_all_chron, 'Empty Inactive Poke Time': inactive_poke_time_empty_chron, 'W Pellet Inactive Poke Time': inactive_poke_time_w_pellet_chron, 
                                               'Pellet Retrieval Time': retrieval_time2, 'Inter-Pellet Interval': ipi_2}
                    export_file_chronological_w = pd.DataFrame(results_chronological_w, columns = ['Time (seconds)', 'FR Step', 'Active port', 
                                                                                                   'ALL Active Pokes', 'Empty Active Pokes', 'W Pellet Active Pokes',
                                                                                                   'ALL Inactive Pokes', 'Empty Inactive Pokes', 'W Pellet Inactive Pokes',
                                                                                                   'ALL Total Pokes', 'Empty Total Pokes', 'W PEllet Total Pokes', 
                                                                                                   '', 'Pellet Count', 
                                                                                                   'ALL Pokes % Active', 'Empty Pokes % Active', 'W Pellet Pokes % Active', '',
                                                                                                   'Active Pokes % Empty', 'Active Pokes % W Pellet',
                                                                                                   '', 'Poke Time', 
                                                                                                   'ALL Active Poke Time', 'Empty Active Poke Time', 'W Pellet Active Poke Time', 
                                                                                                   'ALL Inactive Poke Time', 'Empty Inactive Poke Time', 'W Pellet Inactive Poke Time', 
                                                                                                   'Pellet Retrieval Time', 'Inter-Pellet Interval'])
                    
                    # Create chronological stepped data with breaks between steps
                    
                    seconds_elapsed_step = []
                    session_type2_step = []
                    active2_step = []
                    
                    active_poke_all_step = []
                    active_poke_empty_step = []
                    active_poke_w_pellet_step = []
                    inactive_poke_all_step = []
                    inactive_poke_empty_step = []
                    inactive_poke_w_pellet_step = []
                    total_pokes_all_step = []
                    total_pokes_empty_step = []
                    total_pokes_w_pellet_step = []
                    
                    pellet_count2_step = []
                    
                    all_percent_active_step = []
                    empty_percent_active_step = []
                    w_pellet_percent_active_step = []
                    
                    active_poke_all_chron_within_step = []
                    active_poke_empty_chron_within_step = []
                    active_poke_w_pellet_chron_within_step = []
                    
                    inactive_poke_all_chron_within_step = []
                    inactive_poke_empty_chron_within_step = []
                    inactive_poke_w_pellet_chron_within_step = []
                    
                    total_pokes_all_chron_within_step = []
                    total_pokes_empty_chron_within_step = []
                    total_pokes_w_pellet_chron_within_step = []
                    
                    all_percent_active_chron_within_step = []
                    empty_percent_active_chron_within_step = []
                    w_pellet_percent_active_chron_within_step = []
                    
                    active_poke_time_all_chron_step = []
                    active_poke_time_empty_chron_step = []
                    active_poke_time_w_pellet_chron_step = []
                    
                    inactive_poke_time_all_chron_step = []
                    inactive_poke_time_empty_chron_step = []
                    inactive_poke_time_w_pellet_chron_step = []
                    
                    retrieval_time2_step = []
                    ipi_2_step = []
                    poke_time2_step = []
                    
                    active_percent_empty_step = []
                    active_percent_empty_chron_within_step = []
                    active_percent_w_pellet_step = []
                    active_percent_w_pellet_chron_within_step = []
                    
                    for i in range (0, (len(session_type2)-1)):
                        
                        if session_type2[i + 1] == session_type2[i]:
                            seconds_elapsed_step.append(seconds_elapsed[i])
                            session_type2_step.append(session_type2[i])
                            active2_step.append(active2[i])
                            active_poke_all_step.append(active_poke_all[i])
                            active_poke_empty_step.append(active_poke_empty[i])
                            active_poke_w_pellet_step.append(active_poke_w_pellet[i])
                            inactive_poke_all_step.append(inactive_poke_all[i])
                            inactive_poke_empty_step.append(inactive_poke_empty[i])
                            inactive_poke_w_pellet_step.append(inactive_poke_w_pellet[i])
                            total_pokes_all_step.append(total_pokes_all[i])
                            total_pokes_empty_step.append(total_pokes_empty[i])
                            total_pokes_w_pellet_step.append(total_pokes_w_pellet[i])
                            pellet_count2_step.append(pellet_count2[i])
                            all_percent_active_step.append(all_percent_active[i])
                            empty_percent_active_step.append(empty_percent_active[i])
                            w_pellet_percent_active_step.append(w_pellet_percent_active[i])
                            active_poke_all_chron_within_step.append(active_poke_all_chron_within[i])
                            active_poke_empty_chron_within_step.append(active_poke_empty_chron_within[i])
                            active_poke_w_pellet_chron_within_step.append(active_poke_w_pellet_chron_within[i])
                            inactive_poke_all_chron_within_step.append(inactive_poke_all_chron_within[i])
                            inactive_poke_empty_chron_within_step.append(inactive_poke_empty_chron_within[i])
                            inactive_poke_w_pellet_chron_within_step.append(inactive_poke_w_pellet_chron_within[i])
                            total_pokes_all_chron_within_step.append(total_pokes_all_chron_within[i])
                            total_pokes_empty_chron_within_step.append(total_pokes_empty_chron_within[i])
                            total_pokes_w_pellet_chron_within_step.append(total_pokes_w_pellet_chron_within[i])
                            all_percent_active_chron_within_step.append(all_percent_active_chron_within[i])
                            empty_percent_active_chron_within_step.append(empty_percent_active_chron_within[i])
                            w_pellet_percent_active_chron_within_step.append(w_pellet_percent_active_chron_within[i])
                            active_poke_time_all_chron_step.append(active_poke_time_all_chron[i])
                            active_poke_time_empty_chron_step.append(active_poke_time_empty_chron[i])
                            active_poke_time_w_pellet_chron_step.append(active_poke_time_w_pellet_chron[i])
                            inactive_poke_time_all_chron_step.append(inactive_poke_time_all_chron[i])
                            inactive_poke_time_empty_chron_step.append(inactive_poke_time_empty_chron[i])
                            inactive_poke_time_w_pellet_chron_step.append(inactive_poke_time_w_pellet_chron[i])
                            retrieval_time2_step.append(retrieval_time2[i])
                            ipi_2_step.append(ipi_2[i])
                            poke_time2_step.append(poke_time2[i])
                            active_percent_empty_step.append(active_percent_empty[i])
                            active_percent_empty_chron_within_step.append(active_percent_empty_chron_within[i])
                            active_percent_w_pellet_step.append(active_percent_w_pellet[i])
                            active_percent_w_pellet_chron_within_step.append(active_percent_w_pellet_chron_within[i])
                            
                        elif session_type2[i + 1] != session_type[i]:
                            seconds_elapsed_step.append(seconds_elapsed[i])
                            session_type2_step.append(session_type2[i])
                            active2_step.append(active2[i])
                            active_poke_all_step.append(active_poke_all[i])
                            active_poke_empty_step.append(active_poke_empty[i])
                            active_poke_w_pellet_step.append(active_poke_w_pellet[i])
                            inactive_poke_all_step.append(inactive_poke_all[i])
                            inactive_poke_empty_step.append(inactive_poke_empty[i])
                            inactive_poke_w_pellet_step.append(inactive_poke_w_pellet[i])
                            total_pokes_all_step.append(total_pokes_all[i])
                            total_pokes_empty_step.append(total_pokes_empty[i])
                            total_pokes_w_pellet_step.append(total_pokes_w_pellet[i])
                            pellet_count2_step.append(pellet_count2[i])
                            all_percent_active_step.append(all_percent_active[i])
                            empty_percent_active_step.append(empty_percent_active[i])
                            w_pellet_percent_active_step.append(w_pellet_percent_active[i])
                            active_poke_all_chron_within_step.append(active_poke_all_chron_within[i])
                            active_poke_empty_chron_within_step.append(active_poke_empty_chron_within[i])
                            active_poke_w_pellet_chron_within_step.append(active_poke_w_pellet_chron_within[i])
                            inactive_poke_all_chron_within_step.append(inactive_poke_all_chron_within[i])
                            inactive_poke_empty_chron_within_step.append(inactive_poke_empty_chron_within[i])
                            inactive_poke_w_pellet_chron_within_step.append(inactive_poke_w_pellet_chron_within[i])
                            total_pokes_all_chron_within_step.append(total_pokes_all_chron_within[i])
                            total_pokes_empty_chron_within_step.append(total_pokes_empty_chron_within[i])
                            total_pokes_w_pellet_chron_within_step.append(total_pokes_w_pellet_chron_within[i])
                            all_percent_active_chron_within_step.append(all_percent_active_chron_within[i])
                            empty_percent_active_chron_within_step.append(empty_percent_active_chron_within[i])
                            w_pellet_percent_active_chron_within_step.append(w_pellet_percent_active_chron_within[i])
                            active_poke_time_all_chron_step.append(active_poke_time_all_chron[i])
                            active_poke_time_empty_chron_step.append(active_poke_time_empty_chron[i])
                            active_poke_time_w_pellet_chron_step.append(active_poke_time_w_pellet_chron[i])
                            inactive_poke_time_all_chron_step.append(inactive_poke_time_all_chron[i])
                            inactive_poke_time_empty_chron_step.append(inactive_poke_time_empty_chron[i])
                            inactive_poke_time_w_pellet_chron_step.append(inactive_poke_time_w_pellet_chron[i])
                            retrieval_time2_step.append(retrieval_time2[i])
                            ipi_2_step.append(ipi_2[i])
                            poke_time2_step.append(poke_time2[i])
                            active_percent_empty_step.append(active_percent_empty[i])
                            active_percent_empty_chron_within_step.append(active_percent_empty_chron_within[i])
                            active_percent_w_pellet_step.append(active_percent_w_pellet[i])
                            active_percent_w_pellet_chron_within_step.append(active_percent_w_pellet_chron_within[i])

                            seconds_elapsed_step.append(np.nan)
                            session_type2_step.append(np.nan)
                            active2_step.append(np.nan)
                            active_poke_all_step.append(np.nan)
                            active_poke_empty_step.append(np.nan)
                            active_poke_w_pellet_step.append(np.nan)
                            inactive_poke_all_step.append(np.nan)
                            inactive_poke_empty_step.append(np.nan)
                            inactive_poke_w_pellet_step.append(np.nan)
                            total_pokes_all_step.append(np.nan)
                            total_pokes_empty_step.append(np.nan)
                            total_pokes_w_pellet_step.append(np.nan)
                            pellet_count2_step.append(np.nan)
                            all_percent_active_step.append(np.nan)
                            empty_percent_active_step.append(np.nan)
                            w_pellet_percent_active_step.append(np.nan)
                            active_poke_all_chron_within_step.append(np.nan)
                            active_poke_empty_chron_within_step.append(np.nan)
                            active_poke_w_pellet_chron_within_step.append(np.nan)
                            inactive_poke_all_chron_within_step.append(np.nan)
                            inactive_poke_empty_chron_within_step.append(np.nan)
                            inactive_poke_w_pellet_chron_within_step.append(np.nan)
                            total_pokes_all_chron_within_step.append(np.nan)
                            total_pokes_empty_chron_within_step.append(np.nan)
                            total_pokes_w_pellet_chron_within_step.append(np.nan)
                            all_percent_active_chron_within_step.append(np.nan)
                            empty_percent_active_chron_within_step.append(np.nan)
                            w_pellet_percent_active_chron_within_step.append(np.nan)
                            active_poke_time_all_chron_step.append(np.nan)
                            active_poke_time_empty_chron_step.append(np.nan)
                            active_poke_time_w_pellet_chron_step.append(np.nan)
                            inactive_poke_time_all_chron_step.append(np.nan)
                            inactive_poke_time_empty_chron_step.append(np.nan)
                            inactive_poke_time_w_pellet_chron_step.append(np.nan)
                            retrieval_time2_step.append(np.nan)
                            ipi_2_step.append(np.nan)
                            poke_time2_step.append(np.nan)
                            active_percent_empty_step.append(np.nan)
                            active_percent_empty_chron_within_step.append(np.nan)
                            active_percent_w_pellet_step.append(np.nan)
                            active_percent_w_pellet_chron_within_step.append(np.nan)

                    
                    seconds_elapsed_step.append(seconds_elapsed[-1])
                    session_type2_step.append(session_type2[-1])
                    active2_step.append(active2[-1])
                    active_poke_all_step.append(active_poke_all[-1])
                    active_poke_empty_step.append(active_poke_empty[-1])
                    active_poke_w_pellet_step.append(active_poke_w_pellet[-1])
                    inactive_poke_all_step.append(inactive_poke_all[-1])
                    inactive_poke_empty_step.append(inactive_poke_empty[-1])
                    inactive_poke_w_pellet_step.append(inactive_poke_w_pellet[-1])
                    total_pokes_all_step.append(total_pokes_all[-1])
                    total_pokes_empty_step.append(total_pokes_empty[-1])
                    total_pokes_w_pellet_step.append(total_pokes_w_pellet[-1])
                    pellet_count2_step.append(pellet_count2[-1])
                    all_percent_active_step.append(all_percent_active[-1])
                    empty_percent_active_step.append(empty_percent_active[-1])
                    w_pellet_percent_active_step.append(w_pellet_percent_active[-1])
                    active_poke_all_chron_within_step.append(active_poke_all_chron_within[-1])
                    active_poke_empty_chron_within_step.append(active_poke_empty_chron_within[-1])
                    active_poke_w_pellet_chron_within_step.append(active_poke_w_pellet_chron_within[-1])
                    inactive_poke_all_chron_within_step.append(inactive_poke_all_chron_within[-1])
                    inactive_poke_empty_chron_within_step.append(inactive_poke_empty_chron_within[-1])
                    inactive_poke_w_pellet_chron_within_step.append(inactive_poke_w_pellet_chron_within[-1])
                    total_pokes_all_chron_within_step.append(total_pokes_all_chron_within[-1])
                    total_pokes_empty_chron_within_step.append(total_pokes_empty_chron_within[-1])
                    total_pokes_w_pellet_chron_within_step.append(total_pokes_w_pellet_chron_within[-1])
                    all_percent_active_chron_within_step.append(all_percent_active_chron_within[-1])
                    empty_percent_active_chron_within_step.append(empty_percent_active_chron_within[-1])
                    w_pellet_percent_active_chron_within_step.append(w_pellet_percent_active_chron_within[-1])
                    active_poke_time_all_chron_step.append(active_poke_time_all_chron[-1])
                    active_poke_time_empty_chron_step.append(active_poke_time_empty_chron[-1])
                    active_poke_time_w_pellet_chron_step.append(active_poke_time_w_pellet_chron[-1])
                    inactive_poke_time_all_chron_step.append(inactive_poke_time_all_chron[-1])
                    inactive_poke_time_empty_chron_step.append(inactive_poke_time_empty_chron[-1])
                    inactive_poke_time_w_pellet_chron_step.append(inactive_poke_time_w_pellet_chron[-1])
                    retrieval_time2_step.append(retrieval_time2[-1])
                    ipi_2_step.append(ipi_2[-1])
                    poke_time2_step.append(poke_time2[-1])
                    active_percent_empty_step.append(active_percent_empty[-1])
                    active_percent_empty_chron_within_step.append(active_percent_empty_chron_within[-1])
                    active_percent_w_pellet_step.append(active_percent_w_pellet[-1])
                    active_percent_w_pellet_chron_within_step.append(active_percent_w_pellet_chron_within[-1])
                    
                    

                    results_chronological_c_step = {'Time (seconds)': seconds_elapsed_step, 'FR Step': session_type2_step, 'Active port': active2_step, 
                                               'ALL Active Pokes': active_poke_all_step, 'Empty Active Pokes': active_poke_empty_step, 'W Pellet Active Pokes': active_poke_w_pellet_step,
                                               'ALL Inactive Pokes': inactive_poke_all_step, 'Empty Inactive Pokes': inactive_poke_empty_step, 'W Pellet Inactive Pokes': inactive_poke_w_pellet_step,
                                               'ALL Total Pokes': total_pokes_all_step, 'Empty Total Pokes': total_pokes_empty_step, 'W Pellet Total Pokes': total_pokes_w_pellet_step, 
                                               'Pellet Count': pellet_count2_step, 
                                               'ALL Pokes % Active': all_percent_active_step, 'Empty Pokes % Active': empty_percent_active_step, 'W Pellet Pokes % Active': w_pellet_percent_active_step,
                                               'Active Pokes % Empty': active_percent_empty_step, 'Active Pokes % W Pellet': active_percent_w_pellet_step,
                                               'Poke Time': poke_time2_step, 
                                               'ALL Active Poke Time': active_poke_time_all_chron_step, 'Empty Active Poke Time': active_poke_time_empty_chron_step, 'W Pellet Active Poke Time': active_poke_time_w_pellet_chron_step,
                                               'ALL Inactive Poke Time': inactive_poke_time_all_chron_step, 'Empty Inactive Poke Time': inactive_poke_time_empty_chron_step, 'W Pellet Inactive Poke Time': inactive_poke_time_w_pellet_chron_step, 
                                               'Pellet Retrieval Time': retrieval_time2_step, 'Inter-Pellet Interval': ipi_2_step}
                    
                    export_file_chronological_c_step = pd.DataFrame(results_chronological_c_step, columns = ['Time (seconds)', 'FR Step', 'Active port', 
                                                                                                   'ALL Active Pokes', 'Empty Active Pokes', 'W Pellet Active Pokes',
                                                                                                   'ALL Inactive Pokes', 'Empty Inactive Pokes', 'W Pellet Inactive Pokes',
                                                                                                   'ALL Total Pokes', 'Empty Total Pokes', 'W PEllet Total Pokes', 
                                                                                                   '', 'Pellet Count', 
                                                                                                   'ALL Pokes % Active', 'Empty Pokes % Active', 'W Pellet Pokes % Active', '',
                                                                                                   'Active Pokes % Empty', 'Active Pokes % W Pellet',
                                                                                                   '', 'Poke Time', 
                                                                                                   'ALL Active Poke Time', 'Empty Active Poke Time', 'W Pellet Active Poke Time', 
                                                                                                   'ALL Inactive Poke Time', 'Empty Inactive Poke Time', 'W Pellet Inactive Poke Time', 
                                                                                                   'Pellet Retrieval Time', 'Inter-Pellet Interval'])

                    results_chronological_w_step = {'Time (seconds)': seconds_elapsed_step, 'FR Step': session_type2_step, 'Active port': active2_step, 
                                               'ALL Active Pokes': active_poke_all_chron_within_step, 'Empty Active Pokes': active_poke_empty_chron_within_step, 'W Pellet Active Pokes': active_poke_w_pellet_chron_within_step,
                                               'ALL Inactive Pokes': inactive_poke_all_chron_within_step, 'Empty Inactive Pokes': inactive_poke_empty_chron_within_step, 'W Pellet Inactive Pokes': inactive_poke_w_pellet_chron_within_step,
                                               'ALL Total Pokes': total_pokes_all_chron_within_step, 'Empty Total Pokes': total_pokes_empty_chron_within_step, 'W Pellet Total Pokes': total_pokes_w_pellet_chron_within_step, 
                                               'Pellet Count': pellet_count2_step, 
                                               'ALL Pokes % Active': all_percent_active_chron_within_step, 'Empty Pokes % Active': empty_percent_active_chron_within_step, 'W Pellet Pokes % Active': w_pellet_percent_active_chron_within_step,
                                               'Active Pokes % Empty': active_percent_empty_chron_within_step, 'Active Pokes % W Pellet': active_percent_w_pellet_chron_within_step,
                                               'Poke Time': poke_time2_step, 
                                               'ALL Active Poke Time': active_poke_time_all_chron_step, 'Empty Active Poke Time': active_poke_time_empty_chron_step, 'W Pellet Active Poke Time': active_poke_time_w_pellet_chron_step,
                                               'ALL Inactive Poke Time': inactive_poke_time_all_chron_step, 'Empty Inactive Poke Time': inactive_poke_time_empty_chron_step, 'W Pellet Inactive Poke Time': inactive_poke_time_w_pellet_chron_step, 
                                               'Pellet Retrieval Time': retrieval_time2_step, 'Inter-Pellet Interval': ipi_2_step}

                    export_file_chronological_w_step = pd.DataFrame(results_chronological_w_step, columns = ['Time (seconds)', 'FR Step', 'Active port', 
                                                                                                   'ALL Active Pokes', 'Empty Active Pokes', 'W Pellet Active Pokes',
                                                                                                   'ALL Inactive Pokes', 'Empty Inactive Pokes', 'W Pellet Inactive Pokes',
                                                                                                   'ALL Total Pokes', 'Empty Total Pokes', 'W Pellet Total Pokes', 
                                                                                                   '', 'Pellet Count', 
                                                                                                   'ALL Pokes % Active', 'Empty Pokes % Active', 'W Pellet Pokes % Active', '',
                                                                                                   'Active Pokes % Empty', 'Active Pokes % W Pellet',
                                                                                                   '', 'Poke Time', 
                                                                                                   'ALL Active Poke Time', 'Empty Active Poke Time', 'W Pellet Active Poke Time', 
                                                                                                   'ALL Inactive Poke Time', 'Empty Inactive Poke Time', 'W Pellet Inactive Poke Time', 
                                                                                                   'Pellet Retrieval Time', 'Inter-Pellet Interval'])


                    
                    # Step/binned data (cumulative and within)
    
                    results_blocked_c = {'Date': aus_date_blocked, 'Seconds Elapsed': seconds_elapsed_blocked, 'Duration': duration_blocked, 'FR Step': step_block, 'Active Port': active_blocked, 
                                        'ALL Active Pokes': active_poke_all_blocked, 'Empty Active Pokes': active_poke_empty_blocked, 'W Pellet Active Pokes': active_poke_w_pellet_blocked,
                                        'ALL Inactive Pokes': inactive_poke_all_blocked, 'Empty Inactive Pokes': inactive_poke_empty_blocked, 'W Pellet Inactive Pokes': inactive_poke_w_pellet_blocked,
                                        'ALL Total Pokes': total_pokes_all_blocked, 'Empty Total Pokes': total_pokes_empty_blocked, 'W Pellet Total Pokes': total_pokes_w_pellet_blocked,
                                        'Pellet Count': pellet_blocked, 'Magazine Checks': mag_check_blocked,
                                        'ALL Pokes % Active': all_percent_active_blocked, 'Empty Pokes % Active': empty_percent_active_blocked, 'W Pellet Pokes % Active': w_pellet_percent_active_blocked,
                                        'ALL Mean Poke Time': blocked_mean_both_all_poke_time, 'Empty Mean Poke Time': blocked_mean_both_empty_poke_time, 'W Pellet Mean Poke Time': blocked_mean_both_w_pellet_poke_time,
                                        'Active Pokes % Empty': active_percent_empty_blocked, 'Active Pokes % W Pellet': active_percent_w_pellet_blocked,
                                        'ALL Mean Active Poke Time': blocked_active_all_poke_time, 'Empty Mean Active Poke Time': blocked_active_empty_poke_time, 'W Pellet Mean Active Poke Time': blocked_active_w_pellet_poke_time,
                                        'ALL Mean Inactive Poke Time': blocked_inactive_all_poke_time, 'Empty Mean Inactive Poke Time': blocked_inactive_empty_poke_time, 'W Pellet Mean Inactive Poke Time': blocked_inactive_w_pellet_poke_time,
                                        'Pellet Retrieval Time': retrieval_time_blocked, 'Inter-Pellet Interval': ipi_blocked}
                    export_file_blocked_c = pd.DataFrame(results_blocked_c, columns = ['Date', 'Seconds Elapsed', 'Duration', 'FR Step', 'Active Port', '',
                                                                                        'ALL Active Pokes', 'Empty Active Pokes', 'W Pellet Active Pokes',
                                                                                        'ALL Inactive Pokes', 'Empty Inactive Pokes', 'W Pellet Inactive Pokes',
                                                                                        'ALL Total Pokes', 'Empty Total Pokes', 'W Pellet Total Pokes',
                                                                                        'Pellet Count', 'Magazine Checks', '',
                                                                                        'ALL Pokes % Active', 'Empty Pokes % Active', 'W Pellet Pokes % Active',
                                                                                        'Active Pokes % Empty', 'Active Pokes % W Pellet', '',
                                                                                        'ALL Mean Poke Time', 'Empty Mean Poke Time', 'W Pellet Mean Poke Time',
                                                                                        'ALL Mean Active Poke Time', 'Empty Mean Active Poke Time', 'W Pellet Mean Active Poke Time',
                                                                                        'ALL Mean Inactive Poke Time', 'Empty Mean Inactive Poke Time', 'W Pellet Mean Inactive Poke Time',
                                                                                        'Pellet Retrieval Time', 'Inter-Pellet Interval'])
                    
                    results_blocked_w = {'Date': aus_date_blocked, 'Seconds Elapsed': seconds_elapsed_blocked_within, 'Duration': duration_blocked_within, 'FR Step': step_block, 'Active Port': active_blocked, 
                                        'ALL Active Pokes': active_poke_all_blocked_within, 'Empty Active Pokes': active_poke_empty_blocked_within, 'W Pellet Active Pokes': active_poke_w_pellet_blocked_within,
                                        'ALL Inactive Pokes': inactive_poke_all_blocked_within, 'Empty Inactive Pokes': inactive_poke_empty_blocked_within, 'W Pellet Inactive Pokes': inactive_poke_w_pellet_blocked_within,
                                        'ALL Total Pokes': total_pokes_all_blocked_within, 'Empty Total Pokes': total_pokes_empty_blocked_within, 'W Pellet Total Pokes': total_pokes_w_pellet_blocked_within,
                                        'Pellet Count': pellet_blocked_within, 'Magazine Checks': mag_check_blocked_within,
                                        'ALL Pokes % Active': all_percent_active_blocked_within, 'Empty Pokes % Active': empty_percent_active_blocked_within, 'W Pellet Pokes % Active': w_pellet_percent_active_blocked_within,
                                        'Active Pokes % Empty': active_percent_empty_blocked_within, 'Active Pokes % W Pellet': active_percent_w_pellet_blocked_within,
                                        'ALL Mean Poke Time': blocked_within_mean_both_all_poke_time, 'Empty Mean Poke Time': blocked_within_mean_both_empty_poke_time, 'W Pellet Mean Poke Time': blocked_within_mean_both_w_pellet_poke_time,
                                        'ALL Mean Active Poke Time': blocked_within_active_all_poke_time, 'Empty Mean Active Poke Time': blocked_within_active_empty_poke_time, 'W Pellet Mean Active Poke Time': blocked_within_active_w_pellet_poke_time,
                                        'ALL Mean Inactive Poke Time': blocked_within_inactive_all_poke_time, 'Empty Mean Inactive Poke Time': blocked_within_inactive_empty_poke_time, 'W Pellet Mean Inactive Poke Time': blocked_within_inactive_w_pellet_poke_time,
                                        'Pellet Retrieval Time': retrieval_time_blocked, 'Inter-Pellet Interval': ipi_blocked}
                    export_file_blocked_w = pd.DataFrame(results_blocked_w, columns = ['Date', 'Seconds Elapsed', 'Duration', 'FR Step', 'Active Port', '',
                                                                                        'ALL Active Pokes', 'Empty Active Pokes', 'W Pellet Active Pokes',
                                                                                        'ALL Inactive Pokes', 'Empty Inactive Pokes', 'W Pellet Inactive Pokes',
                                                                                        'ALL Total Pokes', 'Empty Total Pokes', 'W Pellet Total Pokes',
                                                                                        'Pellet Count', 'Magazine Checks', '',
                                                                                        'ALL Pokes % Active', 'Empty Pokes % Active', 'W Pellet Pokes % Active',
                                                                                        'Active Pokes % Empty', 'Active Pokes % W Pellet', '',
                                                                                        'ALL Mean Poke Time', 'Empty Mean Poke Time', 'W Pellet Mean Poke Time',
                                                                                        'ALL Mean Active Poke Time', 'Empty Mean Active Poke Time', 'W Pellet Mean Active Poke Time',
                                                                                        'ALL Mean Inactive Poke Time', 'Empty Mean Inactive Poke Time', 'W Pellet Mean Inactive Poke Time',
                                                                                        'Pellet Retrieval Time', 'Inter-Pellet Interval'])

                    
                    # Export to excel
                    
                    if export_results == 'Y':
                    
                        from openpyxl import Workbook
                    
                        wb = Workbook()
                        
                        ws1 = wb.active
                        ws1.title = 'Summary'
                        
                        ws2 = wb.create_sheet()
                        ws2.title = 'FR Step Summary'
                        
                        ws3 = wb.create_sheet()
                        ws3.title = 'Chronological Cumulative'
                        
                        ws4 = wb.create_sheet()
                        ws4.title = 'Chronological C Step'
                        
                        ws5 = wb.create_sheet()
                        ws5.title = 'Chronological Within'
                        
                        ws6 = wb.create_sheet()
                        ws6.title = 'Chronological W Step'
                        
                        ws7 = wb.create_sheet()
                        ws7.title = 'FR Step Cumulative'
                        
                        ws8 = wb.create_sheet()
                        ws8.title = 'FR Step Within'
                        
                        results_to_export = [export_file_summary, export_file_steps, export_file_chronological_c, export_file_chronological_c_step, 
                                              export_file_chronological_w, export_file_chronological_w_step, 
                                              export_file_blocked_c, export_file_blocked_w]
                        
                        sheets_to_export = wb.sheetnames
                        
                        with pd.ExcelWriter(export_destination) as writer:
                            
                            for i in range(len(sheets_to_export)):
                                results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
    
        
        # ##########---------- Joined Session Summaries ----------##########
        
        # results_joined_summary = {'Date': Date_summary, 'Task': Task_summary, 'Duration': Duration_summary, 'Active Port': Active_port_summary, 
        #                           'ALL Total Pokes': ALL_Total_Pokes_summary, 'ALL Active Pokes': ALL_Active_Pokes_summary, 'ALL Inactive Pokes': ALL_Inactive_Pokes_summary, 'ALL Pokes % Active': ALL_Pokes_Percent_Active_summary, 
        #                           'Empty Total Pokes': Empty_Total_Pokes_summary, 'Empty Active Pokes': Empty_Active_Pokes_summary, 'Empty Inactive Pokes': Empty_Inactive_Pokes_summary, 'Empty Pokes % Active': Empty_Pokes_Percent_Active_summary, 
        #                           'W Pellet Total Pokes': W_Pellet_Total_Pokes_summary, 'W Pellet Active Pokes': W_Pellet_Active_Pokes_summary, 'W Pellet Inactive Pokes': W_Pellet_Inactive_Pokes_summary, 'W Pellet Pokes % Active': W_Pellet_Pokes_Percent_Active_summary, 
        #                           'Active Pokes % Empty': Active_Pokes_Percent_Empty_summary, 'Active Pokes % W Pellet': Active_Pokes_Percent_W_Pellet_summary, 'Pellets': Pellets_summary, 'Magazine Checks': Magazine_Checks_summary}
        # export_file_joined_summary = pd.DataFrame(results_joined_summary, columns = ['Date', 'Task', 'Duration', 'Active Port', 
        #                         'ALL Total Pokes', 'ALL Active Pokes', 'ALL Inactive Pokes', 'ALL Pokes % Active',
        #                         'Empty Total Pokes', 'Empty Active Pokes', 'Empty Inactive Pokes', 'Empty Pokes % Active',
        #                         'W Pellet Total Pokes', 'W Pellet Active Pokes', 'W Pellet Inactive Pokes', 'W Pellet Pokes % Active', 
        #                         'Active Pokes % Empty', 'Active Pokes % W Pellet', 'Pellets', 'Magazine Checks'])

        # from openpyxl import Workbook
                                
        # wb = Workbook()
        
        # ws1 = wb.active
        # ws1.title = 'Session Summary Joined'
        
        # sheets_to_export = wb.sheetnames
                        
        # results_to_export = [export_file_joined_summary]
        
        # joined_name = folder + ' Summary Joined.xlsx'
        # joined_destination = export_location + folder + '/' + joined_name
        # joined_cohort_destination = cohort_export_location + joined_name
                    
        # with pd.ExcelWriter(joined_destination) as writer:
                        
        #     for i in range(len(sheets_to_export)):
        #         results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
                
        # with pd.ExcelWriter(joined_cohort_destination) as writer:
                        
        #     for i in range(len(sheets_to_export)):
        #         results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
        
        # print('Session summaries successfully joined')
        
        # ##########---------- Overview Sheets----------##########
        
        # from openpyxl import Workbook
                
        # wb = Workbook()
        
        # ws = wb.active
        # ws.title = 'Summary Overview'
        
        # ws2 = wb.create_sheet()
        # ws2.title = 'FR Step Frequency Overview'
        
        # ws3 = wb.create_sheet()
        # ws3.title = 'FR Step Inactive Overview'
        
        # ws4 = wb.create_sheet()
        # ws4.title = 'FR Step Magazine Overview'
        
        # value_column = 'Value'
        # variable_column = 'Variable'
        
        # FR_column = 'FR'
        # frequency_column = 'Frequency'
        # mean_inactive_pokes_column = 'Mean Inactive Pokes'
        # mean_magazine_checks_column = 'Mean Magazine Checks'
        
        # counter = 1
        
        # # export_location_folder = export_location + folder + '/'
        
        # for filename in sorted(os.listdir(os.path.join(export_location, folder))):

        #     if filename.endswith("Random Ratio.xlsx"):
        #         print(filename)
        #         if counter == 1:
                    
        #             export_name = filename
        #             export_destination = export_location + folder + '/' + export_name
                    
        #             df_overview = pd.read_excel(export_destination, sheet_name = 'Summary')
                    
        #             df_overview.drop(columns='Value', inplace=True)
                    
        #             df_FR_freq = pd.read_excel(ratio_list_location)
        #             df_FR_inactive = pd.read_excel(ratio_list_location)
        #             df_FR_magazine = pd.read_excel(ratio_list_location)
                    
        #         export_name = filename
        #         export_destination = export_location + folder + '/' + export_name
                
        #         name = 'Session ' + str(counter)
                
        #         df = pd.read_excel(export_destination, sheet_name = 'Summary')
                
        #         values = df[value_column].tolist()
                
        #         df_overview.insert(counter, name, values)
                
        #         df_FR = pd.read_excel(export_destination, sheet_name = 'FR Step Summary')
                
        #         frequency = df_FR[frequency_column].tolist()
        #         inactive_pokes = df_FR[mean_inactive_pokes_column].tolist()
        #         magazine_checks = df_FR[mean_magazine_checks_column].tolist()
                
        #         if values[2] == 'RRatio5':
        #             count = 6
        #             while count < 21:
        #                 frequency.append(np.nan)
        #                 inactive_pokes.append(np.nan)
        #                 magazine_checks.append(np.nan)
        #                 count += 1
        #         elif values[2] == 'RRatio10':
        #             count = 1
        #             while count < 6:
        #                 frequency.insert(0, np.nan)
        #                 inactive_pokes.insert(0, np.nan)
        #                 magazine_checks.insert(0, np.nan)
        #                 count += 1
        #             count = 11
        #             while count > 10 and count < 21:
        #                 frequency.append(np.nan)
        #                 inactive_pokes.append(np.nan)
        #                 magazine_checks.append(np.nan)
        #                 count += 1
        #         elif values[2] == 'RRatio20':
        #             count = 1
        #             while count < 11:
        #                 frequency.insert(0, np.nan)
        #                 inactive_pokes.insert(0, np.nan)
        #                 magazine_checks.insert(0, np.nan)
        #                 count += 1
                
        #         df_FR_freq.insert(counter, name, frequency)
        #         df_FR_inactive.insert(counter, name, inactive_pokes)
        #         df_FR_magazine.insert(counter, name, magazine_checks)
                
        #         counter += 1
                
        #     else:
        #         continue
                    
        #     overview_name = folder + ' Random Ratio Overview.xlsx'
        #     overview_destination = export_location + folder + '/' + overview_name
                       
        #     results_to_export = [df_overview, df_FR_freq, df_FR_inactive, df_FR_magazine]
            
        #     sheets_to_export = wb.sheetnames
            
        #     with pd.ExcelWriter(overview_destination) as writer:
                            
        #         for i in range(len(sheets_to_export)):
        #             results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
            
        #     cohort_overview_destination = cohort_export_location + overview_name
            
        #     with pd.ExcelWriter(cohort_overview_destination) as writer:
                            
        #         for i in range(len(sheets_to_export)):
        #             results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
            
        # print(filename, 'Overview complete')
        # break

                
                    