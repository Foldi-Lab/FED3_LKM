#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Dec 15 13:25:16 2021

@author: lauramilton
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Aug  3 12:51:22 2021

@author: lauramilton
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon May 17 13:27:16 2021

@author: lauramilton
"""

time_column = 'MM:DD:YYYY hh:mm:ss'
event_column = 'Event'
active_poke_column = 'Active_Poke'
session_type_column = 'Session_type'
left_poke_column = 'Left_Poke_Count'
right_poke_column = 'Right_Poke_Count'


retrieval_time_column = 'Retrieval_Time'
ipi_column = 'InterPelletInterval'
poke_time_column = 'Poke_Time'


seconds_in_bins1 = '600' # Enter how long you want the time bins to be in seconds; if not using bins enter '' which will give duration in seconds
seconds_in_bins2 = '1800'
initiation_poke = True

# Enter the FED import folder, FED export folder, cohort/experiment export folder and the FED number

import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/LKM FED 011121/'
export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/LKM FED 011121/'
# cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/Jade Honours/FED210825/DatafilesPython/FR training Overview/'
# FED_num = 'Subset'

dates = ['111821']


#-----------------------------------------------------------------------------

# Import the revelant data: time, FR ratio, event, active port, left poke, right poke, and pellet count.

import pandas as pd
import numpy as np
import os
import openpyxl

#-----

for folder in sorted(os.listdir(import_location)):
    if folder.startswith('FED21'):

        for filename in sorted(os.listdir(os.path.join(import_location, folder))):
            
            if filename.endswith(".CSV"):
                
                for i in range(0, len(dates)):
                    if dates[i] in filename:
                        print(filename)
                        # Import the csv data

                        import_name = filename
                        import_destination = import_location + folder + '/' + import_name
                        export_name = 'Extinction ' + import_name.strip('.CSV') + '.xlsx'
                        export_destination = export_location + folder + '/' + export_name
                
                #-----
                
                        df = pd.read_csv(import_destination)
                        
                        time = df[time_column].tolist()
                        session_type = df[session_type_column].tolist()
                        event = df[event_column].tolist()
                        active = df[active_poke_column].tolist()
                        left_poke = df[left_poke_column].tolist()
                        right_poke = df[right_poke_column].tolist()
                        
                        retrieval_time = df[retrieval_time_column].tolist()
                        ipi = df[ipi_column].tolist()
                        poke_time = df[poke_time_column].tolist()

                        
                        # Change the time column from strings to datetime format for calculating durations between timestamps and creating time bins
                        
                        import datetime as dt
                        
                        time = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time]
                        
                        # Start time of the session is the first timestamp
                        
                        start_time = time[0]
                        
                        # Remove the initiation poke/s and pellet data (if required)
                        
                        if initiation_poke == True:
                        
                            left_count = left_poke[0]
                            right_count = right_poke[0]
                            
                            del time[:1]
                            del active[:1]
                            del session_type[:1]
                            del event[:1]
                            del left_poke[:1]
                            del right_poke[:1]
                                
                                        
                            # Subtract the poke and pellet from the subsequent cumulative data
                            
                            left_poke_shifted = []
                            right_poke_shifted = []
                            
                            for i in range(0, len(left_poke)):
                                left_poke_shifted.append(left_poke[i] - left_count)
                                right_poke_shifted.append(right_poke[i] - right_count)
                                                               
                        
                        # Create a new column that is the total number of pokes
                        
                        total_pokes = []
                        
                        for i in range(0, len(active)):
                            total_pokes.append(left_poke[i] + right_poke[i])
                        
                        # Create new column that is % of pokes that are into the active port. Needs to be based on active_poke value (i.e. Left or Right)
                        
                        percent_active = []
                        
                        for i in range(0, len(active)):
                            if active[i] == 'Left':
                                percent_active.append(left_poke[i] / total_pokes[i] * 100)
                            else:
                                percent_active.append(right_poke[i] / total_pokes[i] * 100)
                                
                        # Calculate how much time (in seconds) has elapsed from start time for each event (nose poke)
                        
                        seconds_elapsed = []
                        
                        for i in range(0, len(time)):
                            time_from_start = time[i] - start_time
                            result = int(time_from_start.total_seconds())
                            seconds_elapsed.append(result)
                        
                        # Create time bins of desired length (input for desired length is at the top of the code)
                        
                        import math
                        
                        if seconds_in_bins1 != '':
                        
                            end_time = seconds_elapsed[-1]
                            
                            # Create end value that is the next greatest multiple of the required time bin duration greater than the end_time so that those data points are retained in the output
                            
                            excess = math.ceil(end_time / int(seconds_in_bins1)) # Rounds up so that the number of bins is the multiple of bin length greater than the time point
                            
                            end_bin = excess * int(seconds_in_bins1) # Creates the end value for the last time bin
                            
                            interval_range = pd.interval_range(start=0, freq=int(seconds_in_bins1), end=end_bin, closed="left") #'left' means that it includes the first data point which is 0 because it's start time
                            
                            time_bins = df[seconds_elapsed] = pd.cut(seconds_elapsed, interval_range, include_lowest=True, ordered=True)
                            
                            # Create dictionary of time bins as keys with bin number as matching entry to label the bins
                            
                            # Create a list of integers of the same length as there are time bins
                            
                            bin_number = []
                            a = 1
                            for i in range(0, len(interval_range)):
                                bin_number.append(a)
                                a += 1
                            
                            make_dictionary = zip(interval_range, bin_number)
                            
                            bin_dictionary = dict(make_dictionary)
                            
                            bin_num = []
                            for i in range(0, len(time_bins)):
                                bin_num.append(bin_dictionary.get((time_bins[i])))

                            # Get the last entry into each bin for the rest of the columns required
                            
                            bin_num_binned1 = []
                            left_binned1 = []
                            right_binned1 = []
                            active_binned1 = []
                            session_type_binned1 = []
                            percent_active_binned1 = []
                            total_pokes_binned1 = []
                            
                            # time binned
                            for i in range(1, len(time_bins)):
                                    if bin_num[i - 1] != bin_num[i]:
                                        bin_num_binned1.append(bin_num[i - 1])
                            bin_num_binned1.append(bin_num[i - 0])
                            
                            # left binned
                            for i in range(1, len(left_poke)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        left_binned1.append(left_poke[i - 1])
                            left_binned1.append(left_poke[i - 0])
                                    
                            # right binned
                            for i in range(1, len(right_poke)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        right_binned1.append(right_poke[i - 1])
                            right_binned1.append(right_poke[i - 0])
                                 
                            
                            # active binned
                            for i in range(1, len(active)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        active_binned1.append(active[i - 1])
                            active_binned1.append(active[i - 0])
                            
                            # session type binned
                            for i in range(1, len(session_type)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        session_type_binned1.append(session_type[i - 1])
                            session_type_binned1.append(session_type[i - 0])
                            
                            # total pokes binned
                            for i in range(1, len(total_pokes)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        total_pokes_binned1.append(total_pokes[i - 1])
                            total_pokes_binned1.append(total_pokes[i - 0])
                            
                            # percent active binned
                            for i in range(0, len(active_binned1)):
                                if active_binned1[i] == 'Left':
                                    percent_active_binned1.append(left_binned1[i] / total_pokes_binned1[i] * 100)
                                else:
                                    percent_active_binned1.append(right_binned1[i] / total_pokes_binned1[i] * 100)
                            
                            # Account for potential empty first bin
                            
                            bin_counter = 1
                            bin_index = 0
                            
                            for i in range(0, len(bin_num)):
                                if bin_num_binned1[bin_index] != bin_counter:
                                    bin_num_binned1.insert(bin_counter - 1, bin_counter)
                                    left_binned1.insert(bin_counter - 1, 0)
                                    right_binned1.insert(bin_counter - 1, 0)
                                    active_binned1.insert(bin_counter - 1, active[i])
                                    session_type_binned1.insert(bin_counter - 1, session_type[0])
                                    total_pokes_binned1.insert(bin_counter - 1, 0)
                                    percent_active_binned1.insert(bin_counter - 1, np.nan)
                            
                            # Enter new row for any bins that have no data in them (for cumulative bins this is a duplication of the preceding row)
                        
                            active_port = active_binned1[0]
                            session_type3 = session_type[0]
                                  
                            for i in bin_dictionary:
                                if bin_dictionary[i] in bin_num_binned1:
                                    continue
                                else:
                                    bin_num_binned1.insert((bin_dictionary[i] - 1), bin_dictionary[i])
                                    left_binned1.insert((bin_dictionary[i] - 1), left_binned1[(bin_dictionary[i] - 2)])
                                    right_binned1.insert((bin_dictionary[i] - 1), right_binned1[(bin_dictionary[i] - 2)])
                                    active_binned1.insert((bin_dictionary[i] - 1), active_port)
                                    percent_active_binned1.insert((bin_dictionary[i] - 1), percent_active_binned1[(bin_dictionary[i] - 2)])
                                    session_type_binned1.insert((bin_dictionary[i] - 1), session_type3)
                                    total_pokes_binned1.insert((bin_dictionary[i] - 1), total_pokes_binned1[(bin_dictionary[i] - 2)])
                                    
                            # Make new counts that are within the bins rather than cumulative
                            
                            left_binned_within1 = []
                            right_binned_within1 = []
                            pellet_binned_within1 = []
                            percent_active_binned_within1 = []
                            total_pokes_binned_within1 = []
                            
                            left_binned_within1.append(left_binned1[0])
                            for i in range(1, len(left_binned1)):
                                left_binned_within1.append(left_binned1[i] - left_binned1[i - 1])
                            
                            right_binned_within1.append(right_binned1[0])
                            for i in range(1, len(right_binned1)):
                                right_binned_within1.append(right_binned1[i] - right_binned1[i - 1])
                            
                            total_pokes_binned_within1.append(total_pokes_binned1[0])
                            for i in range (1, len(total_pokes_binned1)):
                                total_pokes_binned_within1.append(total_pokes_binned1[i] - total_pokes_binned1[i - 1])
                                
                            for i in range(0, len(active_binned1)):
                                if active_binned1[i] == 'Left':
                                    if left_binned_within1[i] + right_binned_within1[i] != 0:
                                        percent_active_binned_within1.append(left_binned_within1[i] / total_pokes_binned_within1[i] * 100)
                                    else:
                                        percent_active_binned_within1.append('NA')
                                else:
                                    if left_binned_within1[i] + right_binned_within1[i] != 0:
                                        percent_active_binned_within1.append(right_binned_within1[i] / total_pokes_binned_within1[i] * 100)
                                    else:
                                        percent_active_binned_within1.append('NA')
                        
                        if seconds_in_bins2 != '':
                        
                            end_time = seconds_elapsed[-1]
                            
                            # Create end value that is the next greatest multiple of the required time bin duration greater than the end_time so that those data points are retained in the output
                            
                            excess = math.ceil(end_time / int(seconds_in_bins2)) # Rounds up so that the number of bins is the multiple of bin length greater than the time point
                            
                            end_bin = excess * int(seconds_in_bins2) # Creates the end value for the last time bin
                            
                            interval_range = pd.interval_range(start=0, freq=int(seconds_in_bins2), end=end_bin, closed="left") #'left' means that it includes the first data point which is 0 because it's start time
                            
                            time_bins = df[seconds_elapsed] = pd.cut(seconds_elapsed, interval_range, include_lowest=True, ordered=True)
                            
                            # Create dictionary of time bins as keys with bin number as matching entry to label the bins
                            
                            # Create a list of integers of the same length as there are time bins
                            
                            bin_number = []
                            a = 1
                            for i in range(0, len(interval_range)):
                                bin_number.append(a)
                                a += 1
                            
                            make_dictionary = zip(interval_range, bin_number)
                            
                            bin_dictionary = dict(make_dictionary)
                            
                            bin_num = []
                            for i in range(0, len(time_bins)):
                                bin_num.append(bin_dictionary.get((time_bins[i])))
                            
                            # Get the last entry into each bin for the rest of the columns required
                            
                            bin_num_binned2 = []
                            left_binned2 = []
                            right_binned2 = []
                            active_binned2 = []
                            session_type_binned2 = []
                            percent_active_binned2 = []
                            total_pokes_binned2 = []
                            
                            # time binned
                            for i in range(1, len(time_bins)):
                                    if bin_num[i - 1] != bin_num[i]:
                                        bin_num_binned2.append(bin_num[i - 1])
                            bin_num_binned2.append(bin_num[i - 0])
                            
                            # left binned
                            for i in range(1, len(left_poke)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        left_binned2.append(left_poke[i - 1])
                            left_binned2.append(left_poke[i - 0])
                                    
                            # right binned
                            for i in range(1, len(right_poke)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        right_binned2.append(right_poke[i - 1])
                            right_binned2.append(right_poke[i - 0])
                                 
                            # active binned
                            for i in range(1, len(active)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        active_binned2.append(active[i - 1])
                            active_binned2.append(active[i - 0])
                            
                            # session type binned
                            for i in range(1, len(session_type)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        session_type_binned2.append(session_type[i - 1])
                            session_type_binned2.append(session_type[i - 0])
                            
                            # total pokes binned
                            for i in range(1, len(total_pokes)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        total_pokes_binned2.append(total_pokes[i - 1])
                            total_pokes_binned2.append(total_pokes[i - 0])
                            
                            # percent active binned
                            for i in range(0, len(active_binned2)):
                                if active_binned2[i] == 'Left':
                                    percent_active_binned2.append(left_binned2[i] / total_pokes_binned2[i] * 100)
                                else:
                                    percent_active_binned2.append(right_binned2[i] / total_pokes_binned2[i] * 100)
                            
                            # Account for potential empty first bin
                            
                            bin_counter = 1
                            bin_index = 0
                            
                            for i in range(0, len(bin_num)):
                                if bin_num_binned2[bin_index] != bin_counter:
                                    bin_num_binned2.insert(bin_counter - 1, bin_counter)
                                    left_binned2.insert(bin_counter - 1, 0)
                                    right_binned2.insert(bin_counter - 1, 0)
                                    active_binned2.insert(bin_counter - 1, active[i])
                                    session_type_binned2.insert(bin_counter - 1, session_type[0])
                                    total_pokes_binned2.insert(bin_counter - 1, 0)
                                    percent_active_binned2.insert(bin_counter - 1, np.nan)
                            
                            # Enter new row for any bins that have no data in them (for cumulative bins this is a duplication of the preceding row)
                        
                            active_port = active_binned2[0]
                            session_type3 = session_type[0]
                                  
                            for i in bin_dictionary:
                                if bin_dictionary[i] in bin_num_binned2:
                                    continue
                                else:
                                    bin_num_binned2.insert((bin_dictionary[i] - 1), bin_dictionary[i])
                                    left_binned2.insert((bin_dictionary[i] - 1), left_binned2[(bin_dictionary[i] - 2)])
                                    right_binned2.insert((bin_dictionary[i] - 1), right_binned2[(bin_dictionary[i] - 2)])
                                    active_binned2.insert((bin_dictionary[i] - 1), active_port)
                                    percent_active_binned2.insert((bin_dictionary[i] - 1), percent_active_binned2[(bin_dictionary[i] - 2)])
                                    session_type_binned2.insert((bin_dictionary[i] - 1), session_type3)
                                    total_pokes_binned2.insert((bin_dictionary[i] - 1), total_pokes_binned2[(bin_dictionary[i] - 2)])
                                    
                            # Make new counts that are within the bins rather than cumulative
                            
                            left_binned_within2 = []
                            right_binned_within2 = []
                            percent_active_binned_within2 = []
                            total_pokes_binned_within2 = []
                            
                            left_binned_within2.append(left_binned2[0])
                            for i in range(1, len(left_binned2)):
                                left_binned_within2.append(left_binned2[i] - left_binned2[i - 1])
                            
                            right_binned_within2.append(right_binned2[0])
                            for i in range(1, len(right_binned2)):
                                right_binned_within2.append(right_binned2[i] - right_binned2[i - 1])
                            
                            total_pokes_binned_within2.append(total_pokes_binned2[0])
                            for i in range (1, len(total_pokes_binned2)):
                                total_pokes_binned_within2.append(total_pokes_binned2[i] - total_pokes_binned2[i - 1])
                                
                            for i in range(0, len(active_binned2)):
                                if active_binned2[i] == 'Left':
                                    if left_binned_within2[i] + right_binned_within2[i] != 0:
                                        percent_active_binned_within2.append(left_binned_within2[i] / total_pokes_binned_within2[i] * 100)
                                    else:
                                        percent_active_binned_within2.append('NA')
                                else:
                                    if left_binned_within2[i] + right_binned_within2[i] != 0:
                                        percent_active_binned_within2.append(right_binned_within2[i] / total_pokes_binned_within2[i] * 100)
                                    else:
                                        percent_active_binned_within2.append('NA')
                            
                            
                        #####----- Create session summary data-----#####
                        
                        task = session_type[0]
                        
                        session_hours = str(math.floor(seconds_elapsed[-1] / 3600))
                        
                        session_mins = math.floor((seconds_elapsed[-1] % 3600) / 60)
                        
                        session_secs = seconds_elapsed[-1] % 60
                        
                        if session_mins < 10:
                            mins = '0' + str(session_mins)
                        else:
                            mins = str(session_mins)
                            
                        if session_secs < 10:
                            secs = '0' + str(session_secs)
                        else:
                            secs = str(session_secs)
                            
                        session_duration = str(session_hours) + ':' + mins + ':' + secs
                        
                        # Assign left and right pokes as active or inactive based on the active port
                        
                        if active[0] == 'Left':
                            active_pokes = left_poke[-1]
                            inactive_pokes = right_poke[-1]
                        elif active[0] == 'Right':
                            active_pokes = right_poke[-1]
                            inactive_pokes = left_poke[-1]
                        
                        active_port = active[0]
                        
                        date = import_name[7:13]
                        aus_date = date[2:4] + '/' + date[0:2] + '/20' + date[4:]
        
                        variable = ['Filename', 'Date', 'Task', 'Duration', 'Active port', 'Total Pokes', 'Active Pokes', 'Inactive Pokes', '% Active Pokes']
                        
                        value = [import_name.strip('.CSV'), aus_date, task, session_duration, active[0], total_pokes[-1], active_pokes, inactive_pokes, percent_active[-1]]
                        
                        #####-----
                        
                        # Create aus_date column for bins
                        
                        if seconds_in_bins1 != '':
                            
                            aus_date_binned1 = []
                            
                            for i in range(0, len(bin_num_binned1)):
                                
                                aus_date_binned1.append(aus_date)
                                
                        if seconds_in_bins2 != '':
                        
                            aus_date_binned2 = []
                            
                            for i in range(0, len(bin_num_binned2)):
                                
                                aus_date_binned2.append(aus_date)
                        
                        #####----- Export the data -----#####
                        
                        export_results = 'Y'
                        
                        if export_results == 'Y':
                            
                            # Always export Summary and Chronological Data
                            
                            results_summary = {'Variable': variable, 'Value': value}
                            export_file_summary = pd.DataFrame(results_summary, columns = ['Variable', 'Value'])
                            
                            # Assign left and right pokes as active or inactive based on the active port
                            
                            if active[0] == 'Left':
                                active_poke = left_poke
                                inactive_poke = right_poke
                                
                                if seconds_in_bins1 != '':
                                    active_poke_binned1 = left_binned1
                                    active_poke_binned_within1 = left_binned_within1
                                    inactive_poke_binned1 = right_binned1
                                    inactive_poke_binned_within1 = right_binned_within1
                                    
                                if seconds_in_bins2 != '':
                                    active_poke_binned2 = left_binned2
                                    active_poke_binned_within2 = left_binned_within2
                                    inactive_poke_binned2 = right_binned2
                                    inactive_poke_binned_within2 = right_binned_within2
                                    
                            elif active[0] == 'Right':
                                active_poke = right_poke
                                inactive_poke = left_poke
                                
                                if seconds_in_bins1 != '':
                                    active_poke_binned1 = right_binned1
                                    active_poke_binned_within1 = right_binned_within1
                                    inactive_poke_binned1 = left_binned1
                                    inactive_poke_binned_within1 = left_binned_within1
                                    
                                if seconds_in_bins2 != '':
                                    active_poke_binned2 = right_binned2
                                    active_poke_binned_within2 = right_binned_within2
                                    inactive_poke_binned2 = left_binned2
                                    inactive_poke_binned_within2 = left_binned_within2
                            
                            
                            results_chronological = {'Time (seconds)': seconds_elapsed, 'Session Type': session_type, 'Active port': active, 'Active Poke': active_poke, 'Inactive Poke': inactive_poke, 
                                                      'Total Poke': total_pokes, '% Active Pokes': percent_active}
                            export_file_chronological = pd.DataFrame(results_chronological, columns = ['Time (seconds)', 'Session Type', 'Active port', 'Active Poke', 'Inactive Poke',
                                                                                                        'Total Poke', '% Active Pokes'])
                
                                
                            # If using time bins also export binned data (cumulative and within)
                            
                            if seconds_in_bins1 != '':
                            
                                results_binned_c1 = {'Date': aus_date_binned1, 'Time bin (' + seconds_in_bins1 + 's each)': bin_num_binned1, 'Session Type': session_type_binned1, 'Active Port': active_binned1, 
                                                    'Active Poke': active_poke_binned1, 'Inactive Poke': inactive_poke_binned1, 'Total Poke': total_pokes_binned1, '% Active Pokes': percent_active_binned1}
                                export_file_binned_c1 = pd.DataFrame(results_binned_c1, columns = ['Date', 'Time bin (' + seconds_in_bins1 + 's each)', 'Session Type', 'Active Port', 
                                                                                                  'Active Poke', 'Inactive Poke', 'Total Poke', '% Active Pokes'])
                                
                                results_binned_w1 = {'Date': aus_date_binned1, 'Time bin (' + seconds_in_bins1 + 's each)': bin_num_binned1, 'Session Type': session_type_binned1, 'Active Port': active_binned1, 
                                                    'Active Poke': active_poke_binned_within1, 'Inactive Poke': inactive_poke_binned_within1, 'Total Poke': total_pokes_binned_within1, '% Active Pokes': percent_active_binned_within1}
                                export_file_binned_w1 = pd.DataFrame(results_binned_w1, columns = ['Date', 'Time bin (' + seconds_in_bins1 + 's each)', 'Session Type', 'Active Port', 
                                                                                                  'Active Poke', 'Inactive Poke', 'Total Poke', '% Active Pokes'])
                            
                            if seconds_in_bins2 != '':
                            
                                results_binned_c2 = {'Date': aus_date_binned2, 'Time bin (' + seconds_in_bins2 + 's each)': bin_num_binned2, 'Session Type': session_type_binned2, 'Active Port': active_binned2, 
                                                    'Active Poke': active_poke_binned2, 'Inactive Poke': inactive_poke_binned2, 'Total Poke': total_pokes_binned2, '% Active Pokes': percent_active_binned2}
                                export_file_binned_c2 = pd.DataFrame(results_binned_c2, columns = ['Date', 'Time bin (' + seconds_in_bins2 + 's each)', 'Session Type', 'Active Port', 
                                                                                                  'Active Poke', 'Inactive Poke', 'Total Poke', '% Active Pokes'])
                                
                                results_binned_w2 = {'Date': aus_date_binned2, 'Time bin (' + seconds_in_bins2 + 's each)': bin_num_binned2, 'Session Type': session_type_binned2, 'Active Port': active_binned2, 
                                                    'Active Poke': active_poke_binned_within2, 'Inactive Poke': inactive_poke_binned_within2, 'Total Poke': total_pokes_binned_within2, '% Active Pokes': percent_active_binned_within2}
                                export_file_binned_w2 = pd.DataFrame(results_binned_w2, columns = ['Date', 'Time bin (' + seconds_in_bins2 + 's each)', 'Session Type', 'Active Port', 
                                                                                                  'Active Poke', 'Inactive Poke', 'Total Poke', '% Active Pokes'])
        
                            
                            # Export to excel
                            
                            from openpyxl import Workbook
                        
                            wb = Workbook()
                            
                            ws1 = wb.active
                            ws1.title = 'Summary'
                            
                            ws2 = wb.create_sheet()
                            ws2.title = 'Chronological'
                            
                            results_to_export = [export_file_summary, export_file_chronological]
                            
                            if seconds_in_bins1 != '':
                                
                                bin_min_length1 = int(seconds_in_bins1) / 60
                                
                                ws3 = wb.create_sheet()
                                ws3.title = str(int(bin_min_length1)) + 'min Binned Cumulative'
                                
                                ws4 = wb.create_sheet()
                                ws4.title = str(int(bin_min_length1)) + 'min Binned Within'
                                
                                results_to_export.append(export_file_binned_c1)
                                results_to_export.append(export_file_binned_w1)
                                
                            if seconds_in_bins2 != '':
                                
                                bin_min_length2 = int(seconds_in_bins2) / 60
                                
                                ws5 = wb.create_sheet()
                                ws5.title = str(int(bin_min_length2)) + 'min Binned Cumulative'
                                
                                ws6 = wb.create_sheet()
                                ws6.title = str(int(bin_min_length2)) + 'min Binned Within'
                                
                                results_to_export.append(export_file_binned_c2)
                                results_to_export.append(export_file_binned_w2)
                            
                            sheets_to_export = wb.sheetnames
                            
                            with pd.ExcelWriter(export_destination) as writer:
                                
                                for i in range(len(sheets_to_export)):
                                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
                    
                            print('complete')
                    
        # ##########---------- Overview Sheets----------##########
        
        # from openpyxl import Workbook
                
        # wb = Workbook()
        
        # ws = wb.active
        # ws.title = 'Summary Overview'
        
        # value_column = 'Value'
        # variable_column = 'Variable'
        
        # value_summary = []
        # session_num = []
        # counter = 1
        
        # # export_location_folder = export_location + folder + '/'
        
        # for filename in sorted(os.listdir(os.path.join(export_location, folder))):
            
        #     if filename.endswith(".xlsx"):
                
        #         if counter == 1:
                    
        #             export_name = filename
        #             export_destination = export_location + folder + '/' + export_name
                    
        #             df_overview = pd.read_excel(export_destination, sheet_name = 'Summary')
                    
        #             df_overview.drop(columns='Value', inplace=True)
                    
        #         export_name = filename
        #         export_destination = export_location + folder + '/' + export_name
                
        #         name = 'Session ' + str(counter)
        #         session_num.append(name)
                
        #         df = pd.read_excel(export_destination, sheet_name = 'Summary')
                
        #         values = df[value_column].tolist()
                
        #         df_overview.insert(counter, name, values)
                
        #         counter += 1
                    
        #     overview_name = folder + ' FR Training Overview.xlsx'
        #     overview_destination = export_location + folder + '/' + overview_name
            
        #     with pd.ExcelWriter(overview_destination) as writer:
                            
        #         df_overview.to_excel(writer, sheet_name='Overview', engine='openpyxl', index=False, header=True)
            
        #     # cohort_overview_destination = cohort_export_location + overview_name
            
        #     # with pd.ExcelWriter(cohort_overview_destination) as writer:
                            
        #     #     df_overview.to_excel(writer, sheet_name='Overview', engine='openpyxl', index=False, header=True)
            
        # print(filename, 'Overview complete')

        