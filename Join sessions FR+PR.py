#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Aug  3 12:33:45 2021

@author: lauramilton
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jul 28 15:27:07 2021

@author: lauramilton
"""

date_column = 'Date'
time_column1 = 'Time bin (600s each)'
time_column2 = 'Time bin (1800s each)'
session_column = 'Session Type'
active_column_sum = 'Active port'
active_column = 'Active Port'
active_poke_column = 'Active Poke'
inactive_poke_column = 'Inactive Poke'
total_poke_column = 'Total Poke'
pellet_column = 'Pellet Count'
percent_active_column = '% Active Pokes'

time_columnPR = 'Seconds Elapsed'
step_column = 'PR Step'

session_length = '180' # session length in minutes
time_bin_length1 = '10'
time_bin_length2 = '30'

import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/FED Data Labelled files/VSG/Output/'
export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/FED Data Labelled files/VSG/Output/'
group_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/FED Data Labelled files/Sham/Output/Joined/'
cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/FED Data Labelled files/Cohort/Individual Joined/'

import pandas as pd
import numpy as np
import os
import openpyxl

num_of_bins1 = int(session_length) / int(time_bin_length1)
num_of_bins2 = int(session_length) / int(time_bin_length2)

for folder in sorted(os.listdir(import_location)):
    if folder.startswith('Mouse ID#'):
        
        # session summary values
        date_summary_all = []
        time_summary_all = []
        session_summary_all = []
        active_summary_all = []
        active_poke_summary_all = []
        inactive_poke_summary_all = []
        total_poke_summary_all = []
        pellet_summary_all = []
        percent_active_summary_all = []
        
        date_summary_FR = []
        time_summary_FR = []
        session_summary_FR = []
        active_summary_FR = []
        active_poke_summary_FR = []
        inactive_poke_summary_FR = []
        total_poke_summary_FR = []
        pellet_summary_FR = []
        percent_active_summary_FR = []

        date_summary_PR = []
        time_summary_PR = []
        session_summary_PR = []
        active_summary_PR = []
        active_poke_summary_PR = []
        inactive_poke_summary_PR = []
        total_poke_summary_PR = []
        pellet_summary_PR = []
        percent_active_summary_PR = []

        # FR time bin 1 within values
        date_1w = []
        time_1w = []
        session_1w = []
        active_1w = []
        active_poke_1w = []
        inactive_poke_1w = []
        total_poke_1w = []
        pellet_1w = []
        percent_active_1w = []
        
        # FR time bin 1 cumulative values
        date_1c = []
        time_1c = []
        session_1c = []
        active_1c = []
        active_poke_1c = []
        inactive_poke_1c = []
        total_poke_1c = []
        pellet_1c = []
        percent_active_1c = []

        # FR time bin 2 within values
        date_2w = []
        time_2w = []
        session_2w = []
        active_2w = []
        active_poke_2w = []
        inactive_poke_2w = []
        total_poke_2w = []
        pellet_2w = []
        percent_active_2w = []
        
        # FR time bin 2 cumulative values
        date_2c = []
        time_2c = []
        session_2c = []
        active_2c = []
        active_poke_2c = []
        inactive_poke_2c = []
        total_poke_2c = []
        pellet_2c = []
        percent_active_2c = []
        
        # PR step cumulative values
        date_PRc = []
        time_PRc = []
        step_PRc = []
        active_PRc = []
        active_poke_PRc = []
        inactive_poke_PRc = []
        total_poke_PRc = []
        pellet_PRc = []
        percent_active_PRc = []

        # PR step within values
        date_PRw = []
        time_PRw = []
        step_PRw = []
        active_PRw = []
        active_poke_PRw = []
        inactive_poke_PRw = []
        total_poke_PRw = []
        pellet_PRw = []
        percent_active_PRw = []

        session_num = 1
        
        for filename in sorted(os.listdir(os.path.join(import_location, folder))):
            
            if filename.endswith(".xlsx"):
                if filename.startswith('Mouse'): # can probs turn it around to be if filename.startswith() and use variable/name from the code that generates those files
                    if not filename.endswith('Overview.xlsx'): # will be able to remove later once those files aren't in the folder anymore (i.e. when file naming procedure is finalized)
                        if not filename.endswith('joined.xlsx'):    
                            print(filename)
                            import_name = filename
                            import_destination = import_location + folder + '/' + import_name
                            
                            if 'FR Training' in filename:
                                df_sum = pd.read_excel(import_destination, sheet_name = 'Chronological')
                                session_sum = df_sum[session_column].tolist()
                            
                            elif 'Progressive Ratio' in filename:
                                df_sum = pd.read_excel(import_destination, sheet_name = 'Chronological Cumulative')
                            
                            active_sum = df_sum[active_column_sum].tolist()
                            active_poke_sum = df_sum[active_poke_column].tolist()
                            inactive_poke_sum = df_sum[inactive_poke_column].tolist()
                            total_poke_sum = df_sum[total_poke_column].tolist()
                            pellet_sum = df_sum[pellet_column].tolist()
                            percent_active_sum = df_sum[percent_active_column].tolist()
                            
    
                            if 'FR Training' in filename:
                            
                                import_sheet1 = str(time_bin_length1) + 'min Binned Within'
                                df1w = pd.read_excel(import_destination, sheet_name = import_sheet1)
                                
                                import_sheet2 = str(time_bin_length1) + 'min Binned Cumulative'
                                df1c = pd.read_excel(import_destination, sheet_name = import_sheet2)
        
                                date1w = df1w[date_column].tolist()
                                time1w = df1w[time_column1].tolist()
                                session1w = df1w[session_column].tolist()
                                active1w = df1w[active_column].tolist()
                                active_poke1w = df1w[active_poke_column].tolist()
                                inactive_poke1w = df1w[inactive_poke_column].tolist()
                                total_poke1w = df1w[total_poke_column].tolist()
                                pellet1w = df1w[pellet_column].tolist()
                                percent_active1w = df1w[percent_active_column].tolist()
                                
                                date1c = df1c[date_column].tolist()
                                time1c = df1c[time_column1].tolist()
                                session1c = df1c[session_column].tolist()
                                active1c = df1c[active_column].tolist()
                                active_poke1c = df1c[active_poke_column].tolist()
                                inactive_poke1c = df1c[inactive_poke_column].tolist()
                                total_poke1c = df1c[total_poke_column].tolist()
                                pellet1c = df1c[pellet_column].tolist()
                                percent_active1c = df1c[percent_active_column].tolist()
                                
                                if time_bin_length2 != '':
                                    import_sheet3 = str(time_bin_length2) + 'min Binned Within'
                                    df2w = pd.read_excel(import_destination, sheet_name = import_sheet3)
                                    import_sheet4 = str(time_bin_length2) + 'min Binned Cumulative'
                                    df2c = pd.read_excel(import_destination, sheet_name = import_sheet4)
    
                                    date2w = df2w[date_column].tolist()
                                    time2w = df2w[time_column2].tolist()
                                    session2w = df2w[session_column].tolist()
                                    active2w = df2w[active_column].tolist()
                                    active_poke2w = df2w[active_poke_column].tolist()
                                    inactive_poke2w = df2w[inactive_poke_column].tolist()
                                    total_poke2w = df2w[total_poke_column].tolist()
                                    pellet2w = df2w[pellet_column].tolist()
                                    percent_active2w = df2w[percent_active_column].tolist()
                                    
                                    date2c = df2c[date_column].tolist()
                                    time2c = df2c[time_column2].tolist()
                                    session2c = df2c[session_column].tolist()
                                    active2c = df2c[active_column].tolist()
                                    active_poke2c = df2c[active_poke_column].tolist()
                                    inactive_poke2c = df2c[inactive_poke_column].tolist()
                                    total_poke2c = df2c[total_poke_column].tolist()
                                    pellet2c = df2c[pellet_column].tolist()
                                    percent_active2c = df2c[percent_active_column].tolist()
                                
                                if len(time1w) > num_of_bins1:
                                    
                                    del date1w[int(num_of_bins1):]
                                    del time1w[int(num_of_bins1):]
                                    del session1w[int(num_of_bins1):]
                                    del active1w[int(num_of_bins1):]
                                    del active_poke1w[int(num_of_bins1):]
                                    del inactive_poke1w[int(num_of_bins1):]
                                    del total_poke1w[int(num_of_bins1):]
                                    del pellet1w[int(num_of_bins1):]
                                    del percent_active1w[int(num_of_bins1):]
                                    
                                    del date1c[int(num_of_bins1):]
                                    del time1c[int(num_of_bins1):]
                                    del session1c[int(num_of_bins1):]
                                    del active1c[int(num_of_bins1):]
                                    del active_poke1c[int(num_of_bins1):]
                                    del inactive_poke1c[int(num_of_bins1):]
                                    del total_poke1c[int(num_of_bins1):]
                                    del pellet1c[int(num_of_bins1):]
                                    del percent_active1c[int(num_of_bins1):]

                                if time_bin_length2 != '':
                                    if len(time2w) > num_of_bins2:
                                        
                                        del date2w[int(num_of_bins2):]
                                        del time2w[int(num_of_bins2):]
                                        del session2w[int(num_of_bins2):]
                                        del active2w[int(num_of_bins2):]
                                        del active_poke2w[int(num_of_bins2):]
                                        del inactive_poke2w[int(num_of_bins2):]
                                        del total_poke2w[int(num_of_bins2):]
                                        del pellet2w[int(num_of_bins2):]
                                        del percent_active2w[int(num_of_bins2):]
                                        
                                        del date2c[int(num_of_bins2):]
                                        del time2c[int(num_of_bins2):]
                                        del session2c[int(num_of_bins2):]
                                        del active2c[int(num_of_bins2):]
                                        del active_poke2c[int(num_of_bins2):]
                                        del inactive_poke2c[int(num_of_bins2):]
                                        del total_poke2c[int(num_of_bins2):]
                                        del pellet2c[int(num_of_bins2):]
                                        del percent_active2c[int(num_of_bins2):]

                                        
                                if len(time1w) < num_of_bins1:
                                    
                                    bin_num1 = len(time1w) + 1
                                    
                                    while len(time1w) < num_of_bins1:
                                        date1w.append(date1w[0])
                                        time1w.append(bin_num1)
                                        session1w.append(np.nan)
                                        active1w.append(np.nan)
                                        active_poke1w.append(np.nan)
                                        inactive_poke1w.append(np.nan)
                                        total_poke1w.append(np.nan)
                                        pellet1w.append(np.nan)
                                        percent_active1w.append(np.nan)
                                        
                                        date1c.append(date1c[0])
                                        time1c.append(bin_num1)
                                        session1c.append(np.nan)
                                        active1c.append(np.nan)
                                        active_poke1c.append(np.nan)
                                        inactive_poke1c.append(np.nan)
                                        total_poke1c.append(np.nan)
                                        pellet1c.append(np.nan)
                                        percent_active1c.append(np.nan)

                                        bin_num1 +=1
                                
                                if time_bin_length2 != '':
                                    if len(time2w) < num_of_bins2:
                                    
                                        bin_num2 = len(time2w) + 1
                                        
                                        while len(time2w) < num_of_bins2:
                                            date2w.append(date2w[0])
                                            time2w.append(bin_num2)
                                            session2w.append(np.nan)
                                            active2w.append(np.nan)
                                            active_poke2w.append(np.nan)
                                            inactive_poke2w.append(np.nan)
                                            total_poke2w.append(np.nan)
                                            pellet2w.append(np.nan)
                                            percent_active2w.append(np.nan)
                                            
                                            date2c.append(date2c[0])
                                            time2c.append(bin_num2)
                                            session2c.append(np.nan)
                                            active2c.append(np.nan)
                                            active_poke2c.append(np.nan)
                                            inactive_poke2c.append(np.nan)
                                            total_poke2c.append(np.nan)
                                            pellet2c.append(np.nan)
                                            percent_active2c.append(np.nan)

                                            bin_num2 +=1
                                            
                                

                                
                                for a, b, c, d, e, f, g, h, i in zip(time1w, session1w, active1w, active_poke1w, inactive_poke1w, total_poke1w, pellet1w, percent_active1w, date1w):
                                        
                                        time_1w.append(str(session_num) + '.' + str(a))
                                        session_1w.append(b)
                                        active_1w.append(c) 
                                        active_poke_1w.append(d)
                                        inactive_poke_1w.append(e)
                                        total_poke_1w.append(f)
                                        pellet_1w.append(g)
                                        percent_active_1w.append(h)
                                        date_1w.append(i)
                                        
                                for a, b, c, d, e, f, g, h, i in zip(time1c, session1c, active1c, active_poke1c, inactive_poke1c, total_poke1c, pellet1c, percent_active1c, date1c):
                                        
                                        time_1c.append(str(session_num) + '.' + str(a))
                                        session_1c.append(b)
                                        active_1c.append(c) 
                                        active_poke_1c.append(d)
                                        inactive_poke_1c.append(e)
                                        total_poke_1c.append(f)
                                        pellet_1c.append(g)
                                        percent_active_1c.append(h)
                                        date_1c.append(i)

                                
                                if time_bin_length2 != '':
                                    for a, b, c, d, e, f, g, h, i in zip(time2w, session2w, active2w, active_poke2w, inactive_poke2w, total_poke2w, pellet2w, percent_active2w, date2w):
                                        
                                            time_2w.append(str(session_num) + '.' + str(a))
                                            session_2w.append(b)
                                            active_2w.append(c) 
                                            active_poke_2w.append(d)
                                            inactive_poke_2w.append(e)
                                            total_poke_2w.append(f)
                                            pellet_2w.append(g)
                                            percent_active_2w.append(h)
                                            date_2w.append(i)
                                    
                                    for a, b, c, d, e, f, g, h, i in zip(time2c, session2c, active2c, active_poke2c, inactive_poke2c, total_poke2c, pellet2c, percent_active2c, date2c):
                                        
                                            time_2c.append(str(session_num) + '.' + str(a))
                                            session_2c.append(b)
                                            active_2c.append(c) 
                                            active_poke_2c.append(d)
                                            inactive_poke_2c.append(e)
                                            total_poke_2c.append(f)
                                            pellet_2c.append(g)
                                            percent_active_2c.append(h)
                                            date_2c.append(i)

                                
                                
                                session_1w = ['NA' if x != x else x for x in session_1w]
                                active_1w = ['NA' if x != x else x for x in active_1w] 
                                active_poke_1w = ['NA' if x != x else x for x in active_poke_1w]
                                inactive_poke_1w = ['NA' if x != x else x for x in inactive_poke_1w]
                                total_poke_1w = ['NA' if x != x else x for x in total_poke_1w]
                                pellet_1w = ['NA' if x != x else x for x in pellet_1w]
                                percent_active_1w = ['NA' if x != x else x for x in percent_active_1w]
                                
                                session_1c = ['NA' if x != x else x for x in session_1c]
                                active_1c = ['NA' if x != x else x for x in active_1c] 
                                active_poke_1c = ['NA' if x != x else x for x in active_poke_1c]
                                inactive_poke_1c = ['NA' if x != x else x for x in inactive_poke_1c]
                                total_poke_1c = ['NA' if x != x else x for x in total_poke_1c]
                                pellet_1c = ['NA' if x != x else x for x in pellet_1c]
                                percent_active_1c = ['NA' if x != x else x for x in percent_active_1c]

                                
                                if time_bin_length2 != '':
                                    session_2w = ['NA' if x != x else x for x in session_2w]
                                    active_2w = ['NA' if x != x else x for x in active_2w] 
                                    active_poke_2w = ['NA' if x != x else x for x in active_poke_2w]
                                    inactive_poke_2w = ['NA' if x != x else x for x in inactive_poke_2w]
                                    total_poke_2w = ['NA' if x != x else x for x in total_poke_2w]
                                    pellet_2w = ['NA' if x != x else x for x in pellet_2w]
                                    percent_active_2w = ['NA' if x != x else x for x in percent_active_2w]
                                    
                                    session_2c = ['NA' if x != x else x for x in session_2c]
                                    active_2c = ['NA' if x != x else x for x in active_2c] 
                                    active_poke_2c = ['NA' if x != x else x for x in active_poke_2c]
                                    inactive_poke_2c = ['NA' if x != x else x for x in inactive_poke_2c]
                                    total_poke_2c = ['NA' if x != x else x for x in total_poke_2c]
                                    pellet_2c = ['NA' if x != x else x for x in pellet_2c]
                                    percent_active_2c = ['NA' if x != x else x for x in percent_active_2c]
                            
                            
                            elif 'Progressive Ratio' in filename:
                                dfPRc = pd.read_excel(import_destination, sheet_name = 'PR Step Cumulative')
                                dfPRw = pd.read_excel(import_destination, sheet_name = 'PR Step Within')
                            
                                datePRc = dfPRc[date_column].tolist()
                                timePRc = dfPRc[time_columnPR].tolist()
                                stepPRc = dfPRc[step_column].tolist()
                                activePRc = dfPRc[active_column].tolist()
                                active_pokePRc = dfPRc[active_poke_column].tolist()
                                inactive_pokePRc = dfPRc[inactive_poke_column].tolist()
                                total_pokePRc = dfPRc[total_poke_column].tolist()
                                pelletPRc = dfPRc[pellet_column].tolist()
                                percent_activePRc = dfPRc[percent_active_column].tolist()

                                datePRw = dfPRw[date_column].tolist()
                                timePRw = dfPRw[time_columnPR].tolist()
                                stepPRw = dfPRw[step_column].tolist()
                                activePRw = dfPRw[active_column].tolist()
                                active_pokePRw = dfPRw[active_poke_column].tolist()
                                inactive_pokePRw = dfPRw[inactive_poke_column].tolist()
                                total_pokePRw = dfPRw[total_poke_column].tolist()
                                pelletPRw = dfPRw[pellet_column].tolist()
                                percent_activePRw = dfPRw[percent_active_column].tolist()
                            
                                for a, b, c, d, e, f, g, h, i in zip(timePRw, stepPRw, activePRw, active_pokePRw, inactive_pokePRw, total_pokePRw, pelletPRw, percent_activePRw, datePRw):
                                        
                                        time_PRw.append(str(session_num) + '.' + str(a))
                                        step_PRw.append(b)
                                        active_PRw.append(c) 
                                        active_poke_PRw.append(d)
                                        inactive_poke_PRw.append(e)
                                        total_poke_PRw.append(f)
                                        pellet_PRw.append(g)
                                        percent_active_PRw.append(h)
                                        date_PRw.append(i)

                                for a, b, c, d, e, f, g, h, i in zip(timePRc, stepPRc, activePRc, active_pokePRc, inactive_pokePRc, total_pokePRc, pelletPRc, percent_activePRc, datePRc):
                                        
                                        time_PRc.append(str(session_num) + '.' + str(a))
                                        step_PRc.append(b)
                                        active_PRc.append(c) 
                                        active_poke_PRc.append(d)
                                        inactive_poke_PRc.append(e)
                                        total_poke_PRc.append(f)
                                        pellet_PRc.append(g)
                                        percent_active_PRc.append(h)
                                        date_PRc.append(i)
                            
                            
                            if 'FR Training' in filename:
                                session_summary_all.append(session_sum[-1])
                                date_summary_all.append(date1c[-1])
                                
                                session_summary_FR.append(session_sum[-1])
                                date_summary_FR.append(date1c[-1])
                                time_summary_FR.append(str(session_num))
                                active_summary_FR.append(active_sum[-1]) 
                                active_poke_summary_FR.append(active_poke_sum[-1])
                                inactive_poke_summary_FR.append(inactive_poke_sum[-1])
                                total_poke_summary_FR.append(total_poke_sum[-1])
                                pellet_summary_FR.append(pellet_sum[-1])
                                percent_active_summary_FR.append(percent_active_sum[-1])



                            elif 'Progressive Ratio' in filename:
                                session_summary_all.append('PR')
                                date_summary_all.append(datePRc[-1])

                                session_summary_PR.append('PR')
                                date_summary_PR.append(datePRc[-1])
                                time_summary_PR.append(str(session_num))
                                active_summary_PR.append(active_sum[-1]) 
                                active_poke_summary_PR.append(active_poke_sum[-1])
                                inactive_poke_summary_PR.append(inactive_poke_sum[-1])
                                total_poke_summary_PR.append(total_poke_sum[-1])
                                pellet_summary_PR.append(pellet_sum[-1])
                                percent_active_summary_PR.append(percent_active_sum[-1])

                            
                            time_summary_all.append(str(session_num))
                            active_summary_all.append(active_sum[-1]) 
                            active_poke_summary_all.append(active_poke_sum[-1])
                            inactive_poke_summary_all.append(inactive_poke_sum[-1])
                            total_poke_summary_all.append(total_poke_sum[-1])
                            pellet_summary_all.append(pellet_sum[-1])
                            percent_active_summary_all.append(percent_active_sum[-1])

                            session_num += 1
                            
                            ##### Export results
                               
                            if 'FR Training' in filename:
                                results_summary_FR = {'Date': date_summary_FR, 'Session Number': time_summary_FR, 'Session Type': session_summary_FR, 'Active Port': active_summary_FR, 'Active Poke': active_poke_summary_FR, 'Inactive Poke': inactive_poke_summary_FR, 'Total Poke': total_poke_summary_FR, 'Pellet Count': pellet_summary_FR, 'Percent Active': percent_active_summary_FR}
                                export_file_summary_FR = pd.DataFrame(results_summary_FR, columns = ['Date', 'Session Number', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
                            
                            elif 'Progressive Ratio' in filename:
                                results_summary_PR = {'Date': date_summary_PR, 'Session Number': time_summary_PR, 'Session Type': session_summary_PR, 'Active Port': active_summary_PR, 'Active Poke': active_poke_summary_PR, 'Inactive Poke': inactive_poke_summary_PR, 'Total Poke': total_poke_summary_PR, 'Pellet Count': pellet_summary_PR, 'Percent Active': percent_active_summary_PR}
                                export_file_summary_PR = pd.DataFrame(results_summary_PR, columns = ['Date', 'Session Number', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])

                            if 'FR Training' in filename:
                                results1w = {'Date': date_1w, 'Time Bin': time_1w, 'Session Type': session_1w, 'Active Port': active_1w, 'Active Poke': active_poke_1w, 'Inactive Poke': inactive_poke_1w, 'Total Poke': total_poke_1w, 'Pellet Count': pellet_1w, 'Percent Active': percent_active_1w}
                                export_file1w = pd.DataFrame(results1w, columns = ['Date', 'Time Bin', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
    
                                results1c = {'Date': date_1c, 'Time Bin': time_1c, 'Session Type': session_1c, 'Active Port': active_1c, 'Active Poke': active_poke_1c, 'Inactive Poke': inactive_poke_1c, 'Total Poke': total_poke_1c, 'Pellet Count': pellet_1c, 'Percent Active': percent_active_1c}
                                export_file1c = pd.DataFrame(results1c, columns = ['Date', 'Time Bin', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
    
                                if time_bin_length2 != '':
                                    results2w = {'Date': date_2w, 'Time Bin': time_2w, 'Session Type': session_2w, 'Active Port': active_2w, 'Active Poke': active_poke_2w, 'Inactive Poke': inactive_poke_2w, 'Total Poke': total_poke_2w, 'Pellet Count': pellet_2w, 'Percent Active': percent_active_2w}
                                    export_file2w = pd.DataFrame(results2w, columns = ['Date', 'Time Bin', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
                                    
                                    results2c = {'Date': date_2c, 'Time Bin': time_2c, 'Session Type': session_2c, 'Active Port': active_2c, 'Active Poke': active_poke_2c, 'Inactive Poke': inactive_poke_2c, 'Total Poke': total_poke_2c, 'Pellet Count': pellet_2c, 'Percent Active': percent_active_2c}
                                    export_file2c = pd.DataFrame(results2c, columns = ['Date', 'Time Bin', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
                            
                            elif 'Progressive Ratio' in filename:
                                results_PRw = {'Date': date_PRw, 'Seconds Elapsed': time_PRw, 'PR Step': step_PRw, 'Active Port': active_PRw, 'Active Poke': active_poke_PRw, 'Inactive Poke': inactive_poke_PRw, 'Total Poke': total_poke_PRw, 'Pellet Count': pellet_PRw, 'Percent Active': percent_active_PRw}
                                export_file_PRw = pd.DataFrame(results_PRw, columns = ['Date', 'Seconds Elapsed', 'PR Step', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
                                
                                results_PRc = {'Date': date_PRc, 'Seconds Elapsed': time_PRc, 'PR Step': step_PRc, 'Active Port': active_PRc, 'Active Poke': active_poke_PRc, 'Inactive Poke': inactive_poke_PRc, 'Total Poke': total_poke_PRc, 'Pellet Count': pellet_PRc, 'Percent Active': percent_active_PRc}
                                export_file_PRc = pd.DataFrame(results_PRc, columns = ['Date', 'Time Bin', 'Seconds Elapsed', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
                            
                            # results_summary_all = {'Date': date_summary, 'Time Bin': time_summary, 'Session Type': session_summary, 'Active Port': active_summary, 'Active Poke': active_poke_summary, 'Inactive Poke': inactive_poke_summary, 'Total Poke': total_poke_summary, 'Pellet Count': pellet_summary, 'Percent Active': percent_active_summary}
                            # export_file_summary_all = pd.DataFrame(results_summary_all, columns = ['Date', 'Time Bin', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])

                            from openpyxl import Workbook
                        
                            wb = Workbook()
                            
                            ws1 = wb.active
                            ws1.title = str(session_length) + 'min summary joined'
                            
                            if 'FR Training' in filename:
                            
                                ws2 = wb.create_sheet()
                                ws2.title = str(time_bin_length1) + 'min within joined'
                                                                
                                ws3 = wb.create_sheet()
                                ws3.title = str(time_bin_length1) + 'min cumulative joined'
    
                                if time_bin_length2 != '':
                                    ws4 = wb.create_sheet()
                                    ws4.title = str(time_bin_length2) + 'min within joined'
                                    
                                    ws5 = wb.create_sheet()
                                    ws5.title = str(time_bin_length2) + 'min cumulative joined'
                            
                            elif 'Progressive Ratio' in filename:
                                
                                ws2 = wb.create_sheet()
                                ws2.title = 'PR step within joined'
                                
                                ws3 = wb.create_sheet()
                                ws3.title = 'PR step cumulative joined'
                            
                            results_to_export = []
                            
                            if 'FR Training' in filename:
                                results_to_export.append(export_file_summary_FR)
                                results_to_export.append(export_file1w)
                                results_to_export.append(export_file1c)
                            
                                if time_bin_length2 != '':
                                    results_to_export.append(export_file2w)
                                    results_to_export.append(export_file2c)
                            
                            elif 'Progressive Ratio' in filename:
                                results_to_export.append(export_file_summary_PR)
                                results_to_export.append(export_file_PRw)
                                results_to_export.append(export_file_PRc)
                            
                            sheets_to_export = wb.sheetnames
                            
                            if 'FR Training' in filename:
                                FR5_name = folder + ' FR5 joined.xlsx'
                                joined_destination = export_location + folder + '/' + FR5_name
                                
                            elif 'Progressive Ratio' in filename:
                                PR_name = folder + ' PR joined.xlsx'
                                joined_destination = export_location + folder + '/' + PR_name
        
                            with pd.ExcelWriter(joined_destination) as writer:
                                
                                for i in range(len(sheets_to_export)):
                                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
        
                            if 'FR Training' in filename:
                                joined_name = FR5_name
                                
                            elif 'Progressive Ratio' in filename:
                                joined_name = PR_name
        
                            group_joined_destination = group_export_location + joined_name
                            
                            with pd.ExcelWriter(group_joined_destination) as writer:
                                
                                for i in range(len(sheets_to_export)):
                                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)    
                                    
                                    
                            cohort_joined_destination = cohort_export_location + joined_name
                            
                            with pd.ExcelWriter(cohort_joined_destination) as writer:
                                
                                for i in range(len(sheets_to_export)):
                                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)    

        results_summary_all = {'Date': date_summary_all, 'Session Number': time_summary_all, 'Session Type': session_summary_all, 'Active Port': active_summary_all, 'Active Poke': active_poke_summary_all, 'Inactive Poke': inactive_poke_summary_all, 'Total Poke': total_poke_summary_all, 'Pellet Count': pellet_summary_all, 'Percent Active': percent_active_summary_all}
        export_file_summary_all = pd.DataFrame(results_summary_all, columns = ['Date', 'Session Number', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])

        from openpyxl import Workbook
    
        wb = Workbook()
        
        ws1 = wb.active
        ws1.title = str(session_length) + 'min summary joined'
        
        results_to_export = export_file_summary_all
        
        sheets_to_export = wb.sheetnames
        
        all_name = folder + ' all session summary joined.xlsx'
        all_destination = export_location + folder + '/' + all_name
        
        with pd.ExcelWriter(all_destination) as writer:
            for i in range(len(sheets_to_export)):
                export_file_summary_all.to_excel(writer, sheet_name = sheets_to_export[i], engine='openpyxl', index=False, header=True)
                
        group_destination = group_export_location + all_name
        
        with pd.ExcelWriter(group_destination) as writer:
            for i in range(len(sheets_to_export)):
                export_file_summary_all.to_excel(writer, sheet_name = sheets_to_export[i], engine='openpyxl', index=False, header=True)
        
            
        cohort_joined_destination = cohort_export_location + all_name
        
        with pd.ExcelWriter(cohort_joined_destination) as writer:
            for i in range(len(sheets_to_export)):
                export_file_summary_all.to_excel(writer, sheet_name = sheets_to_export[i], engine='openpyxl', index=False, header=True)

