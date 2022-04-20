#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Aug 12 15:24:48 2021

@author: lauramilton
"""

import pandas as pd
from matplotlib import pyplot as plt
from pylab import rcParams
rcParams['figure.figsize'] = 7,7 
import seaborn as sns
import numpy as np
sns.set(color_codes=True, font_scale=1.2)

# from heatmap import heatmap, corrplot
import openpyxl
from openpyxl import Workbook

#####-----

def heatmap(x, y, **kwargs):
    if 'color' in kwargs:
        color = kwargs['color']
    else:
        color = [1]*len(x)

    if 'palette' in kwargs:
        palette = kwargs['palette']
        # n_colors = len(palette)
        n_colors = 256
    else:
        n_colors = 256 # Use 256 colors for the diverging color palette
        palette = sns.color_palette("icefire", n_colors) 

    if 'color_range' in kwargs:
        color_min, color_max = kwargs['color_range']
    else:
        color_min, color_max = min(color), max(color) # Range of values that will be mapped to the palette, i.e. min and max possible correlation

    def value_to_color(val):
        if color_min == color_max:
            return palette[-1]
        else:
            val_position = float((val - color_min)) / (color_max - color_min) # position of value in the input range, relative to the length of the input range
            val_position = min(max(val_position, 0), 1) # bound the position betwen 0 and 1
            ind = int(val_position * (n_colors - 1)) # target index in the color palette
            return palette[ind]

    if 'size' in kwargs:
        size = kwargs['size']
    else:
        size = [1]*len(x)

    if 'size_range' in kwargs:
        size_min, size_max = kwargs['size_range'][0], kwargs['size_range'][1]
    else:
        size_min, size_max = min(size), max(size)

    size_scale = kwargs.get('size_scale', 500)

    def value_to_size(val):
        if size_min == size_max:
            return 1 * size_scale
        else:
            val_position = (val - size_min) * 0.99 / (size_max - size_min) + 0.01 # position of value in the input range, relative to the length of the input range
            val_position = min(max(val_position, 0), 1) # bound the position betwen 0 and 1
            return val_position * size_scale
    if 'x_order' in kwargs: 
        x_names = [t for t in kwargs['x_order']]
    else:
        x_names = [t for t in sorted(set([v for v in x]))]
    x_to_num = {p[1]:p[0] for p in enumerate(x_names)}

    if 'y_order' in kwargs: 
        y_names = [t for t in kwargs['y_order']]
    else:
        y_names = [t for t in sorted(set([v for v in y]))]
    y_to_num = {p[1]:p[0] for p in enumerate(y_names)}

    plot_grid = plt.GridSpec(1, 15, hspace=0.2, wspace=0.1) # Setup a 1x10 grid
    ax = plt.subplot(plot_grid[:,:-1]) # Use the left 14/15ths of the grid for the main plot

    marker = kwargs.get('marker', 's')

    kwargs_pass_on = {k:v for k,v in kwargs.items() if k not in [
          'color', 'palette', 'color_range', 'size', 'size_range', 'size_scale', 'marker', 'x_order', 'y_order', 'xlabel', 'ylabel'
    ]}

    ax.scatter(
        x=[x_to_num[v] for v in x],
        y=[y_to_num[v] for v in y],
        marker=marker,
        s=[value_to_size(v) for v in size], 
        c=[value_to_color(v) for v in color],
        **kwargs_pass_on
    )
    ax.set_xticks([v for k,v in x_to_num.items()])
    ax.set_xticklabels([k for k in x_to_num], rotation=45, horizontalalignment='right')
    ax.set_yticks([v for k,v in y_to_num.items()])
    ax.set_yticklabels([k for k in y_to_num])

    ax.grid(False, 'major')
    ax.grid(True, 'minor')
    ax.set_xticks([t + 0.5 for t in ax.get_xticks()], minor=True)
    ax.set_yticks([t + 0.5 for t in ax.get_yticks()], minor=True)

    ax.set_xlim([-0.5, max([v for v in x_to_num.values()]) + 0.5])
    ax.set_ylim([-0.5, max([v for v in y_to_num.values()]) + 0.5])
    ax.set_facecolor('#F1F1F1')

    ax.set_xlabel(kwargs.get('xlabel', ''))
    ax.set_ylabel(kwargs.get('ylabel', ''))

    # Add color legend on the right side of the plot
    if color_min < color_max:
        ax = plt.subplot(plot_grid[:,-1]) # Use the rightmost column of the plot

        col_x = [0]*len(palette) # Fixed x coordinate for the bars
        bar_y=np.linspace(color_min, color_max, n_colors) # y coordinates for each of the n_colors bars

        bar_height = bar_y[1] - bar_y[0]
        ax.barh(
            y=bar_y,
            width=[5]*len(palette), # Make bars 5 units wide
            left=col_x, # Make bars start at 0
            height=bar_height,
            color=palette,
            linewidth=0
        )
        ax.set_xlim(1, 2) # Bars are going from 0 to 5, so lets crop the plot somewhere in the middle
        ax.grid(False) # Hide grid
        ax.set_facecolor('white') # Make background white
        ax.set_xticks([]) # Remove horizontal ticks
        ax.set_yticks(np.linspace(min(bar_y), max(bar_y), 3)) # Show vertical ticks for min, middle and max
        ax.yaxis.tick_right() # Show vertical ticks on the right 


def corrplot(data, size_scale=500, marker='s'):
    corr = pd.melt(data.reset_index(), id_vars='index').replace(np.nan, 0)
    corr.columns = ['x', 'y', 'value']
    heatmap(
        corr['x'], corr['y'],
        color=corr['value'], color_range=[-1, 1],
        # palette=sns.diverging_palette(260, 30, n=256),
        palette=sns.color_palette(palette="Spectral", n_colors=256),
        size=corr['value'].abs(), size_range=[0,1],
        marker=marker,
        x_order=data.columns,
        y_order=data.columns[::-1],
        size_scale=size_scale
    )

# #####-----#####

import_file = r'/Users/lauramilton/Desktop/RO 2021/Behavioural C1+2 correlations.xlsx'
plot_save = r'/Users/lauramilton/Desktop/RO 2021/'

# get the sheetnames

import xlrd
xls = xlrd.open_workbook(import_file, on_demand=True)
sheetnames = xls.sheet_names()

# for each sheet (group/s) create the following graphs

for sheet in sheetnames:

    if 'Graph ABA R' in sheet:
        
        data = pd.read_excel(import_file, sheet_name=sheet)
        
        cols = data.columns.tolist()
        
        # Plot of JUST EPM data
        
        epm_cols_to_graph = []
        
        for i in cols:
            if i.startswith('EPM'):
                epm_cols_to_graph.append(i)
        
        epm_graph_data = pd.read_excel(import_file, sheet_name=sheet, usecols=epm_cols_to_graph)

        plt.figure(figsize=(5, 5))
        corrplot(epm_graph_data.corr(), size_scale=300);
        
        plt.savefig(plot_save + 'Beh C1+2 EPM' + sheet.strip('Graph'), bbox_inches = 'tight')
        
        # Plot of EPM and ABA data
        
        epm_aba_cols_to_graph = []
        
        for i in cols:
            if i.startswith('EPM'):
                epm_aba_cols_to_graph.append(i)
            elif i.startswith('ABA'):
                epm_aba_cols_to_graph.append(i)
        
        epm_aba_graph_data = pd.read_excel(import_file, sheet_name=sheet, usecols=epm_aba_cols_to_graph)

        plt.figure(figsize=(7, 7))
        corrplot(epm_aba_graph_data.corr(), size_scale=300);
        
        plt.savefig(r'/Users/lauramilton/Desktop/RO 2021/Beh C1+2 EPM w ABA' + sheet.strip('Graph'), bbox_inches = 'tight')
        
        # Plot of JUST OF data
        
        of_cols_to_graph = []
        
        for i in cols:
            if i.startswith('OF'):
                of_cols_to_graph.append(i)
        
        of_graph_data = pd.read_excel(import_file, sheet_name=sheet, usecols=of_cols_to_graph)

        plt.figure(figsize=(3, 3))
        corrplot(of_graph_data.corr(), size_scale=300);
        
        plt.savefig(r'/Users/lauramilton/Desktop/RO 2021/Beh C1+2 OF' + sheet.strip('Graph'), bbox_inches = 'tight')
        
        # Plot of OF and ABA data
        
        of_aba_cols_to_graph = []
        
        for i in cols:
            if i.startswith('OF'):
                of_aba_cols_to_graph.append(i)
            elif i.startswith('ABA'):
                of_aba_cols_to_graph.append(i)
        
        of_aba_graph_data = pd.read_excel(import_file, sheet_name=sheet, usecols=of_aba_cols_to_graph)

        plt.figure(figsize=(6, 6))
        corrplot(of_aba_graph_data.corr(), size_scale=300);
        
        plt.savefig(r'/Users/lauramilton/Desktop/RO 2021/Beh C1+2 OF w ABA' + sheet.strip('Graph'), bbox_inches = 'tight')
        
        # Plot of JUST MBT data
        
        mbt_cols_to_graph = []
        
        for i in cols:
            if i.startswith('MBT'):
                mbt_cols_to_graph.append(i)
        
        mbt_graph_data = pd.read_excel(import_file, sheet_name=sheet, usecols=mbt_cols_to_graph)

        plt.figure(figsize=(1.5, 1.5))
        corrplot(mbt_graph_data.corr(), size_scale=300);
        
        plt.savefig(r'/Users/lauramilton/Desktop/RO 2021/Beh C1+2 MBT' + sheet.strip('Graph'), bbox_inches = 'tight')
        
        # Plot of MBT and ABA data
        
        mbt_aba_cols_to_graph = []
        
        for i in cols:
            if i.startswith('MBT'):
                mbt_aba_cols_to_graph.append(i)
            elif i.startswith('ABA'):
                mbt_aba_cols_to_graph.append(i)
        
        mbt_aba_graph_data = pd.read_excel(import_file, sheet_name=sheet, usecols=mbt_aba_cols_to_graph)

        plt.figure(figsize=(4, 4))
        corrplot(mbt_aba_graph_data.corr(), size_scale=300);
        
        plt.savefig(r'/Users/lauramilton/Desktop/RO 2021/Beh C1+2 MBT w ABA' + sheet.strip('Graph'), bbox_inches = 'tight')
        
        # Plot of JUST ABA data
        
        aba_cols_to_graph = []
        
        for i in cols:
            if i.startswith('ABA'):
                aba_cols_to_graph.append(i)
        
        aba_graph_data = pd.read_excel(import_file, sheet_name=sheet, usecols=aba_cols_to_graph)

        plt.figure(figsize=(3, 3))
        corrplot(aba_graph_data.corr(), size_scale=300);
        
        plt.savefig(r'/Users/lauramilton/Desktop/RO 2021/Beh C1+2 ABA' + sheet.strip('Graph'), bbox_inches = 'tight')
                    
        # Plot of ALL data
        
        plt.figure(figsize=(12, 12))
        corrplot(data.corr(), size_scale=300);

        plt.savefig(r'/Users/lauramilton/Desktop/RO 2021/Beh C1+2 ALL' + sheet.strip('Graph'), bbox_inches = 'tight')
        