#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu May 27 13:39:01 2021

@author: lauramilton
"""

time_column = 'MM:DD:YYYY hh:mm:ss'
event_column = 'Event'
active_poke_column = 'Active_Poke'
session_type_column = 'Session_type'
left_poke_column = 'Left_Poke_Count'
right_poke_column = 'Right_Poke_Count'
pellet_count_column = 'Pellet_Count'
retrieval_time_column = 'Retrieval_Time'
poke_time_column = 'Poke_Time'

seconds_in_bins = '1800' # Enter how long you want the time bins to be in seconds; if not using bins enter '' which will give duration in seconds
session_duration_mins = '180'

reversal_block_length = '10'
blocked = 'Y' # Enter Y if you want the data put into discrete reversal blocks, enter N if you want continuous data
initiation_poke_active = True
initiation_poke_inactive = False
schedule_used = 'New Reversal'

# Enter the FED import folder, FED export folder, cohort/experiment export folder and the FED number

import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/NewRev/'
export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/NewRev/'
cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/NewRev/Cohort/'

# dates = ['091121']
export_results = 'N'
export_overview = 'N'
export_joined = 'N'

#-----------------------------------------------------------------------------

# Import the revelant data: time, FR ratio (session type), left poke, right poke, and pellet count.

import pandas as pd
import numpy as np
import os
import math

#-----

for folder in sorted(os.listdir(import_location)):
    if folder.startswith('FED5'):
                date_comp_all = []
                
                attempted_comp = []
                achieved_comp = []
                completed_comp = []
                totalpokes_comp = []
                activepokes_comp = []
                inactivepokes_comp = []
                percentactivepokes_comp = []
                pellets_comp = []
                leftpokes_comp = []
                leftactivepokes_comp = []
                leftinactivepokes_comp = []
                rightpokes_comp = []
                rightactivepokes_comp = []
                rightinactivepokes_comp = []
                inactiveperreversal_comp = []
                leftinactiveperreversal_comp = []
                rightinactiveperreversal_comp = []
                winstay_comp = []
                loseshift_comp = []
                
                attempted_all = []
                achieved_all = []
                completed_all = []
                totalpokes_all = []
                activepokes_all = []
                inactivepokes_all = []
                percentactivepokes_all = []
                pellets_all = []
                leftpokes_all = []
                leftactivepokes_all = []
                leftinactivepokes_all = []
                rightpokes_all = []
                rightactivepokes_all = []
                rightinactivepokes_all = []
                inactiveperreversal_all = []
                leftinactiveperreversal_all = []
                rightinactiveperreversal_all = []
                winstay_all = []
                loseshift_all = []



                for filename in sorted(os.listdir(os.path.join(import_location, folder))):
                    
                    if filename.endswith(".CSV"):
                                               
                                            
                        # Import the data

                        import_name = filename
                        import_destination = import_location + folder + '/' + import_name
                        export_name = folder + ' New Reversal ' + import_name.strip('.CSV') + '.xlsx'
                        export_destination = export_location + folder + '/' + export_name
                        print(filename)
                #-----        
                        
                        df = pd.read_csv(import_destination)
                        
                        time = df[time_column].tolist()
                        session_type = df[session_type_column].tolist()
                        event = df[event_column].tolist()
                        active = df[active_poke_column].tolist()
                        left_poke = df[left_poke_column].tolist()
                        right_poke = df[right_poke_column].tolist()
                        pellet_count = df[pellet_count_column].tolist()
                        retrieval_time = df[retrieval_time_column].tolist()
                        poke_time = df[poke_time_column].tolist()
                        
                        # Change the time column from strings to datetime format for calculating durations between timestamps and creating time bins
                        
                        import datetime as dt

                        time_list = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time]
                        
                        # Determine session start time which is the first timestamp generated by our poke to the/an acitve port to deliver a pellet
                        
                        session_start = time_list[0]
                        
                        # Remove the initiation poke/s and pellet data (if required)
                        
                        if initiation_poke_active == True:
                                        
                            left_count = left_poke[0]
                            right_count = right_poke[0]
                            
                            del time[:2]
                            del session_type[:2]
                            del event[:2]
                            del active[:2]
                            del left_poke[:2]
                            del right_poke[:2]
                            del pellet_count[:2]
                            del retrieval_time[:2]
                            del poke_time[:2]
                            
                            # Subtract the poke and pellet from the subsequent cumulative data
                            
                            left_poke_shifted = []
                            right_poke_shifted = []
                            pellet_count_shifted = []
                            
                            for i in range(0, len(left_poke)):
                                left_poke_shifted.append(left_poke[i] - left_count)
                                right_poke_shifted.append(right_poke[i] - right_count)
                                pellet_count_shifted.append(pellet_count[i] - 1)
                            
                            # Shift the retrieval_time and pellet_count lists one step backwards so that the count/time is in the same index as the corresponding poke
                            
                            retrieval_time.pop(0)
                            pellet_count_shifted.pop(0)
                            
                            # Add in a value to the end of each list to maintain list length so as not to lose the last row of other data
                            # for pellet_count this is a duplicate of preceding row, for retrieval_time this is np.nan
                            
                            retrieval_time.append(np.nan)
                            pellet_count_shifted.append((pellet_count[-1] - 1))

                            #####-----##### CHRONOLOGICAL DATA #####-----#####
                            
                            # Create new lists that only include data from the lines where the event is a Poke, in ReversalTask this means != 'Pellet' (i.e. Poke, Left, or Right)
                            
                            time2 = []
                            session_type2 = []
                            event2 = []
                            active2 = []
                            left_poke2 = []
                            right_poke2 = []
                            pellet_count2 = []
                            retrieval_time2 = []
                            poke_time2 = []
                            
                            for a, b, c, d, e, f, g, h, i in zip(event, left_poke_shifted, right_poke_shifted, pellet_count_shifted, time, session_type, active, retrieval_time, poke_time):
                                if a != 'Pellet':
                                    event2.append(a)
                                    left_poke2.append(b)
                                    right_poke2.append(c) 
                                    pellet_count2.append(d)
                                    time2.append(e)
                                    session_type2.append(f)
                                    active2.append(g)
                                    retrieval_time2.append(h)
                                    poke_time2.append(i)
                        
                        elif initiation_poke_inactive == True: # only remove 1 line of data (the inactive initiaton poke) as there is no pellet line, also do NOT subtract 1 from pellet count, i.e. no need for pellet_count_shifted
                                        
                            left_count = left_poke[0]
                            right_count = right_poke[0]
                            
                            del time[:1]
                            del session_type[:1]
                            del event[:1]
                            del active[:1]
                            del left_poke[:1]
                            del right_poke[:1]
                            del pellet_count[:1]
                            del retrieval_time[:1]
                            del poke_time[:1]
                            
                            # Subtract the poke from the subsequent cumulative data
                            
                            left_poke_shifted = []
                            right_poke_shifted = []
                            
                            for i in range(0, len(left_poke)):
                                left_poke_shifted.append(left_poke[i] - left_count)
                                right_poke_shifted.append(right_poke[i] - right_count)
                            
                            # Shift the retrieval_time and pellet_count lists one step backwards so that the count/time is in the same index as the corresponding poke
                            
                            retrieval_time.pop(0)
                            pellet_count.pop(0)
                            
                            # Add in a value to the end of each list to maintain list length so as not to lose the last row of other data
                            # for pellet_count this is a duplicate of preceding row, for retrieval_time this is np.nan
                            
                            retrieval_time.append(np.nan)
                            pellet_count.append(pellet_count[-1])
                            
                            #####-----##### CHRONOLOGICAL DATA #####-----#####
                            
                            # Create new lists that only include data from the lines where the event is a Poke, in ReversalTask this means != 'Pellet' (i.e. Poke, Left, or Right)
                            
                            time2 = []
                            session_type2 = []
                            event2 = []
                            active2 = []
                            left_poke2 = []
                            right_poke2 = []
                            pellet_count2 = []
                            retrieval_time2 = []
                            poke_time2 = []
                            
                            for a, b, c, d, e, f, g, h, i in zip(event, left_poke_shifted, right_poke_shifted, pellet_count, time, session_type, active, retrieval_time, poke_time):
                                if a != 'Pellet':
                                    event2.append(a)
                                    left_poke2.append(b)
                                    right_poke2.append(c) 
                                    pellet_count2.append(d)
                                    time2.append(e)
                                    session_type2.append(f)
                                    active2.append(g)
                                    retrieval_time2.append(h)
                                    poke_time2.append(i)
                        
                        else:
                            # Shift the retrieval_time and pellet_count lists one step backwards so that the count/time is in the same index as the corresponding poke
                            
                            retrieval_time.pop(0)
                            pellet_count.pop(0)
                            
                            # Add in a value to the end of each list to maintain list length so as not to lose the last row of other data
                            # for pellet_count this is a duplicate of preceding row, for retrieval_time this is np.nan
                            
                            retrieval_time.append(np.nan)
                            pellet_count.append(pellet_count[-1])
                            
                            #####-----##### CHRONOLOGICAL DATA #####-----#####
                            
                            # Create new lists that only include data from the lines where the event is a Poke, in ReversalTask this means != 'Pellet' (i.e. Poke, Left, or Right)
                            
                            time2 = []
                            session_type2 = []
                            event2 = []
                            active2 = []
                            left_poke2 = []
                            right_poke2 = []
                            pellet_count2 = []
                            retrieval_time2 = []
                            poke_time2 = []
                            
                            for a, b, c, d, e, f, g, h, i in zip(event, left_poke, right_poke, pellet_count, time, session_type, active, retrieval_time, poke_time):
                                if a != 'Pellet':
                                    event2.append(a)
                                    left_poke2.append(b)
                                    right_poke2.append(c) 
                                    pellet_count2.append(d)
                                    time2.append(e)
                                    session_type2.append(f)
                                    active2.append(g)
                                    retrieval_time2.append(h)
                                    poke_time2.append(i)
                            
                        
                        # Change the time column from strings to datetime format for calculating durations between timestamps and creating time bins
                        
                        time2 = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time2]
                        
                        # Calculate how much time (in seconds) has elapsed from start time for each event (nose poke)
                        
                        seconds_elapsed = []
                        
                        for i in range(0, len(time2)):
                            time_from_start = time2[i] - session_start
                            result = int(time_from_start.total_seconds())
                            seconds_elapsed.append(result)
                        
                        # Transform seconds_elapsed into duration in h:mm:ss
                        
                        duration_from_start = []
                        
                        for i in range(0, len(seconds_elapsed)):
                            time_hours = str(math.floor(seconds_elapsed[i] / 3600))
                            
                            time_mins = math.floor((seconds_elapsed[i] % 3600) / 60)
                            
                            time_secs = seconds_elapsed[i] % 60
                                
                            if time_mins < 10:
                                t_mins = '0' + str(time_mins)
                            else:
                                t_mins = str(time_mins)
                                
                            if time_secs < 10:
                                t_secs = '0' + str(time_secs)
                            else:
                                t_secs = str(time_secs)
                                
                            duration_from_start.append(str(time_hours) + ':' + t_mins + ':' + t_secs)
                        
                        ##### Remove any data that is from after max session duration
                        
                        end_seconds = int(session_duration_mins) * 60
                        
                        rows_to_keep = []                        
                        for i in range(0, len(seconds_elapsed)):
                            if seconds_elapsed[i] < end_seconds:
                                rows_to_keep.append(seconds_elapsed[i])
                        
                        rows_to_remove = len(seconds_elapsed) - len(rows_to_keep)

                        seconds_elapsed = seconds_elapsed[: (len(seconds_elapsed) - rows_to_remove)]
                        duration_from_start = duration_from_start[: (len(duration_from_start) - rows_to_remove)]
                        event2 = event2[: (len(event2) - rows_to_remove)]
                        left_poke2 = left_poke2[: (len(left_poke2) - rows_to_remove)]
                        right_poke2 = right_poke2[: (len(right_poke2) - rows_to_remove)]
                        pellet_count2 = pellet_count2[: (len(pellet_count2) - rows_to_remove)]
                        time2 = time2[: (len(time2) - rows_to_remove)]
                        session_type2 = session_type2[: (len(session_type2) - rows_to_remove)]
                        active2 = active2[: (len(active2) - rows_to_remove)]
                        retrieval_time2 = retrieval_time2[: (len(retrieval_time2) - rows_to_remove)]
                        poke_time2 = poke_time2[: (len(poke_time2) - rows_to_remove)]
                        
                        #####
                        
                        # Create total pokes column
                        
                        total_poke = []
                        
                        for i in range(0, len(active2)):
                            total_poke.append(left_poke2[i] + right_poke2[i])
                        
                        # Create binary poke columns
                        
                        left_poke_binary = []
                        left_a_poke_binary = []
                        left_i_poke_binary = []
                        
                        if left_poke2[0] == 1:
                            left_poke_binary.append(1)
                        else:
                            left_poke_binary.append(0)
                        
                        for i in range(1, len(left_poke2)):
                            if left_poke2[i - 1] != left_poke2[i]:
                                left_poke_binary.append(1)
                            else:
                                left_poke_binary.append(0)
                        
                        
                        right_poke_binary = []
                        right_a_poke_binary = []
                        right_i_poke_binary = []
                            
                        if right_poke2[0] == 1:
                            right_poke_binary.append(1)
                        else:
                            right_poke_binary.append(0)
                        
                        for i in range(1, len(right_poke2)):
                            if right_poke2[i - 1] != right_poke2[i]:
                                right_poke_binary.append(1)
                            else:
                                right_poke_binary.append(0)
                        
                        
                        for i in range(0, len(left_poke_binary)):
                            if left_poke_binary[i] == 1:
                                if active2[i] == 'Left':
                                    left_a_poke_binary.append(1)
                                    left_i_poke_binary.append(0)
                                    right_i_poke_binary.append(0)
                                    right_a_poke_binary.append(0)
                                elif active2[i] == 'Right':
                                    left_a_poke_binary.append(0)
                                    left_i_poke_binary.append(1)
                                    right_i_poke_binary.append(0)
                                    right_a_poke_binary.append(0)
                            elif left_poke_binary[i] == 0:
                                if active2[i] == 'Left':
                                    left_a_poke_binary.append(0)
                                    left_i_poke_binary.append(0)
                                    right_i_poke_binary.append(1)
                                    right_a_poke_binary.append(0)
                                elif active2[i] == 'Right':
                                    left_a_poke_binary.append(0)
                                    left_i_poke_binary.append(0)
                                    right_i_poke_binary.append(0)
                                    right_a_poke_binary.append(1)
                        
                        # Create active and inactive poke columns
                        
                        active_poke_binary = []
                        inactive_poke_binary = [] 
                                
                        for i in range(0, len(active2)):
                            if active2[i] == 'Left':
                                active_poke_binary.append(left_poke_binary[i])
                                inactive_poke_binary.append(right_poke_binary[i])
                            elif active2[i] == 'Right':
                                active_poke_binary.append(right_poke_binary[i])
                                inactive_poke_binary.append(left_poke_binary[i])
                        
                        active_poke = [] # cumulative
                        inactive_poke = [] # cumulative
                        
                        active_counter = 0
                        inactive_counter = 0
                        
                        for i in range (0, len(active_poke_binary)):
                            if active_poke_binary[i] == 0:
                                active_poke.append(active_counter)
                            elif active_poke_binary[i] == 1:
                                active_counter += 1
                                active_poke.append(active_counter)
                                
                        for i in range (0, len(inactive_poke_binary)):
                            if inactive_poke_binary[i] == 0:
                                inactive_poke.append(inactive_counter)
                            elif inactive_poke_binary[i] == 1:
                                inactive_counter += 1
                                inactive_poke.append(inactive_counter)
                        
                        # create list for pellet count that only included those with recorded retrieval time
                        
                        timed_pellet_count = []
                        
                        counter = 0
                        
                        for i in range(0, len(pellet_count2)):
                            if active_poke_binary[i] == 1:
                                if retrieval_time2[i] == 'Timed_out':
                                    timed_pellet_count.append(counter)
                                elif retrieval_time2[i] != 'Timed_out':
                                    counter += 1
                                    timed_pellet_count.append(counter)
                            elif active_poke_binary[i] == 0:
                                timed_pellet_count.append(counter)

                        # Create percent active column that accounts for the swapping of active ports
                        
                        percent_active = []
                        
                        for i in range(0, len(active_poke)):
                            percent_active.append(active_poke[i] / total_poke[i] * 100)
                        
                        # Create binary column indicating active port for graphing
                        
                        active_port_binary = []
                        
                        for i in range(0, len(active2)):
                            if active2[i] == 'Left':
                                active_port_binary.append(1)
                            elif active2[i] == 'Right':
                                active_port_binary.append(0)
                        
                        # Create Green (correct/active) and Red (incorrect/inactive) binary columns for graphing where 1 and 0 as correct/incorrect is determined by active_port_binary
                        
                        green = []
                        red = []
                        
                        for i in range(0, len(left_poke_binary)):
                            if active_port_binary[i] == 1:
                                if left_poke_binary[i] == 1:
                                    green.append(1)
                                    red.append(np.nan)
                                elif left_poke_binary[i] == 0:
                                    green.append(np.nan)
                                    red.append(0)
                            elif active_port_binary[i] == 0:
                                if left_poke_binary[i] == 0:
                                        green.append(0)
                                        red.append(np.nan)
                                elif left_poke_binary[i] ==1:
                                        green.append(np.nan)
                                        red.append(1)
                        
                        # Create chronological timing data
                        
                        #####-----##### Chronological Poke Time Data #####-----#####
                        
                        l_poke_time = [] # poke time for all left pokes
                        r_poke_time = [] # poke time for all right pokes
                        
                        l_poke_time_chron = [] # poke time for all left pokes with nan if it was a right poke to maintain array length the same as all pokes
                        r_poke_time_chron = [] # poke time for all right pokes with nan if it was a left poke to maintain array length the same as all pokes
                        
                        for i in range(0, len(poke_time2)): # Collates the poke_time by side (regardless of port status)
                           if event2[i] == 'Left':
                               l_poke_time.append(poke_time2[i])
                               l_poke_time_chron.append(poke_time2[i])
                               r_poke_time_chron.append(np.nan)
                           else:
                               r_poke_time.append(poke_time2[i])
                               r_poke_time_chron.append(poke_time2[i])
                               l_poke_time_chron.append(np.nan)
                
                        a_poke_time = [] # poke time for all active pokes
                        a_l_poke_time = [] # poke time for all left pokes when the left port is active
                        a_r_poke_time = [] # poke time for all right pokes when the right poke is active
                        
                        i_poke_time = [] # poke time for all inactive pokes
                        i_l_poke_time = [] # poke time for all left pokes when the right port is active
                        i_r_poke_time = [] # poke time for all right pokes when the left port is active
                        
                        a_poke_time_chron = [] # all chron lists put nan for a poke that was not one of the status that the list is tracking to maintain array length the same as all pokes
                        a_l_poke_time_chron = []
                        a_r_poke_time_chron = []
                        
                        i_poke_time_chron = []
                        i_l_poke_time_chron = []
                        i_r_poke_time_chron = []
                        
                        for i in range(0, len(poke_time2)): # Collates poke_time by status, and status + side
                            if active_poke_binary[i] == 1:                  # if it's an active poke
                                a_poke_time.append(poke_time2[i])           # add time to active poke time lists and nan to inactive list
                                a_poke_time_chron.append(poke_time2[i])
                                i_poke_time_chron.append(np.nan)
                                if active2[i] == 'Left':                    # if the left port is active
                                    a_l_poke_time.append(poke_time2[i])     # add the time to the left active lists and nan to left inactive and right lists
                                    a_l_poke_time_chron.append(poke_time2[i])
                                    i_l_poke_time_chron.append(np.nan)
                                    a_r_poke_time_chron.append(np.nan)
                                    i_r_poke_time_chron.append(np.nan)
                                    
                                    
                                elif active2[i] == 'Right':                 # if the right port is active
                                    a_r_poke_time.append(poke_time2[i])     # add the time to the right active lists and nan to right inactive and left lists
                                    a_r_poke_time_chron.append(poke_time2[i])
                                    i_r_poke_time_chron.append(np.nan)
                                    a_l_poke_time_chron.append(np.nan)
                                    i_l_poke_time_chron.append(np.nan)
                            else:                                           # if it's an inactive poke
                                i_poke_time.append(poke_time2[i])           # add the time to the inactive poke time lists and non to active list
                                i_poke_time_chron.append(poke_time2[i])
                                a_poke_time_chron.append(np.nan)
                                if active2[i] == 'Left':                    # if the left port is active, and therefore the inactive poke is a right poke
                                    i_r_poke_time.append(poke_time2[i])     # add time to the right inactive lists and non to right active and left lists
                                    i_r_poke_time_chron.append(poke_time2[i])
                                    a_r_poke_time_chron.append(np.nan)
                                    i_l_poke_time_chron.append(np.nan)
                                    a_l_poke_time_chron.append(np.nan)
                                elif active2[i] == 'Right':                 # if the right port is active, and therefore the inactive poke is a left poke
                                    i_l_poke_time.append(poke_time2[i])     # add the time to the left inactive lists and nan to left active and right lists
                                    i_l_poke_time_chron.append(poke_time2[i])
                                    a_l_poke_time_chron.append(np.nan)
                                    i_r_poke_time_chron.append(np.nan)
                                    a_r_poke_time_chron.append(np.nan)
                        
                        #####-----##### CREATE BLOCKED DATA #####-----#####
                        
                        # Create blocks (i.e. bins) determined by the active port
                        
                        seconds_block = []
                        duration_block = []
                        session_block = []
                        active_block = []
                        left_block = []
                        right_block = []
                        active_poke_block = []
                        inactive_poke_block = []
                        total_block = []
                        pellet_block = []
                        
                        for i in range(1, len(active2)):
                                if active2[i - 1] != active2[i]:
                                    seconds_block.append(seconds_elapsed[i - 1])
                                    duration_block.append(duration_from_start[i - 1])
                                    session_block.append(session_type2[i - 1])
                                    active_block.append(active2[i - 1])
                                    left_block.append(left_poke2[i - 1])
                                    right_block.append(right_poke2[i - 1])
                                    active_poke_block.append(active_poke[i - 1])
                                    inactive_poke_block.append(inactive_poke[i - 1])
                                    total_block.append(total_poke[i - 1])
                                    pellet_block.append(pellet_count2[i - 1])
                        seconds_block.append(seconds_elapsed[i - 0])
                        duration_block.append(duration_from_start[i - 0])
                        session_block.append(session_type2[i - 0])
                        active_block.append(active2[i - 0])
                        left_block.append(left_poke2[i - 0])
                        right_block.append(right_poke2[i - 0])
                        active_poke_block.append(active_poke[i - 0])
                        inactive_poke_block.append(inactive_poke[i - 0])
                        total_block.append(total_poke[i - 0])
                        pellet_block.append(pellet_count2[i - 0])
                        
                        
                        # Create percent active column
                        
                        percent_active_block = []
                        
                        for i in range(0, len(active_poke_block)):
                            if total_block[i] != 0:
                                percent_active_block.append(active_poke_block[i] / total_block[i] * 100)
                            else:
                                percent_active_block.append('NA')
                        
                        
                        ##### Make the counts within the blocks rather than cumulative
                        
                        seconds_block_within = []
                        
                        seconds_block_within.append(seconds_block[0])
                        for i in range(1, len(seconds_block)):
                            seconds_block_within.append(seconds_block[i] - seconds_block[i - 1])
                            
                        # Transform seconds_elapsed into duration in h:mm:ss
                        
                        duration_of_block_within = []
                        
                        for i in range(0, len(seconds_block_within)):
                            time_hours = str(math.floor(seconds_block_within[i] / 3600))
                            
                            time_mins = math.floor((seconds_block_within[i] % 3600) / 60)
                            
                            time_secs = seconds_block_within[i] % 60
                                
                            if time_mins < 10:
                                t_mins = '0' + str(time_mins)
                            else:
                                t_mins = str(time_mins)
                                
                            if time_secs < 10:
                                t_secs = '0' + str(time_secs)
                            else:
                                t_secs = str(time_secs)
                                
                            duration_of_block_within.append(str(time_hours) + ':' + t_mins + ':' + t_secs)
                        
                        left_block_within = []
                        right_block_within = []
                        active_poke_block_within = []
                        inactive_poke_block_within = []
                        total_block_within = []
                        pellet_block_within = []
                        percent_active_block_within = []
                        
                        left_block_within.append(left_block[0])
                        for i in range(1, len(left_block)):
                            left_block_within.append(left_block[i] - left_block[i - 1])
                        
                        right_block_within.append(right_block[0])
                        for i in range(1, len(right_block)):
                            right_block_within.append(right_block[i] - right_block[i - 1])
                        
                        active_poke_block_within.append(active_poke_block[0])
                        for i in range(1, len(active_poke_block)):
                            active_poke_block_within.append(active_poke_block[i] - active_poke_block[i - 1])
                        
                        inactive_poke_block_within.append(inactive_poke_block[0])
                        for i in range(1, len(inactive_poke_block)):
                            inactive_poke_block_within.append(inactive_poke_block[i] - inactive_poke_block[i - 1])
                        
                        total_block_within.append(total_block[0])
                        for i in range(1, len(total_block)):
                            total_block_within.append(total_block[i] - total_block[i - 1])
                        
                        pellet_block_within.append(pellet_block[0])
                        for i in range(1, len(pellet_block)):
                            pellet_block_within.append(pellet_block[i] - pellet_block[i - 1])
                        
                        # Create percent active pokes column    
                        
                        for i in range(0, len(active_block)):
                            if total_block_within[i] != 0:
                                percent_active_block_within.append(active_poke_block_within[i] / total_block_within[i] * 100)
                            else:
                                percent_active_block_within.append('NA')
                        
                        
                        # Create mean poke and retrieval time columns for blocks
                        
                        block_sum_all_poke_time = [] # cumulative sum of all poke times
                        block_sum_a_poke_time = [] # cumulative sum of active poke times
                        block_sum_i_poke_time = [] # cumulative sum of inactive poke times
                        block_mean_all_poke_time = [] # cumulative mean of all poke times
                        block_mean_a_poke_time = [] # cumulative mean of active poke times
                        block_mean_i_poke_time = [] # cumulative mean of inactive poke times
                        
                        
                        sum_all_poke_time = 0
                        sum_a_poke_time = 0
                        sum_i_poke_time = 0
                        index = 0
                        poke_index = 0

                        for i in range(0, len(poke_time2)):
                            if int(seconds_elapsed[i]) < int(seconds_block[index]):
                                
                                sum_all_poke_time += poke_time2[poke_index]
                                if active_poke_binary[i] == 1:
                                    sum_a_poke_time += poke_time2[poke_index]
                                elif inactive_poke_binary[i] == 1:
                                    sum_i_poke_time += poke_time2[poke_index]
                                poke_index += 1
                                
                            elif int(seconds_elapsed[i]) == int(seconds_block[index]):
                                sum_all_poke_time += poke_time2[poke_index]
                                if active_poke_binary[i] == 1:
                                    sum_a_poke_time += poke_time2[poke_index]
                                elif inactive_poke_binary[i] == 1:
                                    sum_i_poke_time += poke_time2[poke_index]
                                poke_index += 1
                                
                                block_sum_all_poke_time.append(sum_all_poke_time)
                                block_sum_a_poke_time.append(sum_a_poke_time)
                                block_sum_i_poke_time.append(sum_i_poke_time)
                                block_mean_all_poke_time.append((sum_all_poke_time / int(total_block[index])))
                                if int(active_poke_block[index]) != 0:
                                    block_mean_a_poke_time.append((sum_a_poke_time / int(active_poke_block[index])))
                                else:
                                    block_mean_a_poke_time.append('N/A')
                                if int(inactive_poke_block[index]) != 0:
                                    block_mean_i_poke_time.append((sum_i_poke_time / int(inactive_poke_block[index])))
                                else:
                                    block_mean_i_poke_time.append('N/A')
                                
                                index += 1
                            
                            elif int(seconds_elapsed[i]) > int(seconds_block[index]):
                                
                                block_sum_all_poke_time.append(sum_all_poke_time)
                                block_sum_a_poke_time.append(sum_a_poke_time)
                                block_sum_i_poke_time.append(sum_i_poke_time)
                                block_mean_all_poke_time.append((sum_all_poke_time / int(total_block[index])))
                                block_mean_a_poke_time.append((sum_a_poke_time / int(active_poke_block[index])))
                                if int(inactive_poke_block[index]) != 0:
                                    block_mean_i_poke_time.append((sum_i_poke_time / int(inactive_poke_block[index])))
                                else:
                                     block_mean_i_poke_time.append('N/A')
                                
                                index += 1
                                
                                sum_all_poke_time += poke_time2[poke_index]
                                if active_poke_binary[i] == 1:
                                    sum_a_poke_time += poke_time2[poke_index]
                                elif inactive_poke_binary[i] == 1:
                                    sum_i_poke_time += poke_time2[poke_index]
                                poke_index += 1
                        
                        block_within_sum_all_poke_time = [] # within block sum of all poke times
                        block_within_sum_a_poke_time = [] # within block sum of active poke times
                        block_within_sum_i_poke_time = [] # within block sum of inactive poke times
                        
                        block_within_sum_all_poke_time.append(block_sum_all_poke_time[0])
                        block_within_sum_a_poke_time.append(block_sum_a_poke_time[0])
                        block_within_sum_i_poke_time.append(block_sum_i_poke_time[0])  
                        
                        for i in range(1, len(block_sum_all_poke_time)):
                            block_within_sum_all_poke_time.append((block_sum_all_poke_time[i] - block_sum_all_poke_time[i - 1]))
                            block_within_sum_a_poke_time.append((block_sum_a_poke_time[i] - block_sum_a_poke_time[i - 1]))
                            block_within_sum_i_poke_time.append((block_sum_i_poke_time[i] - block_sum_i_poke_time[i - 1]))  
                        
                        block_within_mean_all_poke_time = [] # within block mean of all poke times
                        block_within_mean_a_poke_time = [] # within block mean of active poke times
                        block_within_mean_i_poke_time = [] # within block mean of inactive poke times
                        
                        for i in range(0, len(block_within_sum_all_poke_time)):
                            block_within_mean_all_poke_time.append((block_within_sum_all_poke_time[i] / total_block_within[i]))
                            if active_poke_block_within[i] != 0:
                                block_within_mean_a_poke_time.append((block_within_sum_a_poke_time[i] / active_poke_block_within[i]))
                            else:
                                block_within_mean_a_poke_time.append('N/A')
                            if inactive_poke_block_within[i] != 0:
                                block_within_mean_i_poke_time.append((block_within_sum_i_poke_time[i] / inactive_poke_block_within[i]))
                            else:
                                block_within_mean_i_poke_time.append('N/A')
                        
                        
                        block_sum_all_pellet_retrieval_time = [] # cumulative sum of all pellet retrieval times
                        block_mean_all_pellet_retrieval_time = [] # cumulative mean of all pellet retrieval times
                        block_within_sum_all_pellet_retrieval_time = [] # within block sum of pellet retrieval times
                        block_within_mean_all_pellet_retrieval_time = [] # within block mean of pellet retrieval times
                        
                        block_mean_timed_pellet_retrieval_time = [] # cumulative mean of timed pellet retrieval times
                        block_within_mean_timed_pellet_retrieval_time = [] # within block mean of timed pellet retrieval times
                        
                        block_timed_pellet_count = []
                        block_within_timed_pellet_count = []
                        
                        if initiation_poke_active == True:
                            block_count = 2 # if the first block only has n-1 pellets because of the initiation poke/pellet so need to start 1 later in the count to fill up the block
                        else:
                            block_count = 1
                            
                        block = 1
                        
                        sum_retrieval_time = 0
                        
                        all_pellet_counter = 0
                        timed_pellet_counter = 0

                        retrieval_time2 = [float(i) if i!='Timed_out' else 0 for i in retrieval_time2]

                        pellet_retrieval_time = [x for x in retrieval_time2 if not math.isnan(x)] # gets rid of all the nan values which are due to inactive pokes

                        for i in range(0, len(pellet_retrieval_time)):
                            if block_count < (block * int(reversal_block_length)):
                                if pellet_retrieval_time[i] != 0:
                                    sum_retrieval_time += pellet_retrieval_time[i]
                                    all_pellet_counter += 1
                                    timed_pellet_counter += 1
                                else:
                                    sum_retrieval_time += 0
                                    all_pellet_counter += 1
                                block_count += 1
                            else:
                                if pellet_retrieval_time[i] != 0:
                                    sum_retrieval_time += pellet_retrieval_time[i]
                                    block_sum_all_pellet_retrieval_time.append(sum_retrieval_time)
                                    all_pellet_counter += 1
                                    timed_pellet_counter += 1
                                    block_timed_pellet_count.append(timed_pellet_counter)
                                else:
                                    sum_retrieval_time += 0
                                    block_sum_all_pellet_retrieval_time.append(sum_retrieval_time)
                                    all_pellet_counter += 1
                                    block_timed_pellet_count.append(timed_pellet_counter)
                                
                                if initiation_poke_active == True:
                                    if sum_retrieval_time != 0:
                                        block_mean_all_pellet_retrieval_time.append(sum_retrieval_time / (block * (int(reversal_block_length) - 1))) # the -1 is to account for the initiation pellet retrieval time being removed
                                    else:
                                        block_mean_all_pellet_retrieval_time.append('N/A')
                                else:
                                    if sum_retrieval_time != 0:
                                        block_mean_all_pellet_retrieval_time.append(sum_retrieval_time / (block * int(reversal_block_length)))
                                    else:
                                        block_mean_all_pellet_retrieval_time.append('N/A')
                                        
                                if timed_pellet_counter != 0:
                                    block_mean_timed_pellet_retrieval_time.append(sum_retrieval_time / timed_pellet_counter)
                                else:
                                    block_mean_timed_pellet_retrieval_time.append('N/A')
                                
                                block_count += 1
                                block += 1
                        
                        # if no reversals are completed block_timed_pellet_count and block_sum_all_pellet_retrieval_time are empty, need to force it to add the current count
                        # also if the final reversal block is incomplete but has active pokes (and thus pellet delivery) in it need to trigger adding the current count
                        
                        if len(block_timed_pellet_count) == 0:
                            block_timed_pellet_count.append(timed_pellet_counter)
                            
                        if len(block_sum_all_pellet_retrieval_time) == 0:
                            block_sum_all_pellet_retrieval_time.append(sum_retrieval_time)
                            
                        if len(block_mean_all_pellet_retrieval_time) == 0:
                            if initiation_poke_active == True:
                                if sum_retrieval_time != 0:
                                        block_mean_all_pellet_retrieval_time.append(sum_retrieval_time / (block * (int(reversal_block_length) - 1))) # the -1 is to account for the initiation pellet retrieval time being removed
                                else:
                                    block_mean_all_pellet_retrieval_time.append('N/A')
                            else:
                                if sum_retrieval_time != 0:
                                    block_mean_all_pellet_retrieval_time.append(sum_retrieval_time / (block * int(reversal_block_length)))
                                else:
                                    block_mean_all_pellet_retrieval_time.append('N/A')
                                        
                        if len(block_mean_timed_pellet_retrieval_time) == 0:
                            if timed_pellet_counter != 0:
                                block_mean_timed_pellet_retrieval_time.append(sum_retrieval_time / timed_pellet_counter)
                            else:
                                block_mean_timed_pellet_retrieval_time.append('N/A')
                        
                        # if the final reversal block is incomplete need to trigger adding the last row with the current count
                        if len(pellet_block) != len(block_timed_pellet_count):                         
                            
                            block_sum_all_pellet_retrieval_time.append(sum_retrieval_time)
                            block_timed_pellet_count.append(timed_pellet_counter)
                            if initiation_poke_active == True:
                                if sum_retrieval_time != 0:
                                    block_mean_all_pellet_retrieval_time.append(sum_retrieval_time / (block * (int(reversal_block_length) - 1))) # the -1 is to account for the initiation pellet retrieval time being removed
                                else:
                                    block_mean_all_pellet_retrieval_time.append('N/A')
                            else:
                                if sum_retrieval_time != 0:
                                    block_mean_all_pellet_retrieval_time.append(sum_retrieval_time / (block * int(reversal_block_length)))
                                else:
                                    block_mean_all_pellet_retrieval_time.append('N/A')
                                    
                            if timed_pellet_counter != 0:
                                block_mean_timed_pellet_retrieval_time.append(sum_retrieval_time / timed_pellet_counter)
                            else:
                                block_mean_timed_pellet_retrieval_time.append('N/A')
                        
                        block_within_sum_all_pellet_retrieval_time.append(block_sum_all_pellet_retrieval_time[0])
                        for i in range(1, len(block_sum_all_pellet_retrieval_time)):
                            block_within_sum_all_pellet_retrieval_time.append(block_sum_all_pellet_retrieval_time[i] - block_sum_all_pellet_retrieval_time[i - 1])        
                        
                        if initiation_poke_active == True:
                            block_within_mean_all_pellet_retrieval_time.append(block_within_sum_all_pellet_retrieval_time[0] / (int(reversal_block_length) - 1)) # first block only contains n-1 because of initiation poke pellet retrieval time removal
                            for i in range(1, len(block_within_sum_all_pellet_retrieval_time)):
                                block_within_mean_all_pellet_retrieval_time.append(block_within_sum_all_pellet_retrieval_time[i] / int(reversal_block_length)) # all other blocks contain n
                        else:
                            for i in range(0, len(block_within_sum_all_pellet_retrieval_time)):
                                if block_within_sum_all_pellet_retrieval_time[i] != 0:
                                    block_within_mean_all_pellet_retrieval_time.append(block_within_sum_all_pellet_retrieval_time[i] / int(reversal_block_length)) # all blocks contain n
                                else:
                                    block_within_mean_all_pellet_retrieval_time.append('N/A')
                            
                        # create within block timed pellet count
                        
                        block_within_timed_pellet_count.append(block_timed_pellet_count[0])
                        for i in range(1, len(block_timed_pellet_count)):
                            block_within_timed_pellet_count.append(block_timed_pellet_count[i] - block_timed_pellet_count[i - 1])
                            
                        for i in range(0, len(block_within_timed_pellet_count)):
                            if block_within_timed_pellet_count[i] != 0:
                                block_within_mean_timed_pellet_retrieval_time.append(block_within_sum_all_pellet_retrieval_time[i] / block_within_timed_pellet_count[i])
                            else:
                                block_within_mean_timed_pellet_retrieval_time.append('N/A')
                                

                        
                        
                        #####-----##### Create time bins of desired length (input for desired length is at the top of the code)
                        
                        if seconds_in_bins != '':
                        
                            end_time = seconds_elapsed[-1]
                            
                            # Create end value that is the next greatest multiple of the required time bin duration greater than the end_time so that those data points are retained in the output
                            
                            excess = math.ceil(end_time / int(seconds_in_bins)) # Rounds up so that the number of bins is the multiple of bin length greater than the time point
                            
                            end_bin = excess * int(seconds_in_bins) # Creates the end value for the last time bin
                            
                            interval_range = pd.interval_range(start=0, freq=int(seconds_in_bins), end=end_bin, closed="left") #'left' means that it includes the first data point which is 0 because it's start time
                            
                            time_bins = df[seconds_elapsed] = pd.cut(seconds_elapsed, interval_range, include_lowest=True, ordered=True)
                        
                            
                            # Create dictionary of time bins as keys with bin number as matching entry to label the bins
                            
                            # Create a list of integers of the same length as there are time bins
                            
                            bin_number = []
                            a = 1
                            for i in range(0, len(interval_range)):
                                bin_number.append(a)
                                a += 1
                            
                            make_dictionary = zip(interval_range, bin_number)
                            
                            bin_dictionary = dict(make_dictionary)
                            
                            bin_num = []
                            for i in range(0, len(time_bins)):
                                bin_num.append(bin_dictionary.get((time_bins[i])))
                                                        
                            bin_num_binned = []
                            active_binned = []
                            inactive_binned = []
                            total_pokes_binned = []
                            pellet_binned = []
                            active_percent_binned = []
                            timed_pellet_binned = []

                            # time binned
                            for i in range(1, len(bin_num)):
                                    if bin_num[i - 1] != bin_num[i]:
                                        bin_num_binned.append(bin_num[i - 1])
                            bin_num_binned.append(bin_num[i - 0])
                            
                            # active binned
                            for i in range(1, len(active_poke)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        active_binned.append(active_poke[i - 1])
                            active_binned.append(active_poke[i - 0])
                                    
                            # inactive binned
                            for i in range(1, len(inactive_poke)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        inactive_binned.append(inactive_poke[i - 1])
                            inactive_binned.append(inactive_poke[i - 0])
                            
                            # total binned
                            for i in range(1, len(total_poke)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        total_pokes_binned.append(total_poke[i - 1])
                            total_pokes_binned.append(total_poke[i - 0])
                                 
                            # pellet binned
                            for i in range(1, len(pellet_count2)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        pellet_binned.append(pellet_count2[i - 1])
                            pellet_binned.append(pellet_count2[i - 0])
                            
                            # active percent binned
                            for i in range(0, len(total_pokes_binned)):
                                active_percent_binned.append(active_binned[i] / total_pokes_binned[i] * 100)
                            
                            # timed pellet binned
                            for i in range(1, len(timed_pellet_count)):
                                    if time_bins[i - 1] != time_bins[i]:
                                        timed_pellet_binned.append(timed_pellet_count[i - 1])
                            timed_pellet_binned.append(timed_pellet_count[i - 0])
                            
                            # Create mean poke and retrieval time columns for bins
                        
                            bin_sum_all_poke_time = [] # cumulative sum of all poke times
                            bin_sum_a_poke_time = [] # cumulative sum of active poke times
                            bin_sum_i_poke_time = [] # cumulative sum of inactive poke times
                            bin_mean_all_poke_time = [] # cumulative mean of all poke times
                            bin_mean_a_poke_time = [] # cumulative mean of active poke times
                            bin_mean_i_poke_time = [] # cumulative mean of inactive poke times

                            bin_counter = 1

                            while bin_num[0] != bin_counter: # fills in any empty leading bins until it gets to the bin with the first data point
                                    bin_num_binned.insert(bin_counter - 1, bin_counter)
                                    active_binned.insert(bin_counter - 1, 0)
                                    inactive_binned.insert(bin_counter - 1, 0)
                                    total_pokes_binned.insert(bin_counter - 1, 0)
                                    pellet_binned.insert(bin_counter - 1, 0)
                                    active_percent_binned.insert(bin_counter - 1, np.nan)
                                    timed_pellet_binned.insert(bin_counter - 1, 0)
                                    bin_counter += 1

                            sum_all_poke_time = 0
                            sum_a_poke_time = 0
                            sum_i_poke_time = 0
                            index = 1
                            comp_bin_index = 1
                            a_poke_counter = 0
                            i_poke_counter = 0

                            while bin_num[0] > index: # if there are leading empty time bins need to fill them in
                                bin_sum_all_poke_time.append(0)
                                bin_sum_a_poke_time.append(0)
                                bin_sum_i_poke_time.append(0)
                                bin_mean_all_poke_time.append(0)
                                bin_mean_a_poke_time.append(0)
                                bin_mean_i_poke_time.append(0)
                                
                                index += 1
                                                        
                            if bin_num[-1] > 1: # if the last data point is not in the first time bin
                            
                                for i in range(0, len(bin_num)):
                                    
                                    if bin_num[i] == index:
                                        sum_all_poke_time += poke_time2[i]
                                        if active_poke_binary[i] == 1:
                                            sum_a_poke_time += poke_time2[i]
                                            a_poke_counter += 1
                                        elif inactive_poke_binary[i] == 1:
                                            sum_i_poke_time += poke_time2[i]
                                            i_poke_counter += 1
                                        
                                    elif bin_num[i] > index:
    
                                        bin_sum_all_poke_time.append(sum_all_poke_time)
                                        bin_sum_a_poke_time.append(sum_a_poke_time)
                                        bin_sum_i_poke_time.append(sum_i_poke_time)
                                        if sum_all_poke_time != 0:
                                            bin_mean_all_poke_time.append((sum_all_poke_time / (a_poke_counter + i_poke_counter)))
                                        else:
                                            bin_mean_all_poke_time.append('N/A')
                                        if sum_a_poke_time != 0:
                                            bin_mean_a_poke_time.append((sum_a_poke_time / a_poke_counter))
                                        else:
                                            bin_mean_a_poke_time.append('N/A')
                                        if sum_i_poke_time != 0:
                                            bin_mean_i_poke_time.append((sum_i_poke_time / i_poke_counter))
                                        else:
                                            bin_mean_i_poke_time.append('N/A')
                                        
                                        index += 1
                                        comp_bin_index += 1
                                    
                                        sum_all_poke_time += poke_time2[i]
                                        if active_poke_binary[i] == 1:
                                            sum_a_poke_time += poke_time2[i]
                                            a_poke_counter += 1
                                        elif inactive_poke_binary[i] == 1:
                                            sum_i_poke_time += poke_time2[i]
                                            i_poke_counter += 1
                                                                               
                                        if i != (len(bin_num) - 1):
                                            if bin_num[i] == bin_num_binned[comp_bin_index - 1] and bin_num[i + 1] > bin_num_binned[comp_bin_index - 1]: # if there is only 1 poke in the bin need to add the data to the lists otherwise it just gets included in the following bin
                                                bin_sum_all_poke_time.append(sum_all_poke_time)
                                                bin_sum_a_poke_time.append(sum_a_poke_time)
                                                bin_sum_i_poke_time.append(sum_i_poke_time)
                                                if sum_all_poke_time != 0:
                                                    bin_mean_all_poke_time.append((sum_all_poke_time / (a_poke_counter + i_poke_counter)))
                                                else:
                                                    bin_mean_all_poke_time.append('N/A')
                                                if sum_a_poke_time != 0:
                                                    bin_mean_a_poke_time.append((sum_a_poke_time / a_poke_counter))
                                                else:
                                                    bin_mean_a_poke_time.append('N/A')
                                                if sum_i_poke_time != 0:
                                                    bin_mean_i_poke_time.append((sum_i_poke_time / i_poke_counter))
                                                else:
                                                    bin_mean_i_poke_time.append('N/A')
                                                    
                                                index += 1
                                        
                                            while bin_num[i + 1] > index: # if the next bin is empty need to increase the index to match the bin count so that it doesn't just put each new entry into the lists as it's greater than the corresponding bin number
                                                index += 1
                                        

                                if int(seconds_elapsed[-1]) <= int(session_duration_mins) * 60: #triggers entry of the last bin if it hasn't exceeded the session length
                                    bin_sum_all_poke_time.append(sum_all_poke_time)
                                    bin_sum_a_poke_time.append(sum_a_poke_time)
                                    bin_sum_i_poke_time.append(sum_i_poke_time)
                                    if sum_all_poke_time != 0:
                                        bin_mean_all_poke_time.append((sum_all_poke_time / (a_poke_counter + i_poke_counter)))
                                    else:
                                        bin_mean_all_poke_time.append('N/A')
                                    if sum_a_poke_time != 0:
                                        bin_mean_a_poke_time.append((sum_a_poke_time / a_poke_counter))
                                    else:
                                        bin_mean_a_poke_time.append('N/A')
                                    if sum_i_poke_time != 0:
                                        bin_mean_i_poke_time.append((sum_i_poke_time / i_poke_counter))
                                    else:
                                        bin_mean_i_poke_time.append('N/A')

                            else: # if there is only data in the first time bin need to trigger it getting added to the lists
                                for i in range(0, len(poke_time2)):
                                    
                                    if int(seconds_elapsed[i]) <= int(seconds_in_bins) * index:
                                        sum_all_poke_time += poke_time2[i]
                                        if active_poke_binary[i] == 1:
                                            sum_a_poke_time += poke_time2[i]
                                            a_poke_counter += 1
                                        elif inactive_poke_binary[i] == 1:
                                            sum_i_poke_time += poke_time2[i]
                                            i_poke_counter += 1
    
                                bin_sum_all_poke_time.append(sum_all_poke_time)
                                bin_sum_a_poke_time.append(sum_a_poke_time)
                                bin_sum_i_poke_time.append(sum_i_poke_time)
                                bin_mean_all_poke_time.append((sum_all_poke_time / (a_poke_counter + i_poke_counter)))
                                if a_poke_counter != 0:
                                    bin_mean_a_poke_time.append((sum_a_poke_time / a_poke_counter))
                                else:
                                    bin_mean_a_poke_time.append(np.nan)
                                if i_poke_counter != 0:
                                    bin_mean_i_poke_time.append((sum_i_poke_time / i_poke_counter))
                                else:
                                    bin_mean_i_poke_time.append(np.nan)
                                
                                index += 1
                            
                            bin_sum_all_pellet_retrieval_time = [] # cumulative sum of all pellet retrieval times
                            bin_mean_all_pellet_retrieval_time = [] # cumulative mean of all pellet retrieval times
                            
                            bin_sum_timed_pellet_retrieval_time = []
                            bin_mean_timed_pellet_retrieval_time = []
                                                        
                            sum_all_retrieval_time = 0
                            sum_timed_retrieval_time = 0
                            timed_pellet_counter = 0
                            index = 1
                            comp_bin_index = 1
                            
                            retrieval_time_forbins = []

                            for i in range(0, len(retrieval_time2)):
                                if retrieval_time2[i] == 'Timed_out':
                                    retrieval_time_forbins.append(0)
                                else:
                                    retrieval_time_forbins.append(float(retrieval_time2[i]))

                            while bin_num[0] > index: # if the first time bin is empty (might need to change this to while rather than if for any cases where first and following bins are empty...?)
                                bin_sum_all_pellet_retrieval_time.append(0)
                                bin_sum_timed_pellet_retrieval_time.append(0)
                                bin_mean_all_pellet_retrieval_time.append(0)
                                bin_mean_timed_pellet_retrieval_time.append(0)
                                
                                index += 1
                                comp_bin_index += 1

                            if bin_num[-1] > index:

                                for i in range(0, len(retrieval_time2)):
                                    if bin_num[i] == index:
                                        if math.isnan(retrieval_time_forbins[i]) == False:
                                            sum_all_retrieval_time += retrieval_time_forbins[i]
                                        if retrieval_time_forbins[i] != 0 and math.isnan(retrieval_time_forbins[i]) == False:
                                            timed_pellet_counter += 1
                                            sum_timed_retrieval_time += retrieval_time_forbins[i]
                                                
                                    else:
                                        if sum_all_retrieval_time != 0:
                                            bin_sum_all_pellet_retrieval_time.append(sum_all_retrieval_time)
                                            bin_mean_all_pellet_retrieval_time.append(sum_all_retrieval_time / active_binned[comp_bin_index - 1])
                                        else:
                                            bin_sum_all_pellet_retrieval_time.append(np.nan)
                                            bin_mean_all_pellet_retrieval_time.append(np.nan)
                                        
                                        if sum_timed_retrieval_time != 0:
                                            bin_sum_timed_pellet_retrieval_time.append(sum_timed_retrieval_time)
                                        else:
                                            bin_sum_timed_pellet_retrieval_time.append(np.nan)
                                            
                                        if timed_pellet_counter != 0:
                                            bin_mean_timed_pellet_retrieval_time.append(sum_timed_retrieval_time / timed_pellet_counter)
                                        else:
                                            bin_mean_timed_pellet_retrieval_time.append(np.nan)
                                        
                                        index += 1 
                                        comp_bin_index += 1
                                        
                                        if math.isnan(retrieval_time_forbins[i]) == False:
                                            sum_all_retrieval_time += retrieval_time_forbins[i]
                                        if retrieval_time_forbins[i] != 0 and math.isnan(retrieval_time_forbins[i]) == False:
                                            timed_pellet_counter += 1
                                            sum_timed_retrieval_time += retrieval_time_forbins[i]
                                            
                                        if i != (len(bin_num) - 1):

                                            if bin_num[i] == bin_num_binned[comp_bin_index - 1] and bin_num[i + 1] > bin_num_binned[comp_bin_index - 1]: # if there is only 1 poke in the bin need to add the data to the lists otherwise it just gets included in the following block
                                                if sum_all_retrieval_time != 0:
                                                    bin_sum_all_pellet_retrieval_time.append(sum_all_retrieval_time)
                                                    bin_mean_all_pellet_retrieval_time.append(sum_all_retrieval_time / active_binned[comp_bin_index - 1])
                                                else:
                                                    bin_sum_all_pellet_retrieval_time.append(np.nan)
                                                    bin_mean_all_pellet_retrieval_time.append(np.nan)
                                                
                                                if sum_timed_retrieval_time != 0:
                                                    bin_sum_timed_pellet_retrieval_time.append(sum_timed_retrieval_time)
                                                else:
                                                    bin_sum_timed_pellet_retrieval_time.append(np.nan)
                                                    
                                                if timed_pellet_counter != 0:
                                                    bin_mean_timed_pellet_retrieval_time.append(sum_timed_retrieval_time / timed_pellet_counter)
                                                else:
                                                    bin_mean_timed_pellet_retrieval_time.append(np.nan)
                                                
                                                index += 1 
                                                comp_bin_index += 1
                                        
                                            while bin_num[i + 1] > index:
                                                index += 1

                                if int(seconds_elapsed[-1]) <= int(session_duration_mins) * 60: # triggers adding of last bin if last time point hasn't exceeded session length
                                    bin_sum_all_pellet_retrieval_time.append(sum_all_retrieval_time)
                                    if active_binned[-1] != 0:
                                        bin_mean_all_pellet_retrieval_time.append(sum_all_retrieval_time / active_binned[-1])
                                    else:
                                        bin_mean_all_pellet_retrieval_time.append(np.nan)
                                    
                                    bin_sum_timed_pellet_retrieval_time.append(sum_timed_retrieval_time)
                                    if timed_pellet_counter != 0:
                                        bin_mean_timed_pellet_retrieval_time.append(sum_timed_retrieval_time / timed_pellet_counter)
                                    else:
                                        bin_mean_timed_pellet_retrieval_time.append(np.nan)
                            else:
                                for i in range(0, len(retrieval_time2)):
                                    if int(seconds_elapsed[i]) <= int(seconds_in_bins) * index:
                                        if math.isnan(retrieval_time_forbins[i]) == False:
                                            sum_all_retrieval_time += retrieval_time_forbins[i]
                                        if retrieval_time_forbins[i] != 0:
                                            if math.isnan(retrieval_time_forbins[i]) == False:
                                                sum_timed_retrieval_time += retrieval_time_forbins[i]
                                                timed_pellet_counter += 1
                                
                                bin_sum_all_pellet_retrieval_time.append(sum_all_retrieval_time)
                                if active_binned[index - 1]!= 0:
                                    bin_mean_all_pellet_retrieval_time.append(sum_all_retrieval_time / active_binned[index - 1])
                                else:
                                    bin_mean_all_pellet_retrieval_time.append('N/A')
                                
                                bin_sum_timed_pellet_retrieval_time.append(sum_timed_retrieval_time)
                                if timed_pellet_counter != 0:
                                    bin_mean_timed_pellet_retrieval_time.append(sum_timed_retrieval_time / timed_pellet_counter)
                                else:
                                    bin_mean_timed_pellet_retrieval_time.append('N/A')
                                        
                                index += 1   
                            
                            # Enter new row for any bins that have no data in them (for cumulative bins this is a duplication of the preceding row)

                            for i in bin_dictionary:
                                if bin_dictionary[i] in bin_num_binned:
                                    continue
                                else:
                                    bin_num_binned.insert((bin_dictionary[i] - 1), bin_dictionary[i])
                                    active_binned.insert((bin_dictionary[i] - 1), active_binned[(bin_dictionary[i] - 2)])
                                    inactive_binned.insert((bin_dictionary[i] - 1), inactive_binned[(bin_dictionary[i] - 2)])
                                    total_pokes_binned.insert((bin_dictionary[i] - 1), total_pokes_binned[(bin_dictionary[i] - 2)])
                                    pellet_binned.insert((bin_dictionary[i] - 1), pellet_binned[(bin_dictionary[i] - 2)])
                                    active_percent_binned.insert((bin_dictionary[i] - 1), active_percent_binned[(bin_dictionary[i] - 2)])
                                    timed_pellet_binned.insert((bin_dictionary[i] - 1), timed_pellet_binned[(bin_dictionary[i] - 2)])
                                    bin_sum_all_poke_time.insert((bin_dictionary[i] - 1), bin_sum_all_poke_time[(bin_dictionary[i] - 2)])
                                    bin_sum_a_poke_time.insert((bin_dictionary[i] - 1), bin_sum_a_poke_time[(bin_dictionary[i] - 2)])
                                    bin_sum_i_poke_time.insert((bin_dictionary[i] - 1), bin_sum_i_poke_time[(bin_dictionary[i] - 2)])
                                    bin_mean_all_poke_time.insert((bin_dictionary[i] - 1), bin_mean_all_poke_time[(bin_dictionary[i] - 2)])
                                    bin_mean_a_poke_time.insert((bin_dictionary[i] - 1), bin_mean_a_poke_time[(bin_dictionary[i] - 2)])
                                    bin_mean_i_poke_time.insert((bin_dictionary[i] - 1), bin_mean_i_poke_time[(bin_dictionary[i] - 2)])
                                    bin_sum_all_pellet_retrieval_time.insert((bin_dictionary[i] - 1), bin_sum_all_pellet_retrieval_time[(bin_dictionary[i] - 2)])
                                    bin_mean_all_pellet_retrieval_time.insert((bin_dictionary[i] - 1), bin_mean_all_pellet_retrieval_time[(bin_dictionary[i] - 2)])
                                    bin_sum_timed_pellet_retrieval_time.insert((bin_dictionary[i] - 1), bin_sum_timed_pellet_retrieval_time[(bin_dictionary[i] - 2)])
                                    bin_mean_timed_pellet_retrieval_time.insert((bin_dictionary[i] - 1), bin_mean_timed_pellet_retrieval_time[(bin_dictionary[i] - 2)])

                            num_of_bins = int(session_duration_mins) / (int(seconds_in_bins) / 60)
                            
                            if len(bin_num_binned) > num_of_bins:
                                
                                del bin_num_binned[int(num_of_bins):]
                                del active_binned[int(num_of_bins):]
                                del inactive_binned[int(num_of_bins):]
                                del total_pokes_binned[int(num_of_bins):]
                                del pellet_binned[int(num_of_bins):]
                                del active_percent_binned[int(num_of_bins):]
                                del timed_pellet_binned[int(num_of_bins):]
                                del bin_sum_all_poke_time[int(num_of_bins):]
                                del bin_sum_a_poke_time[int(num_of_bins):]
                                del bin_sum_i_poke_time[int(num_of_bins):]
                                del bin_mean_all_poke_time[int(num_of_bins):]
                                del bin_mean_a_poke_time[int(num_of_bins):]
                                del bin_mean_i_poke_time[int(num_of_bins):]
                                del bin_sum_all_pellet_retrieval_time[int(num_of_bins):]
                                del bin_mean_all_pellet_retrieval_time[int(num_of_bins):]
                                del bin_sum_timed_pellet_retrieval_time[int(num_of_bins):]
                                del bin_mean_timed_pellet_retrieval_time[int(num_of_bins):]
                            
                            # Make the counts within the bins rather than cumulative
                            
                            active_binned_within = []
                            inactive_binned_within = []
                            total_pokes_binned_within = []
                            pellet_binned_within = []
                            active_percent_binned_within = []
                            timed_pellet_binned_within = []
                            
                            active_binned_within.append(active_binned[0])
                            for i in range(1, len(active_binned)):
                                active_binned_within.append(active_binned[i] - active_binned[i - 1])
                            
                            inactive_binned_within.append(inactive_binned[0])
                            for i in range(1, len(inactive_binned)):
                                inactive_binned_within.append(inactive_binned[i] - inactive_binned[i - 1])
                            
                            total_pokes_binned_within.append(total_pokes_binned[0])
                            for i in range(1, len(total_pokes_binned)):
                                total_pokes_binned_within.append(total_pokes_binned[i] - total_pokes_binned[i - 1])
                            
                            pellet_binned_within.append(pellet_binned[0])
                            for i in range(1, len(pellet_binned)):
                                pellet_binned_within.append(pellet_binned[i] - pellet_binned[i - 1])
                                
                            for i in range(0, len(total_pokes_binned_within)):
                                if active_binned_within[i] + inactive_binned_within[i] != 0:
                                    active_percent_binned_within.append(active_binned_within[i] / total_pokes_binned_within[i] * 100)
                                else:
                                    active_percent_binned_within.append(np.nan)
                            
                            timed_pellet_binned_within.append(timed_pellet_binned[0])
                            for i in range(1, len(timed_pellet_binned)):
                                timed_pellet_binned_within.append(timed_pellet_binned[i] - timed_pellet_binned[i - 1])
                                
    
                            bin_within_sum_all_poke_time = [] # within bin sum of all poke times
                            bin_within_sum_a_poke_time = [] # within bin sum of active poke times
                            bin_within_sum_i_poke_time = [] # within bin sum of inactive poke times
                            
                            bin_within_sum_all_poke_time.append(bin_sum_all_poke_time[0])
                            bin_within_sum_a_poke_time.append(bin_sum_a_poke_time[0])
                            bin_within_sum_i_poke_time.append(bin_sum_i_poke_time[0])  

                            for i in range(1, len(bin_sum_all_poke_time)):
                                bin_within_sum_all_poke_time.append((bin_sum_all_poke_time[i] - bin_sum_all_poke_time[i - 1]))
                                bin_within_sum_a_poke_time.append((bin_sum_a_poke_time[i] - bin_sum_a_poke_time[i - 1]))
                                bin_within_sum_i_poke_time.append((bin_sum_i_poke_time[i] - bin_sum_i_poke_time[i - 1]))  
                            
                            bin_within_mean_all_poke_time = [] # within bin mean of all poke times
                            bin_within_mean_a_poke_time = [] # within bin mean of active poke times
                            bin_within_mean_i_poke_time = [] # within bin mean of inactive poke times

                            for i in range(0, len(bin_within_sum_all_poke_time)):
                                
                                if total_pokes_binned_within[i] != 0:
                                    bin_within_mean_all_poke_time.append((bin_within_sum_all_poke_time[i] / total_pokes_binned_within[i]))
                                else:
                                    bin_within_mean_all_poke_time.append(np.nan)
                                
                                if active_binned_within[i] != 0:
                                    bin_within_mean_a_poke_time.append((bin_within_sum_a_poke_time[i] / active_binned_within[i]))
                                else:
                                    bin_within_mean_a_poke_time.append(np.nan)
                                
                                if inactive_binned_within[i] != 0:
                                    bin_within_mean_i_poke_time.append((bin_within_sum_i_poke_time[i] / inactive_binned_within[i]))
                                else:
                                    bin_within_mean_i_poke_time.append(np.nan)
                            
                            bin_within_sum_all_pellet_retrieval_time = [] # within bin sum of pellet retrieval times
                            bin_within_mean_all_pellet_retrieval_time = [] # within bin mean of pellet retrieval times
                            
                            bin_within_sum_timed_pellet_retrieval_time = []
                            bin_within_mean_timed_pellet_retrieval_time = []
                            
                            bin_within_sum_all_pellet_retrieval_time.append(bin_sum_all_pellet_retrieval_time[0])
                            bin_within_sum_timed_pellet_retrieval_time.append(bin_sum_timed_pellet_retrieval_time[0])
                            for i in range(1, len(bin_sum_all_pellet_retrieval_time)):
                                bin_within_sum_all_pellet_retrieval_time.append(bin_sum_all_pellet_retrieval_time[i] - bin_sum_all_pellet_retrieval_time[i - 1])        
                                if bin_sum_timed_pellet_retrieval_time[i] != np.nan:
                                    if bin_sum_timed_pellet_retrieval_time[i - 1] != np.nan:
                                        bin_within_sum_timed_pellet_retrieval_time.append(bin_sum_timed_pellet_retrieval_time[i] - bin_sum_timed_pellet_retrieval_time[i - 1])
                                    else:
                                        bin_within_sum_timed_pellet_retrieval_time.append(np.nan)
    
                            for i in range(0, len(bin_within_sum_all_pellet_retrieval_time)):
                                if active_binned_within[i] != 0:
                                    bin_within_mean_all_pellet_retrieval_time.append(bin_within_sum_all_pellet_retrieval_time[i] / active_binned_within[i])
                                else:
                                    bin_within_mean_all_pellet_retrieval_time.append(np.nan)
                                
                                if timed_pellet_binned_within[i] != 0:
                                    bin_within_mean_timed_pellet_retrieval_time.append(bin_within_sum_timed_pellet_retrieval_time[i] / timed_pellet_binned_within[i])
                                else:
                                    bin_within_mean_timed_pellet_retrieval_time.append(np.nan)

                            
                        ######-----##### Create session summary data ######-----#####
                        
                        task = schedule_used
                        
                        session_duration = duration_from_start[-1] 
                        
                        date = import_name[7:13]
                        aus_date = date[2:4] + '/' + date[0:2] + '/20' + date[4:]
                        
                        if initiation_poke_active == True:
                            num_rev = ((pellet_block[-1] + 1) / int(reversal_block_length)) # the +1 accounts for the removal of the initiation poke pellet
                        else:
                            num_rev = (pellet_block[-1] / int(reversal_block_length))

                        num_comprev = math.floor(num_rev) # gives number of completed reversals
                        
                        # Create win-stay and lose-shift data (for all blocks)
                        
                        win_stay_all = 0
                        win_shift_all = 0
                        lose_shift_all = 0
                        lose_stay_all = 0
                        
                        for i in range (0, (len(active_poke_binary) - 1)):
                            if active_poke_binary[i] == 1:
                                if active_poke_binary[i + 1] == 1:
                                    win_stay_all += 1
                                elif active_poke_binary[i + 1] == 0:
                                    win_shift_all += 1
                            elif active_poke_binary[i] == 0:
                                if active_poke_binary[i + 1] == 1:
                                    lose_shift_all += 1
                                elif active_poke_binary[i + 1] == 0:
                                    lose_stay_all += 1

                        if active_poke_binary[-1] == 1:
                            if active_poke[-1] != 1 and active_poke[-1] != 0:
                                win_stay_proportion_all = win_stay_all / (active_poke[-1] - 1)
                            else:
                                win_stay_proportion_all = 'N/A'
                            if inactive_poke[-1] != 0:
                                lose_shift_proportion_all = lose_shift_all / inactive_poke[-1]
                            else:
                                lose_shift_proportion_all = 'N/A'
                        elif active_poke_binary[-1] == 0:
                            if active_poke[-1] != 0:
                                win_stay_proportion_all = win_stay_all / active_poke[-1]
                            else:
                                win_stay_proportion_all = 'N/A'
                            if inactive_poke[-1] != 1 and inactive_poke[-1] != 0:
                                lose_shift_proportion_all = lose_shift_all / (inactive_poke[-1] - 1)
                            else:
                                lose_shift_proportion_all = 'N/A'
                        
                        # If the final reversal block is incomplete: Determine how many pokes are in it
    
                        if active_poke_block_within[-1] != int(reversal_block_length):
                            excess_pokes_to_remove = total_block_within[-1]
                        else:
                            excess_pokes_to_remove = 0
                        
                        # Remove corresponding data from the active_poke_binary list for calculating win-stay and lose-shift values for only completed reversals in session
                        
                        active_poke_binary_comp = active_poke_binary[: (len(active_poke_binary) - excess_pokes_to_remove)]
                        active_poke_comp = active_poke[: (len(active_poke) - excess_pokes_to_remove)]
                        inactive_poke_comp = inactive_poke[: (len(inactive_poke) - excess_pokes_to_remove)]
                        
                        win_stay_comp = 0
                        win_shift_comp = 0
                        lose_shift_comp = 0
                        lose_stay_comp = 0
                        
                        for i in range (0, (len(active_poke_binary_comp) - 1)):
                            if active_poke_binary_comp[i] == 1:
                                if active_poke_binary_comp[i + 1] == 1:
                                    win_stay_comp += 1
                                elif active_poke_binary_comp[i + 1] == 0:
                                    win_shift_comp += 1
                            elif active_poke_binary_comp[i] == 0:
                                if active_poke_binary_comp[i + 1] == 1:
                                    lose_shift_comp += 1
                                elif active_poke_binary_comp[i + 1] == 0:
                                    lose_stay_comp += 1
                                    
                        if num_comprev != 0:
                            if active_poke_binary_comp[-1] == 1:
                                if active_poke_comp[-1] != 1 and active_poke_comp[-1] != 0:
                                    win_stay_proportion_comp = win_stay_comp / (active_poke_comp[-1] - 1)
                                else:
                                    win_stay_proportion_comp = 'N/A'
                                if inactive_poke_comp[-1]!= 0:
                                    lose_shift_proportion_comp = lose_shift_comp / inactive_poke_comp[-1]
                                else:
                                    lose_shift_proportion_comp = 'N/A'
                            elif active_poke_binary_comp[-1] == 0:
                                if active_poke_comp[-1] != 0:
                                    win_stay_proportion_comp = win_stay_comp / active_poke_comp[-1]
                                else:
                                    win_stay_proportion_comp = 'N/A'
                                if inactive_poke_comp[-1] != 1 and inactive_poke_comp[-1] != 0:
                                    lose_shift_proportion_comp = lose_shift_comp / (inactive_poke_comp[-1] - 1)
                                else:
                                    lose_shift_proportion_comp = 'N/A'
                        else:
                            win_stay_proportion_comp = 'N/A'
                            lose_shift_proportion_comp = 'N/A'
                        
                        # Remove corresponding excess data from the poke_time and retrieval_time lists for calculating mean time for all completed reversals in session
                        
                        poke_time3 = poke_time2[: (len(poke_time2) - excess_pokes_to_remove)] # chronological list
                        retrieval_time3 = retrieval_time2[: (len(retrieval_time2) - excess_pokes_to_remove)] # chronological list
                        
                        # Also remove from l/r and a/i poke times
                        
                        l_poke_time_chron2 = l_poke_time_chron[: (len(l_poke_time_chron) - excess_pokes_to_remove)]
                        r_poke_time_chron2 = r_poke_time_chron[: (len(r_poke_time_chron) - excess_pokes_to_remove)]
                        a_poke_time_chron2 = a_poke_time_chron[: (len(a_poke_time_chron) - excess_pokes_to_remove)]
                        a_l_poke_time_chron2 = a_l_poke_time_chron[: (len(a_l_poke_time_chron) - excess_pokes_to_remove)]
                        a_r_poke_time_chron2 = a_r_poke_time_chron[: (len(a_r_poke_time_chron) - excess_pokes_to_remove)]
                        i_poke_time_chron2 = i_poke_time_chron[: (len(i_poke_time_chron) - excess_pokes_to_remove)]
                        i_l_poke_time_chron2 = i_l_poke_time_chron[: (len(i_l_poke_time_chron) - excess_pokes_to_remove)]
                        i_r_poke_time_chron2 = i_r_poke_time_chron[: (len(i_r_poke_time_chron) - excess_pokes_to_remove)]
                        
                        # Remove nan values for calculating mean values only for completed reversals in session
                        
                        retrieval_time4 = [x for x in retrieval_time3 if not math.isnan(x)] 
                        l_poke_time_chron3 = [x for x in l_poke_time_chron2 if not math.isnan(x)]
                        r_poke_time_chron3 = [x for x in r_poke_time_chron2 if not math.isnan(x)]
                        a_poke_time_chron3 = [x for x in a_poke_time_chron2 if not math.isnan(x)]
                        a_l_poke_time_chron3 = [x for x in a_l_poke_time_chron2 if not math.isnan(x)]
                        a_r_poke_time_chron3 = [x for x in a_r_poke_time_chron2 if not math.isnan(x)]
                        i_poke_time_chron3 = [x for x in i_poke_time_chron2 if not math.isnan(x)]
                        i_l_poke_time_chron3 = [x for x in i_l_poke_time_chron2 if not math.isnan(x)]
                        i_r_poke_time_chron3 = [x for x in i_r_poke_time_chron2 if not math.isnan(x)]
                        
                        # Remove nan values for calculating mean values for ALL ATTEMPTED reversals in session
                        
                        retrieval_time_all = [x for x in retrieval_time2 if not math.isnan(x)]
                        l_poke_time_chron_all = [x for x in l_poke_time_chron if not math.isnan(x)]
                        r_poke_time_chron_all = [x for x in r_poke_time_chron if not math.isnan(x)]
                        a_poke_time_chron_all = [x for x in a_poke_time_chron if not math.isnan(x)]
                        a_l_poke_time_chron_all = [x for x in a_l_poke_time_chron if not math.isnan(x)]
                        a_r_poke_time_chron_all = [x for x in a_r_poke_time_chron if not math.isnan(x)]
                        i_poke_time_chron_all = [x for x in i_poke_time_chron if not math.isnan(x)]
                        i_l_poke_time_chron_all = [x for x in i_l_poke_time_chron if not math.isnan(x)]
                        i_r_poke_time_chron_all = [x for x in i_r_poke_time_chron if not math.isnan(x)]
                        
                        # Determine how many pellets were earned from completed reversals
                            
                        if pellet_block_within[-1] != int(reversal_block_length): # if the final block is incomplete
                            if pellet_count_shifted[-1] > int(reversal_block_length): # if at least one block was completed
                                pellet_comp = pellet_block[-2]
                                timed_pellet_comp = block_timed_pellet_count[-2]
                            else: # if the first block was incomplete
                                pellet_comp = 0
                                timed_pellet_comp = 0
                            
                        else:
                            pellet_comp = pellet_block[-1]
                            timed_pellet_comp = block_timed_pellet_count[-1]
                        
                        # Remove excess pokes from end session poke counts as well (if required), and collate pokes by side and port status

                        if num_rev != num_comprev or excess_pokes_to_remove != 0: # If the session was cut off in the middle of a reversal block we want to remove the data from the incomplete final block (can probably be just the last part of the condition, will double check)
                            
                            if num_rev >= 1: # if at least 1 reversal was completed
                            
                                l_pokes = left_block[-2] # poke totals are the cumulative total at the completion of the final completed reversal block
                                r_pokes = right_block[-2]
                                a_pokes = active_poke_block[-2]
                                i_pokes = inactive_poke_block[-2]
                                
                                l_active = 0
                                l_inactive = 0
                                r_active = 0
                                r_inactive = 0
                                
                                for i in range(0, (len(active_block) - 1)): # Collates the pokes by side and port status and EXCLUDES the final incomplete block
                                    if active_block[i] == 'Left':
                                        l_active = l_active + active_poke_block_within[i]
                                        r_inactive = r_inactive + inactive_poke_block_within[i]
                                    elif active_block[i] == 'Right':
                                        r_active = r_active + active_poke_block_within[i]
                                        l_inactive = l_inactive + inactive_poke_block_within[i]
                            
                            else: # if the first block is incomplete, i.e. no reversals completed
                        
                                l_pokes = left_block[-1]
                                r_pokes = right_block[-1]
                                a_pokes = active_poke_block[-1]
                                i_pokes = inactive_poke_block[-1]
                                
                                l_active = 0
                                l_inactive = 0
                                r_active = 0
                                r_inactive = 0
                                
                                for i in range(0, len(active_block)): # Collates the pokes by side and port status
                                    if active_block[i] == 'Left':
                                        l_active = l_active + active_poke_block_within[i]
                                        r_inactive = r_inactive + inactive_poke_block_within[i]
                                    elif active_block[i] == 'Right':
                                        r_active = r_active + active_poke_block_within[i]
                                        l_inactive = l_inactive + inactive_poke_block_within[i]
                            
                        else: # if the final block was complete it is kept/included in the data
                        
                            l_pokes = left_block[-1]
                            r_pokes = right_block[-1]
                            a_pokes = active_poke_block[-1]
                            i_pokes = inactive_poke_block[-1]
                            
                            l_active = 0
                            l_inactive = 0
                            r_active = 0
                            r_inactive = 0
                            
                            for i in range(0, len(active_block)): # Collates the pokes by side and port status
                                if active_block[i] == 'Left':
                                    l_active = l_active + active_poke_block_within[i]
                                    r_inactive = r_inactive + inactive_poke_block_within[i]
                                elif active_block[i] == 'Right':
                                    r_active = r_active + active_poke_block_within[i]
                                    l_inactive = l_inactive + inactive_poke_block_within[i]
                        
                        # Values for summary ALL, i.e. includes final block even if incomplete
                        
                        l_pokes_all = left_block[-1]
                        r_pokes_all = right_block[-1]
                        a_pokes_all = active_poke_block[-1]
                        i_pokes_all = inactive_poke_block[-1]
                        
                        l_active_all = 0
                        l_inactive_all = 0
                        r_active_all = 0
                        r_inactive_all = 0
                        
                        for i in range(0, len(active_block)): # Collates the pokes by side and port status
                            if active_block[i] == 'Left':
                                l_active_all = l_active_all + active_poke_block_within[i]
                                r_inactive_all = r_inactive_all + inactive_poke_block_within[i]
                            elif active_block[i] == 'Right':
                                r_active_all = r_active_all + active_poke_block_within[i]
                                l_inactive_all = l_inactive_all + inactive_poke_block_within[i]
                        
                        # Summary values for all COMPLETED reversals
                        
                        t_pokes = l_pokes + r_pokes
                        
                        if l_pokes != 0:
                            l_active_prop = l_active / l_pokes * 100
                            l_inactive_prop = l_inactive / l_pokes * 100
                        else:
                            l_active_prop = 'N/A'
                            l_inactive_prop = 'N/A'
                        
                        if r_pokes != 0:
                            r_active_prop = r_active / r_pokes * 100
                            r_inactive_prop = r_inactive / r_pokes * 100
                        else:
                            r_active_prop = 'N/A'
                            r_inactive_prop = 'N/A'
                        
                        if a_pokes != 0:
                            a_l_prop = l_active / a_pokes * 100
                            a_r_prop = r_active / a_pokes * 100
                        else:
                            a_l_prop = 'N/A'
                            a_r_prop = 'N/A'
                        
                        if i_pokes != 0:
                            i_l_prop = l_inactive / i_pokes * 100
                            i_r_prop = r_inactive / i_pokes * 100
                        else:
                            i_l_prop = 'N/A'
                            i_r_prop = 'N/A'
                        
                        if num_comprev >= 1:
                        
                            mean_i_per_comprev = i_pokes / num_comprev
                            if r_inactive != 0:
                                mean_i_per_rev_l_act = r_inactive / math.ceil(num_comprev / 2)
                            else:
                                mean_i_per_rev_l_act = 'N/A'
                            
                            if l_inactive != 0:
                                mean_i_per_rev_r_act = l_inactive / math.floor(num_comprev / 2)
                            else:
                                mean_i_per_rev_r_act = 'N/A'
                                
                            mean_poke_time = sum(poke_time3) / len(poke_time3)
                            
                            if len(l_poke_time_chron3) != 0:
                                mean_l_poke_time = sum(l_poke_time_chron3) / len(l_poke_time_chron3)
                            else:
                                mean_l_poke_time = 'N/A'
                            
                            if len(r_poke_time_chron3) != 0:
                                mean_r_poke_time = sum(r_poke_time_chron3) / len(r_poke_time_chron3)
                            else:
                                mean_r_poke_time = 'N/A'
                            
                            if len(a_poke_time_chron3) != 0:
                                mean_a_poke_time = sum(a_poke_time_chron3) / len(a_poke_time_chron3)
                            else:
                                mean_a_poke_time = 'N/A'
                            
                            if len(a_l_poke_time_chron3) != 0:
                                mean_a_l_poke_time = sum(a_l_poke_time_chron3) / len(a_l_poke_time_chron3)
                            else:
                                mean_a_l_poke_time = 'N/A'
                            
                            if len(a_r_poke_time_chron3) != 0:
                                mean_a_r_poke_time = sum(a_r_poke_time_chron3) / len(a_r_poke_time_chron3)
                            else:
                                mean_a_r_poke_time = 'N/A'
                            
                            if len(i_poke_time_chron3) != 0:
                                mean_i_poke_time = sum(i_poke_time_chron3) / len(i_poke_time_chron3)
                            else:
                                mean_i_poke_time = 'N/A'
                                
                            if len(i_l_poke_time_chron3) != 0:
                                mean_i_l_poke_time = sum(i_l_poke_time_chron3) / len(i_l_poke_time_chron3)
                            else:
                                mean_i_l_poke_time = 'N/A'
                                
                            if len(i_r_poke_time_chron3) != 0:
                                mean_i_r_poke_time = sum(i_r_poke_time_chron3) / len(i_r_poke_time_chron3)
                            else:
                                mean_i_r_poke_time = 'N/A'
                                
                            mean_all_retrieval_time = sum(retrieval_time4) / len(retrieval_time4)
    
                            if excess_pokes_to_remove == 0:
                                mean_timed_retrieval_time = float(block_sum_all_pellet_retrieval_time[-1]) / block_timed_pellet_count[-1]
                            else:
                                if num_comprev == 1:
                                    mean_timed_retrieval_time = float(block_sum_all_pellet_retrieval_time[0]) / block_timed_pellet_count[0]
                                else:
                                    mean_timed_retrieval_time = float(block_sum_all_pellet_retrieval_time[-2]) / block_timed_pellet_count[-2]
                        
                        else:
                            mean_i_per_comprev = 'N/A'
                            mean_i_per_rev_l_act = 'N/A'
                            mean_i_per_rev_r_act = 'N/A'   
                            mean_poke_time = 'N/A'
                            mean_l_poke_time = 'N/A'
                            mean_r_poke_time = 'N/A'
                            mean_a_poke_time = 'N/A'
                            mean_a_l_poke_time = 'N/A'
                            mean_a_r_poke_time = 'N/A'
                            mean_i_poke_time = 'N/A'
                            mean_i_l_poke_time = 'N/A'
                            mean_i_r_poke_time = 'N/A'
                            mean_all_retrieval_time = 'N/A'
                            mean_timed_retrieval_time = 'N/A'
                            
                        
                        # num_rev determines progress in reversals based on pellets, so is restricted to counting active pokes
                        # but there may be an attempt to complete a reversal that is restricted to inactive pokes and this needs to be counted as an attempted reversal block
                        # therefore. determine if there are any additional inactive pokes indicating attempt at another reversal

                        if total_poke[-1] != t_pokes: # if the total number of pokes following the last recorded poke is not the same as the total number of pokes at the end of the last completed reversal...
                           
                            num_revattempt = num_comprev + 1 # gives number of attempted reversal blocks
                           
                        else: # if no reversals are completed need to indicate that the first reversal block was attempted
                            if num_comprev >= 1: 
                                num_revattempt = num_comprev
                            else:
                                num_revattempt = 1
                                
                        # Summary values for ALL reversals (including final block even if incomplete)
                        
                        t_pokes_all = l_pokes_all + r_pokes_all
                        
                        if l_pokes_all != 0:
                            l_active_prop_all = l_active_all / l_pokes_all * 100
                            l_inactive_prop_all = l_inactive_all / l_pokes_all * 100
                        else:
                            l_active_prop_all = 'N/A'
                            l_inactive_prop_all = 'N/A'
                        
                        if r_pokes_all != 0:
                            r_active_prop_all = r_active_all / r_pokes_all * 100
                            r_inactive_prop_all = r_inactive_all / r_pokes_all * 100
                        else:
                            r_active_prop_all = 'N/A'
                            r_inactive_prop_all = 'N/A'
                        
                        if a_pokes_all != 0:
                            a_l_prop_all = l_active_all / a_pokes_all * 100
                            a_r_prop_all = r_active_all / a_pokes_all * 100
                        else:
                            a_l_prop_all = 'N/A'
                            a_r_prop_all = 'N/A'
                        
                        if i_pokes_all != 0:
                            i_l_prop_all = l_inactive_all / i_pokes_all * 100
                            i_r_prop_all = r_inactive_all / i_pokes_all * 100
                        else:
                            i_l_prop_all = 'N/A'
                            i_r_prop_all = 'N/A'
                        
                        mean_i_per_revattempt = i_pokes_all / num_revattempt
                        
                        if r_inactive_all != 0:
                            mean_i_per_rev_l_act_all = r_inactive_all / math.ceil(num_revattempt / 2)
                        else:
                            mean_i_per_rev_l_act_all = 'N/A'
                        
                        if l_inactive_all != 0:
                            mean_i_per_rev_r_act_all = l_inactive_all / math.floor(num_revattempt / 2)
                        else:
                            mean_i_per_rev_r_act_all = 'N/A'
                            
                        mean_poke_time_all = sum(poke_time2) / len(poke_time2)
                        
                        if len(l_poke_time_chron_all) != 0:
                            mean_l_poke_time_all = sum(l_poke_time_chron_all) / len(l_poke_time_chron_all)
                        else:
                            mean_l_poke_time_all = 'N/A'
                        
                        if len(r_poke_time_chron_all) != 0:
                            mean_r_poke_time_all = sum(r_poke_time_chron_all) / len(r_poke_time_chron_all)
                        else:
                            mean_r_poke_time_all = 'N/A'
                        
                        if len(a_poke_time_chron_all) != 0:
                            mean_a_poke_time_all = sum(a_poke_time_chron_all) / len(a_poke_time_chron_all)
                        else:
                            mean_a_poke_time_all = 'N/A'
                        
                        if len(a_l_poke_time_chron_all) != 0:
                            mean_a_l_poke_time_all = sum(a_l_poke_time_chron_all) / len(a_l_poke_time_chron_all)
                        else:
                            mean_a_l_poke_time_all = 'N/A'
                        
                        if len(a_r_poke_time_chron_all) != 0:
                            mean_a_r_poke_time_all = sum(a_r_poke_time_chron_all) / len(a_r_poke_time_chron_all)
                        else:
                            mean_a_r_poke_time_all = 'N/A'
                        
                        if len(i_poke_time_chron_all) != 0:
                            mean_i_poke_time_all = sum(i_poke_time_chron_all) / len(i_poke_time_chron_all)
                        else:
                            mean_i_poke_time_all = 'N/A'
                            
                        if len(i_l_poke_time_chron_all) != 0:
                            mean_i_l_poke_time_all = sum(i_l_poke_time_chron_all) / len(i_l_poke_time_chron_all)
                        else:
                            mean_i_l_poke_time_all = 'N/A'
                            
                        if len(i_r_poke_time_chron_all) != 0:
                            mean_i_r_poke_time_all = sum(i_r_poke_time_chron_all) / len(i_r_poke_time_chron_all)
                        else:
                            mean_i_r_poke_time_all = 'N/A'
                            
                        if sum(retrieval_time_all) != 0:
                            mean_all_retrieval_time_all = sum(retrieval_time_all) / len(retrieval_time_all)
                        else:
                            mean_all_retrieval_time_all = 'N/A'

                        if float(block_sum_all_pellet_retrieval_time[-1]) == 0 or block_timed_pellet_count[-1] == 0:
                            mean_timed_retrieval_time_all = 'N/A'
                        else:
                            mean_timed_retrieval_time_all = float(block_sum_all_pellet_retrieval_time[-1]) / block_timed_pellet_count[-1]
                        
            
                        variable = ['Filename', 'Date', 'Task', 'Duration', 'Reversal Block Length', '',
                                    'Number of Reversals Attempted','Number of Reversals Achieved', 'Number of Reversals Completed', '',
                                    'Win-Stay', 'Lose-Shift', '',
                                    'Total Pokes', 'Mean Poke Time', 'Pellets', 'Timed Pellets', 'Mean ALL Pellet Retrieval Time', 'Mean TIMED Pellet Retrieval Time', '',
                                    'Total Left Pokes', 'Mean Left Poke Time', 'Left Active', 'Left Active %', 'Mean Active Left Poke Time', 'Left Inactive', 'Left Inactive %', 'Mean Inactive Left Poke Time', '',
                                    'Total Right Pokes', 'Mean Right Poke Time', 'Right Active', 'Right Active %', 'Mean Active Right Poke Time', 'Right Inactive', 'Right Inactive %', 'Mean Inactive Right Poke Time', '',
                                    'Total Active Pokes', 'Mean Active Poke Time', 'Active Left %', 'Active Right %', '',
                                    'Total Inactive Pokes', 'Mean Inactive Poke Time', 'Inactive Left proportion (%)', 'Inactive Right proportion (%)', '',
                                    'Mean Inactive pokes per Reversal', 'Mean Left Inactive pokes per Reversal', 'Mean Right Inactive pokes per Reversal']
                        value_comp = [import_name.strip('.CSV'), aus_date, task, session_duration, int(reversal_block_length), '',
                                  num_revattempt, num_rev, num_comprev, '',
                                  win_stay_proportion_comp, lose_shift_proportion_comp, '',
                                  t_pokes, mean_poke_time, pellet_comp, timed_pellet_comp, mean_all_retrieval_time, mean_timed_retrieval_time, '',
                                  l_pokes, mean_l_poke_time, l_active, l_active_prop, mean_a_l_poke_time, l_inactive, l_inactive_prop, mean_i_l_poke_time, '',
                                  r_pokes, mean_r_poke_time, r_active, r_active_prop, mean_a_r_poke_time, r_inactive, r_inactive_prop, mean_i_r_poke_time, '',
                                  a_pokes, mean_a_poke_time, a_l_prop, a_r_prop, '',
                                  i_pokes, mean_i_poke_time, i_l_prop, i_r_prop, '',
                                  mean_i_per_comprev, mean_i_per_rev_r_act, mean_i_per_rev_l_act]
                        
                        value_all = [import_name.strip('.CSV'), aus_date, task, session_duration, int(reversal_block_length), '',
                                  num_revattempt, num_rev, num_comprev, '',
                                  win_stay_proportion_all, lose_shift_proportion_all, '',
                                  t_pokes_all, mean_poke_time_all, pellet_block[-1], block_timed_pellet_count[-1], mean_all_retrieval_time_all, mean_timed_retrieval_time_all, '',
                                  l_pokes_all, mean_l_poke_time_all, l_active_all, l_active_prop_all, mean_a_l_poke_time_all, l_inactive_all, l_inactive_prop_all, mean_i_l_poke_time_all, '',
                                  r_pokes_all, mean_r_poke_time_all, r_active_all, r_active_prop_all, mean_a_r_poke_time_all, r_inactive_all, r_inactive_prop_all, mean_i_r_poke_time_all, '',
                                  a_pokes_all, mean_a_poke_time_all, a_l_prop_all, a_r_prop_all, '',
                                  i_pokes_all, mean_i_poke_time_all, i_l_prop_all, i_r_prop_all, '',
                                  mean_i_per_revattempt, mean_i_per_rev_r_act_all, mean_i_per_rev_l_act_all]
                        
                        percentactivepokescomp = (a_pokes / t_pokes * 100)
                        percentactivepokesall = (a_pokes_all / t_pokes_all * 100)
                        
                        date_comp_all.append(aus_date)
                        
                        attempted_comp.append(num_revattempt)
                        achieved_comp.append(num_rev)
                        completed_comp.append(num_comprev)
                        totalpokes_comp.append(t_pokes)
                        activepokes_comp.append(a_pokes)
                        inactivepokes_comp.append(i_pokes)
                        percentactivepokes_comp.append(percentactivepokescomp)
                        pellets_comp.append(pellet_comp)
                        leftpokes_comp.append(l_pokes)
                        leftactivepokes_comp.append(l_active)
                        leftinactivepokes_comp.append(l_inactive)
                        rightpokes_comp.append(r_pokes)
                        rightactivepokes_comp.append(r_active)
                        rightinactivepokes_comp.append(r_inactive)
                        inactiveperreversal_comp.append(mean_i_per_comprev)
                        leftinactiveperreversal_comp.append(mean_i_per_rev_r_act)
                        rightinactiveperreversal_comp.append(mean_i_per_rev_l_act)
                        winstay_comp.append(win_stay_proportion_comp)
                        loseshift_comp.append(lose_shift_proportion_comp)
                        
                        attempted_all.append(num_revattempt)
                        achieved_all.append(num_rev)
                        completed_all.append(num_comprev)
                        totalpokes_all.append(t_pokes_all)
                        activepokes_all.append(a_pokes_all)
                        inactivepokes_all.append(i_pokes_all)
                        percentactivepokes_all.append(percentactivepokesall)
                        pellets_all.append(pellet_block[-1])
                        leftpokes_all.append(l_pokes_all)
                        leftactivepokes_all.append(l_active_all)
                        leftinactivepokes_all.append(l_inactive_all)
                        rightpokes_all.append(r_pokes_all)
                        rightactivepokes_all.append(r_active_all)
                        rightinactivepokes_all.append(r_inactive_all)
                        inactiveperreversal_all.append(mean_i_per_revattempt)
                        leftinactiveperreversal_all.append(mean_i_per_rev_r_act_all)
                        rightinactiveperreversal_all.append(mean_i_per_rev_l_act_all)
                        winstay_all.append(win_stay_proportion_all)
                        loseshift_all.append(lose_shift_proportion_all)
                
                        
                        #####-----
                        
                        # print(len(seconds_block), len(duration_block), len(session_block), len(active_block), len(left_block), len(right_block))
                        # print(len(active_poke_block), len(inactive_poke_block), len(total_block), len(percent_active_block))
                        # print(len(pellet_block), len(block_timed_pellet_count))
                        # print(len(block_mean_all_poke_time), len(block_mean_a_poke_time), len(block_mean_i_poke_time))
                        # print(len(block_mean_all_pellet_retrieval_time), len(block_mean_timed_pellet_retrieval_time))
                        
                        # print(len(seconds_block), len(seconds_block_within), len(duration_block), len(session_block), len(active_block), len(left_block_within), len(right_block_within))
                        # print(len(active_poke_block_within), len(inactive_poke_block_within), len(total_block_within), len(percent_active_block_within))
                        # print(len(pellet_block_within), len(block_within_timed_pellet_count))
                        # print(len(block_within_mean_all_poke_time), len(block_within_mean_a_poke_time), len(block_within_mean_i_poke_time))
                        # print(len(block_within_mean_all_pellet_retrieval_time), len(block_within_mean_timed_pellet_retrieval_time))
                        
                        #####-----
                        print(filename + ' check passed')
                        # Export the data.
                        
                        # export_results = 'Y'
                        
                        if export_results == 'Y':
                        
                            # Always export Summary, Chronological and Blocked data    
                            
                            results_summary = {'Variable': variable, 'Value COMP': value_comp, 'Value ALL': value_all}
                            export_file_summary = pd.DataFrame(results_summary, columns = ['Variable', 'Value COMP', 'Value ALL'])
                            
                            
                            results_chronological = {'Time (sec from start)': seconds_elapsed, 'Duration': duration_from_start, 'Session Type': session_type2, 'Active port': active2, 'Left Poke': left_poke2, 'Right Poke': right_poke2, 'Active Poke': active_poke, 'Inactive Poke': inactive_poke, 'Total Poke': total_poke, '% Active Pokes': percent_active, 'Poke Time': poke_time2, 
                                                      'Pellet Count': pellet_count2, 'Pellet Retrieval Time': retrieval_time2,
                                                      'Active': green, 'Inactive': red, 'Active Port': active_port_binary}
                            export_file_chronological = pd.DataFrame(results_chronological, columns = ['Time (sec from start)', 'Duration', 'Session Type', 'Active port', 'Left Poke', 'Right Poke', 'Active Poke', 'Inactive Poke', 'Total Poke', '% Active Pokes', 'Poke Time', 
                                                                              '', 'Pellet Count', 'Pellet Retrieval Time',
                                                                              '', 'Active', 'Inactive', 'Active Port'])
                            
                            results_timing = {'Time (sec from start)': seconds_elapsed, 'Duration': duration_from_start, 'Active Port': active2,
                                              'Binary Active': active_poke_binary, 'Binary Inactive': inactive_poke_binary, 'Binary Left': left_poke_binary, 'Binary Right': right_poke_binary,
                                              'Left Active Binary': left_a_poke_binary, 'Left Inactive Binary': left_i_poke_binary, 'Right Active Binary': right_a_poke_binary, 'Right Inactive Binary': right_i_poke_binary,
                                              'Poke Time': poke_time2, 'Left Poke Time': l_poke_time_chron, 'Right Poke Time': r_poke_time_chron, 
                                              'Active Poke Time': a_poke_time_chron, 'Active Left Poke Time': a_l_poke_time_chron, 'Active Right Poke Time': a_r_poke_time_chron, 
                                              'Inactive Poke Time': i_poke_time_chron, 'Inactive Left Poke Time': i_l_poke_time_chron, 'Inactive Right Poke Time': i_r_poke_time_chron}
                            export_file_timing = pd.DataFrame(results_timing, columns = ['Time (sec from start)', 'Duration', 'Active Port', 
                                                                                          '', 'Binary Active', 'Binary Inactive', 'Binary Left', 'Binary Right',
                                                                                          '', 'Left Active Binary', 'Left Inactive Binary', 'Right Active Binary', 'Right Inactive Binary',
                                                                                          '', 'Poke Time', 'Left Poke Time', 'Right Poke Time',
                                                                                          '', 'Active Poke Time', 'Active Left Poke Time', 'Active Right Poke Time',
                                                                                          '', 'Inactive Poke Time', 'Inactive Left Poke Time', 'Inactive Right Poke Time'])
                            
                            # Always export blocked data
                            
                            results_blocked_c = {'Time (sec from start)': seconds_block, 'Duration': duration_block, 'Session Type': session_block, 'Active port': active_block, 
                                                  'Left Poke': left_block, 'Right Poke': right_block, 'Active Poke': active_poke_block, 'Inactive Poke': inactive_poke_block, 'Total Poke': total_block, '% Active Pokes': percent_active_block,
                                                  'Pellet Count': pellet_block, 'Timed Pellet Count': block_timed_pellet_count,
                                                  'Mean Poke Time': block_mean_all_poke_time, 'Mean Active Poke Time': block_mean_a_poke_time, 'Mean Inactive Poke Time': block_mean_i_poke_time, 
                                                  'Mean ALL Pellet Retrieval Time': block_mean_all_pellet_retrieval_time, 'Mean TIMED Pellet Retrieval Time': block_mean_timed_pellet_retrieval_time}
                            export_file_blocked_c = pd.DataFrame(results_blocked_c, columns = ['Time (sec from start)', 'Duration', 'Session Type', 'Active port',
                                                                                                'Left Poke', 'Right Poke', 'Active Poke', 'Inactive Poke', 'Total Poke', '% Active Pokes',
                                                                                                'Pellet Count', 'Timed Pellet Count',
                                                                                                '', 'Mean Poke Time', 'Mean Active Poke Time', 'Mean Inactive Poke Time', 
                                                                                                'Mean ALL Pellet Retrieval Time', 'Mean TIMED Pellet Retrieval Time'])
                            
                            results_blocked_w = {'Time (sec from start)': seconds_block, 'Time (sec within block)': seconds_block_within, 'Duration': duration_block, 'Duration of Block': duration_of_block_within, 'Session Type': session_block, 'Active port': active_block, 
                                                  'Left Poke': left_block_within, 'Right Poke': right_block_within, 'Active Poke': active_poke_block_within, 'Inactive Poke': inactive_poke_block_within, 'Total Poke': total_block_within, '% Active Pokes': percent_active_block_within,
                                                  'Pellet Count': pellet_block_within, 'Timed Pellet Count': block_within_timed_pellet_count,
                                                  'Mean Poke Time': block_within_mean_all_poke_time, 'Mean Active Poke Time': block_within_mean_a_poke_time, 'Mean Inactive Poke Time': block_within_mean_i_poke_time, 
                                                  'Mean ALL Pellet Retrieval Time': block_within_mean_all_pellet_retrieval_time, 'Mean TIMED Pellet Retrieval Time': block_within_mean_timed_pellet_retrieval_time}
                            export_file_blocked_w = pd.DataFrame(results_blocked_w, columns = ['Time (sec from start)', 'Time (sec within block)', 'Duration', 'Duration of Block', 'Session Type', 'Active port', 
                                                                                                'Left Poke', 'Right Poke', 'Active Poke', 'Inactive Poke', 'Total Poke', '% Active Pokes',
                                                                                                'Pellet Count', 'Timed Pellet Count',
                                                                                                '', 'Mean Poke Time', 'Mean Active Poke Time', 'Mean Inactive Poke Time', 
                                                                                                'Mean ALL Pellet Retrieval Time', 'Mean TIMED Pellet Retrieval Time'])

                            
                            if seconds_in_bins != '':       
                                
                                results_binned_c = {'Time bin (' + str(seconds_in_bins) + 's each)': bin_num_binned, 
                                                    'Active Poke': active_binned, 'Inactive Poke': inactive_binned, 'Total Poke': total_pokes_binned, '% Active Pokes': active_percent_binned, 
                                                    'Mean Active Poke Time': bin_mean_a_poke_time, 'Mean Inactive Poke Time': bin_mean_i_poke_time, 'Mean Poke Time': bin_mean_all_poke_time,
                                                    'Pellet Count': pellet_binned, 'Mean ALL Retrieval Time': bin_mean_all_pellet_retrieval_time, 'Timed Pellet Count': timed_pellet_binned, 'Mean TIMED Retrieval Time': bin_mean_timed_pellet_retrieval_time}
                                export_file_binned_c = pd.DataFrame(results_binned_c, columns = ['Time bin (' + seconds_in_bins + 's each)', 
                                                                                                 'Active Poke', 'Inactive Poke', 'Total Poke', '% Active Pokes', 
                                                                                                 'Mean Active Poke Time', 'Mean Inactive Poke Time', 'Mean Poke Time',
                                                                                                 'Pellet Count', 'Mean ALL Retrieval Time', 'Timed Pellet Count', 'Mean TIMED Retrieval Time'])
                                
                                results_binned_w = {'Time bin (' + str(seconds_in_bins) + 's each)': bin_num_binned, 
                                                    'Active Poke': active_binned_within, 'Inactive Poke': inactive_binned_within, 'Total Poke': total_pokes_binned_within, '% Active Pokes': active_percent_binned_within, 
                                                    'Mean Active Poke Time': bin_within_mean_a_poke_time, 'Mean Inactive Poke Time': bin_within_mean_i_poke_time, 'Mean Poke Time': bin_within_mean_all_poke_time,
                                                    'Pellet Count': pellet_binned_within, 'Mean ALL Retrieval Time': bin_within_mean_all_pellet_retrieval_time, 'Timed Pellet Count': timed_pellet_binned_within, 'Mean TIMED Retrieval Time': bin_within_mean_timed_pellet_retrieval_time}
                                export_file_binned_w = pd.DataFrame(results_binned_w, columns = ['Time bin (' + seconds_in_bins + 's each)', 
                                                                                                 'Active Poke', 'Inactive Poke', 'Total Poke', '% Active Pokes', 
                                                                                                 'Mean Active Poke Time', 'Mean Inactive Poke Time', 'Mean Poke Time',
                                                                                                 'Pellet Count', 'Mean ALL Retrieval Time', 'Timed Pellet Count', 'Mean TIMED Retrieval Time'])
                                
                                
                            # Export to excel
                            
                            from openpyxl import Workbook
                            
                            wb = Workbook()
                            
                            ws1 = wb.active
                            ws1.title = 'Summary'
                            
                            ws2 = wb.create_sheet()
                            ws2.title ='Chronological'
                            
                            ws3 = wb.create_sheet()
                            ws3.title ='Chronological Timing'
                            
                            ws4 = wb.create_sheet()
                            ws4.title ='Blocked Cumulative'
                            
                            ws4 = wb.create_sheet()
                            ws4.title ='Blocked Within'
                            
                            if seconds_in_bins != '':
                                ws5 = wb.create_sheet()
                                ws5.title = 'Binned Cumulative'
                                
                                ws6 = wb.create_sheet()
                                ws6.title = 'Binned Within'
                            
                            sheets_to_export = wb.sheetnames
                                
                            results_to_export = [export_file_summary, export_file_chronological, export_file_timing, export_file_blocked_c, export_file_blocked_w]
                            
                            if seconds_in_bins != '':
                                results_to_export.append(export_file_binned_c)
                                results_to_export.append(export_file_binned_w)
                            
                            
                            with pd.ExcelWriter(export_destination) as writer:
                                
                                for i in range(len(sheets_to_export)):
                                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
                
                
                ##########---------- Overview Sheets ----------##########
                
                # export_overview = 'Y'
                
                if export_overview == 'Y':
                
                    from openpyxl import Workbook
                            
                    wb = Workbook()
                    
                    ws1 = wb.active
                    ws1.title = 'Summary Completed Overview'
                    
                    ws2 = wb.create_sheet()
                    ws2.title = 'Summary ALL Overview'
                    
                    variable_column = 'Variable'
                    value_comp_column = 'Value COMP'
                    value_all_column = 'Value ALL'
                    
                    value_summary = []
                    session_num = []
                    counter = 1
                    
                    for filename in sorted(os.listdir(os.path.join(export_location, folder))):

                        # if filename.endswith(".xlsx"):
                            
                        if 'New Reversal' in filename:

                            if 'Overview' not in filename:    

                                if counter == 1:
                                    
                                    export_name = filename
                                    export_destination = export_location + folder + '/' + export_name

                                    df_overview_comp = pd.read_excel(export_destination, sheet_name = 'Summary')
                                    df_overview_all = pd.read_excel(export_destination, sheet_name = 'Summary')
                                    
                                    df_overview_comp.drop(columns=['Value COMP', 'Value ALL'], inplace=True)
                                    df_overview_all.drop(columns=['Value COMP', 'Value ALL'], inplace=True)
                                    
                                export_name = filename
                                export_destination = export_location + folder + '/' + export_name
                                
                                name = 'Session ' + str(counter)
                                session_num.append(name)
                                
                                df_comp = pd.read_excel(export_destination, sheet_name = 'Summary')
                                
                                values_comp = df_comp[value_comp_column].tolist()
                                
                                df_overview_comp.insert(counter, name, values_comp)
                                
                                df_all = pd.read_excel(export_destination, sheet_name = 'Summary')
                                
                                values_all = df_all[value_all_column].tolist()
                                
                                df_overview_all.insert(counter, name, values_all)
                                
                                counter += 1
                            
                    overview_name = folder + ' New Reversal Overview.xlsx'
                    overview_destination = export_location + folder + '/' + overview_name
                    
                    sheets_to_export = wb.sheetnames
                    results_to_export = [df_overview_comp, df_overview_all]
                    
                    with pd.ExcelWriter(overview_destination) as writer:
                        for i in range(len(sheets_to_export)):          
                            results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
                    
                    cohort_overview_destination = cohort_export_location + overview_name
                    
                    with pd.ExcelWriter(cohort_overview_destination) as writer:
                                    
                        for i in range(len(sheets_to_export)):          
                            results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)     
                            
                    print('Overview completed')
                    
                ##### -----
                if export_joined == 'Y':
                    results_joined_comp = {'Date': date_comp_all, 'Reversals Attempted': attempted_comp, 'Reversals Achieved': achieved_comp, 'Reversals Completed': completed_comp, 
                                           'Win-Stay': winstay_comp, 'Lose-Shift': loseshift_comp,
                                           'Total Pokes': totalpokes_comp, 'Active Pokes': activepokes_comp, 'Inactive Pokes': inactivepokes_comp, 'Percent Active': percentactivepokes_comp, 'Pellets': pellets_comp, 
                                           'Left Pokes': leftpokes_comp, 'Left Active Pokes': leftactivepokes_comp, 'Left Inactive Pokes': leftinactivepokes_comp, 
                                           'Right Pokes': rightpokes_comp, 'Right Active Pokes': rightactivepokes_comp, 'Right Inactive Pokes': rightinactivepokes_comp, 
                                           'Inactive Pokes Per Reversal': inactiveperreversal_comp, 'Left Inactive Pokes Per Reversal': leftinactiveperreversal_comp, 'Right Inactive Pokes Per Reversal': rightinactiveperreversal_comp}
                    export_file_joined_comp = pd.DataFrame(results_joined_comp, columns = ['Date', 'Reversals Attempted', 'Reversals Achieved', 'Reversals Completed', 
                                           'Win-Stay', 'Lose-Shift',
                                           'Total Pokes', 'Active Pokes', 'Inactive Pokes', 'Percent Active', 'Pellets', 
                                           'Left Pokes', 'Left Active Pokes', 'Left Inactive Pokes', 
                                           'Right Pokes', 'Right Active Pokes', 'Right Inactive Pokes', 
                                           'Inactive Pokes Per Reversal', 'Left Inactive Pokes Per Reversal', 'Right Inactive Pokes Per Reversal'])
                    
                    results_joined_all = {'Date': date_comp_all, 'Reversals Attempted': attempted_all, 'Reversals Achieved': achieved_all, 'Reversals Completed': completed_all, 
                                           'Win-Stay': winstay_all, 'Lose-Shift': loseshift_all,
                                           'Total Pokes': totalpokes_all, 'Active Pokes': activepokes_all, 'Inactive Pokes': inactivepokes_all,'Percent Active': percentactivepokes_all, 'Pellets': pellets_all, 
                                           'Left Pokes': leftpokes_all, 'Left Active Pokes': leftactivepokes_all, 'Left Inactive Pokes': leftinactivepokes_all, 
                                           'Right Pokes': rightpokes_all, 'Right Active Pokes': rightactivepokes_all, 'Right Inactive Pokes': rightinactivepokes_all, 
                                           'Inactive Pokes Per Reversal': inactiveperreversal_all, 'Left Inactive Pokes Per Reversal': leftinactiveperreversal_all, 'Right Inactive Pokes Per Reversal': rightinactiveperreversal_all}
                    export_file_joined_all = pd.DataFrame(results_joined_all, columns = ['Date', 'Reversals Attempted', 'Reversals Achieved', 'Reversals Completed', 
                                           'Win-Stay', 'Lose-Shift',
                                           'Total Pokes', 'Active Pokes', 'Inactive Pokes', 'Percent Active', 'Pellets', 
                                           'Left Pokes', 'Left Active Pokes', 'Left Inactive Pokes', 
                                           'Right Pokes', 'Right Active Pokes', 'Right Inactive Pokes', 
                                           'Inactive Pokes Per Reversal', 'Left Inactive Pokes Per Reversal', 'Right Inactive Pokes Per Reversal'])
                    
                    from openpyxl import Workbook
                                
                    wb = Workbook()
                    
                    ws1 = wb.active
                    ws1.title = 'Summary Completed Joined'
                    
                    ws2 = wb.create_sheet()
                    ws2.title = 'Summary ALL Joined'
                    
                    sheets_to_export = wb.sheetnames
                                    
                    results_to_export = [export_file_joined_comp, export_file_joined_all]
                    
                    joined_name = folder + ' Summary Joined.xlsx'
                    joined_destination = export_location + folder + '/' + joined_name
                    joined_cohort_destination = cohort_export_location + joined_name
                                
                    with pd.ExcelWriter(joined_destination) as writer:
                                    
                        for i in range(len(sheets_to_export)):
                            results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
                            
                    with pd.ExcelWriter(joined_cohort_destination) as writer:
                                    
                        for i in range(len(sheets_to_export)):
                            results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)