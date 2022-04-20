#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Aug  4 15:39:11 2021

@author: lauramilton
"""

dates_to_join = '02/07/2021, 06/07/2021, 07/07/2021, 08/07/2021'

dates_to_join_list = dates_to_join.split(", ")

##########----------

import pandas as pd
import numpy as np
import os
import math

import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 Reversal/FED210721/DatafilesPython/TEST/'
export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 Reversal/FED210721/DatafilesPython/TEST/'
cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 Reversal/FED210721/DatafilesPython/TEST/Joined/'

date_column = 'Date'
time_column1 = 'Time bin (600s each)'
time_column2 = 'Time bin (1800s each)'
session_column = 'Session Type'
active_column = 'Active Port'
active_poke_column = 'Active Poke'
inactive_poke_column = 'Inactive Poke'
total_poke_column = 'Total Poke'
pellet_column = 'Pellet Count'
percent_active_column = '% Active Pokes'

session_length = '180' # session length in minutes
time_bin_length1 = '10'
time_bin_length2 = '30'

num_of_bins1 = int(session_length) / int(time_bin_length1)
num_of_bins2 = int(session_length) / int(time_bin_length2)

for folder in sorted(os.listdir(import_location)):
    if folder.startswith('FED'):
        
        files_to_join = []        
        
        for filename in sorted(os.listdir(os.path.join(import_location, folder))):
            
            if filename.endswith(".xlsx"):
                if not filename.startswith('FED'): # can probs turn it around to be if filename.startswith() and use variable/name from the code that generates those files
                    if not filename.startswith('10'): # will be able to remove later once those files aren't in the folder anymore (i.e. when file naming procedure is finalized)
                        
                        print(filename)

                        value_column = 'Value'
                        
                        import_destination = import_location + folder + '/' + filename
                        df = pd.read_excel(import_destination, sheet_name='Summary')
                        
                        values = df[value_column].tolist()
                        
                        if values[1] in dates_to_join_list:
                            files_to_join.append(import_destination)
        
        date_chron1 = []
        time_chron1 = []
        session_chron1 = []
        active_chron1 = []
        active_poke_chron1 = []
        inactive_poke_chron1 = []
        total_poke_chron1 = []
        pellet_chron1 = []
        percent_active_chron1 = []
        
        date_chron2 = []
        time_chron2 = []
        session_chron2 = []
        active_chron2 = []
        active_poke_chron2 = []
        inactive_poke_chron2 = []
        total_poke_chron2 = []
        pellet_chron2 = []
        percent_active_chron2 = []

        session_num = 1
                
        for i in files_to_join:
                
                import_sheet1 = str(time_bin_length1) + 'min Binned Within'
                
                if time_bin_length2 != '':
                    import_sheet2 = str(time_bin_length2) + 'min Binned Within'
                        
                df1 = pd.read_excel(i, sheet_name = import_sheet1)

                if time_bin_length2 != '':
                    df2 = pd.read_excel(i, sheet_name = import_sheet2)

                date1 = df1[date_column].tolist()
                time1 = df1[time_column1].tolist()
                session1 = df1[session_column].tolist()
                active1 = df1[active_column].tolist()
                active_poke1 = df1[active_poke_column].tolist()
                inactive_poke1 = df1[inactive_poke_column].tolist()
                total_poke1 = df1[total_poke_column].tolist()
                pellet1 = df1[pellet_column].tolist()
                percent_active1 = df1[percent_active_column].tolist()
                
                if time_bin_length2 != '':
                    date2 = df2[date_column].tolist()
                    time2 = df2[time_column2].tolist()
                    session2 = df2[session_column].tolist()
                    active2 = df2[active_column].tolist()
                    active_poke2 = df2[active_poke_column].tolist()
                    inactive_poke2 = df2[inactive_poke_column].tolist()
                    total_poke2 = df2[total_poke_column].tolist()
                    pellet2 = df2[pellet_column].tolist()
                    percent_active2 = df2[percent_active_column].tolist()
                
                if len(time1) > num_of_bins1:
                    
                    del date1[int(num_of_bins1):]
                    del time1[int(num_of_bins1):]
                    del session1[int(num_of_bins1):]
                    del active1[int(num_of_bins1):]
                    del active_poke1[int(num_of_bins1):]
                    del inactive_poke1[int(num_of_bins1):]
                    del total_poke1[int(num_of_bins1):]
                    del pellet1[int(num_of_bins1):]
                    del percent_active1[int(num_of_bins1):]
                    
                if time_bin_length2 != '':
                    if len(time2) > num_of_bins2:
                        
                        del date2[int(num_of_bins2):]
                        del time2[int(num_of_bins2):]
                        del session2[int(num_of_bins2):]
                        del active2[int(num_of_bins2):]
                        del active_poke2[int(num_of_bins2):]
                        del inactive_poke2[int(num_of_bins2):]
                        del total_poke2[int(num_of_bins2):]
                        del pellet2[int(num_of_bins2):]
                        del percent_active2[int(num_of_bins2):]
                
                if len(time1) < num_of_bins1:
                    
                    bin_num1 = len(time1) + 1
                    
                    while len(time1) < num_of_bins1:
                        date1.append(date1[0])
                        time1.append(bin_num1)
                        session1.append(np.nan)
                        active1.append(np.nan)
                        active_poke1.append(np.nan)
                        inactive_poke1.append(np.nan)
                        total_poke1.append(np.nan)
                        pellet1.append(np.nan)
                        percent_active1.append(np.nan)
                        bin_num1 +=1
                
                if time_bin_length2 != '':
                    if len(time2) < num_of_bins2:
                    
                        bin_num2 = len(time2) + 1
                        
                        while len(time2) < num_of_bins2:
                            date2.append(date2[0])
                            time2.append(bin_num2)
                            session2.append(np.nan)
                            active2.append(np.nan)
                            active_poke2.append(np.nan)
                            inactive_poke2.append(np.nan)
                            total_poke2.append(np.nan)
                            pellet2.append(np.nan)
                            percent_active2.append(np.nan)
                            bin_num2 +=1
                
                for a, b, c, d, e, f, g, h, i in zip(time1, session1, active1, active_poke1, inactive_poke1, total_poke1, pellet1, percent_active1, date1):
                        
                        time_chron1.append(str(session_num) + '.' + str(a))
                        session_chron1.append(b)
                        active_chron1.append(c) 
                        active_poke_chron1.append(d)
                        inactive_poke_chron1.append(e)
                        total_poke_chron1.append(f)
                        pellet_chron1.append(g)
                        percent_active_chron1.append(h)
                        date_chron1.append(i)
                        
                if time_bin_length2 != '':
                    for a, b, c, d, e, f, g, h, i in zip(time2, session2, active2, active_poke2, inactive_poke2, total_poke2, pellet2, percent_active2, date2):
                        
                            time_chron2.append(str(session_num) + '.' + str(a))
                            session_chron2.append(b)
                            active_chron2.append(c) 
                            active_poke_chron2.append(d)
                            inactive_poke_chron2.append(e)
                            total_poke_chron2.append(f)
                            pellet_chron2.append(g)
                            percent_active_chron2.append(h)
                            date_chron2.append(i)
                
                session_num += 1
                
                session_chron1 = ['NA' if x != x else x for x in session_chron1]
                active_chron1 = ['NA' if x != x else x for x in active_chron1] 
                active_poke_chron1 = ['NA' if x != x else x for x in active_poke_chron1]
                inactive_poke_chron1 = ['NA' if x != x else x for x in inactive_poke_chron1]
                total_poke_chron1 = ['NA' if x != x else x for x in total_poke_chron1]
                pellet_chron1 = ['NA' if x != x else x for x in pellet_chron1]
                percent_active_chron1 = ['NA' if x != x else x for x in percent_active_chron1]
                
                if time_bin_length2 != '':
                    session_chron2 = ['NA' if x != x else x for x in session_chron2]
                    active_chron2 = ['NA' if x != x else x for x in active_chron2] 
                    active_poke_chron2 = ['NA' if x != x else x for x in active_poke_chron2]
                    inactive_poke_chron2 = ['NA' if x != x else x for x in inactive_poke_chron2]
                    total_poke_chron2 = ['NA' if x != x else x for x in total_poke_chron2]
                    pellet_chron2 = ['NA' if x != x else x for x in pellet_chron2]
                    percent_active_chron2 = ['NA' if x != x else x for x in percent_active_chron2]

                ##### Export results

                results1 = {'Date': date_chron1, 'Time Bin': time_chron1, 'Session Type': session_chron1, 'Active Port': active_chron1, 'Active Poke': active_poke_chron1, 'Inactive Poke': inactive_poke_chron1, 'Total Poke': total_poke_chron1, 'Pellet Count': pellet_chron1, 'Percent Active': percent_active_chron1}
                export_file1 = pd.DataFrame(results1, columns = ['Date', 'Time Bin', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])

                if time_bin_length2 !='':
                    results2 = {'Date': date_chron2, 'Time Bin': time_chron2, 'Session Type': session_chron2, 'Active Port': active_chron2, 'Active Poke': active_poke_chron2, 'Inactive Poke': inactive_poke_chron2, 'Total Poke': total_poke_chron2, 'Pellet Count': pellet_chron2, 'Percent Active': percent_active_chron2}
                    export_file2 = pd.DataFrame(results2, columns = ['Date', 'Time Bin', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
                
                from openpyxl import Workbook
            
                wb = Workbook()
                
                ws1 = wb.active
                ws1.title = str(time_bin_length1) + 'min binned chron joined'
                
                if time_bin_length2 != '':
                    ws2 = wb.create_sheet()
                    ws2.title = str(time_bin_length2) + 'min binned chron joined'
                
                results_to_export = [export_file1]
                
                if time_bin_length2 != '':
                    results_to_export.append(export_file2)
                
                sheets_to_export = wb.sheetnames
                
                chron_name = folder + ' chron sessions joined.xlsx'
                chron_destination = export_location + folder + '/' + chron_name

                with pd.ExcelWriter(chron_destination) as writer:
                    
                    for i in range(len(sheets_to_export)):
                        results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)

                cohort_destination = cohort_export_location + chron_name
                
                with pd.ExcelWriter(cohort_destination) as writer:
                    
                    for i in range(len(sheets_to_export)):
                        results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)                  