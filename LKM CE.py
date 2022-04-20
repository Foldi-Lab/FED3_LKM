#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Sep 20 11:14:54 2021

@author: lauramilton
"""

# PROGRESSIVE RATIO TASK

time_column = 'MM:DD:YYYY hh:mm:ss'
event_column = 'Event'
active_poke_column = 'Active_Poke'
session_type_column = 'FR'
left_poke_column = 'Left_Poke_Count'
right_poke_column = 'Right_Poke_Count'
pellet_count_column = 'Pellet_Count'
block_pellet_count_column = 'Block_Pellet_Count'

retrieval_time_column = 'Retrieval_Time'
ipi_column = 'InterPelletInterval'
poke_time_column = 'Poke_Time'

import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/LKM FED 011121/'
export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/LKM FED 011121/FED32/Output/'
# cohort_export_location = 

initiation_poke_active = True
initiation_poke_inactive = False

export_results = 'Y'

#-----------------------------------------------------------------------------

# Import the revelant data: time, FR ratio, event, active port, left poke, right poke, and pellet count.

import pandas as pd
import numpy as np
import os
import openpyxl
import math


#-----

for folder in sorted(os.listdir(import_location)):
    if folder.startswith('FED32'):

        for filename in sorted(os.listdir(os.path.join(import_location, folder))):
            
            if filename.endswith(".CSV"):
                
                if '110921' in filename or '110921' in filename:
                        
            # Import the csv data
            
                    import_name = filename
                    import_destination = import_location + folder + '/' + import_name
                    export_name = folder + ' Closed economy PR ' + import_name.strip('.CSV') + '.xlsx'
                    export_destination = export_location + export_name
                    print(filename)
            #-----
            
                    df = pd.read_csv(import_destination)
                    
                    time = df[time_column].tolist()
                    session_type = df[session_type_column].tolist()
                    event = df[event_column].tolist()
                    active = df[active_poke_column].tolist()
                    left_poke = df[left_poke_column].tolist()
                    right_poke = df[right_poke_column].tolist()
                    pellet_count = df[pellet_count_column].tolist()
                    block_pellet_count = df[block_pellet_count_column].tolist()
                    retrieval_time = df[retrieval_time_column].tolist()
                    ipi = df[ipi_column].tolist()
                    poke_time = df[poke_time_column].tolist()

                    # Change the time column from strings to datetime format for calculating durations between timestamps and creating time bins
                    
                    import datetime as dt
                    
                    time_list = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time]
                    
                    # Start time of the session is the first timestamp
                    
                    start_time = time_list[0]
                    
                    # Create column indicating which CE/PR block each row is in
                    
                    block_count = []
                    counter = 1
                    
                    block_count.append(counter)
                    for i in range(1, len(block_pellet_count)):
                        if block_pellet_count[i] > block_pellet_count[i - 1]:
                            block_count.append(counter)
                        elif block_pellet_count[i] == block_pellet_count[i - 1]:
                            block_count.append(counter)
                        elif block_pellet_count[i] < block_pellet_count[i - 1]:
                            counter += 1
                            block_count.append(counter)
                    
                    # Remove the initiation poke/s and pellet data (if required)
                                
                    if initiation_poke_active == True:
                                    
                        left_count = left_poke[0]
                        right_count = right_poke[0]
                        
                        del time[:2]
                        del session_type[:2]
                        del event[:2]
                        del active[:2]
                        del left_poke[:2]
                        del right_poke[:2]
                        del pellet_count[:2]
                        del block_pellet_count[:2]
                        del block_count[:2]
                        del retrieval_time[:2]
                        del poke_time[:2]
                        del ipi[:2]
                        
                        # Subtract the poke and pellet from the subsequent cumulative data
                        
                        left_poke_shifted = []
                        right_poke_shifted = []
                        pellet_count_shifted = []
                        
                        for i in range(0, len(left_poke)):
                            left_poke_shifted.append(left_poke[i] - left_count)
                            right_poke_shifted.append(right_poke[i] - right_count)
                            pellet_count_shifted.append(pellet_count[i] - 1)
                        
                        
                        # Shift the retrieval_time and pellet_count lists one step backwards so that the count/time is in the same index as the corresponding poke
                        
                        retrieval_time.pop(0)
                        pellet_count_shifted.pop(0)
                        ipi.pop(0)
                        block_pellet_count.pop(0)
                        
                        # Add in a value to the end of each list to maintain list length so as not to lose the last row of other data
                        # for pellet_count this is a duplicate of preceding row, for retrieval_time and ipi this is np.nan
                        
                        retrieval_time.append(np.nan)
                        pellet_count_shifted.append(pellet_count[-1])
                        ipi.append(np.nan)
                        block_pellet_count.append(block_pellet_count[-1])
                        
                        #####-----##### CHRONOLOGICAL DATA #####-----#####
                        
                        # Create new lists that only include data from the lines where the event is a Poke, in ReversalTask this means != 'Pellet' (i.e. Poke, Left, or Right)
                        
                        time2 = []
                        session_type2 = []
                        event2 = []
                        active2 = []
                        left_poke2 = []
                        right_poke2 = []
                        pellet_count2 = []
                        retrieval_time2 = []
                        poke_time2 = []
                        ipi_2 = []
                        block_pellet_count2 = []
                        block_count2 = []

                        for a, b, c, d, e, f, g, h, i, j, k, l in zip(event, left_poke_shifted, right_poke_shifted, pellet_count_shifted, time, session_type, active, retrieval_time, poke_time, ipi, block_pellet_count, block_count):
                            if a != 'Pellet':
                                event2.append(a)
                                left_poke2.append(b)
                                right_poke2.append(c) 
                                pellet_count2.append(d)
                                time2.append(e)
                                session_type2.append(f)
                                active2.append(g)
                                retrieval_time2.append(h)
                                poke_time2.append(i)
                                ipi_2.append(j)
                                block_pellet_count2.append(k)
                                block_count2.append(l)
                    
                        
                        

                    
                    elif initiation_poke_inactive == True: # only remove 1 line of data (the inactive initiaton poke) as there is no pellet line, also do NOT subtract 1 from pellet count, i.e. no need for pellet_count_shifted
                                    
                        left_count = left_poke[0]
                        right_count = right_poke[0]
                        
                        del time[:1]
                        del session_type[:1]
                        del event[:1]
                        del active[:1]
                        del left_poke[:1]
                        del right_poke[:1]
                        del pellet_count[:1]
                        del retrieval_time[:1]
                        del poke_time[:1]
                        del ipi[:2]
                        
                        # Subtract the poke from the subsequent cumulative data
                        
                        left_poke_shifted = []
                        right_poke_shifted = []
                        
                        for i in range(0, len(left_poke)):
                            left_poke_shifted.append(left_poke[i] - left_count)
                            right_poke_shifted.append(right_poke[i] - right_count)
                        
                        # Shift the retrieval_time and pellet_count lists one step backwards so that the count/time is in the same index as the corresponding poke
                        
                        retrieval_time.pop(0)
                        pellet_count.pop(0)
                        ipi.pop(0)
                        block_pellet_count.pop(0)
                        
                        # Add in a value to the end of each list to maintain list length so as not to lose the last row of other data
                        # for pellet_count this is a duplicate of preceding row, for retrieval_time this is np.nan
                        
                        retrieval_time.append(np.nan)
                        pellet_count.append(pellet_count[-1])
                        ipi.append(np.nan)
                        block_pellet_count.append(block_pellet_count[-1])
                        
                        #####-----##### CHRONOLOGICAL DATA #####-----#####
                        
                        # Create new lists that only include data from the lines where the event is a Poke, in ReversalTask this means != 'Pellet' (i.e. Poke, Left, or Right)
                        
                        time2 = []
                        session_type2 = []
                        event2 = []
                        active2 = []
                        left_poke2 = []
                        right_poke2 = []
                        pellet_count2 = []
                        retrieval_time2 = []
                        poke_time2 = []
                        ipi_2 = []
                        block_pellet_count2 = []
                        block_count2 = []

                        for a, b, c, d, e, f, g, h, i, j, k, l in zip(event, left_poke_shifted, right_poke_shifted, pellet_count_shifted, time, session_type, active, retrieval_time, poke_time, ipi, block_pellet_count, block_count):
                            if a != 'Pellet':
                                event2.append(a)
                                left_poke2.append(b)
                                right_poke2.append(c) 
                                pellet_count2.append(d)
                                time2.append(e)
                                session_type2.append(f)
                                active2.append(g)
                                retrieval_time2.append(h)
                                poke_time2.append(i)
                                ipi_2.append(j)
                                block_pellet_count2.append(k)
                                block_count2.append(l)
                    
                    else:
                        # Shift the retrieval_time and pellet_count lists one step backwards so that the count/time is in the same index as the corresponding poke
                        
                        retrieval_time.pop(0)
                        pellet_count.pop(0)
                        ipi.pop(0)
                        block_pellet_count.pop(0)

                        # Add in a value to the end of each list to maintain list length so as not to lose the last row of other data
                        # for pellet_count this is a duplicate of preceding row, for retrieval_time this is np.nan
                        
                        retrieval_time.append(np.nan)
                        pellet_count.append(pellet_count[-1])
                        ipi.append(np.nan)
                        block_pellet_count.append(block_pellet_count[-1])

                        #####-----##### CHRONOLOGICAL DATA #####-----#####
                        
                        # Create new lists that only include data from the lines where the event is a Poke, in ReversalTask this means != 'Pellet' (i.e. Poke, Left, or Right)
                        
                        time2 = []
                        session_type2 = []
                        event2 = []
                        active2 = []
                        left_poke2 = []
                        right_poke2 = []
                        pellet_count2 = []
                        retrieval_time2 = []
                        poke_time2 = []
                        ipi_2 = []
                        block_pellet_count2 = []
                        block_count2 = []

                        for a, b, c, d, e, f, g, h, i, j, k, l in zip(event, left_poke_shifted, right_poke_shifted, pellet_count_shifted, time, session_type, active, retrieval_time, poke_time, ipi, block_pellet_count, block_count):
                            if a != 'Pellet':
                                event2.append(a)
                                left_poke2.append(b)
                                right_poke2.append(c) 
                                pellet_count2.append(d)
                                time2.append(e)
                                session_type2.append(f)
                                active2.append(g)
                                retrieval_time2.append(h)
                                poke_time2.append(i)
                                ipi_2.append(j)
                                block_pellet_count2.append(k)
                                block_count2.append(l)
                    
                    # use the 0's in the list of block_pellet_count that occur with CE/PR reset to put a 0.00 into the retrieval_time2 and ipi_2 lists to maintain array length and show that these cells are empty
                    
                    for i in range(0, len(block_pellet_count2)):
                        if block_pellet_count2[i] == 0:
                            del retrieval_time2[i]
                            retrieval_time2.insert(i, 0.00)
                            del ipi_2[i]
                            ipi_2.insert(i, 0.00)

                    # replace the 0's in the list of block_pellet_count that occur with CE/PR reset with the previous value in the list
                        
                    for idx,val in enumerate(block_pellet_count2):
                        if idx==0: continue
                        if val == 0:
                            block_pellet_count2[idx] = block_pellet_count2[idx-1]
                            

                    # if initiation_poke_active is True remove the pellet resulting from the initiation poke from the values in ONLY THE FIRST BLOCK of block_pellet_count
                    
                    block_pellet_count_shifted = []
                    
                    if initiation_poke_active == True:
                        for i in range(0, len(block_pellet_count2)):
                            if block_count2[i] == 1:
                                block_pellet_count_shifted.append(block_pellet_count2[i] - 1)
                            else:
                                block_pellet_count_shifted.append(block_pellet_count2[i])
                    
                    # print(len(time), len(session_type), len(event), len(active), len(left_poke), len(right_poke), len(pellet_count), len(retrieval_time),len(poke_time), len(ipi), len(block_pellet_count), len(block_count))                
                    # print(len(time2), len(session_type2), len(event2), len(active2), len(left_poke2), len(right_poke2), len(pellet_count2), len(retrieval_time2),len(poke_time2), len(ipi_2), len(block_pellet_count_shifted), len(block_count2))                
                    
                    ####################################################################################################
                    ####################################################################################################
                    
                    # Change the time2 column from strings to datetime format for calculating durations between timestamps
                    
                    time2 = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time2]
                    
                    # Calculate how much time (in seconds) has elapsed from start time for each event (nose poke)
                    
                    seconds_elapsed = []
                    
                    for i in range(0, len(time2)):
                        time_from_start = time2[i] - start_time
                        result = int(time_from_start.total_seconds())
                        seconds_elapsed.append(result)
                        
                    # Transform seconds_elapsed into duration in h:mm:ss
                                
                    duration_from_start = []
                    
                    for i in range(0, len(seconds_elapsed)):
                        time_hours = str(math.floor(seconds_elapsed[i] / 3600))
                        
                        time_mins = math.floor((seconds_elapsed[i] % 3600) / 60)
                        
                        time_secs = seconds_elapsed[i] % 60
                            
                        if time_mins < 10:
                            t_mins = '0' + str(time_mins)
                        else:
                            t_mins = str(time_mins)
                            
                        if time_secs < 10:
                            t_secs = '0' + str(time_secs)
                        else:
                            t_secs = str(time_secs)
                            
                        duration_from_start.append(str(time_hours) + ':' + t_mins + ':' + t_secs)
                    
                    # Create a new column that is the total number of pokes
                    
                    total_pokes = []
                    
                    for i in range(0, len(active2)):
                        total_pokes.append(left_poke2[i] + right_poke2[i])
                    
                    # Create new column that is % of pokes that are into the active port. Needs to be based on active_poke value (i.e. Left or Right)
                    
                    percent_active = []
                    
                    for i in range(0, len(active2)):
                        if active2[i] == 'Left':
                            percent_active.append(left_poke2[i] / total_pokes[i] * 100)
                        else:
                            percent_active.append(right_poke2[i] / total_pokes[i] * 100)
                        
                    # Chronological Poke Time Data split by side
                    
                    l_poke_time = [] # poke time for all left pokes
                    r_poke_time = [] # poke time for all right pokes
                    
                    l_poke_time_chron = [] # poke time for all left pokes with nan if it was a right poke to maintain array length the same as all pokes
                    r_poke_time_chron = [] # poke time for all right pokes with nan if it was a left poke to maintain array length the same as all pokes
                    
                    for i in range(0, len(poke_time2)): # Collates the poke_time by side
                       if event2[i] == 'Left':
                           l_poke_time.append(poke_time2[i])
                           l_poke_time_chron.append(poke_time2[i])
                           r_poke_time_chron.append(np.nan)
                       else:
                           r_poke_time.append(poke_time2[i])
                           r_poke_time_chron.append(poke_time2[i])
                           l_poke_time_chron.append(np.nan)
                    
                    ####################################################################################################
                    ####################################################################################################

                    # Create bins based on PR steps - These are cumulative
                    
                    step_bin = []
                    seconds_elapsed_binned = []
                    duration_binned = []
                    left_binned = []
                    right_binned = []
                    pellet_binned = []
                    active_binned = []
                    percent_active_binned = []
                    total_pokes_binned = []
                    block_pellet_binned = []
                    block_count_binned = []

                    for i in range(1, len(session_type2)):
                        if session_type2[i - 1] != session_type2[i]:
                            step_bin.append(session_type2[i - 1])
                            seconds_elapsed_binned.append(seconds_elapsed[i - 1])
                            duration_binned.append(duration_from_start[i - 1])
                            left_binned.append(left_poke2[i - 1])
                            right_binned.append(right_poke2[i - 1])
                            pellet_binned.append(pellet_count2[i - 1])
                            active_binned.append(active2[i - 1])
                            total_pokes_binned.append(total_pokes[i - 1])
                            block_pellet_binned.append(block_pellet_count_shifted[i - 1])
                            block_count_binned.append(block_count2[i - 1])
                    step_bin.append(session_type2[i - 0])
                    seconds_elapsed_binned.append(seconds_elapsed[i - 0])
                    duration_binned.append(duration_from_start [i - 0])
                    left_binned.append(left_poke2[i - 0])
                    right_binned.append(right_poke2[i - 0])
                    pellet_binned.append(pellet_count2[i - 0])
                    active_binned.append(active2[i - 0])
                    total_pokes_binned.append(total_pokes[i - 0])
                    block_pellet_binned.append(block_pellet_count_shifted[i - 0])
                    block_count_binned.append(block_count2[i - 0])
                    
                    # percent active binned
                    for i in range(0, len(active_binned)):
                        if active_binned[i] == 'Left':
                            percent_active_binned.append(left_binned[i] / total_pokes_binned[i] * 100)
                        else:
                            percent_active_binned.append(right_binned[i] / total_pokes_binned[i] * 100)
                    
                    # Create new bins where counts are within bins rather than cumulative
                    
                    seconds_elapsed_binned_within = []
                    left_binned_within = []
                    right_binned_within = []
                    pellet_binned_within = []
                    percent_active_binned_within = []
                    total_pokes_binned_within = []
                    
                    seconds_elapsed_binned_within.append(seconds_elapsed_binned[0])
                    for i in range(1, len(seconds_elapsed_binned)):
                        seconds_elapsed_binned_within.append(seconds_elapsed_binned[i] - seconds_elapsed_binned[i - 1])
                    
                    # Transform seconds_elapsed into duration in h:mm:ss
                                
                    duration_binned_within = []
                    
                    for i in range(0, len(seconds_elapsed_binned_within)):
                        time_hours = str(math.floor(seconds_elapsed_binned_within[i] / 3600))
                        
                        time_mins = math.floor((seconds_elapsed_binned_within[i] % 3600) / 60)
                        
                        time_secs = seconds_elapsed_binned_within[i] % 60
                            
                        if time_mins < 10:
                            t_mins = '0' + str(time_mins)
                        else:
                            t_mins = str(time_mins)
                            
                        if time_secs < 10:
                            t_secs = '0' + str(time_secs)
                        else:
                            t_secs = str(time_secs)
                            
                        duration_binned_within.append(str(time_hours) + ':' + t_mins + ':' + t_secs)

                    left_binned_within.append(left_binned[0])
                    for i in range(1, len(left_binned)):
                        left_binned_within.append(left_binned[i] - left_binned[i - 1])
                    
                    right_binned_within.append(right_binned[0])
                    for i in range(1, len(right_binned)):
                        right_binned_within.append(right_binned[i] - right_binned[i - 1])
                    
                    pellet_binned_within.append(pellet_binned[0])
                    for i in range(1, len(pellet_binned)):
                        pellet_binned_within.append(pellet_binned[i] - pellet_binned[i - 1])
                        
                    total_pokes_binned_within.append(total_pokes_binned[0])
                    for i in range (1, len(total_pokes_binned)):
                        total_pokes_binned_within.append(total_pokes_binned[i] - total_pokes_binned[i - 1])
                        
                    for i in range(0, len(active_binned)):
                        if active_binned[i] == 'Left':
                            if left_binned_within[i] + right_binned_within[i] != 0:
                                percent_active_binned_within.append(left_binned_within[i] / total_pokes_binned_within[i] * 100)
                            else:
                                percent_active_binned_within.append('NA')
                        else:
                            if left_binned_within[i] + right_binned_within[i] != 0:
                                percent_active_binned_within.append(right_binned_within[i] / total_pokes_binned_within[i] * 100)
                            else:
                                percent_active_binned_within.append('NA')
                    
                    # Binned poke time, retrieval time and ipi data
                    
                    # Create mean poke and retrieval time columns for bins
                                
                    binned_sum_all_poke_time = [] # cumulative sum of all poke times
                    binned_sum_l_poke_time = [] # cumulative sum of active poke times
                    binned_sum_r_poke_time = [] # cumulative sum of inactive poke times
                    binned_mean_all_poke_time = [] # cumulative mean of all poke times
                    binned_mean_l_poke_time = [] # cumulative mean of active poke times
                    binned_mean_r_poke_time = [] # cumulative mean of inactive poke times
                    
                    
                    sum_all_poke_time = 0
                    sum_l_poke_time = 0
                    sum_r_poke_time = 0
                    index = 0
                    poke_index = 0
                    # print(seconds_elapsed, seconds_elapsed_binned, left_poke2, poke_time2)
                    # print(len(seconds_elapsed), len(poke_time2), len(seconds_elapsed_binned))
                    for i in range(0, len(poke_time2)):
                        if int(seconds_elapsed[i]) < int(seconds_elapsed_binned[index]):
                            
                            sum_all_poke_time += poke_time2[poke_index]
                            if event2[i] == 'Left':
                                sum_l_poke_time += poke_time2[poke_index]
                            elif event2[i] == 'Right':
                                sum_r_poke_time += poke_time2[poke_index]
                            poke_index += 1
                            
                        elif int(seconds_elapsed[i]) == int(seconds_elapsed_binned[index]):
                            sum_all_poke_time += poke_time2[poke_index]
                            if event2[i] == 'Left':
                                sum_l_poke_time += poke_time2[poke_index]
                            elif event2[i] == 'Right':
                                sum_r_poke_time += poke_time2[poke_index]
                            poke_index += 1
                            
                            binned_sum_all_poke_time.append(sum_all_poke_time)
                            binned_sum_l_poke_time.append(sum_l_poke_time)
                            binned_sum_r_poke_time.append(sum_r_poke_time)
                            binned_mean_all_poke_time.append((sum_all_poke_time / int(total_pokes_binned[index])))
                            if int(left_binned[index]) != 0:
                                binned_mean_l_poke_time.append((sum_l_poke_time / int(left_binned[index])))
                            else:
                                binned_mean_l_poke_time.append('N/A')
                            if int(right_binned[index]) != 0:
                                binned_mean_r_poke_time.append((sum_r_poke_time / int(right_binned[index])))
                            else:
                                binned_mean_r_poke_time.append('N/A')
                            
                            index += 1
                        
                        elif int(seconds_elapsed[i]) > int(seconds_elapsed_binned[index]):
                            
                            binned_sum_all_poke_time.append(sum_all_poke_time)
                            binned_sum_l_poke_time.append(sum_l_poke_time)
                            binned_sum_r_poke_time.append(sum_r_poke_time)
                            binned_mean_all_poke_time.append((sum_all_poke_time / int(total_pokes_binned[index])))
                            if int(left_binned[index]) != 0:
                                binned_mean_l_poke_time.append((sum_l_poke_time / int(left_binned[index])))
                            else:
                                binned_mean_l_poke_time.append('N/A')
                            if int(right_binned[index]) != 0:
                                binned_mean_r_poke_time.append((sum_r_poke_time / int(right_binned[index])))
                            else:
                                binned_mean_r_poke_time.append('N/A')
                            
                            index += 1
                            
                            sum_all_poke_time += poke_time2[poke_index]
                            if event2[i] == 'Left':
                                sum_l_poke_time += poke_time2[poke_index]
                            elif event2[i] == 'Right':
                                sum_r_poke_time += poke_time2[poke_index]
                            poke_index += 1
                    
                    binned_within_sum_all_poke_time = [] # within binned sum of all poke times
                    binned_within_sum_l_poke_time = [] # within binned sum of active poke times
                    binned_within_sum_r_poke_time = [] # within binned sum of inactive poke times
                    
                    binned_within_sum_all_poke_time.append(binned_sum_all_poke_time[0])
                    binned_within_sum_l_poke_time.append(binned_sum_l_poke_time[0])
                    binned_within_sum_r_poke_time.append(binned_sum_r_poke_time[0])  
                    
                    for i in range(1, len(binned_sum_all_poke_time)):
                        binned_within_sum_all_poke_time.append((binned_sum_all_poke_time[i] - binned_sum_all_poke_time[i - 1]))
                        binned_within_sum_l_poke_time.append((binned_sum_l_poke_time[i] - binned_sum_l_poke_time[i - 1]))
                        binned_within_sum_r_poke_time.append((binned_sum_r_poke_time[i] - binned_sum_r_poke_time[i - 1]))  
                    
                    binned_within_mean_all_poke_time = [] # within binned mean of all poke times
                    binned_within_mean_l_poke_time = [] # within binned mean of active poke times
                    binned_within_mean_r_poke_time = [] # within binned mean of inactive poke times
                    
                    for i in range(0, len(binned_within_sum_all_poke_time)):
                        binned_within_mean_all_poke_time.append((binned_within_sum_all_poke_time[i] / total_pokes_binned_within[i]))
                        if left_binned_within[i] != 0:
                            binned_within_mean_l_poke_time.append((binned_within_sum_l_poke_time[i] / left_binned_within[i]))
                        else:
                            binned_within_mean_l_poke_time.append('N/A')
                        if right_binned_within[i] != 0:
                            binned_within_mean_r_poke_time.append((binned_within_sum_r_poke_time[i] / right_binned_within[i]))
                        else:
                            binned_within_mean_r_poke_time.append('N/A')
                    
                    # retrieval time and ipi bins will just be the lists with the nans removed as there is one value for each bin (PR step)
                    
                    retrieval_time_binned = [x for x in retrieval_time2 if not math.isnan(x)] # gets rid of all the nan values
                    ipi_binned = [x for x in ipi_2 if not math.isnan(x)] # gets rid of all the nan values
                                        
                    # replace the 0.00 entered as placekeeper in retrieval_time2, retrieval_time_binned, ipi_2, and ipi_binned with 'N/A'
                    
                    retrieval_time2 = [np.nan if x == 0 else x for x in retrieval_time2]
                    ipi_2 = [np.nan if x == 0 else x for x in ipi_2]
                    retrieval_time_binned = [np.nan if x == 0 else x for x in retrieval_time_binned]
                    ipi_binned = [np.nan if x == 0 else x for x in ipi_binned]
                    
                    # if final step is incomplete need to add nan to end or retrieval_time_binned and ipi_binned
                    
                    if event[-1] != 'Pellet':
                        retrieval_time_binned.append(np.nan)
                        ipi_binned.append(np.nan)
                    
                    ####################################################################################################
                    ####################################################################################################

                    # Create chronological data that is within each PR step (the poke time data has already been done above)
                    
                    # create binary poke columns
                                
                    left_poke_binary = []
                    
                    if left_poke2[0] == 1:
                        left_poke_binary.append(1)
                    else:
                        left_poke_binary.append(0)
                    
                    for i in range(1, len(left_poke2)):
                        if left_poke2[i - 1] != left_poke2[i]:
                            left_poke_binary.append(1)
                        else:
                            left_poke_binary.append(0)
                    
                    
                    right_poke_binary = []
                        
                    if right_poke2[0] == 1:
                        right_poke_binary.append(1)
                    else:
                        right_poke_binary.append(0)
                    
                    for i in range(1, len(right_poke2)):
                        if right_poke2[i - 1] != right_poke2[i]:
                            right_poke_binary.append(1)
                        else:
                            right_poke_binary.append(0)
                    
                    left_poke_chron_within = []
                    left_poke_chron_within_binary = []
                    left_poke_counter = 0
                    
                    for i in range (0, (len(left_poke_binary)-1)):
                        
                        if session_type2[i + 1] == session_type2[i]:
                            if left_poke_binary[i] == 0:
                                left_poke_chron_within.append(left_poke_counter)
                                left_poke_chron_within_binary.append(np.nan)
                            elif left_poke_binary[i] == 1:
                                left_poke_counter += 1
                                left_poke_chron_within.append(left_poke_counter)
                                left_poke_chron_within_binary.append(left_poke_counter)
                            
                        elif session_type2[i + 1] != session_type2[i]:
                            if left_poke_binary[i] == 0:
                                left_poke_chron_within.append(left_poke_counter)
                                left_poke_chron_within_binary.append(np.nan)
                            elif left_poke_binary[i] == 1:
                                left_poke_counter += 1
                                left_poke_chron_within.append(left_poke_counter)
                                left_poke_chron_within_binary.append(left_poke_counter)
                            left_poke_counter = 0
                            # left_poke_chron_within.append(np.nan)
                            # left_poke_chron_within_binary.append(np.nan)
                    
                    left_poke_chron_within.append(left_binned_within[-1])
                    
                    if left_poke_chron_within[-1] != left_poke_chron_within[-2]:
                        left_poke_chron_within_binary.append(left_binned_within[-1])
                    else:
                        left_poke_chron_within_binary.append(np.nan)
                    
                    
                    right_poke_chron_within = []
                    right_poke_chron_within_binary = []
                    right_poke_counter = 0
                    
                    for i in range (0, (len(right_poke_binary)-1)):
                        
                        if session_type2[i + 1] == session_type2[i]:
                            if right_poke_binary[i] == 0:
                                right_poke_chron_within.append(right_poke_counter)
                                right_poke_chron_within_binary.append(np.nan)
                            elif right_poke_binary[i] == 1:
                                right_poke_counter += 1
                                right_poke_chron_within.append(right_poke_counter)
                                right_poke_chron_within_binary.append(right_poke_counter)
                            
                        elif session_type2[i + 1] != session_type2[i]:
                            if right_poke_binary[i] == 0:
                                right_poke_chron_within.append(right_poke_counter)
                                right_poke_chron_within_binary.append(np.nan)
                            elif right_poke_binary[i] == 1:
                                right_poke_counter += 1
                                right_poke_chron_within.append(right_poke_counter)
                                right_poke_chron_within_binary.append(right_poke_counter)
                            right_poke_counter = 0
                            # right_poke_chron_within.append(np.nan)
                            # right_poke_chron_within_binary.append(np.nan)
                    
                    right_poke_chron_within.append(right_binned_within[-1])
                    
                    if right_poke_chron_within[-1] != right_poke_chron_within[-2]:
                        right_poke_chron_within_binary.append(right_binned_within[-1])
                    else:
                        right_poke_chron_within_binary.append(np.nan)
    
                    # create chronological within data for total pokes and percent active
                    
                    total_pokes_chron_within = []
                    percent_active_chron_within = []
                    
                    for i in range(0, len(left_poke_chron_within)):
                        total_pokes_chron_within.append(left_poke_chron_within[i] + right_poke_chron_within[i])
                        if active2[0] == 'Left':
                            percent_active_chron_within.append((left_poke_chron_within[i] / (left_poke_chron_within[i] + right_poke_chron_within[i]) * 100))
                        elif active2[0] == 'Right':
                            percent_active_chron_within.append((right_poke_chron_within[i] / (left_poke_chron_within[i] + right_poke_chron_within[i]) * 100))
    
                    ###########
    
                    #####----- Create session summary data-----#####
                    
                    
                    task = 'Progressive Ratio'
                    
                    session_duration = duration_from_start[-1]
                    
                    # Identify the PR steps and their completion or lackthereof
                                    
                    if event[-1] == 'Pellet':
                        final_step_comp = True
                    else:
                        final_step_comp = False
                    
                    step = []
                                    
                    for i in range(1, len(session_type)):
                        if session_type[i - 1] != session_type[i]:
                            step.append('PR' + str(session_type[i - 1]))
                    
                    if final_step_comp == True:
                        PR_step_comp = step[-1]
                        next_PR_step = 'N/A'
                        pokes_into_next_step = 'N/A'
                    elif final_step_comp == False:
                        step.append(session_type[-1])
                        PR_step_comp = step[-2]
                        next_PR_step = step[-1]
                        if active2[0] == 'Left':
                            pokes_into_next_step = left_binned_within[-1]
                        elif active2[0] == 'Right':
                            pokes_into_next_step = right_binned_within[-1]
                        
                    # Assign left and right pokes as active or inactive based on the active port
                    
                    if active2[0] == 'Left':
                        active_pokes = left_poke2[-1]
                        inactive_pokes = right_poke2[-1]
                    elif active2[0] == 'Right':
                        active_pokes = right_poke2[-1]
                        inactive_pokes = left_poke2[-1]
                    
                    active_port = active2[0]
                    
                    date = import_name[7:13]
                    aus_date = date[2:4] + '/' + date[0:2] + '/20' + date[4:]
    
                    variable = ['Filename', 'Date', 'Task', 'Duration', 'Active port', 'Last completed PR step', 'Next PR step', 'Active pokes into next step', 'Total Pokes', 'Active Pokes', 'Inactive Pokes', '% Active Pokes', 'Pellets']
                    
                    value = [import_name.strip('.CSV'), aus_date, task, session_duration, active2[0], PR_step_comp, next_PR_step, pokes_into_next_step, total_pokes[-1], active_pokes, inactive_pokes, percent_active[-1], pellet_count2[-1]]
                    
                    #####-----
                    
                    # Create aus_date column for PR steps
                    
                    aus_date_binned = []
                        
                    for i in range(0, len(step_bin)):
                        aus_date_binned.append(aus_date)
                   
                    # Always export Summary, Chronological and step/binned data
                    
                    # Summary data
                    
                    results_summary = {'Variable': variable, 'Value': value}
                    export_file_summary = pd.DataFrame(results_summary, columns = ['Variable', 'Value'])
                    
                    # Chronological data
                    
                    # Assign left and right pokes as active or inactive based on the active port
                        
                    if active2[0] == 'Left':
                        active_poke = left_poke2
                        active_poke_chron_within = left_poke_chron_within
                        active_poke_chron_within_binary = left_poke_chron_within_binary
                        active_poke_time_chron = l_poke_time_chron
                        inactive_poke = right_poke2
                        inactive_poke_chron_within = right_poke_chron_within
                        inactive_poke_chron_within_binary = right_poke_chron_within_binary
                        inactive_poke_time_chron = r_poke_time_chron
                        active_poke_binned = left_binned
                        active_poke_binned_within = left_binned_within
                        inactive_poke_binned = right_binned
                        inactive_poke_binned_within = right_binned_within
                        binned_active_poke_time = binned_mean_l_poke_time
                        binned_inactive_poke_time = binned_mean_r_poke_time
                        binned_within_active_poke_time = binned_within_mean_l_poke_time
                        binned_within_inactive_poke_time = binned_within_mean_r_poke_time
                   
                            
                    elif active2[0] == 'Right':
                        active_poke = right_poke2
                        active_poke_chron_within = right_poke_chron_within
                        active_poke_chron_within_binary = right_poke_chron_within_binary
                        active_poke_time_chron = r_poke_time_chron
                        inactive_poke = left_poke2
                        inactive_poke_chron_within = left_poke_chron_within
                        inactive_poke_chron_within_binary = left_poke_chron_within_binary
                        inactive_poke_time_chron = l_poke_time_chron
                        active_poke_binned = right_binned
                        active_poke_binned_within = right_binned_within
                        inactive_poke_binned = left_binned
                        inactive_poke_binned_within = left_binned_within
                        binned_active_poke_time = binned_mean_r_poke_time
                        binned_inactive_poke_time = binned_mean_l_poke_time
                        binned_within_active_poke_time = binned_within_mean_r_poke_time
                        binned_within_inactive_poke_time = binned_within_mean_l_poke_time


                    
                    
                    results_chronological_c = {'Time (seconds)': seconds_elapsed, 'PR Step': session_type2, 'CE Block': block_count2, 'Active port': active2, 'Active Poke': active_poke, 'Inactive Poke': inactive_poke, 
                                              'Total Poke': total_pokes, '% Active Pokes': percent_active, 'Pellet Count': pellet_count2, 'Block pellet count': block_pellet_count_shifted,
                                              'Poke Time': poke_time2, 'Active Poke Time': active_poke_time_chron, 'Inactive Poke Time': inactive_poke_time_chron, 'Pellet Retrieval Time': retrieval_time2, 'Inter-Pellet Interval': ipi_2}
                    export_file_chronological_c = pd.DataFrame(results_chronological_c, columns = ['Time (seconds)', 'PR Step', 'CE Block', 'Active port', 'Active Poke', 'Inactive Poke',
                                                                                                'Total Poke', '% Active Pokes', 'Pellet Count', 'Block pellet count', '',
                                                                                                'Poke Time', 'Active Poke Time', 'Inactive Poke Time', 'Pellet Retrieval Time', 'Inter-Pellet Interval'])
                    
                    
                    results_chronological_w = {'Time (seconds)': seconds_elapsed, 'PR Step': session_type2, 'CE Block': block_count2,'Active port': active2, 'Active Poke': active_poke_chron_within, 'Inactive Poke': inactive_poke_chron_within, 
                                              'Total Poke': total_pokes_chron_within, '% Active Pokes': percent_active_chron_within, 'Pellet Count':pellet_count2, 'Block pellet count': block_pellet_count_shifted,
                                              'Poke Time': poke_time2, 'Active Poke Time': active_poke_time_chron, 'Inactive Poke Time': inactive_poke_time_chron, 'Pellet Retrieval Time': retrieval_time2, 'Inter-Pellet Interval': ipi_2}
                    export_file_chronological_w = pd.DataFrame(results_chronological_w, columns = ['Time (seconds)', 'PR Step', 'CE Block', 'Active port', 'Active Poke', 'Inactive Poke',
                                                                                                'Total Poke', '% Active Pokes',  'Pellet Count', 'Block pellet count', '',
                                                                                                'Poke Time', 'Active Poke Time', 'Inactive Poke Time', 'Pellet Retrieval Time', 'Inter-Pellet Interval'])
                    
                    
                    # Create chronological stepped data with breaks between steps
                    
                    seconds_elapsed_step = []
                    session_type2_step = []
                    active2_step = []
                    active_poke_step = []
                    inactive_poke_step = []
                    total_pokes_step = []
                    pellet_count2_step = []
                    percent_active_step = []
                    active_poke_chron_within_step = []
                    inactive_poke_chron_within_step = []
                    total_pokes_chron_within_step = []
                    percent_active_chron_within_step = []
                    active_poke_time_chron_step = []
                    inactive_poke_time_chron_step = []
                    retrieval_time2_step = []
                    ipi_2_step = []
                    poke_time2_step = []
                    block_count2_step = []
                    block_pellet_count_shifted_step = []
                    
                    for i in range (0, (len(session_type2)-1)):
                        
                        if session_type2[i + 1] == session_type2[i]:
                            seconds_elapsed_step.append(seconds_elapsed[i])
                            session_type2_step.append(session_type2[i])
                            active2_step.append(active2[i])
                            active_poke_step.append(active_poke[i])
                            inactive_poke_step.append(inactive_poke[i])
                            total_pokes_step.append(total_pokes[i])
                            pellet_count2_step.append(pellet_count2[i])
                            percent_active_step.append(percent_active[i])
                            active_poke_chron_within_step.append(active_poke_chron_within[i])
                            inactive_poke_chron_within_step.append(inactive_poke_chron_within[i])
                            total_pokes_chron_within_step.append(total_pokes_chron_within[i])
                            percent_active_chron_within_step.append(percent_active_chron_within[i])
                            active_poke_time_chron_step.append(active_poke_time_chron[i])
                            inactive_poke_time_chron_step.append(inactive_poke_time_chron[i])
                            retrieval_time2_step.append(retrieval_time2[i])
                            ipi_2_step.append(ipi_2[i])
                            poke_time2_step.append(poke_time2[i])
                            block_count2_step.append(block_count2[i])
                            block_pellet_count_shifted_step.append(block_pellet_count_shifted[i])

                        elif session_type2[i + 1] != session_type[i]:
                            seconds_elapsed_step.append(seconds_elapsed[i])
                            session_type2_step.append(session_type2[i])
                            active2_step.append(active2[i])
                            active_poke_step.append(active_poke[i])
                            inactive_poke_step.append(inactive_poke[i])
                            total_pokes_step.append(total_pokes[i])
                            pellet_count2_step.append(pellet_count2[i])
                            percent_active_step.append(percent_active[i])
                            active_poke_chron_within_step.append(active_poke_chron_within[i])
                            inactive_poke_chron_within_step.append(inactive_poke_chron_within[i])
                            total_pokes_chron_within_step.append(total_pokes_chron_within[i])
                            percent_active_chron_within_step.append(percent_active_chron_within[i])
                            active_poke_time_chron_step.append(active_poke_time_chron[i])
                            inactive_poke_time_chron_step.append(inactive_poke_time_chron[i])
                            retrieval_time2_step.append(retrieval_time2[i])
                            ipi_2_step.append(ipi_2[i])
                            poke_time2_step.append(poke_time2[i])
                            block_count2_step.append(block_count2[i])
                            block_pellet_count_shifted_step.append(block_pellet_count_shifted[i])



                            seconds_elapsed_step.append(np.nan)
                            session_type2_step.append(np.nan)
                            active2_step.append(np.nan)
                            active_poke_step.append(np.nan)
                            inactive_poke_step.append(np.nan)
                            total_pokes_step.append(np.nan)
                            pellet_count2_step.append(np.nan)
                            percent_active_step.append(np.nan)
                            active_poke_chron_within_step.append(np.nan)
                            inactive_poke_chron_within_step.append(np.nan)
                            total_pokes_chron_within_step.append(np.nan)
                            percent_active_chron_within_step.append(np.nan)
                            active_poke_time_chron_step.append(np.nan)
                            inactive_poke_time_chron_step.append(np.nan)
                            retrieval_time2_step.append(np.nan)
                            ipi_2_step.append(np.nan)
                            poke_time2_step.append(np.nan)
                            block_count2_step.append(np.nan)
                            block_pellet_count_shifted_step.append(np.nan)


                    
                    seconds_elapsed_step.append(seconds_elapsed[-1])
                    session_type2_step.append(session_type2[-1])
                    active2_step.append(active2[-1])
                    active_poke_step.append(active_poke[-1])
                    inactive_poke_step.append(inactive_poke[-1])
                    total_pokes_step.append(total_pokes[-1])
                    pellet_count2_step.append(pellet_count2[-1])
                    percent_active_step.append(percent_active[-1])
                    active_poke_chron_within_step.append(active_poke_chron_within[-1])
                    inactive_poke_chron_within_step.append(inactive_poke_chron_within[-1])
                    total_pokes_chron_within_step.append(total_pokes_chron_within[-1])
                    percent_active_chron_within_step.append(percent_active_chron_within[-1])
                    active_poke_time_chron_step.append(active_poke_time_chron[-1])
                    inactive_poke_time_chron_step.append(inactive_poke_time_chron[-1])
                    retrieval_time2_step.append(retrieval_time2[-1])
                    ipi_2_step.append(ipi_2[-1])
                    poke_time2_step.append(poke_time2[-1])
                    block_count2_step.append(block_count2[-1])
                    block_pellet_count_shifted_step.append(block_pellet_count_shifted[-1])

                    
                    print('check passed')
                    
                    results_chronological_c_step = {'Time (seconds)': seconds_elapsed_step, 'PR Step': session_type2_step, 'CE Block': block_count2_step, 'Active port': active2_step, 'Active Poke': active_poke_step, 'Inactive Poke': inactive_poke_step, 
                                              'Total Poke': total_pokes_step, '% Active Pokes': percent_active_step, 'Pellet Count': pellet_count2_step, 'Block pellet count': block_pellet_count_shifted_step,
                                              'Poke Time': poke_time2_step, 'Active Poke Time': active_poke_time_chron_step, 'Inactive Poke Time': inactive_poke_time_chron_step, 'Pellet Retrieval Time': retrieval_time2_step, 'Inter-Pellet Interval': ipi_2_step}
                    export_file_chronological_c_step = pd.DataFrame(results_chronological_c_step, columns = ['Time (seconds)', 'PR Step', 'CE Block', 'Active port', 'Active Poke', 'Inactive Poke',
                                                                                                'Total Poke', '% Active Pokes', 'Pellet Count', 'Block pellet count', '',
                                                                                                'Poke Time', 'Active Poke Time', 'Inactive Poke Time', 'Pellet Retrieval Time', 'Inter-Pellet Interval'])
                    
                    results_chronological_w_step = {'Time (seconds)': seconds_elapsed_step, 'PR Step': session_type2_step, 'CE Block': block_count2_step, 'Active port': active2_step, 'Active Poke': active_poke_chron_within_step, 'Inactive Poke': inactive_poke_chron_within_step, 
                                              'Total Poke': total_pokes_chron_within_step, '% Active Pokes': percent_active_chron_within_step, 'Pellet Count': pellet_count2_step, 'Block pellet count': block_pellet_count_shifted_step,
                                              'Poke Time': poke_time2_step, 'Active Poke Time': active_poke_time_chron_step, 'Inactive Poke Time': inactive_poke_time_chron_step, 'Pellet Retrieval Time': retrieval_time2_step, 'Inter-Pellet Interval': ipi_2_step}
                    export_file_chronological_w_step = pd.DataFrame(results_chronological_w_step, columns = ['Time (seconds)', 'PR Step', 'CE Block', 'Active port', 'Active Poke', 'Inactive Poke',
                                                                                                'Total Poke', '% Active Pokes', 'Pellet Count', 'Block pellet count', '',
                                                                                                'Poke Time', 'Active Poke Time', 'Inactive Poke Time', 'Pellet Retrieval Time', 'Inter-Pellet Interval'])
    
                    # Create CE/PR binned data with breaks in it indicating the CE blocks
                    
                    aus_date_binned_step = []
                    seconds_elapsed_binned_step = []
                    duration_binned_step = []
                    step_bin_step = []
                    active_binned_step = []
                    active_poke_binned_step = []
                    inactive_poke_binned_step = []
                    total_pokes_binned_step = []
                    pellet_binned_step = []
                    percent_active_binned_step = []
                    binned_mean_all_poke_time_step = []
                    binned_active_poke_time_step = []
                    binned_inactive_poke_time_step = []
                    retrieval_time_binned_step = []
                    ipi_binned_step = []
                    seconds_elapsed_binned_within_step = []
                    duration_binned_within_step = []
                    active_poke_binned_within_step = []
                    inactive_poke_binned_within_step = []
                    total_pokes_binned_within_step = []
                    pellet_binned_within_step = []
                    percent_active_binned_within_step = []
                    binned_within_mean_all_poke_time_step = []
                    binned_within_active_poke_time_step = []
                    binned_within_inactive_poke_time_step = []
                    block_count_binned_step = []
                    block_pellet_binned_step = []
                    
                    for i in range (0, (len(step_bin)-1)):
                        
                        if step_bin[i + 1] >= step_bin[i]:
                            aus_date_binned_step.append(aus_date_binned[i])
                            seconds_elapsed_binned_step.append(seconds_elapsed_binned[i])
                            duration_binned_step.append(duration_binned[i])
                            step_bin_step.append(step_bin[i])
                            active_binned_step.append(active_binned[i])
                            active_poke_binned_step.append(active_poke_binned[i])
                            inactive_poke_binned_step.append(inactive_poke_binned[i])
                            total_pokes_binned_step.append(total_pokes_binned[i])
                            pellet_binned_step.append(pellet_binned[i])
                            percent_active_binned_step.append(percent_active_binned[i])
                            binned_mean_all_poke_time_step.append(binned_mean_all_poke_time[i])
                            binned_active_poke_time_step.append(binned_active_poke_time[i])
                            binned_inactive_poke_time_step.append(binned_inactive_poke_time[i])
                            retrieval_time_binned_step.append(retrieval_time_binned[i])
                            ipi_binned_step.append(ipi_binned[i])
                            seconds_elapsed_binned_within_step.append(seconds_elapsed_binned_within[i])
                            duration_binned_within_step.append(duration_binned_within[i])
                            active_poke_binned_within_step.append(active_poke_binned_within[i])
                            inactive_poke_binned_within_step.append(inactive_poke_binned_within[i])
                            total_pokes_binned_within_step.append(total_pokes_binned_within[i])
                            pellet_binned_within_step.append(pellet_binned_within[i])
                            percent_active_binned_within_step.append(percent_active_binned_within[i])
                            binned_within_mean_all_poke_time_step.append(binned_within_mean_all_poke_time[i])
                            binned_within_active_poke_time_step.append(binned_within_active_poke_time[i])
                            binned_within_inactive_poke_time_step.append(binned_within_inactive_poke_time[i])
                            block_count_binned_step.append(block_count_binned[i])
                            block_pellet_binned_step.append(block_pellet_binned[i])

                        elif step_bin[i + 1] < step_bin[i]:
                            aus_date_binned_step.append(aus_date_binned[i])
                            seconds_elapsed_binned_step.append(seconds_elapsed_binned[i])
                            duration_binned_step.append(duration_binned[i])
                            step_bin_step.append(step_bin[i])
                            active_binned_step.append(active_binned[i])
                            active_poke_binned_step.append(active_poke_binned[i])
                            inactive_poke_binned_step.append(inactive_poke_binned[i])
                            total_pokes_binned_step.append(total_pokes_binned[i])
                            pellet_binned_step.append(pellet_binned[i])
                            percent_active_binned_step.append(percent_active_binned[i])
                            binned_mean_all_poke_time_step.append(binned_mean_all_poke_time[i])
                            binned_active_poke_time_step.append(binned_active_poke_time[i])
                            binned_inactive_poke_time_step.append(binned_inactive_poke_time[i])
                            retrieval_time_binned_step.append(retrieval_time_binned[i])
                            ipi_binned_step.append(ipi_binned[i])
                            seconds_elapsed_binned_within_step.append(seconds_elapsed_binned_within[i])
                            duration_binned_within_step.append(duration_binned_within[i])
                            active_poke_binned_within_step.append(active_poke_binned_within[i])
                            inactive_poke_binned_within_step.append(inactive_poke_binned_within[i])
                            total_pokes_binned_within_step.append(total_pokes_binned_within[i])
                            pellet_binned_within_step.append(pellet_binned_within[i])
                            percent_active_binned_within_step.append(percent_active_binned_within[i])
                            binned_within_mean_all_poke_time_step.append(binned_within_mean_all_poke_time[i])
                            binned_within_active_poke_time_step.append(binned_within_active_poke_time[i])
                            binned_within_inactive_poke_time_step.append(binned_within_inactive_poke_time[i])
                            block_count_binned_step.append(block_count_binned[i])
                            block_pellet_binned_step.append(block_pellet_binned[i])

                            
                            aus_date_binned_step.append(np.nan)
                            seconds_elapsed_binned_step.append(np.nan)
                            duration_binned_step.append(np.nan)
                            step_bin_step.append(np.nan)
                            active_binned_step.append(np.nan)
                            active_poke_binned_step.append(np.nan)
                            inactive_poke_binned_step.append(np.nan)
                            total_pokes_binned_step.append(np.nan)
                            pellet_binned_step.append(np.nan)
                            percent_active_binned_step.append(np.nan)
                            binned_mean_all_poke_time_step.append(np.nan)
                            binned_active_poke_time_step.append(np.nan)
                            binned_inactive_poke_time_step.append(np.nan)
                            retrieval_time_binned_step.append(np.nan)
                            ipi_binned_step.append(np.nan)
                            seconds_elapsed_binned_within_step.append(np.nan)
                            duration_binned_within_step.append(np.nan)
                            active_poke_binned_within_step.append(np.nan)
                            inactive_poke_binned_within_step.append(np.nan)
                            total_pokes_binned_within_step.append(np.nan)
                            pellet_binned_within_step.append(np.nan)
                            percent_active_binned_within_step.append(np.nan)
                            binned_within_mean_all_poke_time_step.append(np.nan)
                            binned_within_active_poke_time_step.append(np.nan)
                            binned_within_inactive_poke_time_step.append(np.nan)
                            block_count_binned_step.append(np.nan)
                            block_pellet_binned_step.append(np.nan)

                    
                    aus_date_binned_step.append(aus_date_binned[-1])
                    seconds_elapsed_binned_step.append(seconds_elapsed_binned[-1])
                    duration_binned_step.append(duration_binned[-1])
                    step_bin_step.append(step_bin[-1])
                    active_binned_step.append(active_binned[-1])
                    active_poke_binned_step.append(active_poke_binned[-1])
                    inactive_poke_binned_step.append(inactive_poke_binned[-1])
                    total_pokes_binned_step.append(total_pokes_binned[-1])
                    pellet_binned_step.append(pellet_binned[-1])
                    percent_active_binned_step.append(percent_active_binned[-1])
                    binned_mean_all_poke_time_step.append(binned_mean_all_poke_time[-1])
                    binned_active_poke_time_step.append(binned_active_poke_time[-1])
                    binned_inactive_poke_time_step.append(binned_inactive_poke_time[-1])
                    retrieval_time_binned_step.append(retrieval_time_binned[-1])
                    ipi_binned_step.append(ipi_binned[-1])
                    seconds_elapsed_binned_within_step.append(seconds_elapsed_binned_within[-1])
                    duration_binned_within_step.append(duration_binned_within[-1])
                    active_poke_binned_within_step.append(active_poke_binned_within[-1])
                    inactive_poke_binned_within_step.append(inactive_poke_binned_within[-1])
                    total_pokes_binned_within_step.append(total_pokes_binned_within[-1])
                    pellet_binned_within_step.append(pellet_binned_within[-1])
                    percent_active_binned_within_step.append(percent_active_binned_within[-1])
                    binned_within_mean_all_poke_time_step.append(binned_within_mean_all_poke_time[-1])
                    binned_within_active_poke_time_step.append(binned_within_active_poke_time[-1])
                    binned_within_inactive_poke_time_step.append(binned_within_inactive_poke_time[-1])
                    block_count_binned_step.append(block_count_binned[-1])
                    block_pellet_binned_step.append(block_pellet_binned[-1])
                            
                    
                    # Step/binned data (cumulative and within)
    
                    results_binned_c = {'Date': aus_date_binned_step, 'Seconds Elapsed': seconds_elapsed_binned_step, 'Duration': duration_binned_step, 'CE Block': block_count_binned_step, 'PR Step': step_bin_step, 'Active Port': active_binned_step, 
                                        'Active Poke': active_poke_binned_step, 'Inactive Poke': inactive_poke_binned_step, 'Total Poke': total_pokes_binned_step, '% Active Pokes': percent_active_binned_step,'Pellet Count': pellet_binned_step, 'Block pellet count': block_pellet_binned_step,
                                        'Mean Poke Time': binned_mean_all_poke_time_step, 'Mean Active Poke Time': binned_active_poke_time_step, 'Mean Inactive Poke Time': binned_inactive_poke_time_step, 'Pellet Retrieval Time': retrieval_time_binned_step, 'Inter-Pellet Interval': ipi_binned_step}
                    export_file_binned_c = pd.DataFrame(results_binned_c, columns = ['Date', 'Seconds Elapsed', 'Duration', 'CE Block', 'PR Step', 'Active Port', 
                                                                                      'Active Poke', 'Inactive Poke', 'Total Poke', '% Active Pokes', 'Pellet Count', 'Block pellet count', '', 
                                                                                      'Mean Poke Time', 'Mean Active Poke Time', 'Mean Inactive Poke Time', 'Pellet Retrieval Time', 'Inter-Pellet Interval'])
                    
                    results_binned_w = {'Date': aus_date_binned_step, 'Seconds Elapsed': seconds_elapsed_binned_within_step, 'Duration': duration_binned_within_step, 'CE Block': block_count_binned_step, 'PR Step': step_bin_step, 'Active Port': active_binned_step, 
                                        'Active Poke': active_poke_binned_within_step, 'Inactive Poke': inactive_poke_binned_within_step, 'Total Poke': total_pokes_binned_within_step, '% Active Pokes': percent_active_binned_within_step, 'Pellet Count': pellet_binned_within_step, 'Block pellet count': block_pellet_binned_step,
                                        'Mean Poke Time': binned_within_mean_all_poke_time_step, 'Mean Active Poke Time': binned_within_active_poke_time_step, 'Mean Inactive Poke Time': binned_within_inactive_poke_time_step, 'Pellet Retrieval Time': retrieval_time_binned_step, 'Inter-Pellet Interval': ipi_binned_step}
                    export_file_binned_w = pd.DataFrame(results_binned_w, columns = ['Date', 'Seconds Elapsed', 'Duration', 'CE Block', 'PR Step', 'Active Port', 
                                                                                      'Active Poke', 'Inactive Poke', 'Total Poke', '% Active Pokes', 'Pellet Count', 'Block pellet count', '', 
                                                                                      'Mean Poke Time', 'Mean Active Poke Time', 'Mean Inactive Poke Time', 'Pellet Retrieval Time', 'Inter-Pellet Interval'])
                    
                    # Export to excel
                    
                    if export_results == 'Y':
                    
                        from openpyxl import Workbook
                    
                        wb = Workbook()
                        
                        ws1 = wb.active
                        ws1.title = 'Summary'
                        
                        ws2 = wb.create_sheet()
                        ws2.title = 'Chronological Cumulative'
                        
                        ws3 = wb.create_sheet()
                        ws3.title = 'Chronological C Step'
                        
                        ws4 = wb.create_sheet()
                        ws4.title = 'Chronological Within'
                        
                        ws5 = wb.create_sheet()
                        ws5.title = 'Chronological W Step'
                        
                        ws6 = wb.create_sheet()
                        ws6.title = 'PR w CE Step Cumulative'
                        
                        ws7 = wb.create_sheet()
                        ws7.title = 'PR w CE Step Within'
                        
                        results_to_export = [export_file_summary, export_file_chronological_c, export_file_chronological_c_step, 
                                              export_file_chronological_w, export_file_chronological_w_step, 
                                              export_file_binned_c, export_file_binned_w]
                        
                        sheets_to_export = wb.sheetnames
                        
                        with pd.ExcelWriter(export_destination) as writer:
                            
                            for i in range(len(sheets_to_export)):
                                results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
    
        
    
        
                    