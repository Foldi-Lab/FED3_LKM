#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Aug  3 12:33:53 2021

@author: lauramilton
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Jul 29 13:48:25 2021

@author: lauramilton
"""

import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/FED Data Labelled files/VSG/Output/Joined/'
export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/FED Data Labelled files/VSG/Output/Joined/'
cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/FED Data Labelled files/Cohort/Group Joined/'
# FED_num = 'FED3'

sheet_to_join = '180min summary joined'

import pandas as pd
import numpy as np
import os
import openpyxl

from openpyxl import Workbook
        
wb = Workbook()

ws = wb.active
ws.title = 'Active Pokes'

ws2 = wb.create_sheet()
ws2.title = 'Inactive Pokes'

ws3 = wb.create_sheet()
ws3.title = 'Total Pokes'

ws4 = wb.create_sheet()
ws4.title = 'Pellets'

ws5 = wb.create_sheet()
ws5.title = 'Percent Active'


date_column = 'Date'
session_number = 'Session Number'
session_type = 'Session type'
active_poke_column = 'Active Poke'
inactive_poke_column = 'Inactive Poke'
total_poke_column = 'Total Poke'
pellet_column = 'Pellet Count'
percent_active_column = 'Percent Active'

value_summary = []
counter = 1

for filename in sorted(list(os.listdir(import_location))):
    
    if filename.endswith(".xlsx"):
        if filename.endswith('summary joined.xlsx'):
            if counter == 1:
                
                import_name = filename
                import_destination = import_location + import_name
                
                df_active = pd.read_excel(import_destination, sheet_name = sheet_to_join)
                df_inactive = pd.read_excel(import_destination, sheet_name = sheet_to_join)
                df_total = pd.read_excel(import_destination, sheet_name = sheet_to_join)
                df_pellet = pd.read_excel(import_destination, sheet_name = sheet_to_join)
                df_percent = pd.read_excel(import_destination, sheet_name = sheet_to_join)
                
                df_active.drop(columns=['Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'], inplace=True)
                df_inactive.drop(columns=['Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'], inplace=True)
                df_total.drop(columns=['Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'], inplace=True)
                df_pellet.drop(columns=['Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'], inplace=True)
                df_percent.drop(columns=['Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'], inplace=True)
                
            import_name = filename
            import_destination = import_location + import_name
            
            name = filename.strip(' all session summary joined.xlsx')
            
            df = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            
            active_poke = df[active_poke_column].tolist()
            inactive_poke = df[inactive_poke_column].tolist()
            total_poke = df[total_poke_column].tolist()
            pellet_count = df[pellet_column].tolist()
            percent_active = df[percent_active_column].tolist()
            
            if len(active_poke) < len(df_active):
                            
                bin_num = len(active_poke) + 1
                
                while len(active_poke) < len(df_active):
                    active_poke.append(np.nan)
                    inactive_poke.append(np.nan)
                    total_poke.append(np.nan)
                    pellet_count.append(np.nan)
                    percent_active.append(np.nan)
                    bin_num +=1
            
            df_active.insert((counter +2), name, active_poke)
            df_inactive.insert((counter +2), name, inactive_poke)
            df_total.insert((counter +2), name, total_poke)
            df_pellet.insert((counter +2), name, pellet_count)
            df_percent.insert((counter +2), name, percent_active)
            
            counter += 1

df_active = df_active.replace(np.nan, 'NA')
df_inactive = df_inactive.replace(np.nan, 'NA')
df_total = df_total.replace(np.nan, 'NA')
df_pellet = df_pellet.replace(np.nan, 'NA')
df_percent = df_percent.replace(np.nan, 'NA')
        
group_name = 'VSG summary Joined.xlsx'
group_destination = export_location + group_name

results_to_export = [df_active, df_inactive, df_total, df_pellet, df_percent]

sheets_to_export = wb.sheetnames

with pd.ExcelWriter(group_destination) as writer:
    
    for i in range(len(sheets_to_export)):
                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
                    
cohort_destination = cohort_export_location + group_name

with pd.ExcelWriter(cohort_destination) as writer:
    
    for i in range(len(sheets_to_export)):
                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
