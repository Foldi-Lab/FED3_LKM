#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Aug  3 12:33:45 2021

@author: lauramilton
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jul 28 15:27:07 2021

@author: lauramilton
"""

date_column = 'Date'
time_column1 = 'Time bin (600s each)'
time_column2 = 'Time bin (1800s each)'
session_column = 'Session Type'
active_column = 'Active Port'
active_poke_column = 'Active Poke'
inactive_poke_column = 'Inactive Poke'
total_poke_column = 'Total Poke'
pellet_column = 'Pellet Count'
percent_active_column = '% Active Pokes'

session_length = '180' # session length in minutes
time_bin_length1 = '10'
time_bin_length2 = '30'

import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/FED Data Labelled files/VSG/Output/'
export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/FED Data Labelled files/VSG/Output/'
cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Oldfield-Lab/VSG 21-22/FED Data and Photos/FED Data Labelled files/VSG/Output/Joined/'

import pandas as pd
import numpy as np
import os
import openpyxl

num_of_bins1 = int(session_length) / int(time_bin_length1)
num_of_bins2 = int(session_length) / int(time_bin_length2)

for folder in sorted(os.listdir(import_location)):
    if folder.startswith('Mouse'):

        date_chron1 = []
        time_chron1 = []
        session_chron1 = []
        active_chron1 = []
        active_poke_chron1 = []
        inactive_poke_chron1 = []
        total_poke_chron1 = []
        pellet_chron1 = []
        percent_active_chron1 = []
        
        date_chron2 = []
        time_chron2 = []
        session_chron2 = []
        active_chron2 = []
        active_poke_chron2 = []
        inactive_poke_chron2 = []
        total_poke_chron2 = []
        pellet_chron2 = []
        percent_active_chron2 = []
        
        date_chron3 = []
        time_chron3 = []
        session_chron3 = []
        active_chron3 = []
        active_poke_chron3 = []
        inactive_poke_chron3 = []
        total_poke_chron3 = []
        pellet_chron3 = []
        percent_active_chron3 = []
        
        date_chron4 = []
        time_chron4 = []
        session_chron4 = []
        active_chron4 = []
        active_poke_chron4 = []
        inactive_poke_chron4 = []
        total_poke_chron4 = []
        pellet_chron4 = []
        percent_active_chron4 = []

        session_num = 1
        
        for filename in sorted(os.listdir(os.path.join(import_location, folder))):
            
            if filename.endswith("Training.xlsx"):
                if filename.startswith('Mouse'): # can probs turn it around to be if filename.startswith() and use variable/name from the code that generates those files
                    if not filename.endswith('Overview.xlsx'): # will be able to remove later once those files aren't in the folder anymore (i.e. when file naming procedure is finalized)
                        if not filename.endswith('joined.xlsx'):    
                            print(filename)
                            import_name = filename
                            import_destination = import_location + folder + '/' + import_name
                            
                            import_sheet1 = str(time_bin_length1) + 'min Binned Within'
                            df1 = pd.read_excel(import_destination, sheet_name = import_sheet1)
                            
                            if time_bin_length2 != '':
                                import_sheet2 = str(time_bin_length2) + 'min Binned Within'
                                df2 = pd.read_excel(import_destination, sheet_name = import_sheet2)
                                import_sheet3 = str(time_bin_length2) + 'min Binned Cumulative'
                                df3 = pd.read_excel(import_destination, sheet_name = import_sheet3)
    
                            date1 = df1[date_column].tolist()
                            time1 = df1[time_column1].tolist()
                            session1 = df1[session_column].tolist()
                            active1 = df1[active_column].tolist()
                            active_poke1 = df1[active_poke_column].tolist()
                            inactive_poke1 = df1[inactive_poke_column].tolist()
                            total_poke1 = df1[total_poke_column].tolist()
                            pellet1 = df1[pellet_column].tolist()
                            percent_active1 = df1[percent_active_column].tolist()
                            
                            if time_bin_length2 != '':
                                date2 = df2[date_column].tolist()
                                time2 = df2[time_column2].tolist()
                                session2 = df2[session_column].tolist()
                                active2 = df2[active_column].tolist()
                                active_poke2 = df2[active_poke_column].tolist()
                                inactive_poke2 = df2[inactive_poke_column].tolist()
                                total_poke2 = df2[total_poke_column].tolist()
                                pellet2 = df2[pellet_column].tolist()
                                percent_active2 = df2[percent_active_column].tolist()
                                
                                date3 = df3[date_column].tolist()
                                time3 = df3[time_column2].tolist()
                                session3 = df3[session_column].tolist()
                                active3 = df3[active_column].tolist()
                                active_poke3 = df3[active_poke_column].tolist()
                                inactive_poke3 = df3[inactive_poke_column].tolist()
                                total_poke3 = df3[total_poke_column].tolist()
                                pellet3 = df3[pellet_column].tolist()
                                percent_active3 = df3[percent_active_column].tolist()
                                
                                date4 = df3[date_column].tolist()
                                time4 = df3[time_column2].tolist()
                                session4 = df3[session_column].tolist()
                                active4 = df3[active_column].tolist()
                                active_poke4 = df3[active_poke_column].tolist()
                                inactive_poke4 = df3[inactive_poke_column].tolist()
                                total_poke4 = df3[total_poke_column].tolist()
                                pellet4 = df3[pellet_column].tolist()
                                percent_active4 = df3[percent_active_column].tolist()
                            
                            if len(time1) > num_of_bins1:
                                
                                del date1[int(num_of_bins1):]
                                del time1[int(num_of_bins1):]
                                del session1[int(num_of_bins1):]
                                del active1[int(num_of_bins1):]
                                del active_poke1[int(num_of_bins1):]
                                del inactive_poke1[int(num_of_bins1):]
                                del total_poke1[int(num_of_bins1):]
                                del pellet1[int(num_of_bins1):]
                                del percent_active1[int(num_of_bins1):]
                                
                            if time_bin_length2 != '':
                                if len(time2) > num_of_bins2:
                                    
                                    del date2[int(num_of_bins2):]
                                    del time2[int(num_of_bins2):]
                                    del session2[int(num_of_bins2):]
                                    del active2[int(num_of_bins2):]
                                    del active_poke2[int(num_of_bins2):]
                                    del inactive_poke2[int(num_of_bins2):]
                                    del total_poke2[int(num_of_bins2):]
                                    del pellet2[int(num_of_bins2):]
                                    del percent_active2[int(num_of_bins2):]
                                    
                                if len(time3) > num_of_bins2:
                                    
                                    del date3[int(num_of_bins2):]
                                    del time3[int(num_of_bins2):]
                                    del session3[int(num_of_bins2):]
                                    del active3[int(num_of_bins2):]
                                    del active_poke3[int(num_of_bins2):]
                                    del inactive_poke3[int(num_of_bins2):]
                                    del total_poke3[int(num_of_bins2):]
                                    del pellet3[int(num_of_bins2):]
                                    del percent_active3[int(num_of_bins2):]
                                    
                                if len(time4) > num_of_bins2:
                                    
                                    del date4[int(num_of_bins2):]
                                    del time4[int(num_of_bins2):]
                                    del session4[int(num_of_bins2):]
                                    del active4[int(num_of_bins2):]
                                    del active_poke4[int(num_of_bins2):]
                                    del inactive_poke4[int(num_of_bins2):]
                                    del total_poke4[int(num_of_bins2):]
                                    del pellet4[int(num_of_bins2):]
                                    del percent_active4[int(num_of_bins2):]
    
                            
                            if len(time1) < num_of_bins1:
                                
                                bin_num1 = len(time1) + 1
                                
                                while len(time1) < num_of_bins1:
                                    date1.append(date1[0])
                                    time1.append(bin_num1)
                                    session1.append(np.nan)
                                    active1.append(np.nan)
                                    active_poke1.append(np.nan)
                                    inactive_poke1.append(np.nan)
                                    total_poke1.append(np.nan)
                                    pellet1.append(np.nan)
                                    percent_active1.append(np.nan)
                                    bin_num1 +=1
                            
                            if time_bin_length2 != '':
                                if len(time2) < num_of_bins2:
                                
                                    bin_num2 = len(time2) + 1
                                    
                                    while len(time2) < num_of_bins2:
                                        date2.append(date2[0])
                                        time2.append(bin_num2)
                                        session2.append(np.nan)
                                        active2.append(np.nan)
                                        active_poke2.append(np.nan)
                                        inactive_poke2.append(np.nan)
                                        total_poke2.append(np.nan)
                                        pellet2.append(np.nan)
                                        percent_active2.append(np.nan)
                                        bin_num2 +=1
                                        
                                if len(time4) < num_of_bins2:
                                
                                    bin_num4 = len(time4) + 1
                                    
                                    while len(time4) < num_of_bins2:
                                        date4.append(date4[0])
                                        time4.append(bin_num4)
                                        session4.append(np.nan)
                                        active4.append(np.nan)
                                        active_poke4.append(np.nan)
                                        inactive_poke4.append(np.nan)
                                        total_poke4.append(np.nan)
                                        pellet4.append(np.nan)
                                        percent_active4.append(np.nan)
                                        bin_num4 +=1
                            
                            for a, b, c, d, e, f, g, h, i in zip(time1, session1, active1, active_poke1, inactive_poke1, total_poke1, pellet1, percent_active1, date1):
                                    
                                    time_chron1.append(str(session_num) + '.' + str(a))
                                    session_chron1.append(b)
                                    active_chron1.append(c) 
                                    active_poke_chron1.append(d)
                                    inactive_poke_chron1.append(e)
                                    total_poke_chron1.append(f)
                                    pellet_chron1.append(g)
                                    percent_active_chron1.append(h)
                                    date_chron1.append(i)
                                    
                            if time_bin_length2 != '':
                                for a, b, c, d, e, f, g, h, i in zip(time2, session2, active2, active_poke2, inactive_poke2, total_poke2, pellet2, percent_active2, date2):
                                    
                                        time_chron2.append(str(session_num) + '.' + str(a))
                                        session_chron2.append(b)
                                        active_chron2.append(c) 
                                        active_poke_chron2.append(d)
                                        inactive_poke_chron2.append(e)
                                        total_poke_chron2.append(f)
                                        pellet_chron2.append(g)
                                        percent_active_chron2.append(h)
                                        date_chron2.append(i)
                                        
                                for a, b, c, d, e, f, g, h, i in zip(time4, session4, active4, active_poke4, inactive_poke4, total_poke4, pellet4, percent_active4, date4):
                                    
                                        time_chron4.append(str(session_num) + '.' + str(a))
                                        session_chron4.append(b)
                                        active_chron4.append(c) 
                                        active_poke_chron4.append(d)
                                        inactive_poke_chron4.append(e)
                                        total_poke_chron4.append(f)
                                        pellet_chron4.append(g)
                                        percent_active_chron4.append(h)
                                        date_chron4.append(i)
    
                                        
                                  
                                time_chron3.append(str(session_num) + '.' + str(num_of_bins2))
                                session_chron3.append(session3[-1])
                                active_chron3.append(active3[-1]) 
                                active_poke_chron3.append(active_poke3[-1])
                                inactive_poke_chron3.append(inactive_poke3[-1])
                                total_poke_chron3.append(total_poke3[-1])
                                pellet_chron3.append(pellet3[-1])
                                percent_active_chron3.append(percent_active3[-1])
                                date_chron3.append(date3[-1])
                            
                            session_num += 1
                            
                            session_chron1 = ['NA' if x != x else x for x in session_chron1]
                            active_chron1 = ['NA' if x != x else x for x in active_chron1] 
                            active_poke_chron1 = ['NA' if x != x else x for x in active_poke_chron1]
                            inactive_poke_chron1 = ['NA' if x != x else x for x in inactive_poke_chron1]
                            total_poke_chron1 = ['NA' if x != x else x for x in total_poke_chron1]
                            pellet_chron1 = ['NA' if x != x else x for x in pellet_chron1]
                            percent_active_chron1 = ['NA' if x != x else x for x in percent_active_chron1]
                            
                            if time_bin_length2 != '':
                                session_chron2 = ['NA' if x != x else x for x in session_chron2]
                                active_chron2 = ['NA' if x != x else x for x in active_chron2] 
                                active_poke_chron2 = ['NA' if x != x else x for x in active_poke_chron2]
                                inactive_poke_chron2 = ['NA' if x != x else x for x in inactive_poke_chron2]
                                total_poke_chron2 = ['NA' if x != x else x for x in total_poke_chron2]
                                pellet_chron2 = ['NA' if x != x else x for x in pellet_chron2]
                                percent_active_chron2 = ['NA' if x != x else x for x in percent_active_chron2]
                                
                                session_chron4 = ['NA' if x != x else x for x in session_chron4]
                                active_chron4 = ['NA' if x != x else x for x in active_chron4] 
                                active_poke_chron4 = ['NA' if x != x else x for x in active_poke_chron4]
                                inactive_poke_chron4 = ['NA' if x != x else x for x in inactive_poke_chron4]
                                total_poke_chron4 = ['NA' if x != x else x for x in total_poke_chron4]
                                pellet_chron4 = ['NA' if x != x else x for x in pellet_chron4]
                                percent_active_chron4 = ['NA' if x != x else x for x in percent_active_chron4]
        
                            ##### Export results
    
    
                            results1 = {'Date': date_chron1, 'Time Bin': time_chron1, 'Session Type': session_chron1, 'Active Port': active_chron1, 'Active Poke': active_poke_chron1, 'Inactive Poke': inactive_poke_chron1, 'Total Poke': total_poke_chron1, 'Pellet Count': pellet_chron1, 'Percent Active': percent_active_chron1}
                            export_file1 = pd.DataFrame(results1, columns = ['Date', 'Time Bin', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
                            
                            if time_bin_length2 != '':
                                results2 = {'Date': date_chron2, 'Time Bin': time_chron2, 'Session Type': session_chron2, 'Active Port': active_chron2, 'Active Poke': active_poke_chron2, 'Inactive Poke': inactive_poke_chron2, 'Total Poke': total_poke_chron2, 'Pellet Count': pellet_chron2, 'Percent Active': percent_active_chron2}
                                export_file2 = pd.DataFrame(results2, columns = ['Date', 'Time Bin', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
                                
                                results_summary = {'Date': date_chron3, 'Time Bin': time_chron3, 'Session Type': session_chron3, 'Active Port': active_chron3, 'Active Poke': active_poke_chron3, 'Inactive Poke': inactive_poke_chron3, 'Total Poke': total_poke_chron3, 'Pellet Count': pellet_chron3, 'Percent Active': percent_active_chron3}
                                export_file_summary = pd.DataFrame(results_summary, columns = ['Date', 'Time Bin', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
    
                                results4 = {'Date': date_chron4, 'Time Bin': time_chron4, 'Session Type': session_chron4, 'Active Port': active_chron4, 'Active Poke': active_poke_chron4, 'Inactive Poke': inactive_poke_chron4, 'Total Poke': total_poke_chron4, 'Pellet Count': pellet_chron4, 'Percent Active': percent_active_chron4}
                                export_file4 = pd.DataFrame(results4, columns = ['Date', 'Time Bin', 'Session Type', 'Active Port', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active'])
    
                            
                            from openpyxl import Workbook
                        
                            wb = Workbook()
                            
                            ws1 = wb.active
                            ws1.title = str(time_bin_length1) + 'min binned chron joined'
                            
                            if time_bin_length2 != '':
                                ws2 = wb.create_sheet()
                                ws2.title = str(time_bin_length2) + 'min binned chron joined'
                                
                                ws3 = wb.create_sheet()
                                ws3.title = str(session_length) + 'min summary joined'
                                
                                ws4 = wb.create_sheet()
                                ws4.title = str(time_bin_length2) + 'min binned cumu joined'
                            
                            results_to_export = [export_file1]
                            
                            if time_bin_length2 != '':
                                results_to_export.append(export_file2)
                                results_to_export.append(export_file_summary)
                                results_to_export.append(export_file4)
                            
                            sheets_to_export = wb.sheetnames
                            
                            chron_name = folder + ' FR5 joined.xlsx'
                            chron_destination = export_location + folder + '/' + chron_name
        
                            with pd.ExcelWriter(chron_destination) as writer:
                                
                                for i in range(len(sheets_to_export)):
                                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
        
                            cohort_destination = cohort_export_location + chron_name
                            
                            with pd.ExcelWriter(cohort_destination) as writer:
                                
                                for i in range(len(sheets_to_export)):
                                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)    
