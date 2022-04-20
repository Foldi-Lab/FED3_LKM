#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Feb 23 17:51:03 2022

@author: lauramilton
"""



import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/FR1Both/Cohort/SAL/'
export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/FR1Both/Cohort/SAL/'
cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/FR1Both/Cohort/'

sheet_to_join = 'Summary Joined'

group_name = 'SAL Summary Joined.xlsx'

import pandas as pd
import numpy as np
import os
import openpyxl

from openpyxl import Workbook
        
wb = Workbook()

ws = wb.active
ws.title = 'Total Pokes'

ws2 = wb.create_sheet()
ws2.title = 'Left Pokes'

ws3 = wb.create_sheet()
ws3.title = 'Left Percent'

ws4 = wb.create_sheet()
ws4.title = 'Right Pokes'

ws5 = wb.create_sheet()
ws5.title = 'Right Percent'

ws6 = wb.create_sheet()
ws6.title = 'Pellets'


date_column = 'Date'
total_column = 'Total Pokes'
left_column = 'Left Pokes'
left_percent_column = 'Left Percent'
right_column = 'Right Pokes'
right_percent_column = 'Right Percent'
pellets_column = 'Pellets'


counter = 1

for filename in sorted(list(os.listdir(import_location))):
    
    if filename.endswith("Summary Joined.xlsx"):
        print(filename)
        if counter == 1:
            
            import_name = filename
            import_destination = import_location + import_name
            
            df_total = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_pellets = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_left = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_left_percent = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_right = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            df_right_percent = pd.read_excel(import_destination, sheet_name = sheet_to_join)
            
            cols_to_drop = ['Filename', 'Total Pokes', 'Left Pokes', 'Left Percent', 'Right Pokes', 'Right Percent', 'Pellets']
            df_total.drop(columns=cols_to_drop, inplace=True)
            df_pellets.drop(columns=cols_to_drop, inplace=True)
            df_left.drop(columns=cols_to_drop, inplace=True)
            df_left_percent.drop(columns=cols_to_drop, inplace=True)
            df_right.drop(columns=cols_to_drop, inplace=True)
            df_right_percent.drop(columns=cols_to_drop, inplace=True)
            
        import_name = filename
        import_destination = import_location + import_name
        
        name = filename.strip(' Summary Joined.xlsx')
        
        df = pd.read_excel(import_destination, sheet_name = sheet_to_join)
        
        date = df[date_column].tolist()
        total = df[total_column].tolist()
        pellets = df[pellets_column].tolist()
        left = df[left_column].tolist()
        left_percent = df[left_percent_column].tolist()
        right = df[right_column].tolist()
        right_percent = df[right_percent_column].tolist()

        df_total.insert((counter), name, total)
        df_pellets.insert((counter), name, pellets)
        df_left.insert((counter), name, left)
        df_left_percent.insert((counter), name, left_percent)
        df_right.insert((counter), name, right)
        df_right_percent.insert((counter), name, right_percent)
            
        counter += 1

# df_active = df_active.replace(np.nan, 'NA')
# df_inactive = df_inactive.replace(np.nan, 'NA')
# df_total = df_total.replace(np.nan, 'NA')
# df_pellet = df_pellet.replace(np.nan, 'NA')
# df_percent = df_percent.replace(np.nan, 'NA')
        
group_destination = export_location + group_name

results_to_export = [df_total, df_left, df_left_percent, df_right, df_right_percent, df_pellets]

sheets_to_export = wb.sheetnames

with pd.ExcelWriter(group_destination) as writer:
    
    for i in range(len(sheets_to_export)):
                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
                    
cohort_destination = cohort_export_location + group_name

with pd.ExcelWriter(cohort_destination) as writer:
    
    for i in range(len(sheets_to_export)):
                    results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)

