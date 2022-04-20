#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Aug  3 12:33:01 2021

@author: lauramilton
"""

#-----------------------------------------------------------------------------

# GUI TUTORIAL

# Type 'pip install PySimpleGUI' into the console.
# If you have done this once, you do not need to do it again.
# Copy, paste and edit each block of code to add a new line to the GUI.

# For each block you should put in the type of input in inp[key_word] = ...
# It should light up in a different colour.
# If it is text, type str
# If it is a number, type float
# If it is a True or False, type bool
# Make sure the key_word variables are unique for each block as well.

#-----------------------------------------------------------------------------

# Import the PySimpleGUI package and choose a theme.
# To view more themes, type sg.theme_previewer() into the console.
# Change the height and width of the window in size_GUI, so it will include the submit button.

import PySimpleGUI as sg
sg.theme("DarkPurple5")
size_GUI = (650,500)
name_window = "GUI"
layout = []
inp = {}

# sg.theme_previewer() # uncomment to preview themes and select a new one to insert above if desired

# (Block 1) Create a buton to choose an import folder.
# Note that there should be no slash at the end of the default import path.

line          = []
display_text  = "Choose the import location PARENT folder"
default_value = "/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 Reversal/FED_Test/Datafiles/FiberGirl"
key_word      = "Import"
inp[key_word] = str

line += [sg.Text(display_text)]
line += [sg.Input(default_text=default_value,key=key_word,enable_events=True)]
line += [sg.FolderBrowse()]
layout += [[sg.T("")]]
layout += [line]

# (Block 2) Create a button to choose an export folder.
# Note that there should be no slash at the end of the default export path.

line          = []
display_text  = "Choose the export location PARENT folder"
default_value = "/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 Reversal/FED_Test/DatafilesPython/FiberGirl"
key_word      = "Export"
inp[key_word] = str

line += [sg.Text(display_text)]
line += [sg.Input(default_text=default_value,key=key_word,enable_events=True)]
line += [sg.FolderBrowse()]
layout += [[sg.T("")]]
layout += [line]

# (Block 3) Create a button to choose a cohort overview export folder.
# Note that there should be no slash at the end of the default export path.

line          = []
display_text  = "Choose the cohort overview export location folder"
default_value = "/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 Reversal/FED_Test/DatafilesPython/Overview"
key_word      = "Cohort_export"
inp[key_word] = str

line += [sg.Text(display_text)]
line += [sg.Input(default_text=default_value,key=key_word,enable_events=True)]
line += [sg.FolderBrowse()]
layout += [[sg.T("")]]
layout += [line]

# (Block 3.5) Create a text entry.

# line          = []
# display_text  = "What is the FED number?"
# default_value = "#"
# key_word      = "FED_num"
# inp[key_word] = str

# line += [sg.Text(display_text)]
# line += [sg.Input(default_text=default_value,key=key_word,enable_events=True)]
# layout += [[sg.T("")]]
# layout += [line]

# (Block 4) Create a dropdown menu.

line          = []
display_text  = "Which schedule was used?"
dropdown_list = ['FR1/3/5 or FR1/3/5_reversed', 'FR1_both', 'New_Reversal']
default_value = ''
key_word      = "Schedule"
inp[key_word] = str

line += [sg.Text(display_text)]
line += [sg.Combo(dropdown_list,default_value=default_value,key=key_word,enable_events=True)]
layout += [[sg.T("")]]
layout += [line]

# (Block 5) Create a number entry.

line          = []
display_text  = "What was the session duration (in minutes)"
default_value = 0
key_word      = "Session_length"
inp[key_word] = float

line += [sg.Text(display_text)]
line += [sg.Input(default_text=default_value,key=key_word,enable_events=True)]
layout += [[sg.T("")]]
layout += [line]

# (Block 7) Create a checkbox.

line          = []
display_text  = "Do you want to time bin the data?"
default_value = False
key_word      = "Time_bins"
inp[key_word] = bool

line += [sg.Text(display_text)]
line += [sg.Checkbox('',default=default_value,key=key_word,enable_events=True)]
layout += [[sg.T("")]]
layout += [line]

# (Block 5) Create a number entry.

line          = []
display_text  = "If yes, what length time bins (in minutes); If no leave blank"
default_value = ''
key_word      = "Bin_length1"
inp[key_word] = str

line += [sg.Text(display_text)]
line += [sg.Input(default_text=default_value,key=key_word,enable_events=True)]
layout += [[sg.T("")]]
layout += [line]

# (Block 5) Create a number entry.

line          = []
display_text  = "OPTIONAL second bin duration, if not desired leave blank"
default_value = ''
key_word      = "Bin_length2"
inp[key_word] = str

line += [sg.Text(display_text)]
line += [sg.Input(default_text=default_value,key=key_word,enable_events=True)]
layout += [[sg.T("")]]
layout += [line]


# (Block 5) Create a number entry.

line          = []
display_text  = "Length of reversal blocks (if schedule is New_Reversal)"
default_value = 0
key_word      = "Block_length"
inp[key_word] = float

line += [sg.Text(display_text)]
line += [sg.Input(default_text=default_value,key=key_word,enable_events=True)]
layout += [[sg.T("")]]
layout += [line]

# (Block 6) Create a checkbox.

line          = []
display_text  = "Were initiation poke/s to deliver a pellet used?"
default_value = False
key_word      = "Initiation"
inp[key_word] = bool

line += [sg.Text(display_text)]
line += [sg.Checkbox('',default=default_value,key=key_word,enable_events=True)]
layout += [[sg.T("")]]
layout += [line]


# Create the GUI.

layout += [[sg.T("")],[sg.Button("Submit")]]
window = sg.Window(name_window, layout, size=size_GUI)
    
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event=="Exit":
        window.close()
        exit()
    elif event == "Submit":
        window.close()
        break
        
inputs = {}
for key in list(inp.keys()):
    inputs[key] = inp[key](values[key])
                           
# Assign each of the inputs to variable names.
# Replace the key_word variables in the square brackets.
# Note that the slash is added back to the end of the import and export locations.

import_location = inputs["Import"] + '/'
export_location = inputs["Export"] + '/'
cohort_export_location = inputs["Cohort_export"] + '/'
# FED_number = inputs["FED_num"]
schedule = inputs["Schedule"]
session_length = inputs['Session_length']
reversal_block_length = inputs["Block_length"]
initiation_poke = inputs["Initiation"]
binned = inputs["Time_bins"]
seconds_in_bins1 = (inputs["Bin_length1"] * 60) # bin length given in minutes so converted to seconds for the code
seconds_in_bins2 = (inputs["Bin_length2"] * 60) # bin length given in minutes so converted to seconds for the code
# FED_num = 'FED' + FED_number

# Verify the inputs by printing them in the console.

print('Import location is', import_location)
print('Export location is', export_location)
print('Cohort export location is', cohort_export_location)
print('Schedule is', schedule)
print('Session length is', session_length + ' minutes')
print('Reversal block length is', reversal_block_length)
print('Initiation pokes is', initiation_poke)
print('Time bins are being used?', binned)
print('Length of time bins is', seconds_in_bins1 + ' seconds', seconds_in_bins2 + ' seconds')

# Paste the rest of the code here

##########----------########## FR1/3/5 and reversed

if schedule == 'FR1/3/5 or FR1/3/5_reversed':

    time_column = 'MM:DD:YYYY hh:mm:ss'
    event_column = 'Event'
    active_poke_column = 'Active_Poke'
    session_type_column = 'Session_Type'
    left_poke_column = 'Left_Poke_Count'
    right_poke_column = 'Right_Poke_Count'
    pellet_count_column = 'Pellet_Count'
    
    #-----------------------------------------------------------------------------
    
    # Import the revelant data: time, FR ratio, event, active port, left poke, right poke, and pellet count.
    
    import pandas as pd
    import numpy as np
    import os
    import openpyxl
    
    #-----
    
    for folder in os.listdir(import_location):
        if folder.startswith('FED'):
    
            for filename in os.listdir(os.path.join(import_location, folder)):
                
                if filename.endswith(".CSV"):
                            
            # Import the csv data
            
                    import_name = filename
                    import_destination = import_location + folder + '/' + import_name
                    export_name = 'FR Training ' + import_name.strip('.CSV') + '.xlsx'
                    export_destination = export_location + folder + '/' + export_name
            
            #-----
            
                    df = pd.read_csv(import_destination)
                    
                    time = df[time_column].tolist()
                    session_type = df[session_type_column].tolist()
                    event = df[event_column].tolist()
                    active = df[active_poke_column].tolist()
                    left_poke = df[left_poke_column].tolist()
                    right_poke = df[right_poke_column].tolist()
                    pellet_count = df[pellet_count_column].tolist()
                    
                    # Change the time column from strings to datetime format for calculating durations between timestamps and creating time bins
                    
                    import datetime as dt
                    
                    time_list = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time]
                    
                    # Start time of the session is the first timestamp
                    
                    start_time = time_list[0]
                    
                    # Remove the initiation poke/s and pellet data (if required)
                    
                    if initiation_poke == True:
                    
                        if session_type[0] == 'FR1' or session_type[0] == 'FR1_reversed':
                        
                            left_count = left_poke[0]
                            right_count = right_poke[0]
                            
                            del time[:2]
                            del session_type[:2]
                            del event[:2]
                            del left_poke[:2]
                            del right_poke[:2]
                            del pellet_count[:2]
                            
                        elif session_type[0] == 'FR3' or session_type[0] == 'FR3_reversed':
                            
                            left_count = left_poke[3]
                            right_count = right_poke[3]
                            
                            del time[:4]
                            del session_type[:4]
                            del event[:4]
                            del left_poke[:4]
                            del right_poke[:4]
                            del pellet_count[:4]
                            
                        elif session_type[0] == 'FR5' or session_type[0] == 'FR5_reversed':
                            
                            left_count = left_poke[5]
                            right_count = right_poke[5]
                            
                            del time[:6]
                            del session_type[:6]
                            del event[:6]
                            del left_poke[:6]
                            del right_poke[:6]
                            del pellet_count[:6]
                                    
                        # Subtract the poke and pellet from the subsequent cumulative data
                        
                        left_poke_shifted = []
                        right_poke_shifted = []
                        pellet_count_shifted = []
                        
                        for i in range(0, len(left_poke)):
                            left_poke_shifted.append(left_poke[i] - left_count)
                            right_poke_shifted.append(right_poke[i] - right_count)
                            pellet_count_shifted.append(pellet_count[i] - 1)
                                    
                        # Shift the pellet_count lists one step backwards so that the count is in the same index as the corresponding poke
                
                        pellet_count_shifted.pop(0)
                        
                        # Create new lists that only include data from the lines where event == 'Poke'
                        
                        time2 = []
                        session_type2 = []
                        event2 = []
                        active2 = []
                        left_poke2 = []
                        right_poke2 = []
                        pellet_count2 = []
                        
                        for a, b, c, d, e, f, g in zip(event, left_poke_shifted, right_poke_shifted, pellet_count_shifted, time, session_type, active):
                            if a == 'Poke':
                                event2.append(a)
                                left_poke2.append(b)
                                right_poke2.append(c) 
                                pellet_count2.append(d)
                                time2.append(e)
                                session_type2.append(f)
                                active2.append(g)
                    
                    else:
                        # Shift the pellet_count lists one step backwards so that the count is in the same index as the corresponding poke
                
                        pellet_count.pop(0)
                        
                        # Create new lists that only include data from the lines where event == 'Poke'
                        
                        time2 = []
                        session_type2 = []
                        event2 = []
                        active2 = []
                        left_poke2 = []
                        right_poke2 = []
                        pellet_count2 = []
                        
                        for a, b, c, d, e, f, g in zip(event, left_poke, right_poke, pellet_count, time, session_type, active):
                            if a == 'Poke':
                                event2.append(a)
                                left_poke2.append(b)
                                right_poke2.append(c) 
                                pellet_count2.append(d)
                                time2.append(e)
                                session_type2.append(f)
                                active2.append(g)
                    
                    # Change the time2 column from strings to datetime format for calculating durations between timestamps and creating time bins
                    
                    time2 = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time2]
                    
                    # Create a new column that is the total number of pokes
                    
                    total_pokes = []
                    
                    for i in range(0, len(active2)):
                        total_pokes.append(left_poke2[i] + right_poke2[i])
                    
                    # Create new column that is % of pokes that are into the active port. Needs to be based on active_poke value (i.e. Left or Right)
                    
                    percent_active = []
                    
                    for i in range(0, len(active2)):
                        if active2[i] == 'Left':
                            percent_active.append(left_poke2[i] / total_pokes[i] * 100)
                        else:
                            percent_active.append(right_poke2[i] / total_pokes[i] * 100)
                            
                    # Calculate how much time (in seconds) has elapsed from start time for each event (nose poke)
                    
                    seconds_elapsed = []
                    
                    for i in range(0, len(time2)):
                        time_from_start = time2[i] - start_time
                        result = int(time_from_start.total_seconds())
                        seconds_elapsed.append(result)
                    
                    # Create time bins of desired length (input for desired length is at the top of the code)
                    
                    import math
                    
                    if seconds_in_bins1 != '':
                    
                        end_time = seconds_elapsed[-1]
                        
                        # Create end value that is the next greatest multiple of the required time bin duration greater than the end_time so that those data points are retained in the output
                        
                        excess = math.ceil(end_time / int(seconds_in_bins1)) # Rounds up so that the number of bins is the multiple of bin length greater than the time point
                        
                        end_bin = excess * int(seconds_in_bins1) # Creates the end value for the last time bin
                        
                        interval_range = pd.interval_range(start=0, freq=int(seconds_in_bins1), end=end_bin, closed="left") #'left' means that it includes the first data point which is 0 because it's start time
                        
                        time_bins = df[seconds_elapsed] = pd.cut(seconds_elapsed, interval_range, include_lowest=True, ordered=True)
                        
                        # Create dictionary of time bins as keys with bin number as matching entry to label the bins
                        
                        # Create a list of integers of the same length as there are time bins
                        
                        bin_number = []
                        a = 1
                        for i in range(0, len(interval_range)):
                            bin_number.append(a)
                            a += 1
                        
                        make_dictionary = zip(interval_range, bin_number)
                        
                        bin_dictionary = dict(make_dictionary)
                        
                        bin_num = []
                        for i in range(0, len(time_bins)):
                            bin_num.append(bin_dictionary.get((time_bins[i])))
                        
                        # Get the last entry into each bin for the rest of the columns required
                        
                        bin_num_binned1 = []
                        left_binned1 = []
                        right_binned1 = []
                        pellet_binned1 = []
                        active_binned1 = []
                        session_type_binned1 = []
                        percent_active_binned1 = []
                        total_pokes_binned1 = []
                        
                        # time binned
                        for i in range(1, len(time_bins)):
                                if bin_num[i - 1] != bin_num[i]:
                                    bin_num_binned1.append(bin_num[i - 1])
                        bin_num_binned1.append(bin_num[i - 0])
                        
                        # left binned
                        for i in range(1, len(left_poke2)):
                                if time_bins[i - 1] != time_bins[i]:
                                    left_binned1.append(left_poke2[i - 1])
                        left_binned1.append(left_poke2[i - 0])
                                
                        # right binned
                        for i in range(1, len(right_poke2)):
                                if time_bins[i - 1] != time_bins[i]:
                                    right_binned1.append(right_poke2[i - 1])
                        right_binned1.append(right_poke2[i - 0])
                             
                        # pellet binned
                        for i in range(1, len(pellet_count2)):
                                if time_bins[i - 1] != time_bins[i]:
                                    pellet_binned1.append(pellet_count2[i - 1])
                        pellet_binned1.append(pellet_count2[i - 0])
                        
                        # active binned
                        for i in range(1, len(active2)):
                                if time_bins[i - 1] != time_bins[i]:
                                    active_binned1.append(active2[i - 1])
                        active_binned1.append(active2[i - 0])
                        
                        # session type binned
                        for i in range(1, len(session_type2)):
                                if time_bins[i - 1] != time_bins[i]:
                                    session_type_binned1.append(session_type2[i - 1])
                        session_type_binned1.append(session_type2[i - 0])
                        
                        # total pokes binned
                        for i in range(1, len(total_pokes)):
                                if time_bins[i - 1] != time_bins[i]:
                                    total_pokes_binned1.append(total_pokes[i - 1])
                        total_pokes_binned1.append(total_pokes[i - 0])
                        
                        # percent active binned
                        for i in range(0, len(active_binned1)):
                            if active_binned1[i] == 'Left':
                                percent_active_binned1.append(left_binned1[i] / total_pokes_binned1[i] * 100)
                            else:
                                percent_active_binned1.append(right_binned1[i] / total_pokes_binned1[i] * 100)
                        
                        # Enter new row for any bins that have no data in them (for cumulative bins this is a duplication of the preceding row)
                    
                        active_port = active_binned1[0]
                        session_type3 = session_type2[0]
                              
                        for i in bin_dictionary:
                            if bin_dictionary[i] in bin_num_binned1:
                                continue
                            else:
                                bin_num_binned1.insert((bin_dictionary[i] - 1), bin_dictionary[i])
                                left_binned1.insert((bin_dictionary[i] - 1), left_binned1[(bin_dictionary[i] - 2)])
                                right_binned1.insert((bin_dictionary[i] - 1), right_binned1[(bin_dictionary[i] - 2)])
                                pellet_binned1.insert((bin_dictionary[i] - 1), pellet_binned1[(bin_dictionary[i] - 2)])
                                active_binned1.insert((bin_dictionary[i] - 1), active_port)
                                percent_active_binned1.insert((bin_dictionary[i] - 1), percent_active_binned1[(bin_dictionary[i] - 2)])
                                session_type_binned1.insert((bin_dictionary[i] - 1), session_type3)
                                total_pokes_binned1.insert((bin_dictionary[i] - 1), total_pokes_binned1[(bin_dictionary[i] - 2)])
                                
                        # Make new counts that are within the bins rather than cumulative
                        
                        left_binned_within1 = []
                        right_binned_within1 = []
                        pellet_binned_within1 = []
                        percent_active_binned_within1 = []
                        total_pokes_binned_within1 = []
                        
                        left_binned_within1.append(left_binned1[0])
                        for i in range(1, len(left_binned1)):
                            left_binned_within1.append(left_binned1[i] - left_binned1[i - 1])
                        
                        right_binned_within1.append(right_binned1[0])
                        for i in range(1, len(right_binned1)):
                            right_binned_within1.append(right_binned1[i] - right_binned1[i - 1])
                        
                        pellet_binned_within1.append(pellet_binned1[0])
                        for i in range(1, len(pellet_binned1)):
                            pellet_binned_within1.append(pellet_binned1[i] - pellet_binned1[i - 1])
                            
                        total_pokes_binned_within1.append(total_pokes_binned1[0])
                        for i in range (1, len(total_pokes_binned1)):
                            total_pokes_binned_within1.append(total_pokes_binned1[i] - total_pokes_binned1[i - 1])
                            
                        for i in range(0, len(active_binned1)):
                            if active_binned1[i] == 'Left':
                                if left_binned_within1[i] + right_binned_within1[i] != 0:
                                    percent_active_binned_within1.append(left_binned_within1[i] / total_pokes_binned_within1[i] * 100)
                                else:
                                    percent_active_binned_within1.append('NA')
                            else:
                                if left_binned_within1[i] + right_binned_within1[i] != 0:
                                    percent_active_binned_within1.append(right_binned_within1[i] / total_pokes_binned_within1[i] * 100)
                                else:
                                    percent_active_binned_within1.append('NA')
                    
                    if seconds_in_bins2 != '':
                    
                        end_time = seconds_elapsed[-1]
                        
                        # Create end value that is the next greatest multiple of the required time bin duration greater than the end_time so that those data points are retained in the output
                        
                        excess = math.ceil(end_time / int(seconds_in_bins2)) # Rounds up so that the number of bins is the multiple of bin length greater than the time point
                        
                        end_bin = excess * int(seconds_in_bins2) # Creates the end value for the last time bin
                        
                        interval_range = pd.interval_range(start=0, freq=int(seconds_in_bins2), end=end_bin, closed="left") #'left' means that it includes the first data point which is 0 because it's start time
                        
                        time_bins = df[seconds_elapsed] = pd.cut(seconds_elapsed, interval_range, include_lowest=True, ordered=True)
                        
                        # Create dictionary of time bins as keys with bin number as matching entry to label the bins
                        
                        # Create a list of integers of the same length as there are time bins
                        
                        bin_number = []
                        a = 1
                        for i in range(0, len(interval_range)):
                            bin_number.append(a)
                            a += 1
                        
                        make_dictionary = zip(interval_range, bin_number)
                        
                        bin_dictionary = dict(make_dictionary)
                        
                        bin_num = []
                        for i in range(0, len(time_bins)):
                            bin_num.append(bin_dictionary.get((time_bins[i])))
                        
                        # Get the last entry into each bin for the rest of the columns required
                        
                        bin_num_binned2 = []
                        left_binned2 = []
                        right_binned2 = []
                        pellet_binned2 = []
                        active_binned2 = []
                        session_type_binned2 = []
                        percent_active_binned2 = []
                        total_pokes_binned2 = []
                        
                        # time binned
                        for i in range(1, len(time_bins)):
                                if bin_num[i - 1] != bin_num[i]:
                                    bin_num_binned2.append(bin_num[i - 1])
                        bin_num_binned2.append(bin_num[i - 0])
                        
                        # left binned
                        for i in range(1, len(left_poke2)):
                                if time_bins[i - 1] != time_bins[i]:
                                    left_binned2.append(left_poke2[i - 1])
                        left_binned2.append(left_poke2[i - 0])
                                
                        # right binned
                        for i in range(1, len(right_poke2)):
                                if time_bins[i - 1] != time_bins[i]:
                                    right_binned2.append(right_poke2[i - 1])
                        right_binned2.append(right_poke2[i - 0])
                             
                        # pellet binned
                        for i in range(1, len(pellet_count2)):
                                if time_bins[i - 1] != time_bins[i]:
                                    pellet_binned2.append(pellet_count2[i - 1])
                        pellet_binned2.append(pellet_count2[i - 0])
                        
                        # active binned
                        for i in range(1, len(active2)):
                                if time_bins[i - 1] != time_bins[i]:
                                    active_binned2.append(active2[i - 1])
                        active_binned2.append(active2[i - 0])
                        
                        # session type binned
                        for i in range(1, len(session_type2)):
                                if time_bins[i - 1] != time_bins[i]:
                                    session_type_binned2.append(session_type2[i - 1])
                        session_type_binned2.append(session_type2[i - 0])
                        
                        # total pokes binned
                        for i in range(1, len(total_pokes)):
                                if time_bins[i - 1] != time_bins[i]:
                                    total_pokes_binned2.append(total_pokes[i - 1])
                        total_pokes_binned2.append(total_pokes[i - 0])
                        
                        # percent active binned
                        for i in range(0, len(active_binned2)):
                            if active_binned2[i] == 'Left':
                                percent_active_binned2.append(left_binned2[i] / total_pokes_binned2[i] * 100)
                            else:
                                percent_active_binned2.append(right_binned2[i] / total_pokes_binned2[i] * 100)
                        
                        # Enter new row for any bins that have no data in them (for cumulative bins this is a duplication of the preceding row)
                    
                        active_port = active_binned2[0]
                        session_type3 = session_type2[0]
                              
                        for i in bin_dictionary:
                            if bin_dictionary[i] in bin_num_binned2:
                                continue
                            else:
                                bin_num_binned2.insert((bin_dictionary[i] - 1), bin_dictionary[i])
                                left_binned2.insert((bin_dictionary[i] - 1), left_binned2[(bin_dictionary[i] - 2)])
                                right_binned2.insert((bin_dictionary[i] - 1), right_binned2[(bin_dictionary[i] - 2)])
                                pellet_binned2.insert((bin_dictionary[i] - 1), pellet_binned2[(bin_dictionary[i] - 2)])
                                active_binned2.insert((bin_dictionary[i] - 1), active_port)
                                percent_active_binned2.insert((bin_dictionary[i] - 1), percent_active_binned2[(bin_dictionary[i] - 2)])
                                session_type_binned2.insert((bin_dictionary[i] - 1), session_type3)
                                total_pokes_binned2.insert((bin_dictionary[i] - 1), total_pokes_binned2[(bin_dictionary[i] - 2)])
                                
                        # Make new counts that are within the bins rather than cumulative
                        
                        left_binned_within2 = []
                        right_binned_within2 = []
                        pellet_binned_within2 = []
                        percent_active_binned_within2 = []
                        total_pokes_binned_within2 = []
                        
                        left_binned_within2.append(left_binned2[0])
                        for i in range(1, len(left_binned2)):
                            left_binned_within2.append(left_binned2[i] - left_binned2[i - 1])
                        
                        right_binned_within2.append(right_binned2[0])
                        for i in range(1, len(right_binned2)):
                            right_binned_within2.append(right_binned2[i] - right_binned2[i - 1])
                        
                        pellet_binned_within2.append(pellet_binned2[0])
                        for i in range(1, len(pellet_binned2)):
                            pellet_binned_within2.append(pellet_binned2[i] - pellet_binned2[i - 1])
                            
                        total_pokes_binned_within2.append(total_pokes_binned2[0])
                        for i in range (1, len(total_pokes_binned2)):
                            total_pokes_binned_within2.append(total_pokes_binned2[i] - total_pokes_binned2[i - 1])
                            
                        for i in range(0, len(active_binned2)):
                            if active_binned2[i] == 'Left':
                                if left_binned_within2[i] + right_binned_within2[i] != 0:
                                    percent_active_binned_within2.append(left_binned_within2[i] / total_pokes_binned_within2[i] * 100)
                                else:
                                    percent_active_binned_within2.append('NA')
                            else:
                                if left_binned_within2[i] + right_binned_within2[i] != 0:
                                    percent_active_binned_within2.append(right_binned_within2[i] / total_pokes_binned_within2[i] * 100)
                                else:
                                    percent_active_binned_within2.append('NA')
                        
                        
                    #####----- Create session summary data-----#####
                    
                    task = session_type2[0]
                    
                    session_hours = str(math.floor(seconds_elapsed[-1] / 3600))
                    
                    session_mins = math.floor((seconds_elapsed[-1] % 3600) / 60)
                    
                    session_secs = seconds_elapsed[-1] % 60
                    
                    if session_mins < 10:
                        mins = '0' + str(session_mins)
                    else:
                        mins = str(session_mins)
                        
                    if session_secs < 10:
                        secs = '0' + str(session_secs)
                    else:
                        secs = str(session_secs)
                        
                    session_duration = str(session_hours) + ':' + mins + ':' + secs
                    
                    # Assign left and right pokes as active or inactive based on the active port
                    
                    if active2[0] == 'Left':
                        active_pokes = left_poke2[-1]
                        inactive_pokes = right_poke2[-1]
                    elif active2[0] == 'Right':
                        active_pokes = right_poke2[-1]
                        inactive_pokes = left_poke2[-1]
                    
                    active_port = active2[0]
                    
                    date = import_name[7:13]
                    aus_date = date[2:4] + '/' + date[0:2] + '/20' + date[4:]
                            
                    variable = ['Filename', 'Date', 'Task', 'Duration', 'Active port', 'Total Pokes', 'Active Pokes', 'Inactive Pokes', '% Active Pokes', 'Pellets']
                    
                    value = [import_name.strip('.CSV'), aus_date, task, session_duration, active2[0], total_pokes[-1], active_pokes, inactive_pokes, percent_active[-1], pellet_count2[-1]]
                    
                    #####----- Export the data -----#####
                    
                    export_results = 'Y'
                    
                    if export_results == 'Y':
                        
                        # Always export Summary and Chronological Data
                        
                        results_summary = {'Variable': variable, 'Value': value}
                        export_file_summary = pd.DataFrame(results_summary, columns = ['Variable', 'Value'])
                        
                        # Assign left and right pokes as active or inactive based on the active port
                        
                        if active2[0] == 'Left':
                            active_poke = left_poke2
                            inactive_poke = right_poke2
                            
                            if seconds_in_bins1 != '':
                                active_poke_binned1 = left_binned1
                                active_poke_binned_within1 = left_binned_within1
                                inactive_poke_binned1 = right_binned1
                                inactive_poke_binned_within1 = right_binned_within1
                                
                            if seconds_in_bins2 != '':
                                active_poke_binned2 = left_binned2
                                active_poke_binned_within2 = left_binned_within2
                                inactive_poke_binned2 = right_binned2
                                inactive_poke_binned_within2 = right_binned_within2
                                
                        elif active2[0] == 'Right':
                            active_poke = right_poke2
                            inactive_poke = left_poke2
                            
                            if seconds_in_bins1 != '':
                                active_poke_binned1 = right_binned1
                                active_poke_binned_within1 = right_binned_within1
                                inactive_poke_binned1 = left_binned1
                                inactive_poke_binned_within1 = left_binned_within1
                                
                            if seconds_in_bins2 != '':
                                active_poke_binned2 = left_binned2
                                active_poke_binned_within2 = left_binned_within2
                                inactive_poke_binned2 = right_binned2
                                inactive_poke_binned_within2 = right_binned_within2
                        
                        
                        results_chronological = {'Time (seconds)': seconds_elapsed, 'Session Type': session_type2, 'Active port': active2, 'Active Poke': active_poke, 'Inactive Poke': inactive_poke, 
                                                  'Total Poke': total_pokes, 'Pellet Count': pellet_count2, '% Active Pokes': percent_active}
                        export_file_chronological = pd.DataFrame(results_chronological, columns = ['Time (seconds)', 'Session Type', 'Active port', 'Active Poke', 'Inactive Poke',
                                                                                                    'Total Poke', 'Pellet Count', '% Active Pokes'])
            
                            
                        # If using time bins also export binned data (cumulative and within)
                        
                        if seconds_in_bins1 != '':
                        
                            results_binned_c1 = {'Time bin (' + seconds_in_bins1 + 's each)': bin_num_binned1, 'Session Type': session_type_binned1, 'Active Port': active_binned1, 
                                                'Active Poke': active_poke_binned1, 'Inactive Poke': inactive_poke_binned1, 'Total Poke': total_pokes_binned1, 'Pellet Count': pellet_binned1, '% Active Pokes': percent_active_binned1}
                            export_file_binned_c1 = pd.DataFrame(results_binned_c1, columns = ['Time bin (' + seconds_in_bins1 + 's each)', 'Session Type', 'Active Port', 
                                                                                              'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', '% Active Pokes'])
                            
                            results_binned_w1 = {'Time bin (' + seconds_in_bins1 + 's each)': bin_num_binned1, 'Session Type': session_type_binned1, 'Active Port': active_binned1, 
                                                'Active Poke': active_poke_binned_within1, 'Inactive Poke': inactive_poke_binned_within1, 'Total Poke': total_pokes_binned_within1, 'Pellet Count': pellet_binned_within1, '% Active Pokes': percent_active_binned_within1}
                            export_file_binned_w1 = pd.DataFrame(results_binned_w1, columns = ['Time bin (' + seconds_in_bins1 + 's each)', 'Session Type', 'Active Port', 
                                                                                              'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', '% Active Pokes'])
                        
                        if seconds_in_bins2 != '':
                        
                            results_binned_c2 = {'Time bin (' + seconds_in_bins2 + 's each)': bin_num_binned2, 'Session Type': session_type_binned2, 'Active Port': active_binned2, 
                                                'Active Poke': active_poke_binned2, 'Inactive Poke': inactive_poke_binned2, 'Total Poke': total_pokes_binned2, 'Pellet Count': pellet_binned2, '% Active Pokes': percent_active_binned2}
                            export_file_binned_c2 = pd.DataFrame(results_binned_c2, columns = ['Time bin (' + seconds_in_bins2 + 's each)', 'Session Type', 'Active Port', 
                                                                                              'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', '% Active Pokes'])
                            
                            results_binned_w2 = {'Time bin (' + seconds_in_bins2 + 's each)': bin_num_binned2, 'Session Type': session_type_binned2, 'Active Port': active_binned2, 
                                                'Active Poke': active_poke_binned_within2, 'Inactive Poke': inactive_poke_binned_within2, 'Total Poke': total_pokes_binned_within2, 'Pellet Count': pellet_binned_within2, '% Active Pokes': percent_active_binned_within2}
                            export_file_binned_w2 = pd.DataFrame(results_binned_w2, columns = ['Time bin (' + seconds_in_bins2 + 's each)', 'Session Type', 'Active Port', 
                                                                                              'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', '% Active Pokes'])
    
                        
                        # Export to excel
                        
                        from openpyxl import Workbook
                    
                        wb = Workbook()
                        
                        ws1 = wb.active
                        ws1.title = 'Summary'
                        
                        ws2 = wb.create_sheet()
                        ws2.title = 'Chronological'
                        
                        results_to_export = [export_file_summary, export_file_chronological]
                        
                        if seconds_in_bins1 != '':
                            
                            bin_min_length1 = int(seconds_in_bins1) / 60
                            
                            ws3 = wb.create_sheet()
                            ws3.title = str(int(bin_min_length1)) + 'min Binned Cumulative'
                            
                            ws4 = wb.create_sheet()
                            ws4.title = str(int(bin_min_length1)) + 'min Binned Within'
                            
                            results_to_export.append(export_file_binned_c1)
                            results_to_export.append(export_file_binned_w1)
                            
                        if seconds_in_bins2 != '':
                            
                            bin_min_length2 = int(seconds_in_bins2) / 60
                            
                            ws5 = wb.create_sheet()
                            ws5.title = str(int(bin_min_length2)) + 'min Binned Cumulative'
                            
                            ws6 = wb.create_sheet()
                            ws6.title = str(int(bin_min_length2)) + 'min Binned Within'
                            
                            results_to_export.append(export_file_binned_c2)
                            results_to_export.append(export_file_binned_w2)
                        
                        sheets_to_export = wb.sheetnames
                        
                        with pd.ExcelWriter(export_destination) as writer:
                            
                            for i in range(len(sheets_to_export)):
                                results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
            
    
            ##########---------- Overview Sheets----------##########
            
            from openpyxl import Workbook
                    
            wb = Workbook()
            
            ws = wb.active
            ws.title = 'Summary Overview'
            
            value_column = 'Value'
            variable_column = 'Variable'
            
            value_summary = []
            session_num = []
            counter = 1
            
            export_location_folder = export_location + folder + '/'
            
            for filename in list(os.listdir(export_location_folder)):
                
                if filename.endswith(".xlsx"):
                    if not filename.startswith('FED') or filename.startswith('10'):
                    
                        if counter == 1:
                            
                            export_name = filename
                            export_destination = export_location_folder + export_name
                            
                            df_overview = pd.read_excel(export_destination, sheet_name = 'Summary')
                            
                            df_overview.drop(columns='Value', inplace=True)
                            
                        export_name = filename
                        export_destination = export_location_folder + export_name
                        
                        name = 'Session ' + str(counter)
                        session_num.append(name)
                        
                        df = pd.read_excel(export_destination, sheet_name = 'Summary')
                        
                        values = df[value_column].tolist()
                        
                        df_overview.insert(counter, name, values)
                        
                        counter += 1
                    
            overview_name = folder + ' FR Training Overview.xlsx'
            overview_destination = export_location_folder + overview_name
            
            with pd.ExcelWriter(overview_destination) as writer:
                            
                df_overview.to_excel(writer, sheet_name='Overview', engine='openpyxl', index=False, header=True)
            
            cohort_overview_destination = cohort_export_location + overview_name
            
            with pd.ExcelWriter(cohort_overview_destination) as writer:
                            
                df_overview.to_excel(writer, sheet_name='Overview', engine='openpyxl', index=False, header=True)

##########----------########## FR1_both

elif schedule == 'FR1_both':

    time_column = 'MM:DD:YYYY hh:mm:ss'
    event_column = 'Event'
    session_type_column = 'Session_Type'
    left_poke_column = 'Left_Poke_Count'
    right_poke_column = 'Right_Poke_Count'
    pellet_count_column = 'Pellet_Count'
    
    # seconds_in_bins = '600' # Enter how long you want the time bins to be in seconds; if not using bins enter '' which will give duration in seconds
    
    # # Enter the FED import folder, FED export folder, cohort/experiment export folder and the FED number
    
    # import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 Reversal/FED_Test/Datafiles/'
    # export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 Reversal/FED_Test/DatafilesPython/'
    # cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 Reversal/FED_Test/DatafilesPython/Overview/'
    # FED_num = 'FED17'
    
    #-----------------------------------------------------------------------------
    
    # Import the revelant data: time, FR ratio, active port, left poke, right poke, and pellet count.
    
    import pandas as pd
    import numpy as np
    import os
    import openpyxl
    
    #-----
    
    for filename in list(os.listdir(import_location)):
        
        if filename.endswith(".CSV"):
            
            # Import the data
            
            import_name = filename
            import_destination = import_location + import_name
            export_name = 'FR1 both Training ' + import_name.strip(".CSV.") + '.xlsx'
            export_destination = export_location + export_name
    
    #-----
    
            df = pd.read_csv(import_destination)
            
            time = df[time_column].tolist()
            session_type = df[session_type_column].tolist()
            event = df[event_column].tolist()
            left_poke = df[left_poke_column].tolist()
            right_poke = df[right_poke_column].tolist()
            pellet_count = df[pellet_count_column].tolist()
            
            # Change the time column from strings to datetime format for calculating durations between timestamps and creating time bins
            
            import datetime as dt
            
            time_list = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time]
            
            # Determine session start time which is the first timestamp generated by our poke to the/an acitve port to deliver a pellet
            
            session_start = time_list[0]
            
            # Remove the initiation poke/s and pellet data (if required)
            
            if initiation_poke == True:
            
                left_count = left_poke[0]
                right_count = right_poke[0]
                
                del time[:2]
                del session_type[:2]
                del event[:2]
                del left_poke[:2]
                del right_poke[:2]
                del pellet_count[:2]
                
                # Subtract the poke and pellet from the subsequent cumulative data
                
                left_poke_shifted = []
                right_poke_shifted = []
                pellet_count_shifted = []
                
                for i in range(0, len(left_poke)):
                    left_poke_shifted.append(left_poke[i] - left_count)
                    right_poke_shifted.append(right_poke[i] - right_count)
                    pellet_count_shifted.append(pellet_count[i] - 1)
                        
                # Shift the pellet count list one step backwards to match the pokes as in the original csv file it's created on a new line which is deleted in this code
                
                pellet_count_shifted.pop(0)
                
                # Create new lists that only include data from the lines where event == 'Poke'
                time2 = []
                session_type2 = []
                event2 = []
                left_poke2 = []
                right_poke2 = []
                pellet_count2 = []
                
                for a, b, c, d, e, f in zip(event, left_poke_shifted, right_poke_shifted, pellet_count_shifted, time, session_type):
                    if a == 'Poke':
                        event2.append(a)
                        left_poke2.append(b)
                        right_poke2.append(c) 
                        pellet_count2.append(d)
                        time2.append(e)
                        session_type2.append(f)
            else:
                # Shift the pellet count list one step backwards to match the pokes as in the original csv file it's created on a new line which is deleted in this code
                
                pellet_count.pop(0)
                
                # Create new lists that only include data from the lines where event == 'Poke'
                time2 = []
                session_type2 = []
                event2 = []
                left_poke2 = []
                right_poke2 = []
                pellet_count2 = []
                
                for a, b, c, d, e, f in zip(event, left_poke, right_poke, pellet_count, time, session_type):
                    if a == 'Poke':
                        event2.append(a)
                        left_poke2.append(b)
                        right_poke2.append(c) 
                        pellet_count2.append(d)
                        time2.append(e)
                        session_type2.append(f)
            
            # Change the time column from strings to datetime format for calculating durations between timestamps and creating time bins
            
            time2 = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time2]
            
            # Create total_poke column
    
            total_pokes = []
            
            for i in range(0, len(left_poke2)):
                total_pokes.append(left_poke2[i] + right_poke2[i])
            
            # Create two new columns that show the percentage of pokes that are left or right
            
            left_percent = []
            right_percent = []
            
            for i in range(0, len(left_poke2)):
                left_percent.append(left_poke2[i] / total_pokes[i] * 100)
                right_percent.append(right_poke2[i] / total_pokes[i] * 100)
                            
            # Calculate how much time (in seconds) has elapsed from session start time for each event (nose poke)
            
            seconds_elapsed = []
            
            for i in range(0, len(time2)):
                time_from_start = time2[i] - session_start
                result = int(time_from_start.total_seconds())
                seconds_elapsed.append(result)
                
            
            # Create time bins of desired length (input for desired length is at the top of the code)
            
            if seconds_in_bins != '':
            
                end_time = seconds_elapsed[-1]
                
                import math
                
                # Create end value that is the next greatest multiple of the required time bin duration greater than the end_time so that those data points are retained in the output
                
                excess = math.ceil(end_time / int(seconds_in_bins)) # Rounds up so that the number of bins is the multiple of bin length greater than the time point
                
                end_bin = excess * int(seconds_in_bins) # Creates the end value for the last time bin
                
                interval_range = pd.interval_range(start=0, freq=int(seconds_in_bins), end=end_bin, closed="left") #'left' means that it includes the first data point which is 0 because it's start time
                
                time_bins = df[seconds_elapsed] = pd.cut(seconds_elapsed, interval_range, include_lowest=True, ordered=True)
                
                # Create dictionary of time bins as keys with bin number as matching entry to label the bins
                
                # Create a list of integers of the same length as there are time bins
                
                bin_number = []
                a = 1
                for i in range(0, len(interval_range)):
                    bin_number.append(a)
                    a += 1
                
                make_dictionary = zip(interval_range, bin_number)
                
                bin_dictionary = dict(make_dictionary)
                
                bin_num = []
                for i in range(0, len(time_bins)):
                    bin_num.append(bin_dictionary.get((time_bins[i])))
                
                # Get the last entry into each bin for the rest of the columns required
                
                bin_num_binned = []
                left_binned = []
                right_binned = []
                total_pokes_binned = []
                pellet_binned = []
                session_type_binned = []
                left_percent_binned = []
                right_percent_binned = []
                
                # time binned
                for i in range(1, len(time_bins)):
                        if bin_num[i - 1] != bin_num[i]:
                            bin_num_binned.append(bin_num[i - 1])
                bin_num_binned.append(bin_num[i - 0])
                
                # left binned
                for i in range(1, len(left_poke2)):
                        if time_bins[i - 1] != time_bins[i]:
                            left_binned.append(left_poke2[i - 1])
                left_binned.append(left_poke2[i - 0])
                        
                # right binned
                for i in range(1, len(right_poke2)):
                        if time_bins[i - 1] != time_bins[i]:
                            right_binned.append(right_poke2[i - 1])
                right_binned.append(right_poke2[i - 0])
                
                # total binned
                for i in range(1, len(total_pokes)):
                        if time_bins[i - 1] != time_bins[i]:
                            total_pokes_binned.append(total_pokes[i - 1])
                total_pokes_binned.append(total_pokes[i - 0])
                     
                # pellet binned
                for i in range(1, len(pellet_count2)):
                        if time_bins[i - 1] != time_bins[i]:
                            pellet_binned.append(pellet_count2[i - 1])
                pellet_binned.append(pellet_count2[i - 0])
                
                # session type binned
                for i in range(1, len(session_type2)):
                        if time_bins[i - 1] != time_bins[i]:
                            session_type_binned.append(session_type2[i - 1])
                session_type_binned.append(session_type2[i - 0])
                
                # left and right percent binned
                for i in range(0, len(total_pokes_binned)):
                    left_percent_binned.append(left_binned[i] / total_pokes_binned[i] * 100)
                    right_percent_binned.append(right_binned[i] / total_pokes_binned[i] * 100)
                
                
                # Enter new row for any bins that have no data in them (for cumulative bins this is a duplication of the preceding row)
            
                session_type3 = session_type2[0]
                
                for i in bin_dictionary:
                    if bin_dictionary[i] in bin_num_binned:
                        continue
                    else:
                        bin_num_binned.insert((bin_dictionary[i] - 1), bin_dictionary[i])
                        left_binned.insert((bin_dictionary[i] - 1), left_binned[(bin_dictionary[i] - 2)])
                        right_binned.insert((bin_dictionary[i] - 1), right_binned[(bin_dictionary[i] - 2)])
                        total_pokes_binned.insert((bin_dictionary[i] - 1), total_pokes_binned[(bin_dictionary[i] - 2)])
                        pellet_binned.insert((bin_dictionary[i] - 1), pellet_binned[(bin_dictionary[i] - 2)])
                        session_type_binned.insert((bin_dictionary[i] - 1), session_type3)
                        left_percent_binned.insert((bin_dictionary[i] - 1), left_percent_binned[(bin_dictionary[i] - 2)])
                        right_percent_binned.insert((bin_dictionary[i] - 1), right_percent_binned[(bin_dictionary[i] - 2)])
    
                            
                # Make the counts within the bins rather than cumulative
                
                left_binned_within = []
                right_binned_within = []
                total_pokes_binned_within = []
                pellet_binned_within = []
                left_percent_binned_within = []
                right_percent_binned_within = []
                
                left_binned_within.append(left_binned[0])
                for i in range(1, len(left_binned)):
                    left_binned_within.append(left_binned[i] - left_binned[i - 1])
                
                right_binned_within.append(right_binned[0])
                for i in range(1, len(right_binned)):
                    right_binned_within.append(right_binned[i] - right_binned[i - 1])
                
                total_pokes_binned_within.append(total_pokes_binned[0])
                for i in range(1, len(total_pokes_binned)):
                    total_pokes_binned_within.append(total_pokes_binned[i] - total_pokes_binned[i - 1])
                
                pellet_binned_within.append(pellet_binned[0])
                for i in range(1, len(pellet_binned)):
                    pellet_binned_within.append(pellet_binned[i] - pellet_binned[i - 1])
                    
                for i in range(0, len(total_pokes_binned_within)):
                    if left_binned_within[i] + right_binned_within[i] != 0:
                        left_percent_binned_within.append(left_binned_within[i] / total_pokes_binned_within[i] * 100)
                        right_percent_binned_within.append(right_binned_within[i] / total_pokes_binned_within[i] * 100)
                    else:
                        left_percent_binned_within.append('NA')
                        right_percent_binned_within.append('NA')
                        
                #####----- Create session summary data-----#####
                
                task = session_type3
            
                session_hours = str(math.floor(seconds_elapsed[-1] / 3600))
            
                session_mins = math.floor((seconds_elapsed[-1] % 3600) / 60)
                
                session_secs = seconds_elapsed[-1] % 60
                
                if session_mins < 10:
                    mins = '0' + str(session_mins)
                else:
                    mins = str(session_mins)
                    
                if session_secs < 10:
                    secs = '0' + str(session_secs)
                else:
                    secs = str(session_secs)
                    
                session_duration = str(session_hours) + ':' + mins + ':' + secs
                
                date = import_name[7:13]
                aus_date = date[2:4] + '/' + date[0:2] + '/20' + date[4:]
                        
                variable = ['Filename', 'Date', 'Task', 'Duration', 'Total Pokes', 'Left Pokes', 'Right Pokes', '% Left Pokes', '% Right Pokes', 'Pellets']
                
                value = [import_name.strip('.CSV'), aus_date, task, session_duration, total_pokes[-1], left_poke2[-1], right_poke2[-1], left_percent_binned[-1], right_percent_binned[-1], pellet_count2[-1]]
    
            # Export the data.
            
            export_results = 'Y'
            
            if export_results == 'Y':
                # Always export Summary and Chronological Data
                
                results_summary = {'Variable': variable, 'Value': value}
                export_file_summary = pd.DataFrame(results_summary, columns = ['Variable', 'Value'])
                
                results_chronological = {'Time (seconds)': seconds_elapsed, 'Session Type': session_type2, 'Left Poke': left_poke2, 'Right Poke': right_poke2, 'Total Poke': total_pokes, '% Left Pokes': left_percent, '% Right Pokes': right_percent, 'Pellet Count': pellet_count2}
                export_file_chronological = pd.DataFrame(results_chronological, columns = ['Time (seconds)', 'Session Type', 'Left Poke', 'Right Poke', 'Total Poke', '% Left Pokes', '% Right Pokes', 'Pellet Count'])
                
                
                if seconds_in_bins != '':
                    results_binned_c = {'Time bin (' + str(seconds_in_bins) + 's each)': bin_num_binned, 'Session Type': session_type_binned, 
                                        'Left Poke': left_binned, 'Right Poke': right_binned, 'Total Poke': total_pokes_binned, '% Left Pokes': left_percent_binned, '% Right Pokes': right_percent_binned, 'Pellet Count': pellet_binned}
                    export_file_binned_c = pd.DataFrame(results_binned_c, columns = ['Time bin (' + seconds_in_bins + 's each)', 'Session Type', 'Left Poke', 'Right Poke', 'Total Poke', '% Left Pokes', '% Right Pokes', 'Pellet Count'])
                    results_binned_w = {'Time bin (' + seconds_in_bins + 's each)': bin_num_binned, 'Session Type': session_type_binned, 
                                        'Left Poke': left_binned_within, 'Right Poke': right_binned_within, 'Total Poke': total_pokes_binned_within, '% Left Pokes': left_percent_binned_within, '% Right Pokes': right_percent_binned_within, 'Pellet Count': pellet_binned_within}
                    export_file_binned_w = pd.DataFrame(results_binned_w, columns = ['Time bin (' + seconds_in_bins + 's each)', 'Session Type', 'Left Poke', 'Right Poke', 'Total Poke', '% Left Pokes', '% Right Pokes', 'Pellet Count'])
                
                
                # Export to excel
                
                from openpyxl import Workbook
            
                wb = Workbook()
                
                ws1 = wb.active
                ws1.title = 'Summary'
                
                ws2 = wb.create_sheet()
                ws2.title = 'Chronological'
                
                results_to_export = [export_file_summary, export_file_chronological]
                
                if seconds_in_bins != '':
                    
                    ws3 = wb.create_sheet()
                    ws3.title ='Binned Cumulative'
                    
                    ws4 = wb.create_sheet()
                    ws4.title ='Binned Within'
                    
                    results_to_export.append(export_file_binned_c)
                    results_to_export.append(export_file_binned_w)
                
                sheets_to_export = wb.sheetnames
                
                with pd.ExcelWriter(export_destination) as writer:
                    
                    for i in range(len(sheets_to_export)):
                        results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
             
    
    ##########---------- Overview Sheets ----------##########
    
    from openpyxl import Workbook
            
    wb = Workbook()
    
    ws = wb.active
    ws.title = 'Summary Overview'
    
    value_column = 'Value'
    variable_column = 'Variable'
    
    value_summary = []
    session_num = []
    counter = 1
    
    for filename in list(os.listdir(export_location)):
        
        if filename.endswith(".xlsx"):
            
            if counter == 1:
                
                export_name = filename
                export_destination = export_location + export_name
                
                df_overview = pd.read_excel(export_destination, sheet_name = 'Summary')
                
                df_overview.drop(columns='Value', inplace=True)
                
            export_name = filename
            export_destination = export_location + export_name
            
            name = 'Session ' + str(counter)
            session_num.append(name)
            
            df = pd.read_excel(export_destination, sheet_name = 'Summary')
            
            values = df[value_column].tolist()
            
            df_overview.insert(counter, name, values)
            
            counter += 1
            
    overview_name = FED_num + ' FR1 Both Training Overview.xlsx'
    overview_destination = export_location + overview_name
    
    with pd.ExcelWriter(overview_destination) as writer:
                    
        df_overview.to_excel(writer, sheet_name='Overview', engine='openpyxl', index=False, header=True)
    
    cohort_overview_destination = cohort_export_location + overview_name
    
    with pd.ExcelWriter(cohort_overview_destination) as writer:
                    
        df_overview.to_excel(writer, sheet_name='Overview', engine='openpyxl', index=False, header=True)   

##########----------########## New_Reversal

elif schedule == 'New_Reversal':

    time_column = 'MM:DD:YYYY hh:mm:ss'
    event_column = 'Event'
    active_poke_column = 'Active_Poke'
    session_type_column = 'Session_type'
    left_poke_column = 'Left_Poke_Count'
    right_poke_column = 'Right_Poke_Count'
    pellet_count_column = 'Pellet_Count'
    retrieval_time_column = 'Retrieval_Time'
    poke_time_column = 'Poke_Time'
    
    # seconds_in_bins = '' # Enter how long you want the time bins to be in seconds; if not using bins enter '' which will give duration in seconds
    # reversal_block_length = '10'
    blocked = 'Y' # Enter Y if you want the data put into discrete reversal blocks, enter N if you want continuous data
    
    # Enter the FED import folder, FED export folder, cohort/experiment export folder and the FED number
    
    # import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 Reversal/FED_Test/Datafiles/NewRev/'
    # export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 Reversal/FED_Test/DatafilesPython/NewRev/'
    # cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 Reversal/FED_Test/DatafilesPython/Overview/'
    # FED_num = 'FED2'
    
    #-----------------------------------------------------------------------------
    
    # Import the revelant data: time, FR ratio (session type), left poke, right poke, and pellet count.
    
    import pandas as pd
    import numpy as np
    import os
    import math
    
    #-----
    
    for filename in list(os.listdir(import_location)):
        
        if filename.endswith(".CSV"):
            
            # Import the data
            
            import_name = filename
            import_destination = import_location + import_name
            export_name = 'New Reversal ' + import_name.strip('.CSV') + '.xlsx'
            export_destination = export_location + export_name
            
    #-----        
            
            df = pd.read_csv(import_destination)
            
            time = df[time_column].tolist()
            session_type = df[session_type_column].tolist()
            event = df[event_column].tolist()
            active = df[active_poke_column].tolist()
            left_poke = df[left_poke_column].tolist()
            right_poke = df[right_poke_column].tolist()
            pellet_count = df[pellet_count_column].tolist()
            retrieval_time = df[retrieval_time_column].tolist()
            poke_time = df[poke_time_column].tolist()
            
            # Change the time column from strings to datetime format for calculating durations between timestamps and creating time bins
            
            import datetime as dt
            
            time_list = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time]
            
            # Determine session start time which is the first timestamp generated by our poke to the/an acitve port to deliver a pellet
            
            session_start = time_list[0]
            
            # Remove the initiation poke/s and pellet data (if required)
            
            if initiation_poke == True:
                            
                left_count = left_poke[0]
                right_count = right_poke[0]
                
                del time[:2]
                del session_type[:2]
                del event[:2]
                del active[:2]
                del left_poke[:2]
                del right_poke[:2]
                del pellet_count[:2]
                del retrieval_time[:2]
                del poke_time[:2]
                
                # Subtract the poke and pellet from the subsequent cumulative data
                
                left_poke_shifted = []
                right_poke_shifted = []
                pellet_count_shifted = []
                
                for i in range(0, len(left_poke)):
                    left_poke_shifted.append(left_poke[i] - left_count)
                    right_poke_shifted.append(right_poke[i] - right_count)
                    pellet_count_shifted.append(pellet_count[i] - 1)
                
                # Shift the retrieval_time and pellet_count lists one step backwards so that the count/time is in the same index as the corresponding poke
                
                retrieval_time.pop(0)
                pellet_count_shifted.pop(0)
                
                
                #####-----##### CHRONOLOGICAL DATA #####-----#####
                
                # Create new lists that only include data from the lines where the event is a Poke, in ReversalTask this means != 'Pellet' (i.e. Poke, Left, or Right)
                
                time2 = []
                session_type2 = []
                event2 = []
                active2 = []
                left_poke2 = []
                right_poke2 = []
                pellet_count2 = []
                retrieval_time2 = []
                poke_time2 = []
                
                for a, b, c, d, e, f, g, h, i in zip(event, left_poke_shifted, right_poke_shifted, pellet_count_shifted, time, session_type, active, retrieval_time, poke_time):
                    if a != 'Pellet':
                        event2.append(a)
                        left_poke2.append(b)
                        right_poke2.append(c) 
                        pellet_count2.append(d)
                        time2.append(e)
                        session_type2.append(f)
                        active2.append(g)
                        retrieval_time2.append(h)
                        poke_time2.append(i)
            else:
                # Shift the retrieval_time and pellet_count lists one step backwards so that the count/time is in the same index as the corresponding poke
                
                retrieval_time.pop(0)
                pellet_count.pop(0)
                
                
                #####-----##### CHRONOLOGICAL DATA #####-----#####
                
                # Create new lists that only include data from the lines where the event is a Poke, in ReversalTask this means != 'Pellet' (i.e. Poke, Left, or Right)
                
                time2 = []
                session_type2 = []
                event2 = []
                active2 = []
                left_poke2 = []
                right_poke2 = []
                pellet_count2 = []
                retrieval_time2 = []
                poke_time2 = []
                
                for a, b, c, d, e, f, g, h, i in zip(event, left_poke, right_poke, pellet_count, time, session_type, active, retrieval_time, poke_time):
                    if a != 'Pellet':
                        event2.append(a)
                        left_poke2.append(b)
                        right_poke2.append(c) 
                        pellet_count2.append(d)
                        time2.append(e)
                        session_type2.append(f)
                        active2.append(g)
                        retrieval_time2.append(h)
                        poke_time2.append(i)
                
            # Change the time column from strings to datetime format for calculating durations between timestamps and creating time bins
            
            time2 = [dt.datetime.strptime(time, '%m/%d/%Y %H:%M:%S') for time in time2]
            
            # Calculate how much time (in seconds) has elapsed from start time for each event (nose poke)
            
            seconds_elapsed = []
            
            for i in range(0, len(time2)):
                time_from_start = time2[i] - session_start
                result = int(time_from_start.total_seconds())
                seconds_elapsed.append(result)
            
            # Transform seconds_elapsed into duration in h:mm:ss
            
            duration_from_start = []
            
            for i in range(0, len(seconds_elapsed)):
                time_hours = str(math.floor(seconds_elapsed[i] / 3600))
                
                time_mins = math.floor((seconds_elapsed[i] % 3600) / 60)
                
                time_secs = seconds_elapsed[i] % 60
                    
                if time_mins < 10:
                    t_mins = '0' + str(time_mins)
                else:
                    t_mins = str(time_mins)
                    
                if time_secs < 10:
                    t_secs = '0' + str(time_secs)
                else:
                    t_secs = str(time_secs)
                    
                duration_from_start.append(str(time_hours) + ':' + t_mins + ':' + t_secs)
            
            # Create new column that is % of pokes that are into the active port. Needs to be based on active_poke value (i.e. Left or Right)
            
            percent_active = []
            
            for i in range(0, len(active2)):
                if active2[i] == 'Left':
                    percent_active.append(left_poke2[i] / (left_poke2[i] + right_poke2[i]) * 100)
                else:
                    percent_active.append(right_poke2[i] / (left_poke2[i] + right_poke2[i]) * 100)
            
            # Create total pokes column
            
            total_poke = []
            
            for i in range(0, len(active2)):
                total_poke.append(left_poke2[i] + right_poke2[i])
            
            # Create binary poke columns
            
            left_poke_binary = []
            left_a_poke_binary = []
            left_i_poke_binary = []
            
            if left_poke2[0] == 1:
                left_poke_binary.append(1)
            else:
                left_poke_binary.append(0)
            
            for i in range(1, len(left_poke2)):
                if left_poke2[i - 1] != left_poke2[i]:
                    left_poke_binary.append(1)
                else:
                    left_poke_binary.append(0)
            
            
            right_poke_binary = []
            right_a_poke_binary = []
            right_i_poke_binary = []
                
            if right_poke2[0] == 1:
                right_poke_binary.append(1)
            else:
                right_poke_binary.append(0)
            
            for i in range(1, len(right_poke2)):
                if right_poke2[i - 1] != right_poke2[i]:
                    right_poke_binary.append(1)
                else:
                    right_poke_binary.append(0)
            
            
            for i in range(0, len(left_poke_binary)):
                if left_poke_binary[i] == 1:
                    if active2[i] == 'Left':
                        left_a_poke_binary.append(1)
                        left_i_poke_binary.append(0)
                        right_i_poke_binary.append(0)
                        right_a_poke_binary.append(0)
                    elif active2[i] == 'Right':
                        left_a_poke_binary.append(0)
                        left_i_poke_binary.append(1)
                        right_i_poke_binary.append(0)
                        right_a_poke_binary.append(0)
                elif left_poke_binary[i] == 0:
                    if active2[i] == 'Left':
                        left_a_poke_binary.append(0)
                        left_i_poke_binary.append(0)
                        right_i_poke_binary.append(1)
                        right_a_poke_binary.append(0)
                    elif active2[i] == 'Right':
                        left_a_poke_binary.append(0)
                        left_i_poke_binary.append(0)
                        right_i_poke_binary.append(0)
                        right_a_poke_binary.append(1)
            
            # Create active and inactive poke columns
            
            active_poke_binary = []
            inactive_poke_binary = [] 
                    
            for i in range(0, len(active2)):
                if active2[i] == 'Left':
                    active_poke_binary.append(left_poke_binary[i])
                    inactive_poke_binary.append(right_poke_binary[i])
                elif active2[i] == 'Right':
                    active_poke_binary.append(right_poke_binary[i])
                    inactive_poke_binary.append(left_poke_binary[i])
            
            active_poke = [] # cumulative
            inactive_poke = [] # cumulative
            
            active_counter = 0
            inactive_counter = 0
            
            for i in range (0, len(active_poke_binary)):
                if active_poke_binary[i] == 0:
                    active_poke.append(active_counter)
                elif active_poke_binary[i] == 1:
                    active_counter += 1
                    active_poke.append(active_counter)
                    
            for i in range (0, len(inactive_poke_binary)):
                if inactive_poke_binary[i] == 0:
                    inactive_poke.append(inactive_counter)
                elif inactive_poke_binary[i] == 1:
                    inactive_counter += 1
                    inactive_poke.append(inactive_counter)
            
            # Create binary column indicating active port for graphing
            
            active_port_binary = []
            
            for i in range(0, len(active2)):
                if active2[i] == 'Left':
                    active_port_binary.append(1)
                elif active2[i] == 'Right':
                    active_port_binary.append(0)
            
            # Create Green (correct/active) and Red (incorrect/inactive) binary columns for graphing where 1 and 0 as correct/incorrect is determined by active_port_binary
            
            green = []
            red = []
            
            for i in range(0, len(left_poke_binary)):
                if active_port_binary[i] == 1:
                    if left_poke_binary[i] == 1:
                        green.append(1)
                        red.append(np.nan)
                    elif left_poke_binary[i] == 0:
                        green.append(np.nan)
                        red.append(0)
                elif active_port_binary[i] == 0:
                    if left_poke_binary[i] == 0:
                            green.append(0)
                            red.append(np.nan)
                    elif left_poke_binary[i] ==1:
                            green.append(np.nan)
                            red.append(1)
            
            # Create chronological timing data
            
            #####-----##### Chronological Poke Time Data #####-----#####
            
            l_poke_time = [] # poke time for all left pokes
            r_poke_time = [] # poke time for all right pokes
            
            l_poke_time_chron = [] # poke time for all left pokes with nan if it was a right poke to maintain array length the same as all pokes
            r_poke_time_chron = [] # poke time for all right pokes with nan if it was a left poke to maintain array length the same as all pokes
            
            for i in range(0, len(poke_time2)): # Collates the poke_time by side (regardless of port status)
               if event2[i] == 'Left':
                   l_poke_time.append(poke_time2[i])
                   l_poke_time_chron.append(poke_time2[i])
                   r_poke_time_chron.append(np.nan)
               else:
                   r_poke_time.append(poke_time2[i])
                   r_poke_time_chron.append(poke_time2[i])
                   l_poke_time_chron.append(np.nan)
    
            a_poke_time = [] # poke time for all active pokes
            a_l_poke_time = [] # poke time for all left pokes when the left port is active
            a_r_poke_time = [] # poke time for all right pokes when the right poke is active
            
            i_poke_time = [] # poke time for all inactive pokes
            i_l_poke_time = [] # poke time for all left pokes when the right port is active
            i_r_poke_time = [] # poke time for all right pokes when the left port is active
            
            a_poke_time_chron = [] # all chron lists put nan for a poke that was not one of the status that the list is tracking to maintain array length the same as all pokes
            a_l_poke_time_chron = []
            a_r_poke_time_chron = []
            
            i_poke_time_chron = []
            i_l_poke_time_chron = []
            i_r_poke_time_chron = []
            
            for i in range(0, len(poke_time2)): # Collates poke_time by status, and status + side
                if active_poke_binary[i] == 1:                  # if it's an active poke
                    a_poke_time.append(poke_time2[i])           # add time to active poke time lists and nan to inactive list
                    a_poke_time_chron.append(poke_time2[i])
                    i_poke_time_chron.append(np.nan)
                    if active2[i] == 'Left':                    # if the left port is active
                        a_l_poke_time.append(poke_time2[i])     # add the time to the left active lists and nan to left inactive and right lists
                        a_l_poke_time_chron.append(poke_time2[i])
                        i_l_poke_time_chron.append(np.nan)
                        a_r_poke_time_chron.append(np.nan)
                        i_r_poke_time_chron.append(np.nan)
                        
                        
                    elif active2[i] == 'Right':                 # if the right port is active
                        a_r_poke_time.append(poke_time2[i])     # add the time to the right active lists and nan to right inactive and left lists
                        a_r_poke_time_chron.append(poke_time2[i])
                        i_r_poke_time_chron.append(np.nan)
                        a_l_poke_time_chron.append(np.nan)
                        i_l_poke_time_chron.append(np.nan)
                else:                                           # if it's an inactive poke
                    i_poke_time.append(poke_time2[i])           # add the time to the inactive poke time lists and non to active list
                    i_poke_time_chron.append(poke_time2[i])
                    a_poke_time_chron.append(np.nan)
                    if active2[i] == 'Left':                    # if the left port is active, and therefore the inactive poke is a right poke
                        i_r_poke_time.append(poke_time2[i])     # add time to the right inactive lists and non to right active and left lists
                        i_r_poke_time_chron.append(poke_time2[i])
                        a_r_poke_time_chron.append(np.nan)
                        i_l_poke_time_chron.append(np.nan)
                        a_l_poke_time_chron.append(np.nan)
                    elif active2[i] == 'Right':                 # if the right port is active, and therefore the inactive poke is a left poke
                        i_l_poke_time.append(poke_time2[i])     # add the time to the left inactive lists and nan to left active and right lists
                        i_l_poke_time_chron.append(poke_time2[i])
                        a_l_poke_time_chron.append(np.nan)
                        i_r_poke_time_chron.append(np.nan)
                        a_r_poke_time_chron.append(np.nan)
            
            
            #####-----##### CREATE BLOCKED DATA #####-----#####
            
            # Create blocks (i.e. bins) determined by the active port
            
            seconds_block = []
            duration_block = []
            session_block = []
            active_block = []
            left_block = []
            right_block = []
            active_poke_block = []
            inactive_poke_block = []
            total_block = []
            pellet_block = []
            
            for i in range(1, len(active2)):
                    if active2[i - 1] != active2[i]:
                        seconds_block.append(seconds_elapsed[i - 1])
                        duration_block.append(duration_from_start[i - 1])
                        session_block.append(session_type2[i - 1])
                        active_block.append(active2[i - 1])
                        left_block.append(left_poke2[i - 1])
                        right_block.append(right_poke2[i - 1])
                        active_poke_block.append(active_poke[i - 1])
                        inactive_poke_block.append(inactive_poke[i - 1])
                        total_block.append(total_poke[i - 1])
                        pellet_block.append(pellet_count2[i - 1])
            seconds_block.append(seconds_elapsed[i - 0])
            duration_block.append(duration_from_start[i - 0])
            session_block.append(session_type2[i - 0])
            active_block.append(active2[i - 0])
            left_block.append(left_poke2[i - 0])
            right_block.append(right_poke2[i - 0])
            active_poke_block.append(active_poke[i - 0])
            inactive_poke_block.append(inactive_poke[i - 0])
            total_block.append(total_poke[i - 0])
            pellet_block.append(pellet_count2[i - 0])
            
            # Create percent active column
            
            percent_active_block = []
            
            for i in range(0, len(active_poke_block)):
                if total_block[i] != 0:
                    percent_active_block.append(active_poke_block[i] / total_block[i] * 100)
                else:
                    percent_active_block.append('NA')
            
            
            # Make the counts within the blocks rather than cumulative
            
            left_block_within = []
            right_block_within = []
            active_poke_block_within = []
            inactive_poke_block_within = []
            total_block_within = []
            pellet_block_within = []
            percent_active_block_within = []
            
            left_block_within.append(left_block[0])
            for i in range(1, len(left_block)):
                left_block_within.append(left_block[i] - left_block[i - 1])
            
            right_block_within.append(right_block[0])
            for i in range(1, len(right_block)):
                right_block_within.append(right_block[i] - right_block[i - 1])
            
            active_poke_block_within.append(active_poke_block[0])
            for i in range(1, len(active_poke_block)):
                active_poke_block_within.append(active_poke_block[i] - active_poke_block[i - 1])
            
            inactive_poke_block_within.append(inactive_poke_block[0])
            for i in range(1, len(inactive_poke_block)):
                inactive_poke_block_within.append(inactive_poke_block[i] - inactive_poke_block[i - 1])
            
            total_block_within.append(total_block[0])
            for i in range(1, len(total_block)):
                total_block_within.append(total_block[i] - total_block[i - 1])
            
            pellet_block_within.append(pellet_block[0])
            for i in range(1, len(pellet_block)):
                pellet_block_within.append(pellet_block[i] - pellet_block[i - 1])
            
            # Create percent active pokes column    
            
            for i in range(0, len(active_block)):
                if total_block_within[i] != 0:
                    percent_active_block_within.append(active_poke_block_within[i] / total_block_within[i] * 100)
                else:
                    percent_active_block_within.append('NA')
            
            
            # Create mean poke and retrieval time columns for blocks
            
            block_sum_all_poke_time = [] # cumulative sum of all poke times
            block_sum_a_poke_time = [] # cumulative sum of active poke times
            block_sum_i_poke_time = [] # cumulative sum of inactive poke times
            block_mean_all_poke_time = [] # cumulative mean of all poke times
            block_mean_a_poke_time = [] # cumulative mean of active poke times
            block_mean_i_poke_time = [] # cumulative mean of inactive poke times
            
            
            sum_all_poke_time = 0
            sum_a_poke_time = 0
            sum_i_poke_time = 0
            index = 0
            poke_index = 0
           
            for i in range(0, len(poke_time2)):
                
                if int(seconds_elapsed[i]) <= int(seconds_block[index]):
                    sum_all_poke_time += poke_time2[poke_index]
                    if active_poke_binary[i] == 1:
                        sum_a_poke_time += poke_time2[poke_index]
                    elif inactive_poke_binary[i] == 1:
                        sum_i_poke_time += poke_time2[poke_index]
                    poke_index += 1
                    
                elif int(seconds_elapsed[i] >= int(seconds_block[index])):
                    
                    block_sum_all_poke_time.append(sum_all_poke_time)
                    block_sum_a_poke_time.append(sum_a_poke_time)
                    block_sum_i_poke_time.append(sum_i_poke_time)
                    block_mean_all_poke_time.append((sum_all_poke_time / int(total_block[index])))
                    block_mean_a_poke_time.append((sum_a_poke_time / int(active_poke_block[index])))
                    block_mean_i_poke_time.append((sum_i_poke_time / int(inactive_poke_block[index])))
                    
                    index += 1
                    sum_all_poke_time += poke_time2[poke_index]
                    if active_poke_binary[i] == 1:
                        sum_a_poke_time += poke_time2[poke_index]
                    elif inactive_poke_binary[i] == 1:
                        sum_i_poke_time += poke_time2[poke_index]
                    poke_index += 1
            
            block_within_sum_all_poke_time = [] # within block sum of all poke times
            block_within_sum_a_poke_time = [] # within block sum of active poke times
            block_within_sum_i_poke_time = [] # within block sum of inactive poke times
            
            block_within_sum_all_poke_time.append(block_sum_all_poke_time[0])
            block_within_sum_a_poke_time.append(block_sum_a_poke_time[0])
            block_within_sum_i_poke_time.append(block_sum_i_poke_time[0])  
            
            for i in range(1, len(block_sum_all_poke_time)):
                block_within_sum_all_poke_time.append((block_sum_all_poke_time[i] - block_sum_all_poke_time[i - 1]))
                block_within_sum_a_poke_time.append((block_sum_a_poke_time[i] - block_sum_a_poke_time[i - 1]))
                block_within_sum_i_poke_time.append((block_sum_i_poke_time[i] - block_sum_i_poke_time[i - 1]))  
            
            block_within_mean_all_poke_time = [] # within block mean of all poke times
            block_within_mean_a_poke_time = [] # within block mean of active poke times
            block_within_mean_i_poke_time = [] # within block mean of inactive poke times
            
            for i in range(0, len(block_within_sum_all_poke_time)):
                block_within_mean_all_poke_time.append((block_within_sum_all_poke_time[i] / total_block_within[i]))
                block_within_mean_a_poke_time.append((block_within_sum_a_poke_time[i] / active_poke_block_within[i]))
                block_within_mean_i_poke_time.append((block_within_sum_i_poke_time[i] / inactive_poke_block_within[i]))
            
            
            block_sum_pellet_retrieval_time = [] # cumulative sum of all pellet retrieval times
            block_mean_pellet_retrieval_time = [] # cumulative mean of all pellet retrieval times
            block_within_sum_pellet_retrieval_time = [] # within block sum of pellet retrieval times
            block_within_mean_pellet_retrieval_time = [] # within block mean of pellet retrieval times
            
            if initiation_poke == True:
                block_count = 2 # if the first block only has 9 pellets because of the initiation poke/pellet so need to start 1 later in the count to fill up the block
            else:
                block_count = 1
                
            block = 1
            
            sum_retrieval_time = 0
            
            pellet_retrieval_time = [x for x in retrieval_time2 if not math.isnan(x)] # gets rid of all the nan values which are due to inactive pokes
                    
            for i in range(0, len(pellet_retrieval_time)):
                if block_count < (block * 10):
                    sum_retrieval_time += pellet_retrieval_time[i]
                    block_count += 1
                else:
                    sum_retrieval_time += pellet_retrieval_time[i]
                    block_sum_pellet_retrieval_time.append(sum_retrieval_time)
                    if initiation_poke == True:
                        block_mean_pellet_retrieval_time.append(sum_retrieval_time / (block * 10 - 1)) # the -1 is to account for the initiation pellet retrieval time being removed
                    else:
                        block_mean_pellet_retrieval_time.append(sum_retrieval_time / (block * 10))
                    block_count += 1
                    block += 1    
            
            block_within_sum_pellet_retrieval_time.append(block_sum_pellet_retrieval_time[0])
            for i in range(1, len(block_sum_pellet_retrieval_time)):
                block_within_sum_pellet_retrieval_time.append(block_sum_pellet_retrieval_time[i] - block_sum_pellet_retrieval_time[i - 1])        
            
            if initiation_poke == True:
                block_within_mean_pellet_retrieval_time.append(block_within_sum_pellet_retrieval_time[0] / 9) # first block only contains 9 because of initiation poke pellet retrieval time removal
                for i in range(1, len(block_within_sum_pellet_retrieval_time)):
                    block_within_mean_pellet_retrieval_time.append(block_within_sum_pellet_retrieval_time[i] / 10) # all other blocks contain 10
            else:
                for i in range(0, len(block_within_sum_pellet_retrieval_time)):
                    block_within_mean_pellet_retrieval_time.append(block_within_sum_pellet_retrieval_time[i] / 10) # all blocks contain 10
                    
            if active_poke_block_within[-1] != reversal_block_length: # if there is an incomplete block NA needs to be added to the bottom of each list to make the arrays the same length for export
                block_sum_all_poke_time.append('NA')
                block_sum_a_poke_time.append('NA')
                block_sum_i_poke_time.append('NA')
                block_within_sum_all_poke_time.append('NA')
                block_within_sum_a_poke_time.append('NA')
                block_within_sum_i_poke_time.append('NA')
                block_mean_all_poke_time.append('NA')
                block_mean_a_poke_time.append('NA')
                block_mean_i_poke_time.append('NA')
                block_within_mean_all_poke_time.append('NA')
                block_within_mean_a_poke_time.append('NA')
                block_within_mean_i_poke_time.append('NA')
                block_sum_pellet_retrieval_time.append('NA')
                block_mean_pellet_retrieval_time.append('NA')
                block_within_sum_pellet_retrieval_time.append('NA')
                block_within_mean_pellet_retrieval_time.append('NA')
                        
            #####-----##### CREATE BINARY COLUMNS FOR GRAPHING #####-----#####
            
            active_poke_binary_g = []
            inactive_poke_binary_g = []
            left_poke_binary_g = []
            right_poke_binary_g = []
            left_a_poke_binary_g = []
            right_a_poke_binary_g = []
            left_i_poke_binary_g = []
            right_i_poke_binary_g = []
            
            
            #####-----##### Create time bins of desired length (input for desired length is at the top of the code)
            
            if seconds_in_bins != '':
            
                end_time = seconds_elapsed[-1]
                
                # Create end value that is the next greatest multiple of the required time bin duration greater than the end_time so that those data points are retained in the output
                
                excess = math.ceil(end_time / int(seconds_in_bins)) # Rounds up so that the number of bins is the multiple of bin length greater than the time point
                
                end_bin = excess * int(seconds_in_bins) # Creates the end value for the last time bin
                
                interval_range = pd.interval_range(start=0, freq=int(seconds_in_bins), end=end_bin, closed="left") #'left' means that it includes the first data point which is 0 because it's start time
                
                time_bins = df[seconds_elapsed] = pd.cut(seconds_elapsed, interval_range, include_lowest=True, ordered=True)
            
                # Create time bins for blocked data
                
                if blocked == 'Y':
                    time_bins = df[seconds_block] = pd.cut(seconds_block, interval_range, include_lowest=True, ordered=True)
                
                # Create dictionary of time bins as keys with bin number as matching entry to label the bins
                
                # Create a list of integers of the same length as there are time bins
                
                bin_number = []
                a = 1
                for i in range(0, len(interval_range)):
                    bin_number.append(a)
                    a += 1
                
                make_dictionary = zip(interval_range, bin_number)
                
                bin_dictionary = dict(make_dictionary)
                
                bin_num = []
                for i in range(0, len(time_bins)):
                    bin_num.append(bin_dictionary.get((time_bins[i])))
                
            ######-----##### Create session summary data ######-----#####
            
            task = session_block[0]
            
            session_duration = duration_from_start[-1] 
            
            if initiation_poke == True:
                num_rev = ((pellet_block[-1] + 1) / 10) # the +1 accounts for the removal of the initiation poke pellet
            else:
                num_rev = (pellet_block[-1] / 10)
                
            num_comprev = math.floor(num_rev) # gives number of completed reversals
            
            # If the final reversal block is incomplete: Determine how many pokes are in it
            
            if active_poke_block_within[-1] != reversal_block_length:
                excess_pokes_to_remove = total_block_within[-1]
            else:
                excess_pokes_to_remove = 0
            
            # Remove corresponding data from the poke_time and retrieval_time lists for calculating mean time for all completed reversals in session
            
            poke_time3 = poke_time2[: (len(poke_time2) - excess_pokes_to_remove)] # chronological list
            retrieval_time3 = retrieval_time2[: (len(retrieval_time2) - excess_pokes_to_remove)] # chronological list
            
            if num_rev != num_comprev or excess_pokes_to_remove != 0: # If the session was cut off in the middle of a reversal block we want to remove the data from the incomplete final block (can probably be just the last part of the condition, will double check)
                
                l_pokes = left_block[-2] # poke totals are the cumulative total at the completion of the final completed reversal block
                r_pokes = right_block[-2]
                a_pokes = active_poke_block[-2]
                i_pokes = inactive_poke_block[-2]
                
                l_active = 0
                l_inactive = 0
                r_active = 0
                r_inactive = 0
                
                for i in range(0, (len(active_block) - 1)): # Collates the pokes by side and port status and EXCLUDES the final incomplete block
                    if active_block[i] == 'Left':
                        l_active = l_active + active_poke_block_within[i]
                        r_inactive = r_inactive + inactive_poke_block_within[i]
                    elif active_block[i] == 'Right':
                        r_active = r_active + active_poke_block_within[i]
                        l_inactive = l_inactive + inactive_poke_block_within[i]
                
            else: # if the final block was complete it is kept/included in the data
            
                l_pokes = left_block[-1]
                r_pokes = right_block[-1]
                a_pokes = active_poke_block[-1]
                i_pokes = inactive_poke_block[-1]
                
                l_active = 0
                l_inactive = 0
                r_active = 0
                r_inactive = 0
                
                for i in range(0, len(active_block)): # Collates the pokes by side and port status
                    if active_block[i] == 'Left':
                        l_active = l_active + active_poke_block_within[i]
                        r_inactive = r_inactive + inactive_poke_block_within[i]
                    elif active_block[i] == 'Right':
                        r_active = r_active + active_poke_block_within[i]
                        l_inactive = l_inactive + inactive_poke_block_within[i]
                
            t_pokes = l_pokes + r_pokes
            
            l_active_prop = l_active / l_pokes * 100
            l_inactive_prop = l_inactive / l_pokes * 100
            r_active_prop = r_active / r_pokes * 100
            r_inactive_prop = r_inactive / r_pokes * 100
            
            a_l_prop = l_active / a_pokes * 100
            a_r_prop = r_active / a_pokes * 100
            
            i_l_prop = l_inactive / i_pokes * 100
            i_r_prop = r_inactive / i_pokes * 100
            
            mean_i_per_comprev = i_pokes / num_comprev
            mean_i_per_rev_l_act = r_inactive / math.ceil(num_comprev / 2)
            mean_i_per_rev_r_act = l_inactive / math.floor(num_comprev / 2)
            
            mean_poke_time = sum(poke_time3) / len(poke_time3)
            
            mean_l_poke_time = sum(l_poke_time) / len(l_poke_time)
            mean_r_poke_time = sum(r_poke_time) / len(r_poke_time)
            
            mean_a_poke_time = sum(a_poke_time) / len(a_poke_time)
            mean_a_l_poke_time = sum(a_l_poke_time) / len(a_l_poke_time)
            mean_a_r_poke_time = sum(a_r_poke_time) / len(a_r_poke_time)
            
            mean_i_poke_time = sum(i_poke_time) / len(i_poke_time)
            mean_i_l_poke_time = sum(i_l_poke_time) / len(i_l_poke_time)
            mean_i_r_poke_time = sum(i_r_poke_time) / len(i_r_poke_time)
                
            retrieval_time4 = []
            
            for i in range(0, len(retrieval_time3)): # Remove 'nan' data points from inactive poke rows
                if str(retrieval_time3[i]) != 'nan':
                    retrieval_time4.append(retrieval_time3[i])
            
            mean_retrieval_time = sum(retrieval_time4) / len(retrieval_time4)
                
            variable = ['Filename', 'Task', 'Duration', '',
                        'Number of Reversals', 'Number of complete Reversals', '',
                        'Total Pokes', 'Mean Poke Time', 'Mean Pellet Retrieval Time', '',
                        'Total Left Pokes', 'Mean Left Poke Time', 'Left Active', 'Left Active %', 'Mean Active Left Poke Time', 'Left Inactive', 'Left Inactive %', 'Mean Inactive Left Poke Time', '',
                        'Total Right Pokes', 'Mean Right Poke Time', 'Right Active', 'Right Active %', 'Mean Active Right Poke Time', 'Right Inactive', 'Right Inactive %', 'Mean Inactive Right Poke Time', '',
                        'Total Active Pokes', 'Mean Active Poke Time', 'Active Left %', 'Active Right %', '',
                        'Total Inactive Pokes', 'Mean Inactive Poke Time', 'Inactive Left proportion (%)', 'Inactive Right proportion (%)', '',
                        'Mean Inactive pokes per Reversal', 'Mean Left Inactive pokes per Reversal', 'Mean Right Inactive pokes per Reversal']
            value = [import_name.strip('.CSV'), task, session_duration, '',
                      num_rev, num_comprev, '',
                      t_pokes, mean_poke_time, mean_retrieval_time, '',
                      l_pokes, mean_l_poke_time, l_active, l_active_prop, mean_a_l_poke_time, l_inactive, l_inactive_prop, mean_i_l_poke_time, '',
                      r_pokes, mean_r_poke_time, r_active, r_active_prop, mean_a_r_poke_time, r_inactive, r_inactive_prop, mean_i_r_poke_time, '',
                      a_pokes, mean_a_poke_time, a_l_prop, a_r_prop, '',
                      i_pokes, mean_i_poke_time, i_l_prop, i_r_prop, '',
                      mean_i_per_comprev, mean_i_per_rev_r_act, mean_i_per_rev_l_act]
            
            #####-----
            
            
            #####-----
            
            # Export the data.
            
            export_results = 'Y'
            
            if export_results == 'Y':
            
                # Always export Summary and Chronological data    
                
                results_summary = {'Variable': variable, 'Value': value}
                export_file_summary = pd.DataFrame(results_summary, columns = ['Variable', 'Value'])
                
                
                results_chronological = {'Time (sec from start)': seconds_elapsed, 'Duration': duration_from_start, 'Session Type': session_type2, 'Active port': active2, 'Left Poke': left_poke2, 'Right Poke': right_poke2, 'Active Poke': active_poke, 'Inactive Poke': inactive_poke, 'Total Poke': total_poke, '% Active Pokes': percent_active, 'Poke Time': poke_time2, 
                                         'Pellet Count': pellet_count2, 'Pellet Retrieval Time': retrieval_time2,
                                         'Active': green, 'Inactive': red, 'Active Port': active_port_binary}
                export_file_chronological = pd.DataFrame(results_chronological, columns = ['Time (sec from start)', 'Duration', 'Session Type', 'Active port', 'Left Poke', 'Right Poke', 'Active Poke', 'Inactive Poke', 'Total Poke', '% Active Pokes', 'Poke Time', 
                                                                  '', 'Pellet Count', 'Pellet Retrieval Time',
                                                                  '', 'Active', 'Inactive', 'Active Port'])
                
                results_timing = {'Time (sec from start)': seconds_elapsed, 'Duration': duration_from_start, 'Active Port': active2,
                                  'Binary Active': active_poke_binary, 'Binary Inactive': inactive_poke_binary, 'Binary Left': left_poke_binary, 'Binary Right': right_poke_binary,
                                  'Left Active Binary': left_a_poke_binary, 'Left Inactive Binary': left_i_poke_binary, 'Right Active Binary': right_a_poke_binary, 'Right Inactive Binary': right_i_poke_binary,
                                  'Poke Time': poke_time2, 'Left Poke Time': l_poke_time_chron, 'Right Poke Time': r_poke_time_chron, 
                                  'Active Poke Time': a_poke_time_chron, 'Active Left Poke Time': a_l_poke_time_chron, 'Active Right Poke Time': a_r_poke_time_chron, 
                                  'Inactive Poke Time': i_poke_time_chron, 'Inactive Left Poke Time': i_l_poke_time_chron, 'Inactive Right Poke Time': i_r_poke_time_chron}
                export_file_timing = pd.DataFrame(results_timing, columns = ['Time (sec from start)', 'Duration', 'Active Port', 
                                                                             '', 'Binary Active', 'Binary Inactive', 'Binary Left', 'Binary Right',
                                                                             '', 'Left Active Binary', 'Left Inactive Binary', 'Right Active Binary', 'Right Inactive Binary',
                                                                             '', 'Poke Time', 'Left Poke Time', 'Right Poke Time',
                                                                             '', 'Active Poke Time', 'Active Left Poke Time', 'Active Right Poke Time',
                                                                             '', 'Inactive Poke Time', 'Inactive Left Poke Time', 'Inactive Right Poke Time'])
                
                # Always export blocked data
                
                if seconds_in_bins == '':       
                    
                    results_blocked_c = {'Time (sec from start)': seconds_block, 'Duration': duration_block, 'Session Type': session_block, 'Active port': active_block, 
                                         'Left Poke': left_block, 'Right Poke': right_block, 'Active Poke': active_poke_block, 'Inactive Poke': inactive_poke_block, 'Total Poke': total_block, 'Pellet Count': pellet_block, '% Active Pokes': percent_active_block,
                                         'Mean Poke Time': block_mean_all_poke_time, 'Mean Active Poke Time': block_mean_a_poke_time, 'Mean Inactive Poke Time': block_mean_i_poke_time, 'Mean Pellet Retrieval Time': block_mean_pellet_retrieval_time}
                    export_file_blocked_c = pd.DataFrame(results_blocked_c, columns = ['Time (sec from start)', 'Duration', 'Session Type', 'Active port',
                                                                                       'Left Poke', 'Right Poke', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', '% Active Pokes',
                                                                                       '', 'Mean Poke Time', 'Mean Active Poke Time', 'Mean Inactive Poke Time', 'Mean Pellet Retrieval Time'])
                    
                    results_blocked_w = {'Time (sec from start)': seconds_block, 'Duration': duration_block, 'Session Type': session_block, 'Active port': active_block, 
                                         'Left Poke': left_block_within, 'Right Poke': right_block_within, 'Active Poke': active_poke_block_within, 'Inactive Poke': inactive_poke_block_within, 'Total Poke': total_block_within, 'Pellet Count': pellet_block_within, '% Active Pokes': percent_active_block_within,
                                         'Mean Poke Time': block_within_mean_all_poke_time, 'Mean Active Poke Time': block_within_mean_a_poke_time, 'Mean Inactive Poke Time': block_within_mean_i_poke_time, 'Mean Pellet Retrieval Time': block_within_mean_pellet_retrieval_time}
                    export_file_blocked_w = pd.DataFrame(results_blocked_w, columns = ['Time (sec from start)', 'Duration', 'Session Type', 'Active port', 
                                                                                       'Left Poke', 'Right Poke', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', '% Active Pokes',
                                                                                       '', 'Mean Poke Time', 'Mean Active Poke Time', 'Mean Inactive Poke Time', 'Mean Pellet Retrieval Time'])
                    
                # If using time bins get the bin numbers - NOTE: Blocks are NOT collapsed into time bins, the only new data is which time bin each block was completed in    
                
                elif seconds_in_bins != '':
                    
                    results_blocked_c = {'Time (sec from start)': seconds_block, 'Duration': duration_block, 'Time Bin': bin_num, 'Session Type': session_block, 'Active port': active_block, 
                                         'Left Poke': left_block, 'Right Poke': right_block, 'Active Poke': active_poke_block, 'Inactive Poke': inactive_poke_block, 'Total Poke': total_block, 'Pellet Count': pellet_block, '% Active Pokes': percent_active_block}
                    export_file_blocked_c = pd.DataFrame(results_blocked_c, columns = ['Time (sec from start)', 'Duration', 'Time Bin', 'Session Type', 'Active port',
                                                                                       'Left Poke', 'Right Poke', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', '% Active Pokes'])
            
                    results_blocked_w = {'Time (sec from start)': seconds_block, 'Duration': duration_block, 'Time Bin': bin_num, 'Session Type': session_block, 'Active port': active_block, 
                                         'Left Poke': left_block_within, 'Right Poke': right_block_within, 'Active Poke': active_poke_block_within, 'Inactive Poke': inactive_poke_block_within, 'Total Poke': total_block_within, 'Pellet Count': pellet_block_within, '% Active Pokes': percent_active_block_within}
                    export_file_blocked_w = pd.DataFrame(results_blocked_w, columns = ['Time (sec from start)', 'Duration', 'Time Bin', 'Session Type', 'Active port', 
                                                                                       'Left Poke', 'Right Poke', 'Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', '% Active Pokes'])
                    
                # Export to excel
                
                from openpyxl import Workbook
                
                wb = Workbook()
                
                ws1 = wb.active
                ws1.title = 'Summary'
                
                ws2 = wb.create_sheet()
                ws2.title ='Chronological'
                
                ws3 = wb.create_sheet()
                ws3.title ='Chronological Timing'
                
                ws4 = wb.create_sheet()
                ws4.title ='Blocked Cumulative'
                
                ws4 = wb.create_sheet()
                ws4.title ='Blocked Within'
                
                sheets_to_export = wb.sheetnames
                    
                results_to_export = [export_file_summary, export_file_chronological, export_file_timing, export_file_blocked_c, export_file_blocked_w]
                
                
                with pd.ExcelWriter(export_destination) as writer:
                    
                    for i in range(len(sheets_to_export)):
                        results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
    
    
    ##########---------- Overview Sheets ----------##########
    
    export_overview = 'Y'
    
    if export_overview == 'Y':
    
        from openpyxl import Workbook
                
        wb = Workbook()
        
        ws = wb.active
        ws.title = 'Summary Overview'
        
        value_column = 'Value'
        variable_column = 'Variable'
        
        value_summary = []
        session_num = []
        counter = 1
        
        for filename in list(os.listdir(export_location)):
            
            if filename.endswith(".xlsx"):
                
                if counter == 1:
                    
                    export_name = filename
                    export_destination = export_location + export_name
                    
                    df_overview = pd.read_excel(export_destination, sheet_name = 'Summary')
                    
                    df_overview.drop(columns='Value', inplace=True)
                    
                export_name = filename
                export_destination = export_location + export_name
                
                name = 'Session ' + str(counter)
                session_num.append(name)
                
                df = pd.read_excel(export_destination, sheet_name = 'Summary')
                
                values = df[value_column].tolist()
                
                df_overview.insert(counter, name, values)
                
                counter += 1
                
        overview_name = FED_num + ' New Reversal Overview.xlsx'
        overview_destination = export_location + overview_name
        
        with pd.ExcelWriter(overview_destination) as writer:
                        
            df_overview.to_excel(writer, sheet_name='Overview', engine='openpyxl', index=False, header=True)
        
        cohort_overview_destination = cohort_export_location + overview_name
        
        with pd.ExcelWriter(cohort_overview_destination) as writer:
                        
            df_overview.to_excel(writer, sheet_name='Overview', engine='openpyxl', index=False, header=True)      