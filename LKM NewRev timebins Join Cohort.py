#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Feb 23 17:51:03 2022

@author: lauramilton
"""



import_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/NewRev/Cohort/SAL/'
export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/NewRev/Cohort/SAL/'
cohort_export_location = r'/Volumes/shared/MNHS-SOBS-Physiology/Foldi-Group/FED3 data/PSI New Rev 220211/NewRev/Cohort/'

sheet_to_join = ['30min within joined', '30min cumulative joined']

group_name = ['SAL 30min W Joined.xlsx', 'SAL 30min C joined.xlsx']

import pandas as pd
import numpy as np
import os
import openpyxl

time_column = 'Time Bin'
total_column = 'Total Poke'
active_column = 'Active Poke'
inactive_column = 'Inactive Poke'
active_percent_column = 'Percent Active'
pellets_column = 'Pellet Count'


for i in range(0, len(sheet_to_join)):
    counter = 1

    print(sheet_to_join[i])
    print(counter)
    from openpyxl import Workbook
        
    wb = Workbook()
    
    ws = wb.active
    ws.title = 'Total Pokes'
    
    ws2 = wb.create_sheet()
    ws2.title = 'Active Pokes'
    
    ws3 = wb.create_sheet()
    ws3.title = 'Inactive Pokes'
    
    ws4 = wb.create_sheet()
    ws4.title = 'Percent Active'
    
    ws5 = wb.create_sheet()
    ws5.title = 'Pellets'
    
    for filename in sorted(list(os.listdir(import_location))):
    
        if 'NewRev' in filename:
            print(filename)
            if counter == 1:
                
                import_name = filename
                import_destination = import_location + import_name
                
                df_total = pd.read_excel(import_destination, sheet_name = sheet_to_join[i])
                df_pellets = pd.read_excel(import_destination, sheet_name = sheet_to_join[i])
                df_active = pd.read_excel(import_destination, sheet_name = sheet_to_join[i])
                df_active_percent = pd.read_excel(import_destination, sheet_name = sheet_to_join[i])
                df_inactive = pd.read_excel(import_destination, sheet_name = sheet_to_join[i])
                
                cols_to_drop = ['Active Poke', 'Inactive Poke', 'Total Poke', 'Pellet Count', 'Percent Active']
                df_total.drop(columns=cols_to_drop, inplace=True)
                df_pellets.drop(columns=cols_to_drop, inplace=True)
                df_active.drop(columns=cols_to_drop, inplace=True)
                df_active_percent.drop(columns=cols_to_drop, inplace=True)
                df_inactive.drop(columns=cols_to_drop, inplace=True)
                
            import_name = filename
            import_destination = import_location + import_name
            
            name = filename.strip(' NewRev Joined.xlsx')
            
            df = pd.read_excel(import_destination, sheet_name = sheet_to_join[i])
            
            time = df[time_column].tolist()
            total = df[total_column].tolist()
            pellets = df[pellets_column].tolist()
            active = df[active_column].tolist()
            active_percent = df[active_percent_column].tolist()
            inactive = df[inactive_column].tolist()
    
            df_total.insert((counter), name, total)
            df_pellets.insert((counter), name, pellets)
            df_active.insert((counter), name, active)
            df_active_percent.insert((counter), name, active_percent)
            df_inactive.insert((counter), name, inactive)
                
            counter += 1
        
    group_destination = export_location + group_name[i]
    cohort_destination = cohort_export_location + group_name[i]
    
    results_to_export = [df_total, df_active, df_inactive, df_active_percent, df_pellets]
    
    sheets_to_export = wb.sheetnames
    
    with pd.ExcelWriter(group_destination) as writer:
        
        for i in range(len(sheets_to_export)):
                        results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
    
    with pd.ExcelWriter(cohort_destination) as writer:
        
        for i in range(len(sheets_to_export)):
                        results_to_export[i].to_excel(writer, sheet_name=str(sheets_to_export[i]), engine='openpyxl', index=False, header=True)
    