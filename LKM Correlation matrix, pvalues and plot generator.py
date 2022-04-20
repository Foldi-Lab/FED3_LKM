#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Aug 13 12:20:42 2021

@author: lauramilton
"""

import pandas as pd
from matplotlib import pyplot as plt
import seaborn as sns
import numpy as np

import_file = r'/Users/lauramilton/Desktop/RO 2021/Behavioural C1+2 correlations cleaned new cyto.xlsx'
export_file = r'/Users/lauramilton/Desktop/RO 2021/Beh C1+2 w cyto correlation matrix cleaned.xlsx'
p_export_file = r'/Users/lauramilton/Desktop/RO 2021/Beh C1+2 w cyto pvalue matrix cleaned2.xlsx'

from scipy.stats import pearsonr

def calculate_pvalues(df):
    df = df.dropna()._get_numeric_data()
    dfcols = pd.DataFrame(columns=df.columns)
    pvalues = dfcols.transpose().join(dfcols, how='outer')
    for r in df.columns:
        for c in df.columns:
            pvalues[r][c] = round(pearsonr(df[r], df[c])[1], 4)
    return pvalues

sig = '#000000'
ns1 = '#9e0142'
ns2 = '#d53e4f'
ns3 = '#f46d43'
ns4 = '#fdae61'
ns5 = '#fee08b'
ns6 = '#e6f598'
ns7 = '#abdda4'
ns8 = '#66c2a5'
ns9 = '#3288bd'
ns10 = '#5e4fa2'

palette_rainbow = [sig, ns1, ns2, ns2, ns3, ns3, ns4, ns4, ns5, ns5, ns6, ns6, ns7, ns7, ns8, ns8, ns9, ns9, ns10, ns10]
sns.set_palette(palette_rainbow)


sig = '#d9ef8b'
trend = '#66bd63'
ns = '#F1F1F1'
palette_sig = [sig, trend, ns, ns, ns, ns, ns, ns, ns, ns, ns, ns, ns, ns, ns, ns, ns, ns, ns, ns]

### Generate correlation matrix file

# first create a new excel document with a copy of the raw data

import xlrd
xls = xlrd.open_workbook(import_file, on_demand=True)
sheetnames = xls.sheet_names()

# data = pd.read_excel(import_file, sheet_name='All data')

# # with pd.ExcelWriter(export_file, engine='xlsxwriter') as writer:
# #     data.to_excel(writer, sheet_name='All data', index=False, header=True)
    
# with pd.ExcelWriter(p_export_file, engine='xlsxwriter') as writer:
#     data.to_excel(writer, sheet_name='All data', index=False, header=True)

# open the document you just created so you can save new sheets into it

from openpyxl import load_workbook

book = load_workbook(export_file)
writer = pd.ExcelWriter(export_file, engine = 'openpyxl')
writer.book = book

# # create a correlation matrix of all data for each of the groups (each group has it's own sheet in the original file)

# for sheet in sheetnames:
#     if 'corr' in sheet:

#         data = pd.read_excel(import_file, sheet_name=sheet)
        
#         corrMatrix = data.corr()
                
#         corrMatrix.to_excel(writer, sheet_name=sheet, engine='openpyxl', index=False, header=True)
        
#         writer.save()


for sheet in sheetnames:
    if 'All cyto' in sheet:

        data = pd.read_excel(import_file, sheet_name=sheet)
        
        corrMatrix = data.corr()
                
        corrMatrix.to_excel(writer, sheet_name=sheet, engine='openpyxl', index=False, header=True)
        
        writer.save()

writer.close()

####-----##### Generate pvalue matrix

import xlrd
xls = xlrd.open_workbook(export_file, on_demand=True)
sheetnames = xls.sheet_names()

from openpyxl import load_workbook

book = load_workbook(p_export_file)
writer = pd.ExcelWriter(p_export_file, engine = 'openpyxl')
writer.book = book

# for sheet in sheetnames:
#     if 'corr' in sheet:

#         data = pd.read_excel(import_file, sheet_name=sheet)

#         pvalues = calculate_pvalues(data)

#         sheetname_new = sheet.strip('corr') + 'pvalues'
        
#         pvalues.to_excel(writer, sheet_name=sheetname_new, engine='openpyxl', index=False, header=True)
        
#         writer.save()
        

for sheet in sheetnames:
    if 'All cyto' in sheet:

        data = pd.read_excel(import_file, sheet_name=sheet)

        pvalues = calculate_pvalues(data)

        sheetname_new = sheet + ' pvalues'
        
        pvalues.to_excel(writer, sheet_name=sheetname_new, engine='openpyxl', index=False, header=True)
        
        writer.save()
        
writer.close()

#####-----#####

# column_labels = pd.read_excel(export_file, sheet_name='All data') # get the column labels

# variable = column_labels.columns.tolist()

# del variable[:2] # remove RAT# and EXP ID column labels as they're not in the correlation data sheets

# ###--- Generate figures ---###

# # Big figure into which each subplot will be placed

# f,(ax1,ax2, axcb) = plt.subplots(1,3, 
#             gridspec_kw={'width_ratios':[1,1,0.1]},
#             figsize=(8, 8))

# x_labels = ['HAB mean daily RWA', 'ABA mean daily RWA', 'Mean daily 90 min FI', 'Lowest BW%', 'Mean daily BW% change']

# # first subplot

# corr_data = pd.read_excel(p_export_file, sheet_name='ABA_S pvalues')
        
# col_to_drop = []

# for i in range(0, len(variable)):
#     if variable[i].startswith('ABA'):
#         continue

#     else:
#         col_to_drop.append(variable[i])

# corr_data.drop(columns=col_to_drop, inplace=True)
# corr_data.drop(index=[19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35], inplace=True)

# del variable[19:]

# g = sns.heatmap(data=corr_data, vmin=0, vmax=1, center=0.5, linewidths=1, cmap=sns.color_palette(palette_sig, 20), ax=ax1, cbar=False, square=False, yticklabels=variable, xticklabels=x_labels, annot=True, fmt='.4f', annot_kws={"fontsize":8})
# # sns.heatmap(data=corr_data, vmin=-1, vmax=1, center=0, linewidths=1, cmap='coolwarm', cbar=False, square=False, ax=ax1, yticklabels=variable, xticklabels=True, annot=True, fmt='.2f', annot_kws={"fontsize":8})
# g.set_xticklabels(g.get_xticklabels(), rotation=45, horizontalalignment='right')
# # second subplot

# variable = column_labels.columns.tolist()

# del variable[:2] # remove RAT# and EXP ID column labels as they're not in the correlation data sheets

# corr_data = pd.read_excel(p_export_file, sheet_name='ABA_R pvalues')
        
# col_to_drop = []

# for i in range(0, len(variable)):
#     if variable[i].startswith('ABA'):
#         continue

#     else:
#         col_to_drop.append(variable[i])

# corr_data.drop(columns=col_to_drop, inplace=True)
# corr_data.drop(index=[19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35], inplace=True)

# del variable[19:]

# g = sns.heatmap(data=corr_data, vmin=0, vmax=1, center=0.5, linewidths=1, cmap=sns.color_palette(palette_sig, 20), ax=ax2, cbar=True, cbar_ax=axcb, square=False, yticklabels=False, xticklabels=True, annot=True, fmt='.4f', annot_kws={"fontsize":8})
# # sns.heatmap(data=corr_data, vmin=-1, vmax=1, center=0, linewidths=1, cmap='coolwarm', cbar=True, cbar_ax=axcb, square=False, ax=ax2, yticklabels=False, xticklabels=True, annot=True, fmt='.2f', annot_kws={"fontsize":8})
# g.set_xticklabels(g.get_xticklabels(), rotation=45, horizontalalignment='right')

# # plt.savefig(r'/Users/lauramilton/Desktop/RO 2021/Beh C1+2 ABA S + R pvalues.png', bbox_inches = 'tight')

# plt.show()

#####----- difference heatmap

# fig, ax = plt.subplots(figsize=(5, 8))

# corr_data = pd.read_excel(export_file, sheet_name='ABA diff')
        
# col_to_drop = []

# for i in range(0, len(variable)):
#     if variable[i].startswith('ABA'):
#         continue

#     else:
#         col_to_drop.append(variable[i])

# corr_data.drop(columns=col_to_drop, inplace=True)
# corr_data.drop(index=[19, 20, 21, 22, 23], inplace=True)

# del variable[19:]

# palette = ['#543005','#8c510a','#bf812d','#dfc27d','#f6e8c3','#c7eae5','#80cdc1','#35978f','#01665e','#003c30']

# sns.heatmap(data=corr_data, vmin=-2, vmax=2, center=0, linewidths=1, cmap=sns.color_palette(palette, 10), cbar=True, square=False, yticklabels=variable, xticklabels=True, annot=True, fmt='.2f', annot_kws={"fontsize":8})

# plt.savefig(r'/Users/lauramilton/Desktop/RO 2021/Beh C1+2 ABA difference.png', bbox_inches = 'tight')
# plt.show()

# #####-----

# # cytokine correlation PRE measures with behavioural test data for ALL rats

# f,(ax1, ax2) = plt.subplots(1, 2, 
#             gridspec_kw={'width_ratios':[1, 1]},
#             figsize=(7, 9))

# column_labels = pd.read_excel(export_file, sheet_name='All data') # get the column labels

# variable = column_labels.columns.tolist()

# del variable[:2] # remove RAT# and EXP ID column labels as they're not in the correlation data sheets

# corr_data = pd.read_excel(export_file, sheet_name='All corr')
        
# col_to_drop = []

# for i in range(0, len(variable)):
#     if variable[i].startswith('RANTES'):
#         continue
#     elif variable[i].startswith('IL'):
#         continue

#     else:
#         col_to_drop.append(variable[i])

# for i in range(0, len(variable)):
#     if 'pre' in variable[i]:
#         continue
#     else:
#         col_to_drop.append(variable[i])
        
# del variable[19:]

# corr_data.drop(columns=col_to_drop, inplace=True)
# corr_data.drop(index=[19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35], inplace=True)

# sns.heatmap(data=corr_data, vmin=-1, vmax=1, center=0, linewidths=1, cmap='coolwarm', ax=ax1, cbar=True, square=False, yticklabels=variable, xticklabels=True, annot=True, fmt='.2f', annot_kws={"fontsize":8})

# # create matching pvalue plot

# variable = column_labels.columns.tolist()

# del variable[:2] # remove RAT# and EXP ID column labels as they're not in the correlation data sheets

# corr_data = pd.read_excel(p_export_file, sheet_name='All pvalues')
        
# col_to_drop = []

# for i in range(0, len(variable)):
#     if variable[i].startswith('RANTES'):
#         continue
#     elif variable[i].startswith('IL'):
#         continue

#     else:
#         col_to_drop.append(variable[i])

# for i in range(0, len(variable)):
#     if 'pre' in variable[i]:
#         continue
#     else:
#         col_to_drop.append(variable[i])
        
# del variable[19:]

# corr_data.drop(columns=col_to_drop, inplace=True)
# corr_data.drop(index=[19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35], inplace=True)

# sns.heatmap(data=corr_data, vmin=0, vmax=1, center=0.5, linewidths=1, cmap=sns.color_palette(palette_sig, 20), ax=ax2, cbar=True, square=False, yticklabels=False, xticklabels=True, annot=True, fmt='.4f', annot_kws={"fontsize":8})

# plt.show()






















#####-----##### using for loop

# for sheet in sheetnames:
    
#     if 'ABA' in sheet:
        
#         corr_data = pd.read_excel(export_file, sheet_name=sheet)
        
#         col_to_drop = []
        
#         for i in range(0, len(variable)):
#             if variable[i].startswith('ABA'):
#                 continue

#             else:
#                 col_to_drop.append(variable[i])
        
#         corr_data.drop(columns=col_to_drop, inplace=True)
#         corr_data.drop(index=[19, 20, 21, 22, 23], inplace=True)
        
#         del variable[19:]
        
#         # print(corr_data)
#         plt.subplot(1, groups, c)
#         plt.title(sheet)
        
#         x_tick_labels = []
        
#         for i in range(0, len(x_labels)):
#             x_tick_labels.append(x_labels[i] + ' ' + sheet.strip('corr'))
        
#         if c == 1:
#             sns.heatmap(data=corr_data, vmin=-1, vmax=1, center=0, linewidths=1, cmap='coolwarm', cbar=False, square=False, yticklabels=variable, xticklabels=x_tick_labels, annot=True, fmt='.2f', annot_kws={"fontsize":8})
        
#         elif c < groups:
#             sns.heatmap(data=corr_data, vmin=-1, vmax=1, center=0, linewidths=1, cmap='coolwarm', cbar=False, square=False, yticklabels=False, xticklabels=x_tick_labels, annot=True, fmt='.2f', annot_kws={"fontsize":8})
        
#         elif c == groups:
#             sns.heatmap(data=corr_data, vmin=-1, vmax=1, center=0, linewidths=1, cmap='coolwarm', cbar=True, cbar_ax=axcb, square=False, yticklabels=False, xticklabels=x_tick_labels, annot=True, fmt='.2f', annot_kws={"fontsize":8})
        
#         c +=1
        
#         # plt.savefig(r'/Users/lauramilton/Desktop/RO 2021/Beh C1+2 ' + sheet + ' all vs ABA tall.png', bbox_inches = 'tight')
        
#         # plt. clf()
        
#         # corr_data_t = corr_data.transpose()
        
#         # print(corr_data_t)
        
#         # plt.figure()
#         # sns.heatmap(data=corr_data_t, vmin=-1, vmax=1, center=0, linewidths=1, cmap='coolwarm', cbar=True, square=True, xticklabels=variable)
        
#         # plt.savefig(r'/Users/lauramilton/Desktop/RO 2021/Beh C1+2 ' + sheet + ' all vs ABA wide.png', bbox_inches = 'tight')
        
#         # plt. clf()

# plt.show()

